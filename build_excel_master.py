#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_excel_master.py v2 — エンジニア向け DB Master を「OLD形式を保持して修正済みデータで再生成」。

方針(2026-06-18, Tatsuki確定):
  OLD `TS24 DB Master Back UP.xlsx` をテンプレートとして読み込み、各シートの
  ヘッダー/書式/列幅/役割を保持したまま、データ行のみ修正済み ts24_unified.db で差し替える。
  → 現場で「不満→過去事例→最適解」を引ける分析ワークブックを、正しい数値で復活させる。

シート扱い:
  [再生成(flat)] RUN_LOG, SESSION_SUMMARY, TYRE_LOG, LAP_TIMES, PERFORMANCE_CORRELATION,
                 DYNAMICS_ANALYSIS, LAP_SUSPENSION, PROBLEM_LIBRARY, BEST_WORST_ANALYSIS
  [保持(static/手入力/別ソース)] SOLUTION_SEARCH(手順), CHASSIS_GEO(手入力), DB_LOG, TREND_ANALYSIS
     ※ DB_LOG/TREND_ANALYSIS は区分レイアウトが複雑なため今回は旧内容を保持(別途再生成を検討)。
"""
import copy
import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

DB = Path(__file__).parent.parent / "02_DATABASE"
TEMPLATE = DB / "TS24 DB Master Back UP.xlsx"
OUT = DB / "TS24 DB Master.xlsx"
con = sqlite3.connect(DB / "ts24_unified.db"); con.row_factory = sqlite3.Row


def fmt_lap(s):
    if s is None: return None
    try: s = float(s)
    except (TypeError, ValueError): return None
    m = int(s // 60); return f"{m}:{s-60*m:06.3f}"


def q(sql, p=()): return con.execute(sql, p).fetchall()


_HFILL = PatternFill("solid", fgColor="1F3864")   # 列ヘッダ濃紺
_GFILL = PatternFill("solid", fgColor="2E5496")   # グループ見出し
_HFONT = Font(color="FFFFFF", bold=True)
_CEN = Alignment(horizontal="center", vertical="center")
_TITLE_FILL = PatternFill("solid", fgColor="0F243E")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAD_FILL = PatternFill("solid", fgColor="F4CCCC")


def reset_sheet(wb, name, index=None):
    """新規ビュー用シートを作り直す。既存rawシートはこの関数で触らない。"""
    if name in wb.sheetnames:
        old_idx = wb.sheetnames.index(name)
        wb.remove(wb[name])
        index = old_idx if index is None else index
    if index is None:
        return wb.create_sheet(name)
    return wb.create_sheet(name, index)


def apply_table_style(ws, header_row, max_col, widths=None, freeze=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = _HFILL
        cell.font = _HFONT
        cell.alignment = _CEN
    if freeze:
        ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"
    for c in range(1, max_col + 1):
        width = widths.get(c) if widths else None
        if width is None:
            width = 14
        ws.column_dimensions[get_column_letter(c)].width = width


def write_note(ws, row, text, span):
    ws.cell(row, 1, text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row, 1)
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True, color="1F3864")
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def top_items(items, limit=3):
    if not items:
        return ""
    return ", ".join(f"{k}({v})" for k, v in Counter(items).most_common(limit))


def compact_text(value, limit=140):
    if not value:
        return ""
    text = " ".join(str(value).split())
    return text if len(text) <= limit else text[:limit - 1] + "…"

def build_clean_sheet(wb, name, groups, headers, rows):
    """空列を持たないクリーンなシートに作り直す(元位置を保持)。
    groups=[(label,span),...] 1行目グループ見出し / headers=2行目列名 / rows=3行目以降データ。"""
    idx = wb.sheetnames.index(name)
    wb.remove(wb[name]); ws = wb.create_sheet(name, idx)
    c = 1
    for label, span in groups:
        ws.cell(1, c, label)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + span - 1)
        cell = ws.cell(1, c); cell.fill = _GFILL; cell.font = _HFONT; cell.alignment = _CEN
        c += span
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h); cell.fill = _HFILL; cell.font = _HFONT; cell.alignment = _CEN
    for r, row in enumerate(rows, 3):
        for i, v in enumerate(row, 1):
            ws.cell(r, i, v)
    ws.freeze_panes = "A3"
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, min(len(str(h)) + 3, 18))
    return len(rows)


def repopulate(ws, data_start, rows, ncols):
    """data_start 行以降を rows(list of list)で差し替え。列書式は元 data_start 行から踏襲。"""
    # 0) データ域(>=data_start)に掛かる結合セルを解除(区分見出し等。MergedCell書込不可対策)
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= data_start:
            ws.unmerge_cells(str(rng))
    # 1) 元 data_start 行の各列スタイルを退避
    styles = {}
    for c in range(1, ncols + 1):
        src = ws.cell(data_start, c)
        styles[c] = (copy.copy(src.font), copy.copy(src.fill), copy.copy(src.border),
                     copy.copy(src.alignment), src.number_format)
    # 2) 既存データ行を全消去(値のみ)
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(r, c).value = None
    # 3) 新データ書込 + スタイル適用
    for i, row in enumerate(rows):
        r = data_start + i
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.value = row[c - 1] if c - 1 < len(row) else None
            fnt, fil, brd, aln, nf = styles[c]
            cell.font = copy.copy(fnt); cell.fill = copy.copy(fil)
            cell.border = copy.copy(brd); cell.alignment = copy.copy(aln)
            cell.number_format = nf
    return len(rows)


wb = openpyxl.load_workbook(TEMPLATE)  # 書式込みで読込(バックアップ自体は不変、保存先は OUT)
report = []

# 不要シート削除 (Tatsuki確定 2026-06-18): BEST_WORST_ANALYSIS / SESSION_SUMMARY / TYRE_LOG / CHASSIS_GEO
for _sh in ["BEST_WORST_ANALYSIS", "SESSION_SUMMARY", "TYRE_LOG", "CHASSIS_GEO"]:
    if _sh in wb.sheetnames:
        wb.remove(wb[_sh]); report.append((_sh + " (削除)", "—"))

# ── RUN_LOG (row4+, 43col): runs 全setup + comment + 意思決定列 ──
RUN_FIELDS = ["run_id","rider","circuit","session","run_no","weather","track_temp","air_temp",
  "fork_type","f_set_c","f_set_r","f_tos_spring","f_tos_length","f_spr_l","f_spr_r","f_preload",
  "f_oil_level","f_comp","f_reb","f_offset","f_offset2","f_hgt_top","f_hgt_bot","shock_type",
  "r_set_c","r_set_r","r_spr","r_preload","r_comp","r_reb","r_tos_spring","r_tos_length","shock_len",
  "link","ride_hgt","swing_arm","tyre_front","tyre_rear","comment"]
# 意思決定列(sparse): problem_log.description / setup_decision_log.rationale/expected_effect/result_eval
pdesc = {r["run_id"]: r["description"] for r in q("SELECT run_id,description FROM problem_log WHERE run_id IS NOT NULL")}
dec = {r["run_id_to"]: r for r in q("SELECT run_id_to,rationale,expected_effect,result_eval FROM setup_decision_log WHERE run_id_to IS NOT NULL")}
rows = []
for r in q(f"SELECT {','.join(RUN_FIELDS)} FROM runs ORDER BY (date IS NULL), date,rider,session,run_no"):
    base = [r[f] for f in RUN_FIELDS]
    rid = r["run_id"]; d = dec.get(rid)
    base += [pdesc.get(rid), d["rationale"] if d else None,
             d["expected_effect"] if d else None, d["result_eval"] if d else None]
    rows.append(base)
report.append(("RUN_LOG", repopulate(wb["RUN_LOG"], 4, rows, 43)))

# ── LAP_TIMES (row3+, 17col) ──
rows = []
for r in q("""SELECT l.run_id,l.lap_id,r.round,r.circuit,r.date,r.session,r.rider,r.run_no,
   l.lap_no,l.lap_time_s,l.is_outlap,r.weather,r.air_temp,r.track_temp,r.tyre_front,r.tyre_rear
   FROM laps l JOIN runs r USING(run_id) ORDER BY r.date,r.rider,r.session,r.run_no,l.lap_no"""):
    rows.append([r["run_id"],r["lap_id"],r["round"],r["circuit"],r["date"],r["session"],r["rider"],
        r["run_no"],r["lap_no"], fmt_lap(r["lap_time_s"]), r["lap_time_s"],
        ("Y" if r["is_outlap"] else ""), r["weather"],r["air_temp"],r["track_temp"],
        r["tyre_front"],r["tyre_rear"]])
report.append(("LAP_TIMES", repopulate(wb["LAP_TIMES"], 3, rows, 17)))

# ── per-run lap_suspension 集計 (非outlap) ── PERFORMANCE_CORRELATION / DYNAMICS_ANALYSIS 共用
agg = {r["run_id"]: r for r in q("""
  SELECT ls.run_id, count(*) n,
    round(avg(ls.apex_susF_avg),2) apex_susf, round(avg(ls.apex_susR_avg),2) apex_susr,
    round(avg(ls.wf_f_apex_n),1) wf_f, round(avg(ls.wf_r_apex_n),1) wf_r,
    round(avg(ls.apex_spd_avg),1) apex_spd, round(avg(ls.apex_count),1) apex_cnt,
    round(avg(ls.brk_susF_avg),2) brk_susf, round(avg(ls.brk_susR_avg),2) brk_susr,
    round(avg(ls.brk_spd_avg),1) brk_spd, round(avg(ls.brk_count),1) brk_cnt,
    round(avg(ls.fullbrk_susF),2) fb_susf, round(avg(ls.fullbrk_susR),2) fb_susr, round(avg(ls.fullbrk_count),1) fb_cnt,
    round(avg(ls.ce_susF_avg),2) ce_susf, round(avg(ls.ce_susR_avg),2) ce_susr,
    round(avg(ls.ce_spd_avg),1) ce_spd, round(avg(ls.ce_count),1) ce_cnt,
    round(avg(ls.f_dive_spd),0) f_dive, round(avg(ls.f_reb_spd),0) f_reb,
    round(avg(ls.r_dive_spd),0) r_dive, round(avg(ls.r_reb_spd),0) r_reb,
    round(avg(ls.rear_light_brk),1) rear_light,
    round(avg(ls.brk_f_dive_spd_avg),0) brk_f_dive_avg, round(avg(ls.brk_f_dive_spd_peak),0) brk_f_dive_peak,
    round(avg(ls.ce_r_spd_avg),0) ce_r_avg, round(avg(ls.ce_r_spd_peak),0) ce_r_peak,
    round(avg(ls.ph12_rear0_s),2) ph12_rear0
  FROM lap_suspension ls JOIN laps l ON ls.lap_id=l.lap_id
  WHERE l.is_outlap=0 GROUP BY ls.run_id""")}

# ── PERFORMANCE_CORRELATION (row4+, 22col) + Rank/Gap/Tier ──
perf = {r["run_id"]: r for r in q("SELECT * FROM performance")}
runmeta = {r["run_id"]: r for r in q("SELECT run_id,rider,circuit,round,session,run_no,date FROM runs")}
lap_run_ids = {r["run_id"] for r in q("SELECT DISTINCT run_id FROM laps")}  # 実走Runのみ(setup-only除外)
# Rank/Gap/Tier: (round,circuit,rider) 内 best_lap_s 昇順
groups = defaultdict(list)
for rid, p in perf.items():
    if p["best_lap_s"] is not None:
        m = runmeta.get(rid)
        if m: groups[(m["round"], m["circuit"], m["rider"])].append((rid, p["best_lap_s"]))
rank_info = {}
for key, lst in groups.items():
    lst.sort(key=lambda x: x[1]); n = len(lst); best = lst[0][1]
    for i, (rid, bl) in enumerate(lst):
        tier = "FAST" if i < (n + 1) // 2 else "SLOW"
        rank_info[rid] = (f"{i+1}/{n}", round(bl - best, 3), tier)
rows = []
for rid, m in sorted(runmeta.items(), key=lambda kv: (kv[1]["date"] or "", kv[1]["rider"], kv[1]["session"])):
    if rid not in lap_run_ids: continue   # 実走(2D)のあるRunのみ
    p = perf.get(rid); a = agg.get(rid); ri = rank_info.get(rid, ("", "", ""))
    if not p: continue
    rows.append([rid, None, m["rider"], m["circuit"], m["date"], m["session"], m["run_no"],
        fmt_lap(p["best_lap_s"]), fmt_lap(p["run_avg_lap_s"]), p["n_laps"],
        a["apex_susf"] if a else None, a["apex_susr"] if a else None,
        a["wf_f"] if a else None, a["wf_r"] if a else None, a["apex_spd"] if a else None,
        None,  # APEX ax: 新スキーマ未保持
        a["brk_susf"] if a else None, a["brk_susr"] if a else None, a["brk_spd"] if a else None,
        ri[0], ri[1], ri[2]])
report.append(("PERFORMANCE_CORRELATION", repopulate(wb["PERFORMANCE_CORRELATION"], 4, rows, 22)))
# Fast/Slow を視覚的に強調: Tier列(22)を緑/赤で塗り、RUN_ID(1)・Tierを太字に
_FAST_F = PatternFill("solid", fgColor="C6EFCE"); _FAST_T = Font(color="006100", bold=True)
_SLOW_F = PatternFill("solid", fgColor="FFC7CE"); _SLOW_T = Font(color="9C0006", bold=True)
_pw = wb["PERFORMANCE_CORRELATION"]
for r in range(4, _pw.max_row + 1):
    tier = _pw.cell(r, 22).value
    if tier == "FAST": fill, fnt = _FAST_F, _FAST_T
    elif tier == "SLOW": fill, fnt = _SLOW_F, _SLOW_T
    else: continue
    for c in (1, 20, 21, 22):   # RUN_ID, Rank, Gap, Tier
        _pw.cell(r, c).fill = fill; _pw.cell(r, c).font = copy.copy(fnt)
    _pw.cell(r, 22).alignment = Alignment(horizontal="center")

# ── DYNAMICS_ANALYSIS: 新3エリア定義で埋まる列のみ(空列 SR/ax/Pit×4/Tyre×8 を廃止)・グループ見出し付き ──
DYN_GROUPS = [("INFO", 8), ("APEX (MID-CORNER)", 4), ("BRAKING ENTRY", 4), ("FULL BRAKING", 3),
              ("CORNER EXIT", 4), ("DAMPING speed mm/s + Rear-light%", 10)]
DYN_HEADERS = ["RUN_ID","Round","Date","Circuit","Session","Rider","Run","Laps",
    "Count","Spd(km/h)","SusF(mm)","SusR(mm)",
    "Count","Spd(km/h)","SusF(mm)","SusR(mm)",
    "Count","SusF(mm)","SusR(mm)",
    "Count","Spd(km/h)","SusF(mm)","SusR(mm)",
    "F-Dive","F-Reb","R-Dive","R-Reb","RearLight%",
    "Brk F-Dive Avg","Brk F-Dive Peak","CE R-Spd Avg","CE R-Spd Peak","PH1-2 Rear@0[s]"]
rows = []
for rid, m in sorted(runmeta.items(), key=lambda kv: (kv[1]["date"] or "", kv[1]["rider"], kv[1]["session"])):
    if rid not in lap_run_ids: continue   # 実走(2D)のあるRunのみ
    a = agg.get(rid); p = perf.get(rid)
    rows.append([rid, m["round"], m["date"], m["circuit"], m["session"], m["rider"], m["run_no"],
        p["n_laps"] if p else None,
        a["apex_cnt"] if a else None, a["apex_spd"] if a else None, a["apex_susf"] if a else None, a["apex_susr"] if a else None,
        a["brk_cnt"] if a else None, a["brk_spd"] if a else None, a["brk_susf"] if a else None, a["brk_susr"] if a else None,
        a["fb_cnt"] if a else None, a["fb_susf"] if a else None, a["fb_susr"] if a else None,
        a["ce_cnt"] if a else None, a["ce_spd"] if a else None, a["ce_susf"] if a else None, a["ce_susr"] if a else None,
        a["f_dive"] if a else None, a["f_reb"] if a else None, a["r_dive"] if a else None, a["r_reb"] if a else None,
        a["rear_light"] if a else None,
        a["brk_f_dive_avg"] if a else None, a["brk_f_dive_peak"] if a else None,
        a["ce_r_avg"] if a else None, a["ce_r_peak"] if a else None,
        a["ph12_rear0"] if a else None])
report.append(("DYNAMICS_ANALYSIS", build_clean_sheet(wb, "DYNAMICS_ANALYSIS", DYN_GROUPS, DYN_HEADERS, rows)))

# ── PROBLEM_LIBRARY (row4+, 7col): problem_library テーブル直 ──
rows = [[r["id"],r["phase_code"],r["phase_name"],r["fase_it"],r["complaint_it"],r["complaint_en"],r["tags"]]
        for r in q("SELECT id,phase_code,phase_name,fase_it,complaint_it,complaint_en,tags FROM problem_library ORDER BY id")]
report.append(("PROBLEM_LIBRARY", repopulate(wb["PROBLEM_LIBRARY"], 4, rows, 7)))

# ── LAP_SUSPENSION (per-lap 全件): 旧シートは雛形のみ。ヘッダ+全データを書く ──
ws = wb["LAP_SUSPENSION"]
LS_COLS = ["lap_id","run_id","round","circuit","session","rider","run_no","lap_no","date","lap_time_s","lap_time_fmt",
  "apex_count","apex_spd_avg","apex_susF_avg","apex_susR_avg","wf_f_apex_n","wf_r_apex_n",
  "brk_count","brk_spd_avg","brk_susF_avg","brk_susR_avg","wf_f_brk_n","wf_r_brk_n",
  "fullbrk_count","fullbrk_susF","fullbrk_susR",
  "ce_count","ce_spd_avg","ce_susF_avg","ce_susR_avg","wf_f_ce_n","wf_r_ce_n",
  "f_dive_spd","f_reb_spd","r_dive_spd","r_reb_spd","rear_light_brk",
  "brk_f_dive_spd_avg","brk_f_dive_spd_peak","ce_r_spd_avg","ce_r_spd_peak","ph12_rear0_s",
  "brk_f_reb_spd_avg","brk_f_reb_spd_peak","brk_r_dive_spd_avg","brk_r_dive_spd_peak","brk_r_reb_spd_avg","brk_r_reb_spd_peak",
  "apex_f_dive_spd_avg","apex_f_dive_spd_peak","apex_f_reb_spd_avg","apex_f_reb_spd_peak","apex_r_dive_spd_avg","apex_r_dive_spd_peak","apex_r_reb_spd_avg","apex_r_reb_spd_peak",
  "ce_f_dive_spd_avg","ce_f_dive_spd_peak","ce_f_reb_spd_avg","ce_f_reb_spd_peak","ce_r_dive_spd_avg","ce_r_dive_spd_peak","ce_r_reb_spd_avg","ce_r_reb_spd_peak",
  "lap_susF_mean","lap_susF_min","lap_susF_max","lap_susR_mean"]
# 旧の title(row1)を保持、row2=ヘッダ、row3+=データ。全LS_COLS分のヘッダを書く
# (旧実装は ws.max_column=34 で頭打ち→ r_dive_spd 以降のヘッダが欠落していた)。
# スタイルは既存ヘッダ cell(2,1) から複製して書式を維持。
from copy import copy as _copyst
_h0 = ws.cell(2, 1)
for c in range(1, len(LS_COLS) + 1):
    cell = ws.cell(2, c)
    cell.value = LS_COLS[c-1]
    if c > 1:
        cell.font = _copyst(_h0.font); cell.fill = _copyst(_h0.fill)
        cell.alignment = _copyst(_h0.alignment); cell.border = _copyst(_h0.border)
ls_rows = [[r[c] for c in LS_COLS] for r in q(f"SELECT {','.join(LS_COLS)} FROM lap_suspension ORDER BY run_id,lap_no")]
report.append(("LAP_SUSPENSION", repopulate(ws, 3, ls_rows, len(LS_COLS))))

# ── WEEKEND_SUMMARY_HELPER: Report Weekend Summary 用の上位ビュー ──
event_runs = defaultdict(list)
for r in q("""SELECT r.*, p.best_lap_s, p.run_avg_lap_s, p.session_position, p.n_laps AS perf_laps
              FROM runs r LEFT JOIN performance p USING(run_id)
              ORDER BY (r.date IS NULL), r.date, r.rider, r.session, r.run_no"""):
    event_runs[(r["round"], r["circuit"], r["rider"])].append(r)

tags_by_run = defaultdict(list)
for r in q("SELECT run_id, tag FROM run_tags ORDER BY tag"):
    tags_by_run[r["run_id"]].append(r["tag"])

dec_by_to = defaultdict(list)
for r in q("""SELECT run_id_to, component, from_value, to_value, expected_effect, actual_effect, result_eval
              FROM setup_decision_log
              WHERE run_id_to IS NOT NULL
              ORDER BY created_at, decision_id"""):
    dec_by_to[r["run_id_to"]].append(r)

event_metric = {r["run_id"]: r for r in q("""
  SELECT ls.run_id,
    round(avg(ls.brk_f_dive_spd_peak),0) brk_f_dive_peak,
    round(avg(ls.ce_r_spd_avg),0) ce_r_spd_avg,
    round(avg(ls.ph12_rear0_s),2) ph12_rear0_s,
    round(avg(ls.rear_light_brk),1) rear_light_brk
  FROM lap_suspension ls JOIN laps l ON ls.lap_id=l.lap_id
  WHERE l.is_outlap=0
  GROUP BY ls.run_id""")}

ws = reset_sheet(wb, "WEEKEND_SUMMARY_HELPER", 1 if "TREND_ANALYSIS" in wb.sheetnames else None)
ws["A1"] = "WEEKEND SUMMARY HELPER"
ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
ws["A1"].fill = _TITLE_FILL
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
write_note(ws, 2, (
    "How to use this sheet: 1) Filter Round/Circuit/Rider for the target weekend. "
    "2) Review Best Lap, Race Pos, Top Problem Tags, Representative Comment, and Key Suspension Signals. "
    "3) Use Setup Actions and Similar Case Hint to open SIMILAR_CASES / SETUP_EFFECTS for supporting evidence. "
    "4) Use Weekend Summary Draft as report wording material only; it is comparison evidence, not an automatic setup decision. "
    "5) Do not treat blank cells as zero; blanks mean no reliable DB value for that field."
), 15)
ws.row_dimensions[2].height = 72
headers = ["Round","Circuit","Rider","Date","Runs","Laps","Best Lap","Avg Best Lap",
           "Race Pos","Top Problem Tags","Representative Comment","Key Suspension Signals",
           "Setup Actions","Similar Case Hint","Weekend Summary Draft"]
for c, h in enumerate(headers, 1):
    ws.cell(4, c, h)
rows_written = 0
for key, runs in sorted(event_runs.items(), key=lambda kv: (max((r["date"] or "") for r in kv[1]), kv[0][2]), reverse=True):
    rnd, circuit, rider = key
    best_vals = [r["best_lap_s"] for r in runs if r["best_lap_s"] is not None]
    race_pos = [r["session_position"] for r in runs if r["session_position"] is not None and str(r["session"]).startswith("RACE")]
    all_tags = [tag for r in runs for tag in tags_by_run.get(r["run_id"], [])]
    comments = [compact_text(r["comment"], 120) for r in runs if r["comment"]]
    metrics = [event_metric.get(r["run_id"]) for r in runs if event_metric.get(r["run_id"])]
    sig_bits = []
    if metrics:
        vals = [m["brk_f_dive_peak"] for m in metrics if m["brk_f_dive_peak"] is not None]
        if vals: sig_bits.append(f"Brk F-Dive Peak avg {round(sum(vals)/len(vals),0)}")
        vals = [m["ce_r_spd_avg"] for m in metrics if m["ce_r_spd_avg"] is not None]
        if vals: sig_bits.append(f"CE R-Spd avg {round(sum(vals)/len(vals),0)}")
        vals = [m["ph12_rear0_s"] for m in metrics if m["ph12_rear0_s"] is not None]
        if vals: sig_bits.append(f"PH1-2 Rear@0 avg {round(sum(vals)/len(vals),2)}s")
        vals = [m["rear_light_brk"] for m in metrics if m["rear_light_brk"] is not None]
        if vals: sig_bits.append(f"RearLightBrk avg {round(sum(vals)/len(vals),1)}%")
    actions = []
    for r in runs:
        for d in dec_by_to.get(r["run_id"], []):
            action = f"{d['component']}: {d['from_value']}→{d['to_value']}"
            if d["result_eval"]:
                action += f" ({d['result_eval']})"
            actions.append(action)
    top_tag_text = top_items(all_tags)
    best_lap = min(best_vals) if best_vals else None
    avg_best = round(sum(best_vals)/len(best_vals), 3) if best_vals else None
    hint = f"{circuit}/{rider}/{top_tag_text}" if top_tag_text else f"{circuit}/{rider}"
    draft = f"{circuit} {rider}: "
    if best_lap:
        draft += f"best {fmt_lap(best_lap)}. "
    if top_tag_text:
        draft += f"Main logged tendencies: {top_tag_text}. "
    if sig_bits:
        draft += "Data reference: " + "; ".join(sig_bits[:3]) + ". "
    if actions:
        draft += "Setup reference: " + "; ".join(actions[:2]) + ". "
    draft += "Use as comparison evidence, not an automatic setup decision."
    row = [rnd, circuit, rider, max((r["date"] or "") for r in runs), len(runs),
           sum((r["perf_laps"] or r["n_laps"] or 0) for r in runs),
           fmt_lap(best_lap), avg_best, min(race_pos) if race_pos else None,
           top_tag_text, comments[0] if comments else "", "; ".join(sig_bits),
           "; ".join(actions[:4]), hint, draft]
    rows_written += 1
    for c, v in enumerate(row, 1):
        ws.cell(4 + rows_written, c, v)
apply_table_style(ws, 4, len(headers), {
    1:10, 2:16, 3:10, 4:11, 5:8, 6:8, 7:11, 8:12, 9:9, 10:28,
    11:34, 12:38, 13:36, 14:28, 15:70
}, "A5")
for r in range(5, ws.max_row + 1):
    for c in range(10, 16):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
report.append(("WEEKEND_SUMMARY_HELPER", rows_written))

# ── SIMILAR_CASES: 過去事例検索ビュー(problem_log + run_tags + setup結果) ──
case_sql = """
WITH tag_case AS (
  SELECT rt.run_id, rt.tag AS problem_tag, NULL AS phase, NULL AS description, 'run_tags' AS source
  FROM run_tags rt
),
problem_case AS (
  SELECT run_id, problem_tag, phase, description, COALESCE(source,'problem_log') AS source
  FROM problem_log
  WHERE run_id IS NOT NULL
)
SELECT r.round,r.circuit,r.date,r.session,r.rider,r.run_no,r.run_id,
       c.problem_tag,c.phase,c.description,c.source,
       p.best_lap_s,p.run_avg_lap_s,p.session_position,
       a.brk_f_dive_peak,a.ce_r_spd_avg,a.ph12_rear0_s,a.rear_light_brk
FROM (SELECT * FROM tag_case UNION ALL SELECT * FROM problem_case) c
JOIN runs r ON r.run_id=c.run_id
LEFT JOIN performance p ON p.run_id=r.run_id
LEFT JOIN (
  SELECT ls.run_id,
    round(avg(ls.brk_f_dive_spd_peak),0) brk_f_dive_peak,
    round(avg(ls.ce_r_spd_avg),0) ce_r_spd_avg,
    round(avg(ls.ph12_rear0_s),2) ph12_rear0_s,
    round(avg(ls.rear_light_brk),1) rear_light_brk
  FROM lap_suspension ls JOIN laps l ON ls.lap_id=l.lap_id
  WHERE l.is_outlap=0
  GROUP BY ls.run_id
) a ON a.run_id=r.run_id
WHERE c.problem_tag IS NOT NULL
ORDER BY r.date DESC, r.circuit, r.rider, c.problem_tag
"""
case_runs = q(case_sql)
case_count = Counter((r["circuit"], r["rider"], r["problem_tag"]) for r in case_runs)
decision_rows = defaultdict(list)
for r in q("""SELECT run_id_from, run_id_to, component, from_value, to_value, expected_effect, actual_effect, result_eval
              FROM setup_decision_log ORDER BY created_at, decision_id"""):
    if r["run_id_from"]:
        decision_rows[r["run_id_from"]].append(r)
    if r["run_id_to"] and r["run_id_to"] != r["run_id_from"]:
        decision_rows[r["run_id_to"]].append(r)

ws = reset_sheet(wb, "SIMILAR_CASES", 2 if "SOLUTION_SEARCH" in wb.sheetnames else None)
ws["A1"] = "SIMILAR CASES"
ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
ws["A1"].fill = _TITLE_FILL
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
write_note(ws, 2, "過去の事例を検索するためのDB由来ビュー。Confidenceは事例数・2D根拠・setup結果の有無で算出。LOWは断定提案に使わない。", 20)
headers = ["Confidence","Case Count","Round","Circuit","Date","Session","Rider","Run","Problem Tag","Phase",
           "Best Lap","Race Pos","Data Evidence","Setup Action","Result","Source",
           "Report Reference Text","run_id","Similarity Keys","Caution"]
for c, h in enumerate(headers, 1):
    ws.cell(4, c, h)
rows_written = 0
for r in case_runs:
    decisions = decision_rows.get(r["run_id"], [])
    action_bits = []
    result_bits = []
    for d in decisions:
        bit = f"{d['component']}: {d['from_value']}→{d['to_value']}"
        if d["expected_effect"]:
            bit += f" / {compact_text(d['expected_effect'], 60)}"
        action_bits.append(bit)
        if d["result_eval"]:
            result_bits.append(str(d["result_eval"]))
    evidence_bits = []
    for label, col, suffix in [
        ("Brk F-Dive Peak", "brk_f_dive_peak", ""),
        ("CE R-Spd Avg", "ce_r_spd_avg", ""),
        ("PH1-2 Rear@0", "ph12_rear0_s", "s"),
        ("RearLightBrk", "rear_light_brk", "%"),
    ]:
        if r[col] is not None:
            evidence_bits.append(f"{label}={r[col]}{suffix}")
    n = case_count[(r["circuit"], r["rider"], r["problem_tag"])]
    has_positive = any(x.upper() == "POSITIVE" for x in result_bits)
    has_data = len(evidence_bits) >= 2
    if n >= 3 and has_data and has_positive:
        conf = "HIGH"
        fill = _GOOD_FILL
        caution = "Report comparison OK"
    elif n >= 2 and has_data:
        conf = "MED"
        fill = _WARN_FILL
        caution = "Use as hypothesis/reference"
    else:
        conf = "LOW"
        fill = _BAD_FILL
        caution = "Do not use for firm setup proposal"
    report_text = (f"{r['circuit']} {r['rider']} {r['session']} R{r['run_no']} had {r['problem_tag']}. "
                   f"Evidence: {'; '.join(evidence_bits[:4]) or 'no 2D metric evidence'}. "
                   f"Action/result: {'; '.join(action_bits[:2]) or 'no setup action recorded'} / {', '.join(result_bits) or 'no result'}.")
    row = [conf, n, r["round"], r["circuit"], r["date"], r["session"], r["rider"], r["run_no"],
           r["problem_tag"], r["phase"], fmt_lap(r["best_lap_s"]), r["session_position"],
           "; ".join(evidence_bits), "; ".join(action_bits[:3]), ", ".join(result_bits),
           r["source"], report_text, r["run_id"],
           f"{r['circuit']}|{r['rider']}|{r['problem_tag']}|{r['phase'] or ''}", caution]
    rows_written += 1
    for c, v in enumerate(row, 1):
        ws.cell(4 + rows_written, c, v)
    ws.cell(4 + rows_written, 1).fill = fill
    ws.cell(4 + rows_written, 1).font = Font(bold=True)
apply_table_style(ws, 4, len(headers), {
    1:11, 2:10, 3:10, 4:15, 5:11, 6:10, 7:9, 8:7, 9:20, 10:10,
    11:11, 12:8, 13:42, 14:46, 15:14, 16:13, 17:72, 18:44, 19:36, 20:28
}, "A5")
for r in range(5, ws.max_row + 1):
    for c in (13, 14, 17, 18, 19, 20):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
report.append(("SIMILAR_CASES", rows_written))

# ── SETUP_EFFECTS: セット変更→結果の一覧。Report根拠と事例検索に使う ──
ws = reset_sheet(wb, "SETUP_EFFECTS", 3 if "PROBLEM_LIBRARY" in wb.sheetnames else None)
ws["A1"] = "SETUP EFFECTS"
ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
ws["A1"].fill = _TITLE_FILL
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
write_note(ws, 2, "setup_decision_log をReport向けに整理。POSITIVEは過去事例の根拠、pendingは未評価、空欄はまだ判断材料不足として扱う。", 18)
headers = ["Result","Round","Circuit","Session","Rider","From Run","To Run","Component","From","To",
           "Expected Effect","Actual Effect","Rationale","Problem Context","Best Lap To",
           "To Run Data Evidence","run_id_from","run_id_to"]
for c, h in enumerate(headers, 1):
    ws.cell(4, c, h)
effect_sql = """
SELECT d.*, rf.comment AS from_comment, rt.comment AS to_comment,
       pf.best_lap_s AS from_best, pt.best_lap_s AS to_best,
       a.brk_f_dive_peak,a.ce_r_spd_avg,a.ph12_rear0_s,a.rear_light_brk
FROM setup_decision_log d
LEFT JOIN runs rf ON rf.run_id=d.run_id_from
LEFT JOIN runs rt ON rt.run_id=d.run_id_to
LEFT JOIN performance pf ON pf.run_id=d.run_id_from
LEFT JOIN performance pt ON pt.run_id=d.run_id_to
LEFT JOIN (
  SELECT ls.run_id,
    round(avg(ls.brk_f_dive_spd_peak),0) brk_f_dive_peak,
    round(avg(ls.ce_r_spd_avg),0) ce_r_spd_avg,
    round(avg(ls.ph12_rear0_s),2) ph12_rear0_s,
    round(avg(ls.rear_light_brk),1) rear_light_brk
  FROM lap_suspension ls JOIN laps l ON ls.lap_id=l.lap_id
  WHERE l.is_outlap=0
  GROUP BY ls.run_id
) a ON a.run_id=d.run_id_to
ORDER BY COALESCE(d.created_at,''), d.round, d.circuit, d.rider
"""
rows_written = 0
for r in q(effect_sql):
    context = compact_text(r["from_comment"] or r["to_comment"], 170)
    evidence = []
    for label, col, suffix in [
        ("Brk F-Dive Peak", "brk_f_dive_peak", ""),
        ("CE R-Spd Avg", "ce_r_spd_avg", ""),
        ("PH1-2 Rear@0", "ph12_rear0_s", "s"),
        ("RearLightBrk", "rear_light_brk", "%"),
    ]:
        if r[col] is not None:
            evidence.append(f"{label}={r[col]}{suffix}")
    row = [r["result_eval"], r["round"], r["circuit"], r["session"], r["rider"],
           r["run_id_from"], r["run_id_to"], r["component"], r["from_value"], r["to_value"],
           r["expected_effect"], r["actual_effect"], r["rationale"], context, fmt_lap(r["to_best"]),
           "; ".join(evidence), r["run_id_from"], r["run_id_to"]]
    rows_written += 1
    for c, v in enumerate(row, 1):
        ws.cell(4 + rows_written, c, v)
    result = (r["result_eval"] or "").upper()
    if result == "POSITIVE":
        ws.cell(4 + rows_written, 1).fill = _GOOD_FILL
    elif result == "NEGATIVE":
        ws.cell(4 + rows_written, 1).fill = _BAD_FILL
    else:
        ws.cell(4 + rows_written, 1).fill = _WARN_FILL
apply_table_style(ws, 4, len(headers), {
    1:11, 2:10, 3:15, 4:10, 5:9, 6:44, 7:44, 8:16, 9:16, 10:16,
    11:36, 12:36, 13:38, 14:42, 15:11, 16:40, 17:44, 18:44
}, "A5")
for r in range(5, ws.max_row + 1):
    for c in (6, 7, 11, 12, 13, 14, 16, 17, 18):
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
report.append(("SETUP_EFFECTS", rows_written))

# ── 保持シート(変更なし): SOLUTION_SEARCH(手順), DB_LOG, TREND_ANALYSIS ──
for sh in ["SOLUTION_SEARCH","DB_LOG","TREND_ANALYSIS"]:
    if sh in wb.sheetnames:
        report.append((sh + " (保持/未再生成)", "—"))

con.close()
wb.save(OUT)
print(f"保存: {OUT}")
print(f"シート: {wb.sheetnames}")
for name, n in report:
    print(f"  {name:32} {n}")
