#!/usr/bin/env python3
"""
lap_suspension_stats.py — TS24 Puccetti Racing ラップ別サスペンション統計
==========================================================================
2D Logger (.MES) ファイルを再処理してラップ単位のサスペンション統計を抽出し、
以下の2か所に書き込む。

  1. TS24 DB Master.xlsx → 「LAP_SUSPENSION」シート (新規作成 or 上書き)
  2. ts24_unified.db     → 「lap_suspension」テーブル + laps テーブルの sus_* 列更新

【抽出データ (1ラップ1行)】
  RUN_ID, LAP_ID, ラップ番号, ラップタイム(s)
  APEX: SusF平均/SusR平均/速度平均 (mm, km/h)
  BRAKE: SusF平均/SusR平均/速度平均
  FULL BRK: SusF平均/SusR平均
  LAP全体: SusF平均/SusR平均/SusF最小(最大圧縮)/SusF最大

実行方法:
  python lap_suspension_stats.py               ← 全 MES 処理
  python lap_suspension_stats.py --dry-run     ← DB/Excel 書き込みなし(確認のみ)
  python lap_suspension_stats.py --mes /path   ← 特定フォルダのみ処理

依存: parse_2d_channels.py と同じフォルダに置いて実行
"""

import sys
import re
import math
import struct
import sqlite3
import importlib.util
from pathlib import Path
from datetime import datetime

import numpy as np

# ─────────────────────────────────────────────────────────
#  パス設定
# ─────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).parent
DATA_2D_ROOT  = SCRIPT_DIR.parent / "DATA 2D"
DB_DIR        = SCRIPT_DIR.parent / "02_DATABASE"
EXCEL_PATH    = DB_DIR / "TS24 DB Master.xlsx"
UNIFIED_DB    = DB_DIR / "ts24_unified.db"
BACKUP_PATH   = DB_DIR / "TS24 DB Master Back UP.xlsx"
OUTPUT_SHEET  = "LAP_SUSPENSION"

# ─────────────────────────────────────────────────────────
#  APEX 定義パラメータ (2026-04-30 チーム確定)
# ─────────────────────────────────────────────────────────
# APEX (新定義 2026-04-30): BRAKE_FRONT -0.6~0.3Bar / GAS 0~6% /
#   dTPS_A 5~50 / SUSP_F 20~140mm / SUSP_R 5~50mm
# 旧3定義(ACC_Y Peak / BRAKE_OFF / THR_ON)は廃止
# dTPS_Aチャンネルが存在しないファイルは旧THR_ON定義にフォールバック

# THR_ON フォールバック用パラメータ (dTPS_A未搭載ファイル向け)
THR_ON_MIN_PCT           = 5.0    # この値以下を「アクセル全閉」とみなす [%]
THR_ON_TARGET_PCT        = 10.0   # この値を超えたらアクセルON確定 [%]
THR_ON_WIN_BEFORE_S      = 0.3    # Apex前から探し始める
THR_ON_WIN_AFTER_S       = 2.0    # Apex後何秒まで探すか

# ─────────────────────────────────────────────────────────
#  parse_2d_channels.py から必要な関数をインポート
# ─────────────────────────────────────────────────────────
_p2d_path = SCRIPT_DIR / "parse_2d_channels.py"
if not _p2d_path.exists():
    print(f"[ERROR] parse_2d_channels.py が見つかりません: {_p2d_path}")
    sys.exit(1)

spec = importlib.util.spec_from_file_location("p2d", _p2d_path)
p2d  = importlib.util.module_from_spec(spec)
spec.loader.exec_module(p2d)

# インポート先: 使用する関数・定数
parse_hed           = p2d.parse_hed
parse_ddd           = p2d.parse_ddd
parse_lap           = p2d.parse_lap
read_channel        = p2d.read_channel
smooth              = p2d.smooth
derivative          = p2d.derivative
susp_at_speed_index = p2d.susp_at_speed_index
susp_mean_in_range  = p2d.susp_mean_in_range
safe_mean           = p2d.safe_mean
detect_apexes       = p2d.detect_apexes
detect_apexes_accy  = p2d.detect_apexes_accy
detect_apex_area    = p2d.detect_apex_area
detect_brake_entries        = p2d.detect_brake_entries
detect_full_braking_sus     = p2d.detect_full_braking_sus
find_tyre_channel           = p2d.find_tyre_channel
_event_key_from_path        = p2d._event_key_from_path
_build_event_meta           = p2d._build_event_meta
_infer_date_for_session     = p2d._infer_date_for_session
find_all_mes                = p2d.find_all_mes
RIDER_PARAMS                = p2d.RIDER_PARAMS
KMH_TO_MS                   = p2d.KMH_TO_MS
G                           = p2d.G
MIN_LAP_DURATION_S          = p2d.MIN_LAP_DURATION_S

# ラウンド名正規化マップ (parse_2d_channels.py の _ROUND_NORM と同期)
_ROUND_NORM: dict[str, str] = {
    "R01": "ROUND1", "R1": "ROUND1",
    "R02": "ROUND2", "R2": "ROUND2",
    "R03": "ROUND3", "R3": "ROUND3",
    "R04": "ROUND4", "R4": "ROUND4",
    "R05": "ROUND5", "R5": "ROUND5",
    "T01": "TEST1",  "T1": "TEST1",
    "T02": "TEST2",  "T2": "TEST2",
    "T03": "TEST3",  "T3": "TEST3",
    "T04": "TEST4",  "T4": "TEST4",
    "T05": "TEST5",  "T5": "TEST5",
    "T06": "TEST6",  "T6": "TEST6",
    "TEST1": "TEST1", "TEST2": "TEST2", "TEST3": "TEST3",
    "TEST4": "TEST4", "TEST5": "TEST5", "TEST6": "TEST6",
    "WORKSHOP": "WORKSHOP",
}

_SESS_MAP = {
    "FP": "FP", "F1": "FP", "F2": "FP",
    "QP": "QP", "Q1": "QP", "Q2": "QP",
    "WU": "WUP", "WU1": "WUP", "WU2": "WUP",
    "WUP": "WUP", "WUP1": "WUP", "WUP2": "WUP",
    "R1": "RACE1", "R2": "RACE2",
    "RACE1": "RACE1", "RACE2": "RACE2",
    "D1": "TEST_D1", "D2": "TEST_D2",
    "L1": "TEST_D1", "L2": "TEST_D2",
    "SP": "SP",
    "INLAPR1": "RACE1", "INLAPR2": "RACE2",
}

_DEFAULT_CIRCUIT = {"PHILLIP ISLAND", "PHILLIPISLAND"}
_DEFAULT_DATE    = "16/02/2026"


# ─────────────────────────────────────────────────────────
#  RUN_ID / LAP_ID 生成 (build_unified_db.py と同一ロジック)
# ─────────────────────────────────────────────────────────
def _clean(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper().strip())


def make_run_id(round_val, circuit, session, rider, run_no):
    r  = _clean(round_val) if round_val else "UNK"
    c  = _clean(circuit)   if circuit   else "UNK"
    s  = _clean(session)   if session   else "UNK"
    ri = _clean(rider)     if rider     else "UNK"
    try:
        n = int(float(str(run_no)))
    except (ValueError, TypeError):
        n = 1
    return f"{r}_{c}_{s}_{ri}_R{n}"


def make_lap_id(run_id, lap_no):
    try:
        n = int(float(str(lap_no)))
    except (ValueError, TypeError):
        n = 0
    return f"{run_id}_L{n}"


def sec_to_laptime(s):
    if s is None:
        return None
    mins = int(s) // 60
    secs = s - mins * 60
    return f"{mins}:{secs:06.3f}"


# ─────────────────────────────────────────────────────────
#  1 MES ファイルのラップ別解析
# ─────────────────────────────────────────────────────────
def analyze_mes_per_lap(mes_path: Path, event_meta: dict) -> list[dict]:
    """
    1 MES フォルダを解析し、ラップ単位の統計リストを返す。
    各要素: {run_id, lap_id, lap_no, lap_time_s, apex_*, brake_*, fullbrk_*, lap_*}
    """
    base = mes_path.name.replace(".MES", "")

    # ── メタデータ ────────────────────────────────────────
    hed = parse_hed(mes_path, base)
    if not hed:
        return []

    # ライダー識別
    rider_num = hed.get("Rider Number", "")
    if "77" in rider_num:
        rider_tag = "DA77"
    elif "52" in rider_num or "JA" in rider_num.upper():
        rider_tag = "JA52"
    else:
        fname_up = mes_path.name.upper()
        par_up   = mes_path.parent.name.upper()
        if "JA52" in fname_up or "JA2" in fname_up or "52" in fname_up or par_up in ("52", "JA52"):
            rider_tag = "JA52"
        elif "77" in fname_up or "DA77" in fname_up or par_up in ("DA77", "77"):
            rider_tag = "DA77"
        else:
            return []

    # ラウンド正規化
    _raw_ekey = _event_key_from_path(mes_path)
    event = _ROUND_NORM.get(_raw_ekey or "", "") if _raw_ekey else ""
    if not event:
        hed_event = hed.get("Event", "").strip()
        _m_evt = re.match(r"^(R\d+|T\d+)(?:[^0-9]|$)", hed_event.upper())
        if _m_evt:
            event = _ROUND_NORM.get(_m_evt.group(1), hed_event)
        else:
            event = _ROUND_NORM.get(hed_event.upper(), hed_event)

    # セッション種別
    fn_prefix_m = re.match(r"^([A-Za-z0-9]+)", base)
    fn_prefix   = fn_prefix_m.group(1).upper() if fn_prefix_m else ""
    raw_sess    = hed.get("Session", "")
    session_type = _SESS_MAP.get(fn_prefix, _SESS_MAP.get(raw_sess.upper(), raw_sess.upper() or fn_prefix))

    # ラン番号
    run_match = re.search(r"-(\d+)$", base)
    run_no    = int(run_match.group(1)) if run_match else int(hed.get("Run", "1") or 1)

    # サーキット/日付 (JA52 補完含む)
    circuit = hed.get("Circuit", "").upper()
    date_s  = hed.get("Date", "")

    if re.match(r"^\d{2}\.\d{4}$", date_s):
        yy, mmdd = date_s[:2], date_s[3:]
        date_s = f"{mmdd[2:]}/{mmdd[:2]}/20{yy}"

    date_fmt = date_s
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            date_fmt = datetime.strptime(date_s, fmt).strftime("%Y-%m-%d")
            break
        except Exception:
            continue

    is_default_meta = (
        (circuit in _DEFAULT_CIRCUIT and date_s == _DEFAULT_DATE
         and raw_sess.upper() in ("L1", "L2", ""))
        or circuit in ("", "?")
    )
    if is_default_meta and event_meta:
        ekey = _event_key_from_path(mes_path)
        if ekey and ekey in event_meta:
            em = event_meta[ekey]
            new_circuit = em.get("circuit", "").strip()
            if new_circuit:
                circuit = new_circuit
            new_date = _infer_date_for_session(fn_prefix, em.get("session_dates", {}))
            if new_date:
                date_s = new_date
                date_fmt = date_s
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
                    try:
                        date_fmt = datetime.strptime(date_s, fmt).strftime("%Y-%m-%d")
                        break
                    except Exception:
                        continue

    _CIRC_NORM = {
        "PHILLIPISLAND": "PHILLIP ISLAND",
        "PHILLIPISISLAND": "PHILLIP ISLAND",
        "PHILLIP ISLAND": "PHILLIP ISLAND",
    }
    circuit = _CIRC_NORM.get(circuit.upper().strip(), circuit.upper().strip())

    # RUN_ID 生成
    run_id = make_run_id(event, circuit, session_type, rider_tag, run_no)

    # ── チャンネル読み込み ────────────────────────────────
    chs = parse_ddd(mes_path, base)
    needed = ["SPEED_FRONT", "SUSP_FRONT", "SUSP_REAR"]
    for ch in needed:
        if ch not in chs or not chs[ch]["ext"]:
            return []

    sf_raw    = read_channel(mes_path, base, chs["SPEED_FRONT"])
    sus_f_raw = read_channel(mes_path, base, chs["SUSP_FRONT"])
    sus_r_raw = read_channel(mes_path, base, chs.get("SUSP_REAR", {}))

    if len(sf_raw) < 10:
        return []

    susp_ratio = max(1, round(len(sus_f_raw) / len(sf_raw))) if len(sus_f_raw) > 0 else 4

    # BRAKE_FRONT (オプション)
    brake_f_raw = np.array([], dtype=np.float32)
    brake_ch_key = next((k for k in chs if k.upper() == "BRAKE_FRONT" and chs[k].get("ext")), None)
    if brake_ch_key:
        brake_f_raw = read_channel(mes_path, base, chs[brake_ch_key])
    brake_scale = len(brake_f_raw) / len(sf_raw) if len(brake_f_raw) > 0 and len(sf_raw) > 0 else 1.0

    # ACC_Y (横G最大点APEX検出 — Primary method)
    acc_y_raw = np.array([], dtype=np.float32)
    acc_y_ch_key = next((k for k in chs if k.upper() == "ACC_Y" and chs[k].get("ext")), None)
    if acc_y_ch_key:
        acc_y_raw = read_channel(mes_path, base, chs[acc_y_ch_key])
    accy_ratio = max(1, round(len(acc_y_raw) / len(sf_raw))) if len(acc_y_raw) > 0 else 1

    # GAS / GAS_SMOOTH (新APEX定義 + THR_ONフォールバック用)
    gas_raw = np.array([], dtype=np.float32)
    gas_ch_key = next(
        (k for k in ("GAS_SMOOTH", "GAS", "TPS", "TPS_A", "SEN_TPSA1")
         if k in chs and chs[k].get("ext")),
        None
    )
    if gas_ch_key:
        gas_raw = read_channel(mes_path, base, chs[gas_ch_key])
    gas_ratio = max(1, round(len(gas_raw) / len(sf_raw))) if len(gas_raw) > 0 else 1

    # dTPS_A (デルタスロットル — 新APEX定義の5条件のひとつ)
    dtps_raw = np.array([], dtype=np.float32)
    dtps_ch_key = next(
        (k for k in chs if k.upper() in ("DTPS_A", "DTPS", "TPS_DELTA", "DELTA_TPS")
         and chs[k].get("ext")),
        None
    )
    if dtps_ch_key:
        dtps_raw = read_channel(mes_path, base, chs[dtps_ch_key])
    has_dtps = len(dtps_raw) > 0

    # ── ラップ境界 ────────────────────────────────────────
    n_laps, lap_times_ms = parse_lap(mes_path, base)

    if n_laps > 0 and lap_times_ms:
        total_ms = lap_times_ms[-1]
        sr = len(sf_raw) / (total_ms / 1000.0) if total_ms > 0 else 100.0
    else:
        sr = 100.0
    sr = max(10.0, min(sr, 500.0))
    dt = 1.0 / sr
    sf_ms = sf_raw * KMH_TO_MS

    if n_laps > 0 and lap_times_ms:
        boundaries = []
        prev_ms = 0
        for t_ms in lap_times_ms:
            s_idx = int(prev_ms / 1000.0 * sr)
            e_idx = int(t_ms   / 1000.0 * sr)
            lap_t_s = (t_ms - prev_ms) / 1000.0
            boundaries.append((s_idx, min(e_idx, len(sf_raw) - 1), lap_t_s))
            prev_ms = t_ms
    else:
        dur_s = len(sf_raw) / sr
        boundaries = [(0, len(sf_raw) - 1, dur_s)]

    # ── ラップ別解析 ─────────────────────────────────────
    results = []

    for lap_idx, (lap_start, lap_end, lap_t_s) in enumerate(boundaries):
        lap_no = lap_idx + 1

        # アウトラップ/フォーメーション除外
        if lap_t_s < MIN_LAP_DURATION_S:
            continue
        if lap_end - lap_start < 20:
            continue

        lap_id = make_lap_id(run_id, lap_no)

        # ── APEX検出 (新定義 2026-04-30) ─────────────────────────────────
        # BRAKE_FRONT -0.6~0.3Bar / GAS 0~6% / dTPS_A 5~50 /
        # SUSP_F 20~140mm / SUSP_R 5~50mm の5条件同時成立区間
        apex_spds, apex_susF, apex_susR = [], [], []

        if len(brake_f_raw) > 0 and len(gas_raw) > 0 and has_dtps:
            # ── 新APEX定義: detect_apex_area() ──
            # brake_fレート基準のラップスライス
            b_start = int(lap_start * brake_scale)
            b_end   = int(min(lap_end * brake_scale, len(brake_f_raw)))
            g_start = lap_start * gas_ratio
            g_end   = min(lap_end * gas_ratio, len(gas_raw))
            d_start = lap_start * gas_ratio
            d_end   = min(lap_end * gas_ratio, len(dtps_raw))
            s_start = lap_start * susp_ratio
            s_end   = min(lap_end * susp_ratio, len(sus_f_raw))

            brake_lap = brake_f_raw[b_start:b_end]
            gas_lap   = gas_raw[g_start:g_end]
            dtps_lap  = dtps_raw[d_start:d_end]
            sus_f_lap = sus_f_raw[s_start:s_end]
            sus_r_lap = sus_r_raw[s_start:s_end] if len(sus_r_raw) > 0 else np.array([], dtype=np.float32)

            apex_areas = detect_apex_area(brake_lap, gas_lap, dtps_lap,
                                          sus_f_lap, sus_r_lap,
                                          gas_ratio=gas_ratio, sus_ratio=susp_ratio)

            for a in apex_areas:
                # midインデックスはbrake_fスペース → speed_rawスペースに戻す
                spd_idx = max(lap_start, min(lap_end - 1,
                              b_start + int(a["mid"] / brake_scale)))
                v = float(sf_raw[spd_idx])
                apex_spds.append(v)
                if not math.isnan(a["susF_avg"]):
                    apex_susF.append(a["susF_avg"])
                if not math.isnan(a["susR_avg"]):
                    apex_susR.append(a["susR_avg"])

        elif len(gas_raw) > 0:
            # ── フォールバック: dTPS_A未搭載ファイル → 旧THR_ON定義 ──
            fallback_apexes = detect_apexes_accy(acc_y_raw, sf_raw,
                                                  lap_start, lap_end, sr, accy_ratio)
            win_gb = int(THR_ON_WIN_BEFORE_S * sr * gas_ratio)
            win_ga = int(THR_ON_WIN_AFTER_S  * sr * gas_ratio)
            for ai in fallback_apexes:
                g0 = int(max(lap_start * gas_ratio, ai * gas_ratio - win_gb))
                g1 = int(min(lap_end   * gas_ratio, ai * gas_ratio + win_ga))
                g1 = min(g1, len(gas_raw))
                if g1 <= g0:
                    continue
                gas_win = gas_raw[g0:g1].astype(np.float64)
                min_idx = int(np.argmin(gas_win))
                if float(gas_win[min_idx]) > THR_ON_TARGET_PCT:
                    continue
                after_min = gas_win[min_idx:]
                crosses = np.where(after_min > THR_ON_TARGET_PCT)[0]
                if len(crosses) == 0:
                    continue
                cross_local = min_idx + int(crosses[0])
                gi_global   = int((g0 + cross_local) / gas_ratio)
                gi_global   = max(lap_start, min(lap_end - 1, gi_global))
                v  = float(sf_raw[gi_global])
                sF = susp_at_speed_index(sus_f_raw, gi_global, susp_ratio)
                sR = susp_at_speed_index(sus_r_raw, gi_global, susp_ratio) if len(sus_r_raw) > 0 else float("nan")
                apex_spds.append(v)
                if not math.isnan(sF): apex_susF.append(sF)
                if not math.isnan(sR): apex_susR.append(sR)

        # ── ブレーキ直前 ────────────────────────────────────────────────
        entries = detect_brake_entries(sf_raw, sf_ms, lap_start, lap_end, dt)
        brk_spds, brk_susF, brk_susR = [], [], []
        for bi in entries:
            v = float(sf_raw[bi]) if bi < len(sf_raw) else 0
            sF = susp_at_speed_index(sus_f_raw, bi, susp_ratio)
            sR = susp_at_speed_index(sus_r_raw, bi, susp_ratio) if len(sus_r_raw) > 0 else float("nan")
            brk_spds.append(v)
            if not math.isnan(sF): brk_susF.append(sF)
            if not math.isnan(sR): brk_susR.append(sR)

        # フルブレーキング
        fb_susF, fb_susR = [], []
        if len(brake_f_raw) > 0:
            fb_events = detect_full_braking_sus(
                brake_f_raw, sf_raw, sus_f_raw, sus_r_raw,
                lap_start, lap_end, brake_scale, susp_ratio, sr
            )
            for ev in fb_events:
                if ev["susF_mm"] is not None: fb_susF.append(ev["susF_mm"])
                if ev["susR_mm"] is not None: fb_susR.append(ev["susR_mm"])

        # 旧BOFF/THRON列はNone（廃止）— 後方互換のためキーは保持
        boff_susF, boff_susR, boff_spds = [], [], []
        thron_susF, thron_susR, thron_spds = apex_susF, apex_susR, apex_spds  # 同値

        # ── ラップ全体 SusF / SusR 統計 ──────────────────────────────────
        si_start = lap_start * susp_ratio
        si_end   = min(lap_end * susp_ratio, len(sus_f_raw))
        lap_susF_seg = sus_f_raw[si_start:si_end]
        lap_susR_seg = sus_r_raw[si_start:si_end] if len(sus_r_raw) > 0 else np.array([])

        lap_susF_mean = round(float(np.mean(lap_susF_seg)), 2) if len(lap_susF_seg) > 0 else None
        lap_susF_min  = round(float(np.min(lap_susF_seg)),  2) if len(lap_susF_seg) > 0 else None
        lap_susF_max  = round(float(np.max(lap_susF_seg)),  2) if len(lap_susF_seg) > 0 else None
        lap_susR_mean = round(float(np.mean(lap_susR_seg)), 2) if len(lap_susR_seg) > 0 else None

        def _mean(lst):
            return round(sum(lst) / len(lst), 2) if lst else None

        results.append({
            "run_id":        run_id,
            "lap_id":        lap_id,
            "round":         event,
            "circuit":       circuit,
            "session":       session_type,
            "rider":         rider_tag,
            "run_no":        run_no,
            "lap_no":        lap_no,
            "date":          date_fmt,
            "lap_time_s":    round(lap_t_s, 3),
            "lap_time_fmt":  sec_to_laptime(lap_t_s),
            # ① APEX — ACC_Y Peak (幾何学的Apex / 純旋回荷重)
            "apex_count":    len(apex_spds),
            "apex_spd_avg":  _mean(apex_spds),
            "apex_susF_avg": _mean(apex_susF),
            "apex_susR_avg": _mean(apex_susR),
            # ② BRAKE_OFF — ブレーキ解放点 (縦+横の複合荷重ピーク)
            "boff_count":    len(boff_spds),
            "boff_spd_avg":  _mean(boff_spds),
            "boff_susF_avg": _mean(boff_susF),
            "boff_susR_avg": _mean(boff_susR),
            # ③ THR_ON — アクセルON点 (ライダー体感Apex)
            "thron_count":   len(thron_spds),
            "thron_spd_avg": _mean(thron_spds),
            "thron_susF_avg":_mean(thron_susF),
            "thron_susR_avg":_mean(thron_susR),
            # BRAKE ENTRY
            "brk_count":     len(brk_spds),
            "brk_spd_avg":   _mean(brk_spds),
            "brk_susF_avg":  _mean(brk_susF),
            "brk_susR_avg":  _mean(brk_susR),
            # FULL BRAKING
            "fullbrk_count": len(fb_susF),
            "fullbrk_susF":  _mean(fb_susF),
            "fullbrk_susR":  _mean(fb_susR),
            # ラップ全体
            "lap_susF_mean": lap_susF_mean,
            "lap_susF_min":  lap_susF_min,
            "lap_susF_max":  lap_susF_max,
            "lap_susR_mean": lap_susR_mean,
        })

    return results


# ─────────────────────────────────────────────────────────
#  Excel 書き込み
# ─────────────────────────────────────────────────────────
HEADERS = [
    "RUN_ID", "LAP_ID", "ROUND", "CIRCUIT", "SESSION", "RIDER",
    "RUN_NO", "LAP_NO", "DATE", "LAP_TIME", "LAP_TIME_S",
    # APEX (新定義 2026-04-30: BRAKE_FRONT+GAS+dTPS_A+SUSP_F+SUSP_R 5条件同時成立)
    "APEX_CNT",   "APEX_SPD_AVG",   "APEX_SUSF_AVG",   "APEX_SUSR_AVG",
    # BOFF (廃止 — 後方互換のため列保持、値はNone)
    "BOFF_CNT",   "BOFF_SPD_AVG",   "BOFF_SUSF_AVG",   "BOFF_SUSR_AVG",
    # THRON (新APEX定義と同値 — dashboard後方互換のため保持)
    "THRON_CNT",  "THRON_SPD_AVG",  "THRON_SUSF_AVG",  "THRON_SUSR_AVG",
    # BRAKE ENTRY / FULL BRK
    "BRK_CNT",    "BRK_SPD_AVG",    "BRK_SUSF_AVG",    "BRK_SUSR_AVG",
    "FULLBRK_CNT", "FULLBRK_SUSF", "FULLBRK_SUSR",
    # LAP OVERALL
    "LAP_SUSF_MEAN", "LAP_SUSF_MIN", "LAP_SUSF_MAX", "LAP_SUSR_MEAN",
]

FIELDS = [
    "run_id", "lap_id", "round", "circuit", "session", "rider",
    "run_no", "lap_no", "date", "lap_time_fmt", "lap_time_s",
    # ①
    "apex_count",  "apex_spd_avg",  "apex_susF_avg",  "apex_susR_avg",
    # ②
    "boff_count",  "boff_spd_avg",  "boff_susF_avg",  "boff_susR_avg",
    # ③
    "thron_count", "thron_spd_avg", "thron_susF_avg", "thron_susR_avg",
    # BRAKE
    "brk_count",   "brk_spd_avg",   "brk_susF_avg",   "brk_susR_avg",
    "fullbrk_count", "fullbrk_susF", "fullbrk_susR",
    # LAP
    "lap_susF_mean", "lap_susF_min", "lap_susF_max", "lap_susR_mean",
]

HDR_BG = "1F3864"   # ダークネイビー
HDR_FG = "FFFFFF"
DAT_BG = "D6E4F0"   # 薄青


def write_to_excel(all_rows: list, path: Path, dry_run: bool = False):
    try:
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

    if dry_run:
        print(f"  [DRY-RUN] Excel への書き込みをスキップ ({len(all_rows)} 行)")
        return

    wb = load_workbook(path)

    # シートが存在すれば削除して再作成
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    ws = wb.create_sheet(OUTPUT_SHEET)

    # ── タイトル行 (row 1) ────────────────────────────────
    ws.merge_cells("A1:Z1")
    tc = ws["A1"]
    tc.value = "LAP_SUSPENSION — ラップ別サスペンション統計 (2D Logger 再処理)"
    tc.font      = Font(bold=True, color=HDR_FG, size=11, name="Arial")
    tc.fill      = PatternFill("solid", start_color=HDR_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── ヘッダー行 (row 2) ───────────────────────────────
    hdr_font = Font(bold=True, color=HDR_FG, size=9, name="Arial")
    hdr_fill = PatternFill("solid", start_color=HDR_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
    ws.row_dimensions[2].height = 30

    # ── データ行 ─────────────────────────────────────────
    dat_fill  = PatternFill("solid", start_color=DAT_BG)
    dat_font  = Font(color="1F3864", size=9, name="Arial")
    dat_align = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(all_rows, 3):
        for ci, field in enumerate(FIELDS, 1):
            val = row.get(field)
            c = ws.cell(row=ri, column=ci, value=val)
            if field in ("run_id", "lap_id"):
                c.font = Font(bold=True, color="1F3864", size=9, name="Arial")
                c.fill = dat_fill
                c.alignment = dat_align
            else:
                c.font = dat_font; c.fill = dat_fill; c.alignment = dat_align

    # ── 列幅 ─────────────────────────────────────────────
    col_widths = {
        1: 35, 2: 40, 3: 9, 4: 12, 5: 9, 6: 6,
        7: 7,  8: 7,  9: 12, 10: 10, 11: 11,
    }
    for ci in range(1, len(HEADERS) + 1):
        w = col_widths.get(ci, 11)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # フィルター
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"

    wb.save(path)
    print(f"  [Excel] {OUTPUT_SHEET} シートに {len(all_rows)} 行書き込み完了")


# ─────────────────────────────────────────────────────────
#  SQLite 書き込み
# ─────────────────────────────────────────────────────────
CREATE_LAP_SUSPENSION = """
DROP TABLE IF EXISTS lap_suspension;
CREATE TABLE lap_suspension (
    lap_id           TEXT PRIMARY KEY,
    run_id           TEXT,
    round            TEXT,
    circuit          TEXT,
    session          TEXT,
    rider            TEXT,
    run_no           INTEGER,
    lap_no           INTEGER,
    date             TEXT,
    lap_time_s       REAL,
    lap_time_fmt     TEXT,
    -- ① ACC_Y Peak : 幾何学的Apex (純旋回荷重)
    apex_count       INTEGER,
    apex_spd_avg     REAL,
    apex_susF_avg    REAL,
    apex_susR_avg    REAL,
    -- ② BRAKE_OFF : ブレーキ解放点 (縦+横の複合荷重ピーク)
    boff_count       INTEGER,
    boff_spd_avg     REAL,
    boff_susF_avg    REAL,
    boff_susR_avg    REAL,
    -- ③ THR_ON : アクセルON点 (ライダー体感Apex)
    thron_count      INTEGER,
    thron_spd_avg    REAL,
    thron_susF_avg   REAL,
    thron_susR_avg   REAL,
    -- ブレーキ進入 / フルブレーキング
    brk_count        INTEGER,
    brk_spd_avg      REAL,
    brk_susF_avg     REAL,
    brk_susR_avg     REAL,
    fullbrk_count    INTEGER,
    fullbrk_susF     REAL,
    fullbrk_susR     REAL,
    -- ラップ全体
    lap_susF_mean    REAL,
    lap_susF_min     REAL,
    lap_susF_max     REAL,
    lap_susR_mean    REAL,
    updated_at       TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_lapsus_run     ON lap_suspension(run_id);
CREATE INDEX IF NOT EXISTS idx_lapsus_circuit ON lap_suspension(circuit);
CREATE INDEX IF NOT EXISTS idx_lapsus_rider   ON lap_suspension(rider);
"""

INSERT_SQL = """
INSERT OR REPLACE INTO lap_suspension (
    lap_id, run_id, round, circuit, session, rider, run_no, lap_no, date,
    lap_time_s, lap_time_fmt,
    apex_count,  apex_spd_avg,  apex_susF_avg,  apex_susR_avg,
    boff_count,  boff_spd_avg,  boff_susF_avg,  boff_susR_avg,
    thron_count, thron_spd_avg, thron_susF_avg, thron_susR_avg,
    brk_count,   brk_spd_avg,   brk_susF_avg,   brk_susR_avg,
    fullbrk_count, fullbrk_susF, fullbrk_susR,
    lap_susF_mean, lap_susF_min, lap_susF_max, lap_susR_mean,
    updated_at
) VALUES (
    :lap_id, :run_id, :round, :circuit, :session, :rider, :run_no, :lap_no, :date,
    :lap_time_s, :lap_time_fmt,
    :apex_count,  :apex_spd_avg,  :apex_susF_avg,  :apex_susR_avg,
    :boff_count,  :boff_spd_avg,  :boff_susF_avg,  :boff_susR_avg,
    :thron_count, :thron_spd_avg, :thron_susF_avg, :thron_susR_avg,
    :brk_count,   :brk_spd_avg,   :brk_susF_avg,   :brk_susR_avg,
    :fullbrk_count, :fullbrk_susF, :fullbrk_susR,
    :lap_susF_mean, :lap_susF_min, :lap_susF_max, :lap_susR_mean,
    datetime('now')
)
"""


def write_to_sqlite(all_rows: list, db_path: Path, dry_run: bool = False):
    if dry_run:
        print(f"  [DRY-RUN] SQLite への書き込みをスキップ ({len(all_rows)} 行)")
        return

    conn = sqlite3.connect(db_path)
    # DROP TABLE + CREATE を一括実行 (スキーマ変更に対応)
    for stmt in CREATE_LAP_SUSPENSION.strip().split(";"):
        stmt = stmt.strip()
        if stmt:
            conn.execute(stmt)
    conn.commit()

    inserted = 0
    for row in all_rows:
        conn.execute(INSERT_SQL, row)
        inserted += 1

    conn.commit()
    conn.close()
    print(f"  [SQLite] lap_suspension テーブルに {inserted} 行 INSERT/REPLACE 完了")


# ─────────────────────────────────────────────────────────
#  メイン
# ─────────────────────────────────────────────────────────
def main():
    dry_run    = "--dry-run" in sys.argv
    single_mes = None
    for arg in sys.argv[1:]:
        if arg.startswith("--mes="):
            single_mes = Path(arg.split("=", 1)[1])
        elif arg not in ("--dry-run",) and not arg.startswith("-"):
            single_mes = Path(arg)

    if not DATA_2D_ROOT.exists():
        print(f"[ERROR] DATA 2D フォルダが見つかりません: {DATA_2D_ROOT}")
        sys.exit(1)

    print("=" * 60)
    print("  LAP SUSPENSION STATS — 2D ラップ別サスペンション解析")
    print("=" * 60)

    # イベントメタ構築 (JA52 HED 補完用)
    print("\n[1/4] イベントメタデータ構築中 (DA77 HED スキャン)...")
    event_meta = _build_event_meta(DATA_2D_ROOT)
    print(f"      {len(event_meta)} イベントキー検出")

    # MES フォルダ収集
    if single_mes:
        mes_list = [single_mes] if single_mes.is_dir() else []
        print(f"\n[2/4] 指定 MES フォルダ: {single_mes}")
    else:
        print("\n[2/4] MES フォルダ収集中...")
        mes_list = find_all_mes(DATA_2D_ROOT)
        print(f"      {len(mes_list)} フォルダ検出")

    # 解析
    print(f"\n[3/4] ラップ別サスペンション解析中...")
    all_rows: list[dict] = []
    skip_cnt = 0
    lap_cnt  = 0

    for i, mes_path in enumerate(mes_list):
        rows = analyze_mes_per_lap(mes_path, event_meta)
        if not rows:
            skip_cnt += 1
            continue
        all_rows.extend(rows)
        lap_cnt += len(rows)
        if (i + 1) % 10 == 0 or i == len(mes_list) - 1:
            print(f"      {i+1}/{len(mes_list)} MES 処理済 | "
                  f"有効: {len(mes_list)-skip_cnt} | ラップ合計: {lap_cnt}")

    if not all_rows:
        print("\n[INFO] 解析結果がありませんでした。")
        return

    # ソート (ROUND, CIRCUIT, SESSION, RIDER, RUN_NO, LAP_NO)
    all_rows.sort(key=lambda r: (
        r.get("round",""), r.get("circuit",""), r.get("session",""),
        r.get("rider",""), r.get("run_no", 0), r.get("lap_no", 0)
    ))

    print(f"\n  解析完了: {len(all_rows)} ラップ ({len(mes_list)-skip_cnt} MES ファイル)")

    # ── 統計サマリー ─────────────────────────────────────
    print("\n  ── サマリー ──────────────────────────────────")
    by_rider: dict = {}
    for r in all_rows:
        rd = r.get("rider","?")
        by_rider.setdefault(rd, []).append(r)
    for rd, rows in sorted(by_rider.items()):
        apex_f = [r["apex_susF_avg"] for r in rows if r.get("apex_susF_avg")]
        brk_f  = [r["brk_susF_avg"]  for r in rows if r.get("brk_susF_avg")]
        avg_lt = [r["lap_time_s"]    for r in rows if r.get("lap_time_s")]
        print(f"  {rd}: {len(rows)} ラップ | "
              f"APEX SusF avg: {round(sum(apex_f)/len(apex_f),1) if apex_f else '—'} mm | "
              f"BRK SusF avg: {round(sum(brk_f)/len(brk_f),1) if brk_f else '—'} mm | "
              f"LapTime avg: {round(sum(avg_lt)/len(avg_lt),3) if avg_lt else '—'} s")

    # ── 書き込み ─────────────────────────────────────────
    print(f"\n[4/4] データ書き込み...")

    if EXCEL_PATH.exists():
        write_to_excel(all_rows, EXCEL_PATH, dry_run)
    else:
        print(f"  [WARN] Excel が見つかりません: {EXCEL_PATH}")

    if UNIFIED_DB.exists():
        write_to_sqlite(all_rows, UNIFIED_DB, dry_run)
    else:
        print(f"  [WARN] SQLite DB が見つかりません: {UNIFIED_DB}")

    # ── JSON エクスポート (Streamlit Cloud用) ────────────────────────
    import json as _json
    JSON_OUT = SCRIPT_DIR / "lap_suspension_data.json"
    if not dry_run:
        # FIELDS → HEADERS のマッピングでキーを大文字化
        field_to_header = dict(zip(FIELDS, HEADERS))
        json_rows = [
            {field_to_header.get(k, k.upper()): v for k, v in row.items()
             if k in field_to_header}
            for row in all_rows
        ]
        JSON_OUT.write_text(
            _json.dumps(json_rows, ensure_ascii=False, indent=None),
            encoding="utf-8"
        )
        print(f"  [JSON] lap_suspension_data.json に {len(json_rows)} 行書き込み完了")
    else:
        print(f"  [DRY-RUN] JSON 書き込みをスキップ ({len(all_rows)} 行)")

    print(f"\n完了 {'(DRY-RUN)' if dry_run else ''}")
    print("=" * 60)


if __name__ == "__main__":
    main()
