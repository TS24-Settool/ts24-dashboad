#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_master_to_excel.py — ts24_master.db から シンプルな TS24 DB Master.xlsx を生成。
重要シートのみ: EVENTS / RUNS / LAPS / LAP_METRICS / PERFORMANCE / TAGS / PROBLEM_LIBRARY
旧シート(DB_LOG, TREND_ANALYSIS, BEST_WORST_ANALYSIS, SOLUTION_SEARCH, SESSION_SUMMARY,
TYRE_LOG, RUN_LOG, CHASSIS_GEO, DYNAMICS_ANALYSIS, LAP_SUSPENSION, LAP_TIMES,
PERFORMANCE_CORRELATION) は廃止し新データに置換。
"""
import sqlite3, shutil
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
DB   = ROOT/"02_DATABASE"/"ts24_master.db"
UNIFIED = ROOT/"02_DATABASE"/"ts24_unified.db"
XLSX = ROOT/"02_DATABASE"/"TS24 DB Master.xlsx"

HDR_BG="1F3864"; HDR_FG="FFFFFF"

def sheet_from_query(wb, name, conn, sql, extra_rows=None):
    cur=conn.execute(sql)
    cols=[d[0] for d in cur.description]
    ws=wb.create_sheet(name)
    ws.append(cols)
    for c in range(1,len(cols)+1):
        cell=ws.cell(1,c); cell.font=Font(bold=True,color=HDR_FG)
        cell.fill=PatternFill("solid",fgColor=HDR_BG); cell.alignment=Alignment(horizontal="center")
    n=0
    for row in cur:
        ws.append([ (v if v is not None else "") for v in row ]); n+=1
    ws.freeze_panes="A2"
    return n

def main():
    if not DB.exists():
        print("ts24_master.db が見つかりません"); return
    # 旧Excelを退避
    if XLSX.exists():
        bak = XLSX.with_name(f"TS24 DB Master PRE_REBUILD_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        shutil.copy2(XLSX, bak); print(f"旧Excel退避: {bak.name}")

    conn=sqlite3.connect(DB)
    wb=openpyxl.Workbook(); wb.remove(wb.active)

    counts={}
    counts["EVENTS"]=sheet_from_query(wb,"EVENTS",conn,
        "SELECT event_id,date,round,rider,circuit,report_file,weekend_summary,start_setup,end_setup FROM events ORDER BY date,rider")
    counts["RUNS"]=sheet_from_query(wb,"RUNS",conn,
        "SELECT run_id,rider,circuit,round,session,run_no,date,source,has_2d,n_laps,best_lap_s,comment,"
        "fork_type,f_spr_l,f_spr_r,f_preload,f_comp,f_reb,f_offset,f_hgt_top,f_hgt_bot,"
        "shock_type,r_spr,r_preload,r_comp,r_reb,shock_len,link,ride_hgt,swing_arm,tyre_front,tyre_rear "
        "FROM runs ORDER BY date,rider,circuit,session,run_no")
    counts["LAPS"]=sheet_from_query(wb,"LAPS",conn,
        "SELECT lap_id,run_id,lap_no,lap_time_s,susf_mean,susf_max,susr_mean,mes_file FROM laps ORDER BY run_id,lap_no")
    counts["LAP_METRICS"]=sheet_from_query(wb,"LAP_METRICS",conn,
        "SELECT lap_id,area,n,susf,susr,speed,brake,thr FROM lap_metrics ORDER BY lap_id,area")
    counts["PERFORMANCE"]=sheet_from_query(wb,"PERFORMANCE",conn,
        "SELECT run_id,rider,circuit,round,session,run_no,best_lap_s,run_avg_lap_s,session_position,n_laps "
        "FROM performance ORDER BY rider,circuit,round,session,run_no")
    counts["TAGS"]=sheet_from_query(wb,"TAGS",conn,"SELECT tag,category,complaint_en FROM tags ORDER BY tag")
    counts["RUN_TAGS"]=sheet_from_query(wb,"RUN_TAGS",conn,"SELECT run_id,tag,source FROM run_tags ORDER BY run_id")
    conn.close()

    # PROBLEM_LIBRARY を旧DBから保全
    try:
        u=sqlite3.connect(UNIFIED)
        counts["PROBLEM_LIBRARY"]=sheet_from_query(wb,"PROBLEM_LIBRARY",u,
            "SELECT * FROM problem_library ORDER BY id")
        u.close()
    except Exception as e:
        print("[WARN] problem_library保全失敗:",e)

    wb.save(XLSX)
    print(f"\n新 TS24 DB Master.xlsx 生成完了 (sheets={wb.sheetnames})")
    for k,v in counts.items(): print(f"  {k}: {v} 行")

if __name__=="__main__":
    main()
