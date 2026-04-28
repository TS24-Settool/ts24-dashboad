#!/usr/bin/env python3
"""
TS24 Puccetti Racing — TREND_ANALYSIS 自動集計スクリプト
=========================================================
RUN_LOG の PROBLEM DESC / COMMENT 列を読み取り、
TREND_ANALYSIS シートを自動更新する。

実行方法:
  python update_trend_analysis.py

更新対象:
  1. サブタイトル行 (セッション数・日付範囲)
  2. TOP PROBLEM TAGS テーブル (全体集計)
  3. RIDER COMPARISON テーブル (ライダー別・サーキット別)
  4. KEY ENGINEER NOTES (COMMENT 列から自動転記)
  5. データソース注記
"""

import re
import sys
import shutil
from copy import copy
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  パス設定
# ─────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
DB_DIR     = SCRIPT_DIR.parent / "02_DATABASE"
EXCEL_PATH = DB_DIR / "TS24 DB Master.xlsx"
BACKUP_PATH= DB_DIR / "TS24 DB Master Back UP.xlsx"

# ─────────────────────────────────────────────
#  問題タグ定義 (タグ名 → (フェーズ, 説明))
#  RUN_LOG の PROBLEM DESC 列に記入するタグ名と一致させること
# ─────────────────────────────────────────────
TAG_DEFS = {
    "chattering_brake": ("PH1",   "F chattering under braking — most frequent issue across both riders"),
    "line_loss_exit":   ("PH4",   "Cannot maintain exit line on throttle — consistent PH4 problem"),
    "nervousness":      ("PH2",   "Overall nervous/twitchy feeling — especially in wet/cold conditions"),
    "push_rear_exit":   ("PH4",   "Rear pushes bike wide on throttle — JA52 frequent complaint"),
    "no_turn_in":       ("PH2",   "Bike resists turn-in — affects both riders especially DA77"),
    "front_dive":       ("PH1",   "Front dives too deep or bottoms out"),
    "understeer_apex":  ("PH3",   "Understeer at apex — mid corner push"),
    "general_nervous":  ("PH1-5", "General nervous feeling throughout lap"),
}

# ─────────────────────────────────────────────
#  セルスタイルヘルパー
# ─────────────────────────────────────────────
def copy_style(src, dst):
    """src セルのスタイルを dst セルにコピー"""
    if src.has_style:
        dst.font       = copy(src.font)
        dst.border     = copy(src.border)
        dst.fill       = copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment  = copy(src.alignment)


# ─────────────────────────────────────────────
#  RUN_LOG 読み取り
# ─────────────────────────────────────────────
def read_run_log(wb):
    """
    RUN_LOG シートを読み取り、行ごとの dict リストを返す。
    ヘッダー行 (row=3) のカラム名をキーに使用。
    """
    ws = wb["RUN_LOG"]

    # 行3がヘッダー
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v:
            key = str(v).replace("\n", " ").strip()
            col_map[key] = c

    required = ["RIDER", "CIRCUIT", "SESSION", "RUN",
                "PROBLEM DESC", "COMMENT", "CHANGE INTENT",
                "EXPECTED EFFECT", "RESULT EVAL"]

    rows = []
    for r in range(4, ws.max_row + 1):
        rider   = ws.cell(row=r, column=col_map.get("RIDER",   1)).value
        circuit = ws.cell(row=r, column=col_map.get("CIRCUIT", 2)).value
        session = ws.cell(row=r, column=col_map.get("SESSION", 3)).value
        run     = ws.cell(row=r, column=col_map.get("RUN",     4)).value

        if not rider or not circuit:
            continue

        prob    = ws.cell(row=r, column=col_map.get("PROBLEM DESC",    39)).value or ""
        comment = ws.cell(row=r, column=col_map.get("COMMENT",         38)).value or ""
        change  = ws.cell(row=r, column=col_map.get("CHANGE INTENT",   40)).value or ""
        effect  = ws.cell(row=r, column=col_map.get("EXPECTED EFFECT", 41)).value or ""
        result  = ws.cell(row=r, column=col_map.get("RESULT EVAL",     42)).value or ""

        rows.append({
            "rider":   str(rider).strip(),
            "circuit": str(circuit).strip().upper(),
            "session": str(session).strip() if session else "",
            "run":     run,
            "prob":    str(prob).strip(),
            "comment": str(comment).strip(),
            "change":  str(change).strip(),
            "effect":  str(effect).strip(),
            "result":  str(result).strip(),
        })
    return rows, col_map


# ─────────────────────────────────────────────
#  タグ解析
# ─────────────────────────────────────────────
def parse_tags(tag_string):
    """
    PROBLEM DESC 文字列からタグリストを抽出。
    区切り文字: カンマ / セミコロン / 改行
    """
    if not tag_string:
        return []
    parts = re.split(r"[,;\n]+", tag_string)
    return [p.strip().lower() for p in parts if p.strip()]


# ─────────────────────────────────────────────
#  集計処理
# ─────────────────────────────────────────────
def aggregate(run_rows):
    """
    RUN_LOG 行から以下を集計して返す:
      - tag_total:   {tag: count}
      - rider_data:  {rider: {"sessions": set(), "tag_circuits": {tag: set(circuits)}}}
      - session_ids: set of (event_id, rider) — COMMENTがある unique イベント
      - note_map:    {event_id: {rider: note_text}} — COMMENT 列から
    """
    tag_total   = {t: 0 for t in TAG_DEFS}
    rider_data  = {}
    note_map    = {}          # e.g. {"20260313-TEST4": {"DA77": "Main issue...", ...}}

    for row in run_rows:
        rider   = row["rider"]
        circuit = row["circuit"]
        tags    = parse_tags(row["prob"])
        comment = row["comment"]

        # ライダー初期化
        if rider not in rider_data:
            rider_data[rider] = {"sessions": set(), "tag_circuits": {t: set() for t in TAG_DEFS}}

        # セッション識別 (タグか COMMENT がある行を「有効セッション」とカウント)
        if tags or comment:
            rider_data[rider]["sessions"].add((circuit, row["session"]))

        # タグ集計
        for tag in tags:
            if tag in TAG_DEFS:
                tag_total[tag]  += 1
                rider_data[rider]["tag_circuits"][tag].add(circuit)

        # COMMENT をイベントノートに集約
        # セッション (circuit+session) を event_id キーとして使用
        if comment:
            # イベントキーを生成 (例: "ASSEN-RACE1-JA52" 形式)
            ev_key = f"{circuit}-{row['session']}-{rider}"
            if ev_key not in note_map:
                note_map[ev_key] = comment

    return tag_total, rider_data, note_map


# ─────────────────────────────────────────────
#  TREND_ANALYSIS シート構造解析
# ─────────────────────────────────────────────
def find_trend_rows(ws):
    """
    TREND_ANALYSIS シートの主要行番号を辞書で返す。
    {
      "subtitle":       2,
      "tag_header":     5,
      "tag_data_start": 6,
      "tag_data_end":   13,
      "rider_ja52_header": 16,
      "rider_ja52_tag_start": 17,  (TAG ヘッダー行)
      "rider_da77_header": 26,
      "rider_da77_tag_start": 27,
      "notes_header": 35,
      "notes_data_start": 36,
      "data_source": 51,
    }
    """
    info = {}
    in_ja52_section = False
    in_da77_section = False
    notes_started   = False

    for r in range(1, ws.max_row + 2):
        v = str(ws.cell(row=r, column=1).value or "").strip()

        if r == 2:
            info["subtitle"] = r

        if "TOP PROBLEM TAGS" in v.upper():
            info["tag_section_header"] = r

        if v.upper() == "TAG" and r < 10:
            info["tag_data_start"] = r + 1

        if "RIDER COMPARISON" in v.upper():
            info["rider_comparison_header"] = r

        if re.match(r"Rider:\s*JA52", v):
            info["rider_ja52_header"] = r
            in_ja52_section = True
            in_da77_section = False

        if v.upper() == "TAG" and in_ja52_section and r > 15:
            info["rider_ja52_tag_row"] = r
            info["rider_ja52_data_start"] = r + 1

        if re.match(r"Rider:\s*DA77", v):
            info["rider_da77_header"] = r
            in_ja52_section = False
            in_da77_section = True

        if v.upper() == "TAG" and in_da77_section:
            info["rider_da77_tag_row"] = r
            info["rider_da77_data_start"] = r + 1

        if "KEY ENGINEER NOTES" in v.upper():
            info["notes_section_header"] = r
            notes_started = True

        if notes_started and r > info.get("notes_section_header", 999):
            if "DATA SOURCE" in v.upper():
                info["data_source"] = r
                notes_started = False

    return info


# ─────────────────────────────────────────────
#  TREND_ANALYSIS 更新
# ─────────────────────────────────────────────
def update_trend_analysis(wb, run_rows, tag_total, rider_data, note_map):
    ws = wb["TREND_ANALYSIS"]
    info = find_trend_rows(ws)
    print(f"  TREND_ANALYSIS 構造: {info}")

    # ── (A) サブタイトル更新 ────────────────────────────────
    # セッション数 (タグ/コメントあり)
    sessions_ja52 = len(rider_data.get("JA52", {}).get("sessions", set()))
    sessions_da77 = len(rider_data.get("DA77", {}).get("sessions", set()))
    total_sessions = sessions_ja52 + sessions_da77

    # 全 RUN_LOG から日付範囲取得 (SESSION_SUMMARY から)
    try:
        import pandas as pd
        df_s = pd.read_excel(EXCEL_PATH, sheet_name="SESSION_SUMMARY", header=1)
        dates = pd.to_datetime(df_s["DATE"].dropna(), errors="coerce").dropna()
        date_min = dates.min().strftime("%Y-%m") if len(dates) else "?"
        date_max = dates.max().strftime("%Y-%m") if len(dates) else "?"
    except Exception:
        date_min, date_max = "?", "?"

    if total_sessions > 0:
        # タグデータあり → 正確な数字で上書き
        subtitle = (
            f"Based on {total_sessions} sessions: "
            f"JA52 ({sessions_ja52} sessions) & DA77 ({sessions_da77} sessions) | "
            f"{date_min} → {date_max}"
        )
        ws.cell(row=info["subtitle"], column=1).value = subtitle
        print(f"  サブタイトル更新: {subtitle[:80]}...")
    else:
        # タグデータなし → 日付範囲末尾のみ更新、セッション数は現状維持
        existing = str(ws.cell(row=info["subtitle"], column=1).value or "")
        updated  = re.sub(r"→ \d{4}-\d{2}", f"→ {date_max}", existing)
        if updated != existing:
            ws.cell(row=info["subtitle"], column=1).value = updated
            print(f"  サブタイトル(日付のみ更新): {updated[:80]}...")
        else:
            print(f"  サブタイトル: 変更なし (タグ未入力)")

    # ── (B) TOP PROBLEM TAGS テーブル更新 ───────────────────
    # tag_data_start 行から順に TAG_DEFS の順序で書き込む
    if total_sessions > 0 and "tag_data_start" in info:
        # 降順ソート
        sorted_tags = sorted(TAG_DEFS.keys(), key=lambda t: tag_total.get(t, 0), reverse=True)
        r = info["tag_data_start"]
        for tag in sorted_tags:
            if tag_total.get(tag, 0) == 0:
                continue
            phase, desc = TAG_DEFS[tag]
            ws.cell(row=r, column=1).value = tag
            ws.cell(row=r, column=2).value = str(tag_total[tag])
            ws.cell(row=r, column=3).value = phase
            ws.cell(row=r, column=4).value = desc
            r += 1
        # 残行をクリア (旧データが長い場合)
        while ws.cell(row=r, column=1).value and r < info.get("rider_comparison_header", 999):
            ws.cell(row=r, column=1).value = None
            ws.cell(row=r, column=2).value = None
            ws.cell(row=r, column=3).value = None
            ws.cell(row=r, column=4).value = None
            r += 1
        print(f"  TOP PROBLEM TAGS 更新完了 ({len([t for t in sorted_tags if tag_total.get(t,0)>0])} タグ)")

    # ── (C) RIDER COMPARISON 更新 ────────────────────────────
    for rider, header_key, data_start_key in [
        ("JA52", "rider_ja52_header",    "rider_ja52_data_start"),
        ("DA77", "rider_da77_header",    "rider_da77_data_start"),
    ]:
        if header_key not in info or data_start_key not in info:
            continue

        rd = rider_data.get(rider, {})
        n_sess = len(rd.get("sessions", set()))

        # ヘッダー行のセッション数更新
        ws.cell(row=info[header_key], column=1).value = f"Rider: {rider}  ({n_sess} sessions)"

        if total_sessions == 0 or not rd.get("tag_circuits"):
            continue

        # タグ→サーキット 降順
        tag_circ = rd.get("tag_circuits", {})
        sorted_rt = sorted(tag_circ.keys(),
                           key=lambda t: len(tag_circ.get(t, [])), reverse=True)

        r = info[data_start_key]
        for tag in sorted_rt:
            circs = tag_circ.get(tag, set())
            if not circs:
                continue
            circ_str = ", ".join(sorted(circs))
            ws.cell(row=r, column=1).value = tag
            ws.cell(row=r, column=2).value = str(len(circs))
            ws.cell(row=r, column=3).value = circ_str
            r += 1
        # 残行クリア
        end_row = (info.get("rider_da77_header") if rider == "JA52"
                   else info.get("notes_section_header", 999))
        while ws.cell(row=r, column=1).value and r < end_row:
            ws.cell(row=r, column=1).value = None
            ws.cell(row=r, column=2).value = None
            ws.cell(row=r, column=3).value = None
            r += 1
        print(f"  RIDER COMPARISON ({rider}) 更新完了")

    # ── (D) KEY ENGINEER NOTES 更新 ──────────────────────────
    # note_map: {ev_key: note_text}
    # 既存の KEY ENGINEER NOTES 行を走査し、一致するセッション IDに上書き
    if note_map and "notes_section_header" in info:
        notes_start = info["notes_section_header"] + 1
        notes_end   = info.get("data_source", ws.max_row)
        for r in range(notes_start, notes_end):
            session_id = ws.cell(row=r, column=1).value
            if not session_id:
                continue
            sid = str(session_id).strip()
            # ev_key は "CIRCUIT-SESSION-RIDER" 形式なので直接マッチは難しい
            # session_id 末尾の RIDER 部分を取得して note_map と照合
            for ev_key, note in note_map.items():
                # ev_key 例: "ASSEN-RACE1-JA52"
                rider_suffix = ev_key.split("-")[-1]
                if rider_suffix in sid:
                    circuit_part = ev_key.split("-")[0]
                    if circuit_part.upper() in sid.upper():
                        existing = ws.cell(row=r, column=3).value or ""
                        if "(not yet recorded)" in str(existing).lower() or not existing:
                            ws.cell(row=r, column=3).value = note
                            print(f"  NOTE 更新: row {r} → {sid}")
                            break

    # ── (E) データソース注記更新 ─────────────────────────────
    if "data_source" in info:
        today = date.today().strftime("%Y-%m-%d")
        ws.cell(row=info["data_source"], column=1).value = (
            f"DATA SOURCE: TS24 DB Master RUN_LOG シートから自動集計 | "
            f"最終更新: {today} | "
            f"セッション数 JA52={sessions_ja52}, DA77={sessions_da77} | "
            f"タグ入力済みラン: {sum(1 for r in run_rows if r['prob'])}"
        )
        print(f"  データソース注記更新")

    return ws


# ─────────────────────────────────────────────
#  メイン
# ─────────────────────────────────────────────
def main():
    print("=" * 55)
    print("TS24 TREND_ANALYSIS 自動集計スクリプト")
    print("=" * 55)

    if not EXCEL_PATH.exists():
        print(f"ERROR: ファイルが見つかりません: {EXCEL_PATH}")
        sys.exit(1)

    print(f"\n1. Excelを読み込み中: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    print("2. RUN_LOG を解析中...")
    run_rows, col_map = read_run_log(wb)
    print(f"   → {len(run_rows)} 行を読み込みました")

    filled_rows = [r for r in run_rows if r["prob"]]
    print(f"   → PROBLEM DESC 入力済み: {len(filled_rows)} 行")

    commented_rows = [r for r in run_rows if r["comment"]]
    print(f"   → COMMENT 入力済み:      {len(commented_rows)} 行")

    print("3. 集計処理中...")
    tag_total, rider_data, note_map = aggregate(run_rows)

    # タグ集計サマリ表示
    if any(v > 0 for v in tag_total.values()):
        print("   タグ集計結果:")
        for tag, count in sorted(tag_total.items(), key=lambda x: -x[1]):
            if count > 0:
                print(f"     {tag}: {count}")
    else:
        print("   ※ PROBLEM DESC にタグ未入力 → カウントは 0 のまま")

    print("4. TREND_ANALYSIS を更新中...")
    update_trend_analysis(wb, run_rows, tag_total, rider_data, note_map)

    print("5. 保存中...")
    wb.save(str(EXCEL_PATH))
    print(f"   → {EXCEL_PATH.name} 保存完了")

    print("6. バックアップ更新中...")
    shutil.copy(str(EXCEL_PATH), str(BACKUP_PATH))
    print(f"   → {BACKUP_PATH.name} 更新完了")

    print()
    print("=" * 55)
    print("完了!")
    print()
    print("【使い方メモ】")
    print("  RUN_LOG の「PROBLEM DESC」列 (Col AO) にタグを入力してください。")
    print("  複数タグはカンマ区切り: chattering_brake, no_turn_in")
    print()
    print("  使用可能なタグ一覧:")
    for tag, (phase, desc) in TAG_DEFS.items():
        print(f"    {tag:<22} ({phase}) — {desc[:45]}")
    print()
    print("  入力後にこのスクリプトを再実行すると TREND_ANALYSIS が更新されます。")
    print("=" * 55)


if __name__ == "__main__":
    main()
