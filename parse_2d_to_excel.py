#!/usr/bin/env python3
"""
TS24 Puccetti Racing — 2D Data Auto-Import Script
==================================================
Scans all round folders for MES directories, extracts:
  - Lap times from .LAP files
  - Setup & session info from .HED files
  - Tyre data from .HED files
Writes results to PUCCETTI_DB_MASTER.xlsx (LAP_TIMES, SESSION_SUMMARY, TYRE_LOG sheets)

Usage:
  python parse_2d_to_excel.py [ROOT_FOLDER] [EXCEL_PATH]

Default ROOT_FOLDER: same directory as this script (looks for *2D DATA* folders)
Default EXCEL_PATH:  ../02_DATABASE/PUCCETTI_DB_MASTER.xlsx
"""

import json
import os
import re
import struct
import sys
from copy import copy
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  Colors & Styles
# ─────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SUB_BG      = "2E75B6"   # medium blue
C_SUB_FG      = "FFFFFF"
C_ALT_BG      = "EBF3FB"   # light blue
C_BEST_BG     = "E2EFDA"   # light green (best lap)
C_OUT_BG      = "FFF2CC"   # light yellow (outlap)
C_ACCENT      = "2E75B6"

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(ws, row, cols, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=fg, size=size)
        c.alignment = _center()
        c.border = _border()

def _style_data_row(ws, row, cols, alt=False, special_fill=None):
    bg = C_ALT_BG if alt else "FFFFFF"
    if special_fill:
        bg = special_fill
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        if not special_fill:
            c.fill = _fill(bg)
        else:
            c.fill = _fill(special_fill)
        c.font = _font(size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────
#  2D File Parsers
# ─────────────────────────────────────────────

def parse_decompress_log(mes_path):
    """Read Decompress2D.TXT to extract actual circuit from Windows path."""
    txt = mes_path / "Decompress2D.TXT"
    if not txt.exists():
        return ""
    try:
        content = txt.read_bytes().decode("latin-1", errors="ignore")
        m = re.search(r"Path for measurement:.*\\([^\\]+)\\[^\\]+\.MES", content)
        if m:
            folder = m.group(1).upper()
            # e.g. R03_ASSEN.26 → ASSEN
            c = re.search(r"R\d+_([A-Z_]+)\.\d+", folder)
            if c:
                return c.group(1).replace("_", " ").title()
    except Exception:
        pass
    return ""


def parse_hed(hed_path):
    """Parse .HED file → dict of session/setup/tyre info"""
    try:
        raw = Path(hed_path).read_bytes().decode("latin-1", errors="ignore")
    except Exception:
        return {}

    def val(key):
        m = re.search(rf"^{re.escape(key)}=(.+)", raw, re.MULTILINE)
        return m.group(1).strip() if m else ""

    # Laptimes section
    lap_section = re.search(r"\[Laptimes\](.*?)(?:\[|\Z)", raw, re.DOTALL)
    fastest = ""
    if lap_section:
        fm = re.search(r"Fastest lap=(.+)", lap_section.group(1))
        fastest = fm.group(1).strip() if fm else ""

    return {
        "date_raw":     val("Date"),
        "event":        val("Event"),
        "circuit":      val("Circuit"),
        "session_code": val("Session"),
        "run":          val("Run"),
        "rider_name":   val("Rider"),
        "rider_num":    val("Rider Number"),
        "condition":    val("Condition"),
        "air_temp":     val("Air Temp"),
        "track_temp":   val("Track Temp"),
        "humidity":     val("Humidity"),
        "chassis":      val("Chassis"),
        "fork":         val("ForkNo"),
        "fork_spec":    val("Spec"),
        "fork_comp":    val("Comp"),
        "fork_reb":     val("Reb"),
        "fork_spring_l":val("Spring Left"),
        "fork_spring_r":val("Spring Right"),
        "fork_preload": val("Preload"),
        "fork_oil_lvl": val("Oil Level"),
        "shock":        val("ShockNo"),
        "shock_spec":   val("Spec"),
        "shock_comp":   val("Comp"),
        "shock_reb":    val("Reb"),
        "shock_spring": val("Spring"),
        "shock_preload":val("Preload"),
        "shock_len":    val("Shock Length"),
        "offset":       val("Offset"),
        "tyre_front":   val("Front Tyre"),
        "tyre_rear":    val("Rear Tyre"),
        "tyre_f_press": val("FPressOUT"),
        "tyre_r_press": val("RPressOUT"),
        "tyre_f_laps":  val("F Laps"),
        "tyre_r_laps":  val("R Laps"),
        "tyre_f_temp":  val("FWTemp"),
        "tyre_r_temp":  val("RWTemp"),
        "f_sprkt":      val("F_Sprkt"),
        "r_sprkt":      val("R_Sprkt"),
        "engine":       val("Engine"),
        "fastest_lap":  fastest,
        "ecu":          val("ECU"),
        "fuel_type":    val("Fuel Type"),
    }


def parse_lap(lap_path):
    """Parse .LAP binary file → list of lap time strings + seconds"""
    try:
        raw = Path(lap_path).read_bytes()
    except Exception:
        return []
    if len(raw) < 12:
        return []

    n_laps = struct.unpack_from("<I", raw, 0)[0]
    if n_laps == 0 or n_laps > 200:
        return []

    timestamps = []
    for i in range(n_laps + 1):
        offset = (i + 2) * 4
        if offset + 4 > len(raw):
            break
        timestamps.append(struct.unpack_from("<I", raw, offset)[0])

    laps = []
    for i in range(len(timestamps) - 1):
        ms = timestamps[i + 1] - timestamps[i]
        t = ms / 1000.0
        m = int(t // 60)
        s = t - m * 60
        laps.append({
            "lap_no":    i,          # 0 = outlap
            "lap_time":  f"{m}:{s:06.3f}",
            "lap_time_s": round(t, 3),
            "is_outlap": i == 0,
        })
    return laps


def _fmt_t(t_s):
    """seconds → M:SS.mmm string"""
    try:
        t = float(t_s)
        m = int(t // 60)
        s = t - m * 60
        return f"{m}:{s:06.3f}"
    except Exception:
        return ""


def _parse_date(raw):
    """Try to parse date from HED date string"""
    for fmt in ("%d/%m/%Y", "%y.%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return raw


def _infer_session_type(session_name, hed_session_code):
    """Infer session type from folder name or HED session code"""
    name = session_name.upper()
    if re.search(r"\bFP\b|F1\b|FREE", name):     return "FP"
    if re.search(r"\bSP\b|SPRINT", name):          return "SP"
    if re.search(r"\bQP\b|QUALI", name):           return "QP"
    if re.search(r"WUP2|WU2\b", name):             return "WUP2"
    if re.search(r"WUP1?|WU1\b", name):            return "WUP1"
    if re.search(r"\bR1\b|RACE1|RA1", name):       return "RACE1"
    if re.search(r"\bR2\b|RACE2|RA2", name):       return "RACE2"
    if "RACE" in name or re.search(r"\bR[D]?\d", name): return "RACE1"
    code = str(hed_session_code).upper()
    if "L1" in code: return "FP"
    if "SP" in code: return "SP"
    if "Q"  in code: return "QP"
    if "W"  in code: return "WUP1"
    if "R1" in code: return "RACE1"
    if "R2" in code: return "RACE2"
    return "FP"


def _infer_rider(session_name, hed_num):
    """Infer rider ID from session name or HED rider number"""
    name = session_name.upper()
    if "52" in name or "JA" in name:  return "JA52"
    if "77" in name or "DA" in name:  return "DA77"
    num = str(hed_num).strip()
    if num == "52": return "JA52"
    if num == "77": return "DA77"
    return f"#{num}" if num else "UNKNOWN"


def _infer_run_no(session_name):
    """Extract run number from session folder name"""
    m = re.search(r"-(\d+)\.MES$", session_name, re.IGNORECASE)
    return int(m.group(1)) if m else 1


def _infer_round(folder_name):
    """Extract round name from containing folder name"""
    name = folder_name.upper()
    m = re.search(r"ROUND\s*(\d+)", name)
    if m:
        return f"ROUND{m.group(1)}"
    m = re.search(r"R(\d+)", name)
    if m:
        return f"ROUND{m.group(1)}"
    return folder_name


# ─────────────────────────────────────────────
#  Scanner
# ─────────────────────────────────────────────

def scan_all_mes(root):
    """
    Recursively scan root for all .MES directories.
    Returns list of (round_name, rider_folder, mes_path) tuples.
    """
    root = Path(root)
    results = []

    if not root.exists():
        return results

    for item in sorted(root.iterdir()):
        if not item.is_dir():
            continue
        if item.name.startswith("."):
            continue

        # Check if this is a round-level folder
        round_name = _infer_round(item.name)

        # Look for rider sub-folders (52, 77) or direct MES dirs
        rider_dirs = [d for d in item.iterdir()
                      if d.is_dir() and not d.name.startswith(".")]

        for rdir in sorted(rider_dirs):
            # Could be rider folder (52/77) or MES folder directly
            if rdir.suffix.upper() == ".MES":
                results.append((round_name, "", rdir))
            else:
                # Assume rider sub-folder
                for mes in sorted(rdir.iterdir()):
                    if mes.is_dir() and mes.suffix.upper() == ".MES":
                        results.append((round_name, rdir.name, mes))

    return results


def extract_session(round_name, mes_path):
    """Parse one MES folder → session dict with laps list"""
    mes_name = mes_path.name  # e.g. FP-JA52-01.MES
    stem     = mes_path.stem  # e.g. FP-JA52-01

    hed_file = mes_path / f"{stem}.HED"
    lap_file = mes_path / f"{stem}.LAP"

    hed = parse_hed(hed_file) if hed_file.exists() else {}
    laps = parse_lap(lap_file) if lap_file.exists() else []

    # Circuit: prefer Decompress2D.TXT path (more reliable than HED template)
    circuit = (parse_decompress_log(mes_path) or hed.get("circuit", "") or hed.get("event", "")).upper()
    date_raw = hed.get("date_raw", "")

    # Date: prefer actual date from HED
    # HED date_raw might be "26.0417" (yyMMdd) or "17/04/2026"
    date_str = _parse_date(date_raw)
    if date_str == date_raw:
        # Try yyMMdd format
        m = re.match(r"(\d{2})\.(\d{2})(\d{2})", date_raw)
        if m:
            date_str = f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"

    session_type = _infer_session_type(stem, hed.get("session_code", ""))
    rider        = _infer_rider(stem, hed.get("rider_num", ""))
    run_no       = _infer_run_no(mes_name)

    # Best lap (exclude outlap)
    timed_laps = [l for l in laps if not l["is_outlap"]]
    best_lap_s = min((l["lap_time_s"] for l in timed_laps), default=None)
    best_lap   = _fmt_t(best_lap_s) if best_lap_s else hed.get("fastest_lap", "")
    avg_lap_s  = (round(sum(l["lap_time_s"] for l in timed_laps) / len(timed_laps), 3)
                  if timed_laps else None)

    return {
        "round":        round_name,
        "circuit":      circuit,
        "date":         date_str,
        "session_type": session_type,
        "rider":        rider,
        "run_no":       run_no,
        "condition":    hed.get("condition", ""),
        "air_temp":     hed.get("air_temp", ""),
        "track_temp":   hed.get("track_temp", ""),
        "best_lap":     best_lap,
        "best_lap_s":   best_lap_s,
        "avg_lap_s":    avg_lap_s,
        "total_laps":   len(timed_laps),
        # Setup
        "chassis":      hed.get("chassis", ""),
        "fork":         hed.get("fork", ""),
        "fork_spec":    hed.get("fork_spec", ""),
        "fork_comp":    hed.get("fork_comp", ""),
        "fork_reb":     hed.get("fork_reb", ""),
        "fork_spr_l":   hed.get("fork_spring_l", ""),
        "fork_spr_r":   hed.get("fork_spring_r", ""),
        "fork_preload": hed.get("fork_preload", ""),
        "fork_oil":     hed.get("fork_oil_lvl", ""),
        "shock":        hed.get("shock", ""),
        "shock_spec":   hed.get("shock_spec", ""),
        "shock_comp":   hed.get("shock_comp", ""),
        "shock_reb":    hed.get("shock_reb", ""),
        "shock_spr":    hed.get("shock_spring", ""),
        "shock_pre":    hed.get("shock_preload", ""),
        "shock_len":    hed.get("shock_len", ""),
        "offset":       hed.get("offset", ""),
        "f_sprkt":      hed.get("f_sprkt", ""),
        "r_sprkt":      hed.get("r_sprkt", ""),
        "engine":       hed.get("engine", ""),
        # Tyres
        "tyre_f":       hed.get("tyre_front", ""),
        "tyre_r":       hed.get("tyre_rear", ""),
        "tyre_f_press": hed.get("tyre_f_press", ""),
        "tyre_r_press": hed.get("tyre_r_press", ""),
        "tyre_f_laps":  hed.get("tyre_f_laps", ""),
        "tyre_r_laps":  hed.get("tyre_r_laps", ""),
        "tyre_f_temp":  hed.get("tyre_f_temp", ""),
        "tyre_r_temp":  hed.get("tyre_r_temp", ""),
        "laps":         laps,
    }


# ─────────────────────────────────────────────
#  Excel Writer
# ─────────────────────────────────────────────

def write_lap_times_sheet(wb, sessions):
    ws = wb.create_sheet("LAP_TIMES")

    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = "TS24 PUCCETTI RACING — All Lap Times (Auto-imported from 2D)"
    ws["A1"].font  = _font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill  = _fill(C_HEADER_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["ROUND", "CIRCUIT", "DATE", "SESSION", "RIDER", "RUN",
               "LAP", "LAP TIME", "TIME (s)", "OUTLAP?",
               "WEATHER", "AIR °C", "TRACK °C", "TYRE F", "TYRE R"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    row = 3
    for s in sessions:
        for lap in s["laps"]:
            is_best = (not lap["is_outlap"] and
                       s["best_lap_s"] and
                       abs(lap["lap_time_s"] - s["best_lap_s"]) < 0.002)
            fill = C_BEST_BG if is_best else (C_OUT_BG if lap["is_outlap"] else None)

            vals = [
                s.get("round",""), s.get("circuit",""), s.get("date",""), s.get("session_type",""),
                s.get("rider",""), s.get("run_no",1),
                lap.get("lap_no",0), lap.get("lap_time",""), lap.get("lap_time_s",""),
                "YES" if lap.get("is_outlap") else "",
                s.get("condition",""), s.get("air_temp",""), s.get("track_temp",""),
                s.get("tyre_f",""), s.get("tyre_r","")
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.fill = _fill(fill) if fill else _fill(C_ALT_BG if row % 2 == 0 else "FFFFFF")
                c.font = _font(bold=is_best, size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _border()
            row += 1

    _set_col_widths(ws, [10, 14, 12, 10, 8, 6, 6, 12, 10, 9, 10, 8, 9, 8, 8])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:O{row - 1}"
    return ws


def write_session_summary_sheet(wb, sessions):
    ws = wb.create_sheet("SESSION_SUMMARY")

    ws.merge_cells("A1:T1")
    ws["A1"] = "TS24 PUCCETTI RACING — Session Summary"
    ws["A1"].font  = _font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill  = _fill(C_HEADER_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    headers = [
        "ROUND", "CIRCUIT", "DATE", "SESSION", "RIDER", "RUN",
        "LAPS", "BEST LAP", "BEST (s)", "AVG LAP (s)",
        "WEATHER", "AIR °C", "TRACK °C",
        "FORK", "FORK SPEC", "F COMP", "F REB",
        "SHOCK", "SHOCK SPEC", "OFFSET",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    for i, s in enumerate(sessions):
        row = i + 3
        vals = [
            s.get("round",""), s.get("circuit",""), s.get("date",""), s.get("session_type",""),
            s.get("rider",""), s.get("run_no",1),
            s.get("total_laps",0), s.get("best_lap",""), s.get("best_lap_s",""), s.get("avg_lap_s",""),
            s.get("condition",""), s.get("air_temp",""), s.get("track_temp",""),
            s.get("fork",""), s.get("fork_spec",""), s.get("fork_comp",""), s.get("fork_reb",""),
            s.get("shock",""), s.get("shock_spec",""), s.get("offset",""),
        ]
        bg = C_ALT_BG if row % 2 == 0 else "FFFFFF"
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()

    _set_col_widths(ws, [10,14,12,10,8,6, 7,12,10,12, 10,8,9, 8,12,8,8, 8,12,8])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:T{len(sessions) + 2}"
    return ws


def write_tyre_log_sheet(wb, sessions):
    ws = wb.create_sheet("TYRE_LOG")

    ws.merge_cells("A1:N1")
    ws["A1"] = "TS24 PUCCETTI RACING — Tyre Usage Log (Auto-imported from 2D)"
    ws["A1"].font  = _font(bold=True, color=C_HEADER_FG, size=13)
    ws["A1"].fill  = _fill(C_HEADER_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    headers = [
        "ROUND", "CIRCUIT", "DATE", "SESSION", "RIDER", "RUN",
        "WEATHER", "TRACK °C",
        "FRONT COMPOUND", "F LAPS USED", "F PRESS OUT",  "F WARM TEMP",
        "REAR COMPOUND",  "R LAPS USED", "R PRESS OUT",  "R WARM TEMP",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 22

    for i, s in enumerate(sessions):
        row = i + 3
        vals = [
            s.get("round",""), s.get("circuit",""), s.get("date",""), s.get("session_type",""),
            s.get("rider",""), s.get("run_no",1),
            s.get("condition",""), s.get("track_temp",""),
            s.get("tyre_f",""), s.get("tyre_f_laps",""), s.get("tyre_f_press",""), s.get("tyre_f_temp",""),
            s.get("tyre_r",""), s.get("tyre_r_laps",""), s.get("tyre_r_press",""), s.get("tyre_r_temp",""),
        ]
        bg = C_ALT_BG if row % 2 == 0 else "FFFFFF"
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()

    _set_col_widths(ws, [10,14,12,10,8,6, 10,9, 16,12,12,12, 16,12,12,12])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:P{len(sessions) + 2}"
    return ws


def copy_sheet(src_wb, dst_wb, sheet_name):
    """Copy sheet from src to dst workbook (data only, no formulas)."""
    if sheet_name not in src_wb.sheetnames:
        return
    src = src_wb[sheet_name]
    dst = dst_wb.create_sheet(sheet_name)

    for row in src.iter_rows():
        for cell in row:
            nc = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    nc.font      = copy(cell.font)
                    nc.fill      = copy(cell.fill)
                    nc.border    = copy(cell.border)
                    nc.alignment = copy(cell.alignment)
                    nc.number_format = cell.number_format
                except Exception:
                    pass

    for col_dim in src.column_dimensions.values():
        dst.column_dimensions[col_dim.index].width = col_dim.width
    for row_dim in src.row_dimensions.values():
        dst.row_dimensions[row_dim.index].height = row_dim.height

    for mc in src.merged_cells.ranges:
        dst.merge_cells(str(mc))

    # Freeze panes: override to safe header rows to prevent scroll lockout
    SAFE_FREEZE = {"RUN_LOG": "A4", "DB_LOG": "A5", "TREND_ANALYSIS": "A4",
                   "SOLUTION_SEARCH": "A4", "REPORT": "A4"}
    dst.freeze_panes = SAFE_FREEZE.get(sheet_name, src.freeze_panes)


def _session_key(s):
    """Composite dedup key for a session dict."""
    return (s.get("rider",""), s.get("date",""), s.get("circuit",""),
            s.get("session_type",""), s.get("run_no",0),
            round(float(s.get("best_lap_s") or 0), 3))


def load_cache(cache_path):
    """Load historical sessions from JSON cache file."""
    p = Path(cache_path)
    if p.exists():
        try:
            with open(p, encoding="utf-8") as f:
                data = json.load(f)
            print(f"Loaded {len(data)} sessions from cache: {p.name}")
            return data
        except Exception as e:
            print(f"WARNING: Could not load cache ({e}), starting fresh")
    return []


def save_cache(sessions, cache_path):
    """Save all sessions to JSON cache file."""
    p = Path(cache_path)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(sessions, f, ensure_ascii=False, indent=2)
    print(f"Cache saved: {p.name}  ({len(sessions)} sessions)")


def build_excel(root_folder, excel_path, cache_path=None):
    root_folder = Path(root_folder)
    excel_path  = Path(excel_path)

    # Default cache alongside the Excel file
    if cache_path is None:
        cache_path = excel_path.parent / "all_sessions.json"
    cache_path = Path(cache_path)

    print(f"\n{'='*60}")
    print(f"  TS24 Puccetti — 2D Data Import")
    print(f"  Root : {root_folder}")
    print(f"  Excel: {excel_path}")
    print(f"  Cache: {cache_path}")
    print(f"{'='*60}\n")

    # ── 1. Load historical sessions from JSON cache ───────────────
    cached_sessions = load_cache(cache_path)
    cached_keys = {_session_key(s) for s in cached_sessions}

    # ── 2. Scan new MES folders ───────────────────────────────────
    mes_list = scan_all_mes(root_folder)
    print(f"\nFound {len(mes_list)} MES folders to scan\n")

    new_sessions = []
    skip_count   = 0
    for round_name, rider_folder, mes_path in mes_list:
        print(f"  Parsing: {round_name} / {mes_path.name} ... ", end="", flush=True)
        try:
            s = extract_session(round_name, mes_path)
            key = _session_key(s)
            if key in cached_keys:
                print(f"SKIP (already in cache)")
                skip_count += 1
            else:
                new_sessions.append(s)
                cached_keys.add(key)
                print(f"NEW  ({s['rider']} {s['session_type']} run{s['run_no']}"
                      f"  {len(s['laps'])} laps  best={s['best_lap']})")
        except Exception as e:
            print(f"ERROR: {e}")

    print(f"\nNew sessions found : {len(new_sessions)}")
    print(f"Skipped (duplicate): {skip_count}")

    # ── 3. Merge ──────────────────────────────────────────────────
    all_sessions = cached_sessions + new_sessions

    # Sort by date → round → rider → session type → run
    def _sort_key(s):
        return (s.get("date",""), s.get("round",""), s.get("rider",""),
                s.get("session_type",""), s.get("run_no",0))
    all_sessions.sort(key=_sort_key)

    total_laps = sum(len(s["laps"]) for s in all_sessions)
    print(f"\nTotal sessions: {len(all_sessions)}  |  Total laps: {total_laps}")

    # ── 4. Save updated cache ─────────────────────────────────────
    if new_sessions:
        save_cache(all_sessions, cache_path)
    else:
        print("No new sessions — cache unchanged")

    # ── 5. Load existing Excel (for RUN_LOG etc.) ─────────────────
    keep_sheets = ["RUN_LOG", "DB_LOG", "TREND_ANALYSIS",
                   "SOLUTION_SEARCH", "REPORT"]
    src_wb = None
    if excel_path.exists():
        print(f"\nLoading existing Excel: {excel_path.name}")
        src_wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # ── 6. Build new workbook ─────────────────────────────────────
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    print("\nWriting LAP_TIMES sheet...")
    write_lap_times_sheet(new_wb, all_sessions)

    print("Writing SESSION_SUMMARY sheet...")
    write_session_summary_sheet(new_wb, all_sessions)

    print("Writing TYRE_LOG sheet...")
    write_tyre_log_sheet(new_wb, all_sessions)

    if src_wb:
        for sname in keep_sheets:
            if sname in src_wb.sheetnames:
                print(f"Copying existing sheet: {sname}")
                copy_sheet(src_wb, new_wb, sname)

    # ── 7. Save Excel ─────────────────────────────────────────────
    print(f"\nSaving → {excel_path}")
    new_wb.save(str(excel_path))
    print(f"\n✅ Done!  {len(all_sessions)} sessions ({len(new_sessions)} new), "
          f"{total_laps} laps written.\n")


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

if __name__ == "__main__":
    script_dir = Path(__file__).parent.resolve()

    # Usage:
    #   parse_2d_to_excel.py [ROOT_FOLDER] [EXCEL_PATH] [CACHE_JSON]
    #
    # ROOT_FOLDER  : folder containing new round data to scan
    #                (pass "NONE" or omit to run cache-only rebuild)
    # EXCEL_PATH   : path to PUCCETTI_DB_MASTER.xlsx
    # CACHE_JSON   : path to all_sessions.json (default: 02_DATABASE/all_sessions.json)

    root_arg = sys.argv[1] if len(sys.argv) >= 2 else None
    excel    = Path(sys.argv[2]) if len(sys.argv) >= 3 else \
               script_dir.parent / "02_DATABASE" / "PUCCETTI_DB_MASTER.xlsx"
    cache    = Path(sys.argv[3]) if len(sys.argv) >= 4 else \
               script_dir.parent / "02_DATABASE" / "all_sessions.json"

    # Determine root scan folder
    if root_arg and root_arg.upper() != "NONE" and Path(root_arg).exists():
        root = Path(root_arg)
    else:
        # No scan folder specified → use a non-existent temp path so scan returns 0
        # (Excel is rebuilt purely from cache)
        root = Path("/tmp/__no_scan__")
        if root_arg and root_arg.upper() != "NONE":
            print(f"WARNING: Scan folder not found: {root_arg} — rebuilding from cache only")

    build_excel(root, excel, cache)
