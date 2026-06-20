#!/usr/bin/env python3
"""
Company BSB レポートインポートスクリプト
=========================================
BSBフォーマット（DH55等）のレポートExcelを SQLite ts24_unified.db にインポートする。

ファイル名フォーマット: YYYYMMDD-ROUNDx-RIDER-BSB.xlsx
  例: 20260501-ROUND1-DH55-BSB.xlsx

使い方:
  python import_company_bsb.py 20260501-ROUND1-DH55-BSB.xlsx
  python import_company_bsb.py  (← 01_REPORTS/COMPANY/ 内の未処理ファイルを自動検出)
"""

import re
import sys
import math
import sqlite3
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ─── パス設定 ─────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
DB_DIR     = SCRIPT_DIR.parent / "02_DATABASE"
REPORT_DIR = SCRIPT_DIR.parent / "01_REPORTS" / "COMPANY"
DB_PATH    = DB_DIR / "ts24_unified.db"

DATA_SCOPE = "COMPANY"  # このスクリプトで挿入する全行に付与するスコープ

# ─── BSB → TS24 セッション名マッピング ────────────────
SESSION_MAP = {
    "FP":     "FP",
    "QP":     "QP",
    "Q":      "QP",
    "SPRINT": "SPRINT",
    "WUP":    "WUP",
    "RACE":   "RACE",
    "RACE1":  "RACE1",
    "RACE2":  "RACE2",
}

# ─── BSB フォーマット: 行インデックス (0-based) ────────
# TS24と同一の行: 10-30 (FORK TYPE 〜 SWING ARM)
# TS24と異なる行: 31以降 (BSBにSD/GR/FUELが追加されるため+3オフセット)
BSB_FIELD_ROWS = {
    #  row_idx  (0-indexed = row番号 - 1)
    'FORK_TYPE':  (10, 'p'),   # row11 primary
    'F_SET':      (11, 'p'),   # row12 primary "C104 R104"
    'F_TOS_SPR':  (12, 'p'),   # row13 primary ("2.7X60")
    'F_TOS_SEC':  (12, 's'),   # row13 secondary (right-leg code or same)
    'F_SPR_L':    (13, 'p'),   # row14 primary
    'F_SPR_R':    (13, 's'),   # row14 secondary
    'F_PRELOAD':  (14, 'p'),   # row15
    'F_OIL':      (15, 'p'),   # row16 ("30CC" → 数値部分のみ抽出)
    'F_COMP':     (16, 'p'),   # row17
    'F_REB':      (17, 'p'),   # row18
    'F_OFFSET':   (18, 'p'),   # row19
    'F_HGT_TOP':  (19, 'p'),   # row20 primary (TOP)
    # BSB row20 には BOTTOM がない（primary=TOP のみ）
    'SHOCK_TYPE': (20, 'p'),   # row21
    'SHOCK_SET':  (21, 'p'),   # row22 "C7 R7"
    'R_SPR':      (22, 'p'),   # row23
    'R_PRELOAD':  (23, 'p'),   # row24
    'R_COMP':     (24, 'p'),   # row25
    'R_REB':      (25, 'p'),   # row26
    # row27: BSBはprimary=length, secondary=spring_dia (TS24と逆!)
    'R_TOS_LEN':  (26, 'p'),   # row27 primary → length (mm)
    'R_TOS_SPR':  (26, 's'),   # row27 secondary → spring code
    'SHOCK_LEN':  (27, 'p'),   # row28
    'LINK':       (28, 'p'),   # row29 (BSBは "OEM" 等テキスト)
    'RIDE_HGT':   (29, 'p'),   # row30
    'SWING_ARM':  (30, 'p'),   # row31
    # ── BSB固有行 (row32-34): スキップ ──
    #   row32: STEERING DAMPER → DBには保存しない（スキーマ外）
    #   row33: GEAR RATIO      → DBには保存しない
    #   row34: FUEL            → DBには保存しない
    # ── タイヤ: BSBはTS24より+3行 ──
    'F_TYRE':     (34, 'p'),   # row35 primary
    'R_TYRE':     (36, 'p'),   # row37 primary
    'LAPS':       (38, 'p'),   # row39
    'FASTEST_LAP':(39, 'p'),   # row40
    # COMMENTS: row51 (0-indexed=50)
    'COMMENT':    (50, 'p'),   # row51
}


# ─── ユーティリティ関数 ────────────────────────────────

def _clean(s):
    """スペース・特殊文字を除去してIDに使える文字列に変換"""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper().strip())


def _f(val):
    """float変換 (失敗→None)"""
    if val is None:
        return None
    try:
        v = float(str(val).replace(",", "").replace("CC", "").strip())
        return None if math.isnan(v) else v
    except (ValueError, TypeError):
        return None


def _i(val):
    """int変換 (失敗→None)"""
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _v(val):
    """None/空文字→None、それ以外は文字列"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_laptime(val):
    """
    ラップタイムを "M:SS.mmm" 形式に正規化する。
    BSBは "1,37,673" (カンマ区切り) または "1.37.673" (ドット区切り) 形式。
    TS24は "1:37.623" 形式。
    """
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # カンマ区切り: "1,37,673" → "1:37.673"
    m = re.match(r'^(\d+)[,.](\d{2})[,.](\d+)$', s)
    if m:
        return f"{m.group(1)}:{m.group(2)}.{m.group(3)}"
    # すでに "1:37.623" 形式
    if re.match(r'^\d+:\d{2}\.\d+$', s):
        return s
    return _v(val)


def laptime_to_seconds(lap_str):
    """ラップタイム文字列を秒数 (float) に変換"""
    if not lap_str:
        return None
    m = re.match(r'^(\d+):(\d{2})\.(\d+)$', str(lap_str))
    if m:
        mins = int(m.group(1))
        secs = int(m.group(2))
        frac = float("0." + m.group(3))
        return mins * 60 + secs + frac
    return None


def parse_setting(setting_str):
    """
    "C104 R104" → (104, 104)
    "C7 R7"     → (7, 7)
    """
    if not setting_str:
        return None, None
    m = re.search(r'C(\d+(?:\.\d+)?)', str(setting_str), re.IGNORECASE)
    c_val = float(m.group(1)) if m else None
    m = re.search(r'R(\d+(?:\.\d+)?)', str(setting_str), re.IGNORECASE)
    r_val = float(m.group(1)) if m else None
    return c_val, r_val


def parse_offset(offset_str):
    """
    "26/-0.5" → (26, -0.5)
    "30"      → (30, None)
    """
    if not offset_str:
        return None, None
    s = str(offset_str).strip()
    m = re.match(r'^([+-]?\d+(?:\.\d+)?)[/,]([+-]?\d+(?:\.\d+)?)$', s)
    if m:
        return _f(m.group(1)), _f(m.group(2))
    return _f(s), None


def extract_rider_from_filename(fname):
    """
    "20260501-ROUND1-DH55-BSB.xlsx" → "DH55"
    ファイル名の3番目のパーツをライダーコードとして使用。
    """
    stem = Path(fname).stem  # "20260501-ROUND1-DH55-BSB"
    parts = stem.split('-')
    if len(parts) >= 3:
        # "-BSB" サフィックスがある場合はその直前
        if parts[-1].upper() == 'BSB' and len(parts) >= 4:
            return parts[-2].upper()
        return parts[2].upper()
    return "UNK"


def make_run_id(round_val, circuit, session, rider, run_no):
    r  = _clean(round_val) if round_val else "UNK"
    c  = _clean(circuit)   if circuit   else "UNK"
    s  = _clean(session)   if session   else "UNK"
    ri = _clean(rider)     if rider     else "UNK"
    try:
        n = int(float(str(run_no)))
    except (ValueError, TypeError):
        n = 1
    return f"{r}_{c}_{s}_{ri}_R{n}"


def make_session_id(round_val, circuit, session, rider):
    r  = _clean(round_val) if round_val else "UNK"
    c  = _clean(circuit)   if circuit   else "UNK"
    s  = _clean(session)   if session   else "UNK"
    ri = _clean(rider)     if rider     else "UNK"
    return f"{r}_{c}_{s}_{ri}"


# ─── BSBシート読み取り ─────────────────────────────────

def parse_run_headers(row7, start_col=3):
    """
    row7 (0-indexed) からrunヘッダーとそのprimary/secondary列を解析。

    BSB FP シートの例:
      col3: "1.0" (primary), col4: "9.40" (secondary)
      col5: "1.1" (primary), col6: None (secondary)
      col7: "2.0" (primary), col8: None (secondary)
      col9: "2.1" (primary), col10: None (secondary)

    BSB QP/SPRINT/WUP/RACE の例:
      col3: "Q1" (primary), col4: "1.0" (secondary)

    返り値: [(run_label, col_primary, col_secondary), ...]
    """
    runs = []
    i = start_col
    while i < len(row7):
        val = row7[i]
        if val is None:
            i += 1
            continue
        val_str = str(val).strip()
        # 末尾チェック (WET/SET などはスキップ)
        if val_str.upper() in ('WET', 'SET', 'WET SET', ''):
            i += 1
            continue
        # primary列として登録
        col_primary = i
        col_secondary = i + 1 if (i + 1) < len(row7) else None
        runs.append((val_str, col_primary, col_secondary))
        i += 2  # primary + secondary = 2列ずつ進む
    return runs


def read_bsb_sheet(ws, round_val, circuit, session_type, rider, date_str):
    """
    BSBシート1枚からランデータを読み取る。
    キャリーフォワードを適用して各ランの完全なセットアップを返す。

    返り値: [dict, dict, ...] (1ランにつき1dict)
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # row7 (idx=6) からrunヘッダーを解析
    row7 = all_rows[6] if len(all_rows) > 6 else []

    # STARTとなるベースライン列 (col3) と実際のランを解析
    run_headers = parse_run_headers(row7, start_col=3)

    if not run_headers:
        return []

    # キャリーフォワード用の初期値 (START列 = 最初のヘッダー)
    carry = {}

    # weather はrow8 (idx=7)
    row8 = all_rows[7] if len(all_rows) > 7 else []

    # track_temp, air_temp はrow9 (idx=8)
    row9 = all_rows[8] if len(all_rows) > 8 else []

    # ─ キャリーフォワードしながら各ランを構築 ─
    results = []

    for run_no, (run_label, col_p, col_s) in enumerate(run_headers, start=1):
        def get_p(row_idx):
            """primary列の値を取得"""
            if row_idx >= len(all_rows):
                return None
            row = all_rows[row_idx]
            return row[col_p] if col_p < len(row) else None

        def get_s(row_idx):
            """secondary列の値を取得"""
            if col_s is None or row_idx >= len(all_rows):
                return None
            row = all_rows[row_idx]
            return row[col_s] if col_s < len(row) else None

        # キャリーフォワード: 空白なら直前のrun値を引き継ぐ
        def cf(key, new_val):
            v = _v(new_val)
            if v is not None:
                carry[key] = v
            return carry.get(key)

        # ── セットアップ読み取り ──
        weather     = cf('weather',    get_p(7))   # row8
        track_temp  = cf('track_temp', all_rows[8][col_p] if len(all_rows) > 8 and col_p < len(all_rows[8]) else None)
        air_temp    = cf('air_temp',   all_rows[8][col_s] if col_s and len(all_rows) > 8 and col_s < len(all_rows[8]) else None)

        fork_type   = cf('fork_type',  get_p(10))
        f_set_raw   = cf('f_set',      get_p(11))
        f_set_c, f_set_r = parse_setting(f_set_raw)

        f_tos_spr   = cf('f_tos_spr',  get_p(12))
        f_tos_len   = None  # BSBはTOS LengthをF_TOSの別列に持たない

        f_spr_l     = cf('f_spr_l',    get_p(13))
        f_spr_r     = cf('f_spr_r',    get_s(13))
        f_preload   = cf('f_preload',  get_p(14))
        f_oil_raw   = cf('f_oil',      get_p(15))
        f_comp      = cf('f_comp',     get_p(16))
        f_reb       = cf('f_reb',      get_p(17))
        f_offset_raw= cf('f_offset_r', get_p(18))
        f_hgt_top   = cf('f_hgt_top',  get_p(19))

        shock_type  = cf('shock_type', get_p(20))
        shock_set   = cf('shock_set',  get_p(21))
        r_set_c, r_set_r = parse_setting(shock_set)

        r_spr       = cf('r_spr',      get_p(22))
        r_preload   = cf('r_preload',  get_p(23))
        r_comp      = cf('r_comp',     get_p(24))
        r_reb       = cf('r_reb',      get_p(25))

        # row27: BSBはprimary=length, secondary=spring_code (TS24と逆!)
        r_tos_len   = cf('r_tos_len',  get_p(26))
        r_tos_spr   = cf('r_tos_spr',  get_s(26))

        shock_len   = cf('shock_len',  get_p(27))
        link_raw    = cf('link',       get_p(28))
        ride_hgt    = cf('ride_hgt',   get_p(29))
        swing_arm   = cf('swing_arm',  get_p(30))

        # タイヤ (row35/37 = idx34/36)
        f_tyre      = cf('f_tyre',     get_p(34))
        r_tyre      = cf('r_tyre',     get_p(36))

        # ラップ情報 (row39/40 = idx38/39)
        laps        = cf('laps',       get_p(38))
        fastest_lap = cf('fastest_lap', get_p(39))

        # コメント (row51 = idx50)
        comment     = cf('comment',    get_p(50))

        # ── 変換・正規化 ──
        f_offset, f_offset2 = parse_offset(f_offset_raw)

        # F_OIL: "30CC" → 30 (数値のみ抽出)
        f_oil_lvl = None
        if f_oil_raw:
            m = re.search(r'(\d+(?:\.\d+)?)', str(f_oil_raw))
            f_oil_lvl = float(m.group(1)) if m else None

        # LINK: "OEM" はテキストなのでNULLに変換 (DBはREAL型)
        link_val = _f(link_raw) if link_raw and link_raw.upper() != 'OEM' else None

        fastest_lap_norm = normalize_laptime(fastest_lap)
        lap_s = laptime_to_seconds(fastest_lap_norm)

        # RUN_ID 生成
        run_id    = make_run_id(round_val, circuit, session_type, rider, run_no)
        session_id = make_session_id(round_val, circuit, session_type, rider)

        results.append({
            'run_id':      run_id,
            'session_id':  session_id,
            'round':       round_val,
            'circuit':     circuit,
            'session':     session_type,
            'rider':       rider,
            'run_no':      run_no,
            'date':        date_str,
            'run_label':   run_label,   # 元のBSBラベル (参考用)
            'weather':     weather,
            'track_temp':  _f(track_temp),
            'air_temp':    _f(air_temp),
            'fork_type':   fork_type,
            'f_set_c':     _f(f_set_c),
            'f_set_r':     _f(f_set_r),
            'f_tos_spr':   f_tos_spr,
            'f_tos_len':   _f(f_tos_len),
            'f_spr_l':     _f(f_spr_l),
            'f_spr_r':     _f(f_spr_r),
            'f_preload':   _f(f_preload),
            'f_oil_lvl':   f_oil_lvl,
            'f_comp':      _f(f_comp),
            'f_reb':       _f(f_reb),
            'f_offset':    _f(f_offset),
            'f_offset2':   _f(f_offset2),
            'f_hgt_top':   _f(f_hgt_top),
            'f_hgt_bot':   None,  # BSBにはBOTTOM値なし
            'shock_type':  shock_type,
            'r_set_c':     _f(r_set_c),
            'r_set_r':     _f(r_set_r),
            'r_spr':       _f(r_spr),
            'r_preload':   _f(r_preload),
            'r_comp':      _f(r_comp),
            'r_reb':       _f(r_reb),
            'r_tos_spr':   _v(r_tos_spr),
            'r_tos_len':   _f(r_tos_len),
            'shock_len':   _f(shock_len),
            'link':        link_val,
            'ride_hgt':    _f(ride_hgt),
            'swing_arm':   _f(swing_arm),
            'tyre_front':  f_tyre,
            'tyre_rear':   r_tyre,
            'perf_best_lap': fastest_lap_norm,
            'perf_best_lap_s': lap_s,
            'perf_n_laps': _i(laps),
            'comment':     comment,
            'data_scope':  DATA_SCOPE,
        })

    return results


def import_bsb_file(filepath):
    """BSBレポートファイル1件をDBにインポートする"""
    filepath = Path(filepath)
    print(f"\n{'='*60}")
    print(f"インポート: {filepath.name}")
    print(f"{'='*60}")

    # ─ ファイル名からメタ情報を抽出 ─
    stem = filepath.stem  # e.g. "20260501-ROUND1-DH55-BSB"
    parts = stem.split('-')
    if len(parts) < 3:
        print(f"  [ERROR] ファイル名フォーマット不正: {filepath.name}")
        print(f"          期待値: YYYYMMDD-ROUNDx-RIDER-BSB.xlsx")
        return False

    date_str  = f"{parts[0][:4]}-{parts[0][4:6]}-{parts[0][6:8]}"
    round_val = parts[1].upper()  # "ROUND1"
    rider     = extract_rider_from_filename(filepath.name)

    print(f"  DATE:  {date_str}")
    print(f"  ROUND: {round_val}")
    print(f"  RIDER: {rider}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets_found = wb.sheetnames

    # ─ CIRCUITをFPシート(またはQP等)から読み取り ─
    circuit_raw = None
    for sname in ['FP', 'QP', 'SPRINT', 'WUP', 'RACE']:
        if sname in sheets_found:
            ws = wb[sname]
            row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            circuit_raw = row2[7] if len(row2) > 7 else None
            if circuit_raw:
                break

    if not circuit_raw:
        print(f"  [WARN] CIRCUIT名を取得できませんでした。ファイルを確認してください。")
        circuit_raw = "UNKNOWN"

    circuit = _clean(circuit_raw)  # "Oulton Park" → "OULTONPARK"
    print(f"  CIRCUIT: {circuit} (元: {circuit_raw})")

    # ─ 全セッションシートを読み取り ─
    all_runs = []
    all_sessions = []

    target_sheets = [s for s in ['FP', 'QP', 'SPRINT', 'WUP', 'RACE'] if s in sheets_found]
    print(f"  処理対象シート: {target_sheets}")

    for sheet_name in target_sheets:
        session_type = SESSION_MAP.get(sheet_name.upper(), sheet_name.upper())
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        sheet_date = str(rows[3][3] or date_str).split(" ")[0] if len(rows) > 3 else date_str
        if sheet_date.startswith("2") and len(sheet_date) == 10:
            pass
        else:
            sheet_date = date_str

        # ライダーIDをシートのFILE NAMEからも確認
        file_name_cell = rows[4][7] if len(rows) > 4 and len(rows[4]) > 7 else None
        if file_name_cell:
            # "DH55" or "HARRISON 55" → "DH55" に変換
            fn_str = str(file_name_cell).strip().upper()
            # ライダーコードっぽい部分を抽出 (2-4文字のアルファベット + 数字)
            m = re.search(r'\b([A-Z]{1,3}\d{1,3})\b', fn_str)
            if m:
                rider_from_sheet = m.group(1)
                if rider_from_sheet != rider:
                    print(f"    [INFO] {sheet_name}: FILE_NAME={fn_str} → ライダー={rider_from_sheet} (ファイル名={rider}を使用)")

        print(f"\n  [{sheet_name}] → session={session_type}, date={sheet_date}")

        runs = read_bsb_sheet(wb[sheet_name], round_val, circuit,
                              session_type, rider, sheet_date)
        print(f"    {len(runs)} ランを読み取り")
        for r in runs:
            print(f"      RUN_ID: {r['run_id']} | F_COMP={r['f_comp']} R_COMP={r['r_comp']} | {r['run_label']}")

        all_runs.extend(runs)

        # SESSION_SUMMARY用データを構築
        best = None
        best_s = None
        total_laps = 0
        for run in runs:
            if run['perf_best_lap_s']:
                if best_s is None or run['perf_best_lap_s'] < best_s:
                    best = run['perf_best_lap']
                    best_s = run['perf_best_lap_s']
            if run['perf_n_laps']:
                total_laps += run['perf_n_laps']

        weather_first = runs[0]['weather'] if runs else None
        track_temp_first = runs[0]['track_temp'] if runs else None
        air_temp_first = runs[0]['air_temp'] if runs else None

        sid = make_session_id(round_val, circuit, session_type, rider)
        all_sessions.append({
            'session_id': sid,
            'event_id':   f"{_clean(round_val)}_{circuit}",
            'round':      round_val,
            'circuit':    circuit,
            'session':    session_type,
            'rider':      rider,
            'date':       sheet_date,
            'total_laps': total_laps or None,
            'best_lap':   best,
            'best_lap_s': best_s,
            'weather':    weather_first,
            'track_temp': track_temp_first,
            'air_temp':   air_temp_first,
            'data_scope': DATA_SCOPE,
        })

    if not all_runs:
        print("\n  [WARN] インポートするランがありません。")
        return False

    # ─ SQLite DB に書き込み ─
    print(f"\n  DB書き込み中: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    inserted_runs = 0
    skipped_runs = 0

    # events テーブル
    event_id = f"{_clean(round_val)}_{circuit}"
    cur.execute("INSERT OR IGNORE INTO events(event_id, round, circuit) VALUES (?,?,?)",
                (event_id, round_val, circuit))

    # ts24_sessions テーブル
    for sess in all_sessions:
        existing = cur.execute("SELECT session_id FROM ts24_sessions WHERE session_id=?",
                               (sess['session_id'],)).fetchone()
        if existing:
            print(f"    [SKIP] session already exists: {sess['session_id']}")
        else:
            cur.execute("""
                INSERT INTO ts24_sessions
                (session_id, event_id, round, circuit, session, rider, date,
                 total_laps, best_lap, best_lap_s, weather, air_temp, track_temp,
                 data_scope)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                sess['session_id'], sess['event_id'], sess['round'], sess['circuit'],
                sess['session'], sess['rider'], sess['date'],
                sess['total_laps'], sess['best_lap'], sess['best_lap_s'],
                sess['weather'], sess['air_temp'], sess['track_temp'],
                DATA_SCOPE,
            ))
            print(f"    [OK] session: {sess['session_id']}")

    # runs テーブル
    for run in all_runs:
        existing = cur.execute("SELECT run_id FROM runs WHERE run_id=?",
                               (run['run_id'],)).fetchone()
        if existing:
            print(f"    [SKIP] run already exists: {run['run_id']}")
            skipped_runs += 1
            continue

        cur.execute("""
            INSERT INTO runs
            (run_id, session_id, round, circuit, session, rider, run_no, date,
             weather, track_temp, air_temp,
             fork_type, f_set_c, f_set_r, f_tos_spr, f_tos_len,
             f_spr_l, f_spr_r, f_preload, f_oil_lvl, f_comp, f_reb,
             f_offset, f_offset2, f_hgt_top, f_hgt_bot,
             shock_type, r_set_c, r_set_r, r_spr, r_preload, r_comp, r_reb,
             r_tos_spr, r_tos_len, shock_len, link, ride_hgt, swing_arm,
             tyre_front, tyre_rear,
             perf_best_lap, perf_n_laps,
             comment, data_scope,
             created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            run['run_id'], run['session_id'], run['round'], run['circuit'],
            run['session'], run['rider'], run['run_no'], run['date'],
            run['weather'], run['track_temp'], run['air_temp'],
            run['fork_type'], run['f_set_c'], run['f_set_r'],
            run['f_tos_spr'], run['f_tos_len'],
            run['f_spr_l'], run['f_spr_r'],
            run['f_preload'], run['f_oil_lvl'],
            run['f_comp'], run['f_reb'],
            run['f_offset'], run['f_offset2'],
            run['f_hgt_top'], run['f_hgt_bot'],
            run['shock_type'], run['r_set_c'], run['r_set_r'],
            run['r_spr'], run['r_preload'], run['r_comp'], run['r_reb'],
            run['r_tos_spr'], run['r_tos_len'],
            run['shock_len'], run['link'], run['ride_hgt'], run['swing_arm'],
            run['tyre_front'], run['tyre_rear'],
            run['perf_best_lap'], run['perf_n_laps'],
            run['comment'], DATA_SCOPE,
            now, now,
        ))
        inserted_runs += 1

    conn.commit()
    conn.close()

    print(f"\n  ✓ 完了: {inserted_runs}ラン挿入, {skipped_runs}ランスキップ")
    return True


def main():
    if len(sys.argv) >= 2:
        # 引数でファイルを指定
        target = Path(sys.argv[1])
        if not target.is_absolute():
            # 相対パスの場合はいくつかの場所を探す
            for base in [Path.cwd(), REPORT_DIR,
                         REPORT_DIR.parent / "JA52",
                         REPORT_DIR.parent / "DA77"]:
                if (base / target).exists():
                    target = base / target
                    break
        if not target.exists():
            print(f"[ERROR] ファイルが見つかりません: {sys.argv[1]}")
            sys.exit(1)
        import_bsb_file(target)
    else:
        # 自動検出: 01_REPORTS/COMPANY/ 内の *.xlsx を処理
        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        files = sorted(REPORT_DIR.glob("*.xlsx"))
        if not files:
            print(f"[INFO] {REPORT_DIR} に処理対象ファイルがありません。")
            print(f"       ファイルを配置するか: python import_company_bsb.py <ファイルパス>")
            sys.exit(0)
        print(f"[INFO] {len(files)} ファイルを処理します。")
        for f in files:
            import_bsb_file(f)

    print("\n完了!")


if __name__ == "__main__":
    main()
