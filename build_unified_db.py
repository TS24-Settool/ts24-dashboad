#!/usr/bin/env python3
"""
TS24 Puccetti Racing — 統合データベース構築スクリプト
======================================================
Excel の全シート (RUN_LOG, TYRE_LOG, SESSION_SUMMARY,
DYNAMICS_ANALYSIS, LAP_TIMES, PERFORMANCE_CORRELATION) を
統一 RUN_ID で紐付け、SQLite に統合する。

また、各 Excel シートに RUN_ID / LAP_ID 列を自動追加する。

実行方法:
  python build_unified_db.py

出力:
  02_DATABASE/ts24_unified.db  ← 統合 SQLite DB
  02_DATABASE/TS24 DB Master.xlsx  ← RUN_ID 列を追加して上書き
"""

import re
import sys
import shutil
import sqlite3
from copy import copy
from pathlib import Path
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install",
                    "openpyxl", "pandas", "openpyxl", "--quiet"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import pandas as pd

# ─────────────────────────────────────────────
#  パス設定
# ─────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
DB_DIR      = SCRIPT_DIR.parent / "02_DATABASE"
EXCEL_PATH  = DB_DIR / "TS24 DB Master.xlsx"
BACKUP_PATH = DB_DIR / "TS24 DB Master Back UP.xlsx"
UNIFIED_DB  = DB_DIR / "ts24_unified.db"

# ID 列のスタイル
ID_COL_BG  = "1F3864"   # dark navy (ヘッダー)
ID_COL_FG  = "FFFFFF"
ID_DATA_BG = "D6E4F0"   # 薄青 (データ)
ID_DATA_FG = "1F3864"


# ─────────────────────────────────────────────
#  RUN_ID 生成
# ─────────────────────────────────────────────
def _clean(s):
    """スペース・特殊文字を除去してIDに使える文字列に変換"""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper().strip())


def make_run_id(round_val, circuit, session, rider, run_no):
    """
    RUN_ID を生成する。
    形式: {ROUND}_{CIRCUIT}_{SESSION}_{RIDER}_R{N}
    例:   R03_ASSEN_FP_JA52_R1
    """
    r  = _clean(round_val)  if round_val  else "UNK"
    c  = _clean(circuit)    if circuit    else "UNK"
    s  = _clean(session)    if session    else "UNK"
    ri = _clean(rider)      if rider      else "UNK"
    try:
        n = int(float(str(run_no)))
    except (ValueError, TypeError):
        n = 1
    return f"{r}_{c}_{s}_{ri}_R{n}"


def make_session_id(round_val, circuit, session, rider):
    """SESSION_ID (ラン番号なし)"""
    r  = _clean(round_val)  if round_val  else "UNK"
    c  = _clean(circuit)    if circuit    else "UNK"
    s  = _clean(session)    if session    else "UNK"
    ri = _clean(rider)      if rider      else "UNK"
    return f"{r}_{c}_{s}_{ri}"


def make_lap_id(run_id, lap_no):
    """LAP_ID: {RUN_ID}_L{N}"""
    try:
        n = int(float(str(lap_no)))
    except (ValueError, TypeError):
        n = 0
    return f"{run_id}_L{n}"


# ─────────────────────────────────────────────
#  SQLite スキーマ定義
# ─────────────────────────────────────────────
SCHEMA_SQL = """
-- ================================================
--  TS24 統合データベース スキーマ
--  中心キー: run_id  例) R03_ASSEN_FP_JA52_R1
-- ================================================

PRAGMA journal_mode = WAL;
PRAGMA foreign_keys = OFF;

-- ── イベント (ラウンド/テスト) ────────────────────
CREATE TABLE IF NOT EXISTS events (
    event_id    TEXT PRIMARY KEY,   -- "R03_ASSEN"
    round       TEXT,               -- "ROUND3", "TEST2"
    circuit     TEXT,               -- "ASSEN"
    event_type  TEXT,               -- "RACE" / "TEST"
    date_start  TEXT,               -- "2026-04-17"
    date_end    TEXT                -- "2026-04-19"
);

-- ── セッション (イベント内の 1セッションタイプ×ライダー) ──
CREATE TABLE IF NOT EXISTS ts24_sessions (
    session_id  TEXT PRIMARY KEY,   -- "R03_ASSEN_FP_JA52"
    event_id    TEXT REFERENCES events(event_id),
    round       TEXT,
    circuit     TEXT,
    session     TEXT,               -- "FP","QP","SP","WUP1","RACE1","RACE2"
    rider       TEXT,               -- "JA52","DA77"
    date        TEXT,               -- "2026-04-17"
    -- SESSION_SUMMARY から
    total_laps  INTEGER,
    best_lap    TEXT,               -- "1:37.455"
    best_lap_s  REAL,
    avg_lap_s   REAL,
    weather     TEXT,
    air_temp    REAL,
    track_temp  REAL
);

-- ── ラン (スティント = 1回のコース走行) ── ★中心テーブル★
CREATE TABLE IF NOT EXISTS runs (
    run_id      TEXT PRIMARY KEY,   -- "R03_ASSEN_FP_JA52_R1"
    session_id  TEXT REFERENCES ts24_sessions(session_id),
    -- 基本
    round       TEXT,
    circuit     TEXT,
    session     TEXT,
    rider       TEXT,
    run_no      INTEGER,
    date        TEXT,
    weather     TEXT,
    track_temp  REAL,
    air_temp    REAL,
    -- ── RUN_LOG: フォーク ──
    fork_type   TEXT,
    f_set_c     REAL,   -- Front Comp setting
    f_set_r     REAL,   -- Front Reb setting
    f_tos_spr   TEXT,   -- TOS Spring dia
    f_tos_len   REAL,   -- TOS Length mm
    f_spr_l     REAL,   -- Front Spring L N/mm
    f_spr_r     REAL,   -- Front Spring R N/mm
    f_preload   REAL,
    f_oil_lvl   REAL,
    f_comp      REAL,
    f_reb       REAL,
    f_offset    REAL,
    f_offset2   REAL,
    f_hgt_top   REAL,
    f_hgt_bot   REAL,
    -- ── RUN_LOG: ショック ──
    shock_type  TEXT,
    r_set_c     REAL,
    r_set_r     REAL,
    r_spr       REAL,
    r_preload   REAL,
    r_comp      REAL,
    r_reb       REAL,
    r_tos_spr   TEXT,
    r_tos_len   REAL,
    shock_len   REAL,
    link        REAL,
    ride_hgt    REAL,
    swing_arm   REAL,
    -- ── TYRE_LOG ──
    tyre_front  TEXT,
    tyre_rear   TEXT,
    f_laps_used TEXT,
    r_laps_used TEXT,
    f_press_out REAL,
    r_press_out REAL,
    f_warm_temp REAL,
    r_warm_temp REAL,
    -- ── DYNAMICS_ANALYSIS ──
    dyn_sr_hz      REAL,
    dyn_laps       INTEGER,
    apex_count     INTEGER,
    apex_spd       REAL,
    apex_sus_f     REAL,
    apex_sus_r     REAL,
    apex_whl_f     REAL,
    apex_whl_r     REAL,
    apex_ax        REAL,
    pit_count      INTEGER,
    pit_spd        REAL,
    pit_sus_f      REAL,
    pit_sus_r      REAL,
    brk_count      INTEGER,
    brk_spd        REAL,
    brk_sus_f      REAL,
    brk_sus_r      REAL,
    fullbrk_cnt    INTEGER,
    fullbrk_sus_f  REAL,
    fullbrk_sus_r  REAL,
    tyre_f_st      REAL,
    tyre_f_en      REAL,
    tyre_f_avg     REAL,
    tyre_f_delta   REAL,
    tyre_r_st      REAL,
    tyre_r_en      REAL,
    tyre_r_avg     REAL,
    tyre_r_delta   REAL,
    -- ── PERFORMANCE_CORRELATION ──
    perf_best_lap  TEXT,
    perf_avg_lap   TEXT,
    perf_n_laps    INTEGER,
    perf_rank      TEXT,
    perf_gap_s     REAL,
    perf_tier      TEXT,
    -- ── RUN_LOG: ノート ──
    comment        TEXT,
    problem_desc   TEXT,
    change_intent  TEXT,
    expected_effect TEXT,
    result_eval    TEXT,
    -- メタ
    created_at     TEXT DEFAULT (datetime('now')),
    updated_at     TEXT DEFAULT (datetime('now'))
);

-- ── ラップ ────────────────────────────────────────
CREATE TABLE IF NOT EXISTS laps (
    lap_id      TEXT PRIMARY KEY,   -- "R03_ASSEN_FP_JA52_R1_L3"
    run_id      TEXT REFERENCES runs(run_id),
    session_id  TEXT REFERENCES ts24_sessions(session_id),
    -- 基本
    round       TEXT,
    circuit     TEXT,
    session     TEXT,
    rider       TEXT,
    run_no      INTEGER,
    lap_no      INTEGER,
    date        TEXT,
    -- タイム
    lap_time    TEXT,               -- "1:37.455"
    lap_time_s  REAL,
    is_outlap   INTEGER DEFAULT 0,
    -- コンディション
    weather     TEXT,
    air_temp    REAL,
    track_temp  REAL,
    tyre_front  TEXT,
    tyre_rear   TEXT
);

-- ── インデックス ──────────────────────────────────
CREATE INDEX IF NOT EXISTS idx_runs_circuit   ON runs(circuit);
CREATE INDEX IF NOT EXISTS idx_runs_rider     ON runs(rider);
CREATE INDEX IF NOT EXISTS idx_runs_session   ON runs(session_id);
CREATE INDEX IF NOT EXISTS idx_laps_run       ON laps(run_id);
CREATE INDEX IF NOT EXISTS idx_laps_circuit   ON laps(circuit);
CREATE INDEX IF NOT EXISTS idx_laps_rider     ON laps(rider);
"""


# ─────────────────────────────────────────────
#  Excel シート読み込み → DataFrameで返す
# ─────────────────────────────────────────────
def load_sheet(wb, sheet_name, header_row=1):
    """openpyxl ワークブックから指定シートを pandas DataFrame に変換"""
    ws = wb[sheet_name]
    data = []
    headers = [ws.cell(row=header_row+1, column=c).value
               for c in range(1, ws.max_column + 1)]
    # ヘッダーの改行・空白を正規化
    headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
               for i, h in enumerate(headers)]
    for r in range(header_row + 2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value
               for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    return pd.DataFrame(data)


def _v(val):
    """None/NaN → None"""
    if val is None:
        return None
    try:
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    return val


def _f(val):
    """float 変換 (失敗→None)"""
    try:
        v = float(str(val).replace(",",""))
        import math
        return None if math.isnan(v) else v
    except Exception:
        return None


def _i(val):
    """int 変換 (失敗→None)"""
    try:
        return int(float(str(val)))
    except Exception:
        return None


# ─────────────────────────────────────────────
#  各シートから run_id 付きデータを構築
# ─────────────────────────────────────────────
def build_run_log_df(wb):
    df = load_sheet(wb, "RUN_LOG", header_row=1)  # row1=title, row2=desc, row3=headers
    # RUN_LOG は header が row 3 → header_row=2
    ws = wb["RUN_LOG"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column+1)]
    headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
               for i, h in enumerate(headers)]
    data = []
    for r in range(4, ws.max_row+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    df = pd.DataFrame(data)
    if df.empty:
        return df
    # round は SESSION_SUMMARY から後で補完するため空欄でOK
    df["round"] = df.get("ROUND", pd.Series([""] * len(df)))
    if "ROUND" not in df.columns:
        df["round"] = ""
    else:
        df["round"] = df["ROUND"].fillna("")
    return df


def build_run_ids_for_sheet(df, round_col, circuit_col, session_col, rider_col, run_col):
    """DataFrame の指定列から RUN_ID 列を生成して追加"""
    run_ids = []
    for _, row in df.iterrows():
        rid = make_run_id(
            row.get(round_col, ""),
            row.get(circuit_col, ""),
            row.get(session_col, ""),
            row.get(rider_col, ""),
            row.get(run_col, 1),
        )
        run_ids.append(rid)
    df = df.copy()
    df["RUN_ID"] = run_ids
    return df


# ─────────────────────────────────────────────
#  Excel シートへ RUN_ID 列を追加
# ─────────────────────────────────────────────
def add_id_column_to_sheet(ws, run_ids, header_row_excel,
                            col_label="RUN_ID", insert_col=1):
    """
    ws に RUN_ID 列を挿入する。
    insert_col: 挿入するカラム位置 (1-based, デフォルト: 先頭)
    """
    # 既に RUN_ID 列がある場合はスキップ
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=header_row_excel, column=c).value or "").upper() == col_label:
            # 上書きモード: 既存の列に値を書き込む
            existing_col = c
            for i, val in enumerate(run_ids, header_row_excel + 1):
                ws.cell(row=i, column=existing_col).value = val
            print(f"    → {col_label} 列を上書き (col {existing_col})")
            return existing_col

    # 新規挿入
    ws.insert_cols(insert_col)
    # タイトル行は結合していることがあるのでそのまま
    hdr_cell = ws.cell(row=header_row_excel, column=insert_col)
    hdr_cell.value = col_label
    hdr_cell.font      = Font(name="Arial", bold=True, color=ID_COL_FG, size=9)
    hdr_cell.fill      = PatternFill("solid", start_color=ID_COL_BG, fgColor=ID_COL_BG)
    hdr_cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    for i, val in enumerate(run_ids, header_row_excel + 1):
        cell = ws.cell(row=i, column=insert_col)
        cell.value = val
        cell.font      = Font(name="Arial", size=8, color=ID_DATA_FG)
        cell.fill      = PatternFill("solid", start_color=ID_DATA_BG,
                                     fgColor=ID_DATA_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # 列幅設定
    ws.column_dimensions[get_column_letter(insert_col)].width = 28
    print(f"    → {col_label} 列を挿入 (col {insert_col})")
    return insert_col


# ─────────────────────────────────────────────
#  統合 SQLite DB 構築
# ─────────────────────────────────────────────
def build_sqlite(conn, wb):
    """全シートを読み込み、統合 DB テーブルを構築する"""
    cur = conn.cursor()

    # ── SESSION_SUMMARY ──────────────────────
    ws_sess = wb["SESSION_SUMMARY"]
    sess_headers = [ws_sess.cell(row=2, column=c).value
                    for c in range(1, ws_sess.max_column+1)]
    sess_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                    for i, h in enumerate(sess_headers)]
    sess_data = []
    for r in range(3, ws_sess.max_row+1):
        row = [ws_sess.cell(row=r, column=c).value
               for c in range(1, ws_sess.max_column+1)]
        if any(v is not None for v in row):
            sess_data.append(dict(zip(sess_headers, row)))
    df_sess = pd.DataFrame(sess_data)

    # ── RUN_LOG ──────────────────────────────
    ws_run = wb["RUN_LOG"]
    run_headers = [ws_run.cell(row=3, column=c).value
                   for c in range(1, ws_run.max_column+1)]
    run_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                   for i, h in enumerate(run_headers)]
    run_data = []
    for r in range(4, ws_run.max_row+1):
        row = [ws_run.cell(row=r, column=c).value
               for c in range(1, ws_run.max_column+1)]
        if any(v is not None for v in row):
            run_data.append(dict(zip(run_headers, row)))
    df_run = pd.DataFrame(run_data)

    # ── TYRE_LOG ─────────────────────────────
    ws_tyre = wb["TYRE_LOG"]
    tyre_headers = [ws_tyre.cell(row=2, column=c).value
                    for c in range(1, ws_tyre.max_column+1)]
    tyre_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                    for i, h in enumerate(tyre_headers)]
    tyre_data = []
    for r in range(3, ws_tyre.max_row+1):
        row = [ws_tyre.cell(row=r, column=c).value
               for c in range(1, ws_tyre.max_column+1)]
        if any(v is not None for v in row):
            tyre_data.append(dict(zip(tyre_headers, row)))
    df_tyre = pd.DataFrame(tyre_data)

    # ── DYNAMICS_ANALYSIS ────────────────────
    ws_dyn = wb["DYNAMICS_ANALYSIS"]
    dyn_headers = [ws_dyn.cell(row=2, column=c).value
                   for c in range(1, ws_dyn.max_column+1)]
    dyn_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                   for i, h in enumerate(dyn_headers)]
    dyn_data = []
    for r in range(3, ws_dyn.max_row+1):
        row = [ws_dyn.cell(row=r, column=c).value
               for c in range(1, ws_dyn.max_column+1)]
        if any(v is not None for v in row):
            dyn_data.append(dict(zip(dyn_headers, row)))
    df_dyn = pd.DataFrame(dyn_data)

    # ── PERFORMANCE_CORRELATION ───────────────
    ws_pc = wb["PERFORMANCE_CORRELATION"]
    pc_headers = [ws_pc.cell(row=2, column=c).value
                  for c in range(1, ws_pc.max_column+1)]
    pc_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                  for i, h in enumerate(pc_headers)]
    pc_data = []
    for r in range(3, ws_pc.max_row+1):
        row = [ws_pc.cell(row=r, column=c).value
               for c in range(1, ws_pc.max_column+1)]
        if any(v is not None for v in row):
            pc_data.append(dict(zip(pc_headers, row)))
    df_pc = pd.DataFrame(pc_data)

    # ── LAP_TIMES ────────────────────────────
    ws_lt = wb["LAP_TIMES"]
    lt_headers = [ws_lt.cell(row=2, column=c).value
                  for c in range(1, ws_lt.max_column+1)]
    lt_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                  for i, h in enumerate(lt_headers)]
    lt_data = []
    for r in range(3, ws_lt.max_row+1):
        row = [ws_lt.cell(row=r, column=c).value
               for c in range(1, ws_lt.max_column+1)]
        if any(v is not None for v in row):
            lt_data.append(dict(zip(lt_headers, row)))
    df_lt = pd.DataFrame(lt_data)

    print(f"  SESSION_SUMMARY: {len(df_sess)} 行")
    print(f"  RUN_LOG:         {len(df_run)} 行")
    print(f"  TYRE_LOG:        {len(df_tyre)} 行")
    print(f"  DYNAMICS:        {len(df_dyn)} 行")
    print(f"  PERF_CORR:       {len(df_pc)} 行")
    print(f"  LAP_TIMES:       {len(df_lt)} 行")

    # ── ts24_sessions テーブル ────────────────
    cur.execute("DELETE FROM ts24_sessions")
    inserted_sessions = set()
    for _, row in df_sess.iterrows():
        rnd  = _v(row.get("ROUND",""))
        circ = _v(row.get("CIRCUIT",""))
        sess = _v(row.get("SESSION",""))
        rider= _v(row.get("RIDER",""))
        if not all([rnd, circ, sess, rider]):
            continue
        sid = make_session_id(rnd, circ, sess, rider)
        if sid in inserted_sessions:
            continue
        inserted_sessions.add(sid)
        eid = f"{_clean(rnd)}_{_clean(circ)}"
        # events テーブルにも挿入
        cur.execute("""
            INSERT OR IGNORE INTO events(event_id, round, circuit)
            VALUES (?,?,?)
        """, (eid, str(rnd), str(circ)))
        date_val = str(row.get("DATE","") or "").split(" ")[0]
        cur.execute("""
            INSERT OR REPLACE INTO ts24_sessions
            (session_id, event_id, round, circuit, session, rider, date,
             total_laps, best_lap, best_lap_s, avg_lap_s, weather, air_temp, track_temp)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            sid, eid, str(rnd), str(circ), str(sess), str(rider), date_val,
            _i(row.get("LAPS")), _v(row.get("BEST LAP")),
            _f(row.get("BEST (s)")), _f(row.get("AVG LAP (s)")),
            _v(row.get("WEATHER")), _f(row.get("AIR °C")), _f(row.get("TRACK °C")),
        ))
    conn.commit()
    print(f"  ts24_sessions: {len(inserted_sessions)} 件挿入")

    # ── runs テーブル ─────────────────────────
    cur.execute("DELETE FROM runs")
    inserted_runs = set()

    # RUN_LOG ベースで runs を作成
    for _, row in df_run.iterrows():
        rnd  = _v(row.get("ROUND",""))  # RUN_LOG には ROUND 列がない場合あり
        circ = _v(row.get("CIRCUIT",""))
        sess = _v(row.get("SESSION",""))
        rider= _v(row.get("RIDER",""))
        rno  = _i(row.get("RUN", 1)) or 1
        if not all([circ, sess, rider]):
            continue
        # ROUND が空の場合は SESSION_SUMMARY から補完
        if not rnd:
            match = df_sess[
                (df_sess["CIRCUIT"].astype(str).str.upper() == str(circ).upper()) &
                (df_sess["SESSION"].astype(str).str.upper() == str(sess).upper()) &
                (df_sess["RIDER"].astype(str).str.upper() == str(rider).upper())
            ]
            rnd = match.iloc[0]["ROUND"] if not match.empty else "UNK"
        run_id = make_run_id(rnd, circ, sess, rider, rno)
        sid    = make_session_id(rnd, circ, sess, rider)
        if run_id in inserted_runs:
            continue
        inserted_runs.add(run_id)
        cur.execute("""
            INSERT OR REPLACE INTO runs
            (run_id, session_id, round, circuit, session, rider, run_no,
             weather, track_temp, air_temp,
             fork_type, f_set_c, f_set_r, f_tos_spr, f_tos_len,
             f_spr_l, f_spr_r, f_preload, f_oil_lvl, f_comp, f_reb,
             f_offset, f_offset2, f_hgt_top, f_hgt_bot,
             shock_type, r_set_c, r_set_r, r_spr, r_preload, r_comp, r_reb,
             r_tos_spr, r_tos_len, shock_len, link, ride_hgt, swing_arm,
             comment, problem_desc, change_intent, expected_effect, result_eval)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            run_id, sid, str(rnd), str(circ), str(sess), str(rider), rno,
            _v(row.get("WEATHER")), _f(row.get("TRACK\nTEMP") or row.get("TRACK TEMP")),
            _f(row.get("AIR\nTEMP") or row.get("AIR TEMP")),
            _v(row.get("FORK\nTYPE") or row.get("FORK TYPE")),
            _f(row.get("F_SET\nC") or row.get("F_SET C")),
            _f(row.get("F_SET\nR") or row.get("F_SET R")),
            _v(row.get("F_TOS\nSPRING") or row.get("F_TOS SPRING")),
            _f(row.get("F_TOS\nLENGTH") or row.get("F_TOS LENGTH")),
            _f(row.get("F_SPR\nL") or row.get("F_SPR L")),
            _f(row.get("F_SPR\nR") or row.get("F_SPR R")),
            _f(row.get("F_PRE\nLOAD") or row.get("F_PRE LOAD")),
            _f(row.get("F_OIL\nLEVEL") or row.get("F_OIL LEVEL")),
            _f(row.get("F_COMP")), _f(row.get("F_REB")),
            _f(row.get("F_OFF\nSET") or row.get("F_OFF SET")),
            _f(row.get("F_OFF\nSET2") or row.get("F_OFF SET2")),
            _f(row.get("F_HGT\nTOP") or row.get("F_HGT TOP")),
            _f(row.get("F_HGT\nBOT") or row.get("F_HGT BOT")),
            _v(row.get("SHOCK\nTYPE") or row.get("SHOCK TYPE")),
            _f(row.get("R_SET\nC") or row.get("R_SET C")),
            _f(row.get("R_SET\nR") or row.get("R_SET R")),
            _f(row.get("R_SPR")), _f(row.get("R_PRE\nLOAD") or row.get("R_PRE LOAD")),
            _f(row.get("R_COMP")), _f(row.get("R_REB")),
            _v(row.get("R_TOS\nSPRING") or row.get("R_TOS SPRING")),
            _f(row.get("R_TOS\nLENGTH") or row.get("R_TOS LENGTH")),
            _f(row.get("SHOCK\nLEN") or row.get("SHOCK LEN")),
            _f(row.get("LINK")), _f(row.get("RIDE\nHGT") or row.get("RIDE HGT")),
            _f(row.get("SWING\nARM") or row.get("SWING ARM")),
            _v(row.get("COMMENT")), _v(row.get("PROBLEM\nDESC") or row.get("PROBLEM DESC")),
            _v(row.get("CHANGE\nINTENT") or row.get("CHANGE INTENT")),
            _v(row.get("EXPECTED\nEFFECT") or row.get("EXPECTED EFFECT")),
            _v(row.get("RESULT\nEVAL") or row.get("RESULT EVAL")),
        ))

    # DYNAMICS から runs を UPDATE
    for _, row in df_dyn.iterrows():
        rnd  = _v(row.get("Round",""))
        circ = _v(row.get("Circuit",""))
        sess = _v(row.get("Session",""))
        rider= _v(row.get("Rider",""))
        rno  = _i(row.get("Run", 1)) or 1
        if not all([circ, sess, rider]):
            continue
        run_id = make_run_id(rnd or "UNK", circ, sess, rider, rno)
        # runs に存在しない場合は INSERT
        cur.execute("INSERT OR IGNORE INTO runs(run_id, session_id, round, circuit, session, rider, run_no) VALUES (?,?,?,?,?,?,?)",
                    (run_id, None, str(rnd or ""), str(circ), str(sess), str(rider), rno))
        cur.execute("""
            UPDATE runs SET
              dyn_sr_hz=?, dyn_laps=?,
              apex_count=?, apex_spd=?, apex_sus_f=?, apex_sus_r=?,
              apex_whl_f=?, apex_whl_r=?, apex_ax=?,
              pit_count=?, pit_spd=?, pit_sus_f=?, pit_sus_r=?,
              brk_count=?, brk_spd=?, brk_sus_f=?, brk_sus_r=?,
              fullbrk_cnt=?, fullbrk_sus_f=?, fullbrk_sus_r=?,
              tyre_f_st=?, tyre_f_en=?, tyre_f_avg=?, tyre_f_delta=?,
              tyre_r_st=?, tyre_r_en=?, tyre_r_avg=?, tyre_r_delta=?,
              updated_at=datetime('now')
            WHERE run_id=?
        """, (
            _f(row.get("SR (Hz)")), _i(row.get("Laps")),
            _i(row.get("APEX Count")), _f(row.get("APEX Spd (km/h)")),
            _f(row.get("APEX SusF (mm)")), _f(row.get("APEX SusR (mm)")),
            _f(row.get("APEX WhlF (N)")), _f(row.get("APEX WhlR (N)")),
            _f(row.get("APEX ax (m/s²)")),
            _i(row.get("Pit Count")), _f(row.get("Pit Spd (km/h)")),
            _f(row.get("Pit SusF (mm)")), _f(row.get("Pit SusR (mm)")),
            _i(row.get("Brk Count")), _f(row.get("Brk Spd (km/h)")),
            _f(row.get("Brk SusF (mm)")), _f(row.get("Brk SusR (mm)")),
            _i(row.get("FullBrk Cnt")),
            _f(row.get("FullBrk SusF")), _f(row.get("FullBrk SusR")),
            _f(row.get("TyreF St(Bar)")), _f(row.get("TyreF En(Bar)")),
            _f(row.get("TyreF Avg(Bar)")), _f(row.get("TyreF Δ(Bar)")),
            _f(row.get("TyreR St(Bar)")), _f(row.get("TyreR En(Bar)")),
            _f(row.get("TyreR Avg(Bar)")), _f(row.get("TyreR Δ(Bar)")),
            run_id,
        ))

    # TYRE_LOG から runs を UPDATE
    for _, row in df_tyre.iterrows():
        rnd  = _v(row.get("ROUND",""))
        circ = _v(row.get("CIRCUIT",""))
        sess = _v(row.get("SESSION",""))
        rider= _v(row.get("RIDER",""))
        rno  = _i(row.get("RUN", 1)) or 1
        if not all([circ, sess, rider]):
            continue
        run_id = make_run_id(rnd or "UNK", circ, sess, rider, rno)
        cur.execute("INSERT OR IGNORE INTO runs(run_id, session_id, round, circuit, session, rider, run_no) VALUES (?,?,?,?,?,?,?)",
                    (run_id, None, str(rnd or ""), str(circ), str(sess), str(rider), rno))
        cur.execute("""
            UPDATE runs SET
              tyre_front=?, tyre_rear=?,
              f_laps_used=?, r_laps_used=?,
              f_press_out=?, r_press_out=?,
              f_warm_temp=?, r_warm_temp=?,
              updated_at=datetime('now')
            WHERE run_id=?
        """, (
            _v(row.get("FRONT COMPOUND")), _v(row.get("REAR COMPOUND")),
            _v(row.get("F LAPS USED")), _v(row.get("R LAPS USED")),
            _f(row.get("F PRESS OUT")), _f(row.get("R PRESS OUT")),
            _f(row.get("F WARM TEMP")), _f(row.get("R WARM TEMP")),
            run_id,
        ))

    # PERF_CORR から runs を UPDATE
    for _, row in df_pc.iterrows():
        rnd  = _v(row.get("Round",""))
        circ = _v(row.get("Circuit",""))
        sess = _v(row.get("Session",""))
        rider= _v(row.get("Rider",""))
        rno  = _i(row.get("Run", 1)) or 1
        if not all([circ, sess, rider]):
            continue
        run_id = make_run_id(rnd or "UNK", circ, sess, rider, rno)
        cur.execute("""
            UPDATE runs SET
              perf_best_lap=?, perf_avg_lap=?, perf_n_laps=?,
              perf_rank=?, perf_gap_s=?, perf_tier=?,
              updated_at=datetime('now')
            WHERE run_id=?
        """, (
            _v(row.get("Best Lap")), _v(row.get("Avg Lap")),
            _i(row.get("N Laps")),
            _v(row.get("Rank")), _f(row.get("Gap(s)")), _v(row.get("Tier")),
            run_id,
        ))

    conn.commit()
    n_runs = cur.execute("SELECT COUNT(*) FROM runs").fetchone()[0]
    print(f"  runs: {n_runs} 件")

    # ── laps テーブル ─────────────────────────
    cur.execute("DELETE FROM laps")
    n_laps = 0
    for _, row in df_lt.iterrows():
        rnd  = _v(row.get("ROUND",""))
        circ = _v(row.get("CIRCUIT",""))
        sess = _v(row.get("SESSION",""))
        rider= _v(row.get("RIDER",""))
        rno  = _i(row.get("RUN", 1)) or 1
        lap  = _i(row.get("LAP", 0))
        date_val = str(row.get("DATE","") or "").split(" ")[0]
        if not all([circ, sess, rider]):
            continue
        run_id = make_run_id(rnd or "UNK", circ, sess, rider, rno)
        lap_id = make_lap_id(run_id, lap)
        sid    = make_session_id(rnd or "UNK", circ, sess, rider)
        is_out = 1 if str(row.get("OUTLAP?","")).upper() in ("YES","Y","1","TRUE") else 0
        cur.execute("""
            INSERT OR REPLACE INTO laps
            (lap_id, run_id, session_id, round, circuit, session, rider, run_no,
             lap_no, date, lap_time, lap_time_s, is_outlap,
             weather, air_temp, track_temp, tyre_front, tyre_rear)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            lap_id, run_id, sid,
            str(rnd or ""), str(circ), str(sess), str(rider), rno,
            lap, date_val,
            _v(row.get("LAP TIME")), _f(row.get("TIME (s)")), is_out,
            _v(row.get("WEATHER")), _f(row.get("AIR °C")), _f(row.get("TRACK °C")),
            _v(row.get("TYRE F")), _v(row.get("TYRE R")),
        ))
        n_laps += 1
    conn.commit()
    print(f"  laps: {n_laps} 件")

    return df_sess, df_run, df_tyre, df_dyn, df_pc, df_lt


# ─────────────────────────────────────────────
#  Excel に RUN_ID / LAP_ID 列を追加
# ─────────────────────────────────────────────
def add_ids_to_excel(wb, df_sess, df_run, df_tyre, df_dyn, df_pc, df_lt):
    """各シートに RUN_ID (または LAP_ID) 列を追加する"""

    def gen_ids(df, round_col, circ_col, sess_col, rider_col, run_col):
        ids = []
        for _, r in df.iterrows():
            ids.append(make_run_id(
                r.get(round_col,""), r.get(circ_col,""),
                r.get(sess_col,""), r.get(rider_col,""), r.get(run_col,1)))
        return ids

    # SESSION_SUMMARY
    print("  SESSION_SUMMARY に RUN_ID 追加...")
    ss_ids = gen_ids(df_sess, "ROUND","CIRCUIT","SESSION","RIDER","RUN")
    add_id_column_to_sheet(wb["SESSION_SUMMARY"], ss_ids, header_row_excel=2, insert_col=1)

    # RUN_LOG
    print("  RUN_LOG に RUN_ID 追加...")
    # RUN_LOG の ROUND 列は空のことが多い → SESSION_SUMMARY から補完済み df_run は使えない
    ws_run = wb["RUN_LOG"]
    run_headers = [ws_run.cell(row=3, column=c).value for c in range(1, ws_run.max_column+1)]
    run_headers = [str(h).replace("\n"," ").strip() if h else f"_col{i}"
                   for i, h in enumerate(run_headers)]
    rl_ids = []
    for r in range(4, ws_run.max_row+1):
        row = {run_headers[c-1]: ws_run.cell(row=r, column=c).value
               for c in range(1, ws_run.max_column+1)}
        circ  = _v(row.get("CIRCUIT","")) or ""
        sess  = _v(row.get("SESSION","")) or ""
        rider = _v(row.get("RIDER",""))   or ""
        rno   = _i(row.get("RUN", 1)) or 1
        # ROUND を SESSION_SUMMARY から検索
        match = df_sess[
            (df_sess["CIRCUIT"].astype(str).str.upper() == str(circ).upper()) &
            (df_sess["SESSION"].astype(str).str.upper() == str(sess).upper()) &
            (df_sess["RIDER"].astype(str).str.upper()   == str(rider).upper())
        ]
        rnd = str(match.iloc[0]["ROUND"]) if not match.empty else "UNK"
        rl_ids.append(make_run_id(rnd, circ, sess, rider, rno))
    add_id_column_to_sheet(wb["RUN_LOG"], rl_ids, header_row_excel=3, insert_col=1)

    # TYRE_LOG
    print("  TYRE_LOG に RUN_ID 追加...")
    ty_ids = gen_ids(df_tyre, "ROUND","CIRCUIT","SESSION","RIDER","RUN")
    add_id_column_to_sheet(wb["TYRE_LOG"], ty_ids, header_row_excel=2, insert_col=1)

    # DYNAMICS_ANALYSIS
    print("  DYNAMICS_ANALYSIS に RUN_ID 追加...")
    dyn_ids = gen_ids(df_dyn, "Round","Circuit","Session","Rider","Run")
    add_id_column_to_sheet(wb["DYNAMICS_ANALYSIS"], dyn_ids, header_row_excel=2, insert_col=1)

    # PERFORMANCE_CORRELATION
    print("  PERFORMANCE_CORRELATION に RUN_ID 追加...")
    pc_ids = gen_ids(df_pc, "Round","Circuit","Session","Rider","Run")
    add_id_column_to_sheet(wb["PERFORMANCE_CORRELATION"], pc_ids, header_row_excel=2, insert_col=1)

    # LAP_TIMES: RUN_ID + LAP_ID の2列追加
    print("  LAP_TIMES に RUN_ID / LAP_ID 追加...")
    lt_run_ids = []
    lt_lap_ids = []
    for _, r in df_lt.iterrows():
        rid = make_run_id(
            r.get("ROUND",""), r.get("CIRCUIT",""),
            r.get("SESSION",""), r.get("RIDER",""), r.get("RUN",1))
        lid = make_lap_id(rid, r.get("LAP",0))
        lt_run_ids.append(rid)
        lt_lap_ids.append(lid)
    add_id_column_to_sheet(wb["LAP_TIMES"], lt_run_ids, header_row_excel=2,
                            col_label="RUN_ID",  insert_col=1)
    add_id_column_to_sheet(wb["LAP_TIMES"], lt_lap_ids, header_row_excel=2,
                            col_label="LAP_ID",  insert_col=2)


# ─────────────────────────────────────────────
#  メイン
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("TS24 統合データベース構築スクリプト")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} が見つかりません")
        sys.exit(1)

    # ── 1. SQLite 初期化 ──────────────────────
    print(f"\n1. SQLite DB 初期化: {UNIFIED_DB.name}")
    conn = sqlite3.connect(str(UNIFIED_DB))
    conn.executescript(SCHEMA_SQL)
    conn.commit()
    print("   スキーマ作成完了")

    # ── 2. Excel 読み込み ─────────────────────
    print(f"\n2. Excel 読み込み: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    # ── 3. SQLite 構築 ────────────────────────
    print("\n3. SQLite テーブルにデータ投入...")
    dfs = build_sqlite(conn, wb)
    df_sess, df_run, df_tyre, df_dyn, df_pc, df_lt = dfs

    # ── 4. Excel に ID 列追加 ─────────────────
    print("\n4. Excel シートに RUN_ID / LAP_ID 列を追加...")
    add_ids_to_excel(wb, df_sess, df_run, df_tyre, df_dyn, df_pc, df_lt)

    # ── 5. Excel 保存 ─────────────────────────
    print("\n5. Excel 保存中...")
    wb.save(str(EXCEL_PATH))
    print(f"   → {EXCEL_PATH.name} 保存完了")

    shutil.copy(str(EXCEL_PATH), str(BACKUP_PATH))
    print(f"   → {BACKUP_PATH.name} バックアップ完了")

    conn.close()

    # ── 6. サマリ表示 ─────────────────────────
    conn2 = sqlite3.connect(str(UNIFIED_DB))
    n_ev  = conn2.execute("SELECT COUNT(*) FROM events").fetchone()[0]
    n_ss  = conn2.execute("SELECT COUNT(*) FROM ts24_sessions").fetchone()[0]
    n_ru  = conn2.execute("SELECT COUNT(*) FROM runs").fetchone()[0]
    n_la  = conn2.execute("SELECT COUNT(*) FROM laps").fetchone()[0]
    conn2.close()

    print()
    print("=" * 60)
    print("完了! 統合 DB サマリ:")
    print(f"  events          : {n_ev:>5} 件")
    print(f"  ts24_sessions   : {n_ss:>5} 件")
    print(f"  runs (中心テーブル) : {n_ru:>5} 件")
    print(f"  laps            : {n_la:>5} 件")
    print(f"  DB ファイル     : {UNIFIED_DB}")
    print()
    print("【活用例 (Python)】")
    print("  import sqlite3")
    print(f"  conn = sqlite3.connect('{UNIFIED_DB.name}')")
    print()
    print("  # ASSENのJA52: セットアップ×ラップタイム×サスペンション 一括取得")
    print("  conn.execute('''")
    print("    SELECT r.run_id, r.f_comp, r.f_reb, r.apex_sus_f,")
    print("           l.lap_no, l.lap_time_s")
    print("    FROM runs r JOIN laps l ON r.run_id = l.run_id")
    print("    WHERE r.circuit='ASSEN' AND r.rider='JA52'")
    print("    ORDER BY l.lap_time_s")
    print("  ''').fetchall()")
    print("=" * 60)


if __name__ == "__main__":
    main()
