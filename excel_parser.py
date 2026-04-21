#!/usr/bin/env python3
"""
excel_parser.py — TS24 Excel Report Parser
============================================================
NEW_EVENT_TEAM_REPORT_TEMPLATE.xlsx を解析して
Supabase の pending_sessions / pending_lap_times 用データに変換する。

使用シート: DAY1 / DAY2 / REPORT
============================================================
"""

import re
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ── ラップタイム変換 ─────────────────────────────────────
def parse_lap_time(val) -> float | None:
    """'1'38,7' / '1:38.7' / 98.7 → float 秒数"""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ['-', '—', 'P6', 'P15', 'DNS', 'DNF', 'NC']:
        return None
    # 1'38,7  or  1:38.7  or  1'38.7
    m = re.match(r"(\d+)[':'](\d+)[,.](\d+)", s)
    if m:
        mins = int(m.group(1))
        secs = int(m.group(2))
        frac = m.group(3)
        return round(mins * 60 + secs + float(f"0.{frac}"), 3)
    # 純粋な数値
    try:
        return round(float(s), 3)
    except Exception:
        return None

def fmt_lap_time(val) -> str | None:
    """生の値をそのまま文字列として保存（表示用）"""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ── メインパーサー ───────────────────────────────────────
def parse_report_excel(file_bytes: bytes, submitted_by: str) -> dict:
    """
    Excelファイルのバイト列を受け取り、以下を返す:
    {
        "sessions": [pending_sessions 用 dict, ...],
        "laps":     [pending_lap_times 用 dict, ...],
        "errors":   [エラーメッセージ, ...]
    }
    """
    sessions_out = []
    laps_out = []
    errors = []

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return {"sessions": [], "laps": [], "errors": [f"Excelを開けませんでした: {e}"]}

    target_sheets = [s for s in ['DAY1', 'DAY2', 'REPORT'] if s in wb.sheetnames]
    if not target_sheets:
        return {"sessions": [], "laps": [],
                "errors": ["DAY1 / DAY2 / REPORT シートが見つかりません"]}

    for sheet_name in target_sheets:
        ws = wb[sheet_name]

        # ── ヘッダー情報 ─────────────────────────────────
        try:
            rider     = str(ws['B1'].value or '').strip() or None   # DA77 / JA52
            circuit   = str(ws['H2'].value or '').strip().upper() or None
            round_id  = str(ws['H3'].value or '').strip().upper() or None
            date_raw  = ws['D4'].value
            bike_model = str(ws['H4'].value or 'ZX-636').strip()

            if isinstance(date_raw, datetime):
                session_date = date_raw.strftime('%Y-%m-%d')
            elif date_raw:
                session_date = str(date_raw)
            else:
                session_date = None

            if not rider or not circuit:
                errors.append(f"[{sheet_name}] ライダー/サーキット情報なし — スキップ")
                continue
        except Exception as e:
            errors.append(f"[{sheet_name}] ヘッダー読み取りエラー: {e}")
            continue

        # ── ベースセットアップ（列D） ─────────────────────
        def v(row, col='D'):
            return ws[f'{col}{row}'].value

        base_setup = {
            'fork_type':  str(v(11) or '').strip() or None,
            'f_spring':   str(v(14) or '').strip() or None,
            'f_preload':  _to_float(v(15)),
            'f_comp':     _to_int(v(17)),
            'f_reb':      _to_int(v(18)),
            'shock_type': str(v(21) or '').strip() or None,
            'r_spring':   _to_float(v(23)),
            'r_preload':  _to_float(v(24)),
            'r_comp':     _to_int(v(25)),
            'r_reb':      _to_int(v(26)),
            'ride_height': _to_float(v(30)),
            'swing_arm':  _to_int(v(31)),
        }

        # ── セッション列マッピング（行7をスキャン） ────────
        SESSION_LABELS = {'FP', 'FP1', 'QP', 'SP', 'WUP', 'WUP1', 'WUP2',
                          'RACE1', 'RACE2', 'RACE', 'TEST'}
        session_cols = {}  # {セッションラベル: 列番号(1-indexed)}

        for cell in ws[7]:
            if cell.value and isinstance(cell.value, str):
                label = cell.value.strip().upper()
                if label in SESSION_LABELS:
                    session_cols[label] = cell.column

        if not session_cols:
            errors.append(f"[{sheet_name}] セッション列が見つかりません（行7）— スキップ")
            continue

        # ── セッションごとにデータを抽出 ─────────────────
        for session_type, col_idx in session_cols.items():
            try:
                total_laps = _to_int(ws.cell(row=36, column=col_idx).value)
                if not total_laps:
                    continue  # データなし

                track_temp = _to_float(ws.cell(row=9, column=col_idx).value)
                air_temp   = _to_float(ws.cell(row=9, column=col_idx + 1).value)
                f_tyre     = str(ws.cell(row=32, column=col_idx).value or '').strip() or None
                r_tyre     = str(ws.cell(row=34, column=col_idx).value or '').strip() or None
                best_lap_raw = ws.cell(row=37, column=col_idx).value
                best_lap_str = fmt_lap_time(best_lap_raw)

                # ── セッションレコード ────────────────────
                session_rec = {
                    'submitted_by':  submitted_by,
                    'session_date':  session_date,
                    'circuit':       circuit,
                    'session_type':  session_type,
                    'rider':         rider,
                    'bike_model':    bike_model,
                    'track_temp':    track_temp,
                    'air_temp':      air_temp,
                    'f_tyre':        f_tyre,
                    'r_tyre':        r_tyre,
                    'best_lap':      best_lap_str,
                    'status':        'pending',
                    **base_setup,
                }
                sessions_out.append(session_rec)

                # ── ラップタイム ─────────────────────────
                rider_num = 77 if '77' in str(rider) else (52 if '52' in str(rider) else 0)

                for lap_row in range(38, 80):
                    lap_raw = ws.cell(row=lap_row, column=col_idx).value
                    if lap_raw is None:
                        break
                    lap_sec = parse_lap_time(lap_raw)
                    if lap_sec is None:
                        continue
                    lap_no = lap_row - 37  # row38=LAP1, row39=LAP2...
                    laps_out.append({
                        'submitted_by': submitted_by,
                        'round_id':     round_id,
                        'circuit':      circuit,
                        'session_type': session_type,
                        'rider_num':    rider_num,
                        'rider_name':   rider,
                        'lap_no':       lap_no,
                        'lap_time':     lap_sec,
                        'speed':        None,
                        'flag':         '',
                        'is_valid':     1,
                        'status':       'pending',
                    })

            except Exception as e:
                errors.append(f"[{sheet_name}] {session_type} 処理エラー: {e}")

    return {"sessions": sessions_out, "laps": laps_out, "errors": errors}


# ── ユーティリティ ───────────────────────────────────────
def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '.'))
    except Exception:
        return None

def _to_int(val) -> int | None:
    f = _to_float(val)
    return int(f) if f is not None else None


# ── CLI テスト ────────────────────────────────────────────
if __name__ == "__main__":
    import sys, json
    path = sys.argv[1] if len(sys.argv) > 1 else \
        "/Users/ts24/Desktop/Data TS24 Claude/01_REPORTS/DA77/20260417-ROUND3-DA77.xlsx"
    with open(path, 'rb') as f:
        result = parse_report_excel(f.read(), "test_user")
    print(f"Sessions: {len(result['sessions'])}")
    print(f"Laps:     {len(result['laps'])}")
    print(f"Errors:   {result['errors']}")
    if result['sessions']:
        print("\nFirst session:")
        print(json.dumps(result['sessions'][0], indent=2, default=str))
