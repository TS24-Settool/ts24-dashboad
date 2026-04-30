#!/usr/bin/env python3
from __future__ import annotations  # Python 3.9 で dict|None / list[X] 型ヒントを有効化
"""
parse_2d_channels.py — 2D Logger Channel Dynamics Analysis
===========================================================
2D Analyser (.MES) フォルダからチャンネルデータを読み込み、
以下の指標を計算して PUCCETTI_DB_MASTER.xlsx の新シート
「DYNAMICS_ANALYSIS」に保存する。

計算項目:
  1. APEX時のF/Rサスペンションストローク量 (mm)
  2. APEX時のホイールフォース (N) — 車体動力学方程式から算出
  3. ピットレーンリミッター発動区間 (Speed F 56〜64 km/h が3秒以上) のF/Rサスストローク平均 (mm)
  4. ブレーキ直前のF/Rサスストローク (mm)

使用チャンネル:
  SPEED_FRONT (km/h)  → APEX検出・低速ゾーン判定・加減速度計算
  SUSP_FRONT  (mm)    → フロントサスストローク
  SUSP_REAR   (mm)    → リアサスストローク
  BRAKE_FRONT (Bar)   → ブレーキ検出補助
  BRAKE_REAR  (Bar)   → ブレーキ検出補助

TS24 パラメータ (Knowledge Base 11.8):
  JA52: WB=1399mm, h=644.9mm, Fw_rate=30.1N/mm, Rw_rate=58.4N/mm
        F-weight=48.92%, R-weight=51.08%
  DA77: WB=1411.5mm, h=651.1mm, Fw_rate=26.1N/mm, Rw_rate=57.8N/mm
        F-weight=48.71%, R-weight=51.29%

実行方法:
  python3 parse_2d_channels.py
"""

import os
import re
import struct
import math
import json
from pathlib import Path
from datetime import datetime
import numpy as np

# ── 定数 ─────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).resolve().parent
DATA_2D_ROOT  = SCRIPT_DIR.parent / "DATA 2D"
EXCEL_PATH    = SCRIPT_DIR.parent / "02_DATABASE" / "TS24 DB Master.xlsx"
OUTPUT_SHEET  = "DYNAMICS_ANALYSIS"

G             = 9.81        # m/s²
KMH_TO_MS     = 1.0 / 3.6  # km/h → m/s

# TS24 バイオメカニクスパラメータ
RIDER_PARAMS = {
    "JA52": {
        "wheelbase_m":    1.399,
        "cog_h_m":        0.6449,
        "f_weight_ratio": 0.4892,
        "fw_rate_Nmm":    30.1,
        "rw_rate_Nmm":    58.4,
    },
    "DA77": {
        "wheelbase_m":    1.4115,
        "cog_h_m":        0.6511,
        "f_weight_ratio": 0.4871,
        "fw_rate_Nmm":    26.1,
        "rw_rate_Nmm":    57.8,
    },
}

# APEX検出パラメータ (局所極小点 + プロミネンスフィルタ方式) — 速度最小点フォールバック用
APEX_SPEED_THRESHOLD_RATIO = 0.92  # ラップ最高速の92%以下のみAPEX候補とする
APEX_MIN_GAP_S             = 0.25  # 隣接APEX間の最小間隔 [秒] (シケイン対応)
APEX_SMOOTH_WINDOW_S       = 0.3   # Speed Fスムージング [秒]
APEX_MIN_PROMINENCE_KMH    = 15.0  # APEX前後 ±4秒の最高速との最低落差 [km/h]
#   → 直線上の微細な速度変動(5〜10km/h)を除外し、真のコーナーのみ検出
APEX_PROMINENCE_WINDOW_S   = 4.0   # プロミネンス判定ウィンドウ [秒]

# APEX検出パラメータ — 横G最大点(ACC_Y Peak)方式 [Primary]
# 物理的根拠: |ACC_Y|最大 ≡ コーナー半径最小 ≡ 幾何学的Apex
#   ACC_Y(body frame) = g·sin(θ) + (v²/R)·cos(θ)
#   → バンク角・旋回半径・速度を統合した不変量
#   → IMU固体センサーのためキャリブレーションズレなし(BIKE_ANGLEより信頼性大)
APEX_ACCY_MIN_MS2          = 1.5   # コーナー判定の最小横G閾値 [m/s²]  (≈0.15g)
APEX_ACCY_SMOOTH_S         = 0.25  # ACC_Yスムージングウィンドウ [秒]
APEX_ACCY_MIN_GAP_S        = 0.40  # 隣接APEX最小間隔 [秒]
APEX_ACCY_PROMINENCE_MS2   = 0.60  # プロミネンス閾値 [m/s²] (直線ノイズ除外)
APEX_ACCY_QUALITY_RATIO    = 0.70  # ACC_Y検出数/速度検出数の最低比率(品質チェック)
MIN_LAP_DURATION_S         = 60.0  # これ未満のラップはアウトラップ/フォーメーション等として除外

# ピットレーンリミッターパラメータ
PIT_LIMITER_SPEED_KMH  = 60.0  # ピットリミッター速度
PIT_LIMITER_MARGIN_KMH = 4.0   # ±マージン → 56〜64 km/h
PIT_LIMITER_MIN_S      = 3.0   # 最小持続時間 (秒)

# ブレーキ検出パラメータ
BRAKE_DECEL_THRESHOLD_MS2 = -8.0   # 減速度閾値 (m/s²) — 負 = 減速 (強めに設定)
BRAKE_CONFIRM_SAMPLES     = 15     # 連続N サンプル以上で確定 (~0.15s @100Hz)
BRAKE_MIN_GAP_SAMPLES     = 80     # 次ブレーキイベントまでの最小間隔 (~0.8s)
BRAKE_MIN_SPEED_KMH       = 80.0   # この速度以上でのみブレーキとみなす
BRAKE_LOOK_AHEAD_SAMPLES  = 8      # ブレーキ開始N サンプル前の値を「直前」とする
BRAKE_SMOOTH_WINDOW       = 7      # Speed F 微分前スムージングウィンドウ

# フルブレーキング検出パラメータ
FULL_BRAKE_THRESHOLD_BAR  = 1.0   # ブレーキ入力有効閾値 (Bar) — ノイズ除去
FULL_BRAKE_PEAK_RATIO     = 0.85  # ピーク圧力の何%以上を「フルブレーキングエリア」とするか
FULL_BRAKE_MIN_SPEED_KMH  = 80.0  # この速度以上でのみ検出 (低速誤検出除外)
FULL_BRAKE_MIN_DUR_S      = 0.05  # 最小持続時間 (秒) — 瞬時入力除外

# タイヤ内圧チャンネル候補 (2Dロガー命名規則の揺れに対応)
TYRE_CH_CANDIDATES_F = [
    "PRESS_TYRE_F", "TYRE_PRESS_F", "TYREPRESS_F", "TYRE_F_PRESS",
    "PRESS_F_TYRE", "PRESSURE_TYRE_F", "TYRE_PRESSURE_F",
    "TPRES_F", "TYREPRES_F", "PRESS_FL", "TIRE_PRESS_F",
    "TYRE_F", "TIRE_F", "PRESS_FRONT_TYRE", "FRONT_TYRE_PRESS",
]
TYRE_CH_CANDIDATES_R = [
    "PRESS_TYRE_R", "TYRE_PRESS_R", "TYREPRESS_R", "TYRE_R_PRESS",
    "PRESS_R_TYRE", "PRESSURE_TYRE_R", "TYRE_PRESSURE_R",
    "TPRES_R", "TYREPRES_R", "PRESS_RL", "TIRE_PRESS_R",
    "TYRE_R", "TIRE_R", "PRESS_REAR_TYRE", "REAR_TYRE_PRESS",
]
TYRE_EDGE_RATIO  = 0.05   # Run開始・終了の何%(最低10サンプル)を代表値とするか
TYRE_VALID_MIN   = 0.8    # 有効タイヤ内圧下限 (Bar) — センサーオフライン除外
TYRE_VALID_MAX   = 5.0    # 有効タイヤ内圧上限 (Bar)

# ── DDD パーサー ──────────────────────────────────────────────────────
def parse_ddd(mes_path: Path, base: str) -> dict:
    """DDD → {channel_name: {ext, offset, scale, signed}}"""
    ddd_path = mes_path / f"{base}.DDD"
    if not ddd_path.exists():
        return {}
    raw  = ddd_path.read_bytes()
    text = raw.decode("latin-1", errors="replace")

    sections, current = {}, None
    for line in text.splitlines():
        line = line.strip()
        m = re.match(r"^\[([0-9A-Fa-f]+)\]$", line)
        if m:
            current = m.group(1)
            sections[current] = {}
            continue
        if current and "=" in line:
            k, _, v = line.partition("=")
            sections[current][k.strip()] = v.strip()

    def hex_float(h):
        try:
            return struct.unpack(">f", bytes.fromhex(h.zfill(8)))[0]
        except Exception:
            return 1.0

    channels = {}
    for _, fields in sections.items():
        line4 = fields.get("4", "")
        line6 = fields.get("6", "")
        line3 = fields.get("3", "")
        name  = line4.split(",")[0].strip() if line4 else ""
        if not name:
            continue
        parts6 = line6.split(",")
        ext    = parts6[3] if len(parts6) > 3 else ""
        parts3 = line3.split(",")
        offset = int(parts3[0]) if parts3 and parts3[0].lstrip("-").isdigit() else 0
        scale  = hex_float(parts3[3]) if len(parts3) > 3 else 1.0
        mult   = hex_float(parts3[6]) if len(parts3) > 6 else 1.0
        signed = line6.startswith("FFFFFFFE") or (len(parts6) > 1 and parts6[1] == "I")
        channels[name] = {
            "ext": ext, "offset": offset,
            "scale": scale, "mult": mult or 1.0,
            "signed": signed,
        }
    return channels


def read_channel(mes_path: Path, base: str, ch: dict) -> np.ndarray:
    """チャンネルファイルを物理値 numpy 配列として読み込む"""
    path = mes_path / f"{base}.{ch['ext']}"
    if not path.exists():
        return np.array([], dtype=np.float32)
    data = path.read_bytes()
    n    = len(data) // 2
    dtype = np.int16 if ch["signed"] else np.uint16
    raw  = np.frombuffer(data[: n * 2], dtype=dtype).astype(np.float32)
    off, sc, mu = ch["offset"], ch["scale"], ch["mult"]
    return (raw - off) * sc * (mu or 1.0)


# ── HED パーサー ──────────────────────────────────────────────────────
def parse_hed(mes_path: Path, base: str) -> dict:
    """HED → セッションメタデータ dict
    セクション別に収集し、フラットキーは [GENERAL] を最優先とする。
    セクション固有キーは 'SECTION.Key' 形式でも保存。
    """
    path = mes_path / f"{base}.HED"
    if not path.exists():
        return {}
    text    = path.read_bytes().decode("latin-1", errors="replace")
    info    = {}
    section = "GENERAL"
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("[") and line.endswith("]"):
            section = line[1:-1]
            continue
        if "=" in line:
            k, _, v = line.partition("=")
            k = k.strip()
            v = v.strip()
            # セクション固有キー (例: MASS.Bike, GENERAL.Date)
            info[f"{section}.{k}"] = v
            # フラットキー: GENERAL 優先、なければ初回のみ
            if k not in info or section == "GENERAL":
                info[k] = v
    return info


# ── LAP パーサー ──────────────────────────────────────────────────────
def parse_lap(mes_path: Path, base: str):
    """LAP → (n_laps, [lap_end_ms, ...])"""
    path = mes_path / f"{base}.LAP"
    if not path.exists():
        return 0, []
    data   = path.read_bytes()
    n_int  = len(data) // 4
    if n_int < 3:
        return 0, []
    vals   = struct.unpack(f"<{n_int}I", data[: n_int * 4])
    n_laps = vals[0]
    times  = list(vals[3: 3 + n_laps])
    return n_laps, times


# ── ユーティリティ ────────────────────────────────────────────────────
def smooth(arr: np.ndarray, w: int = 5) -> np.ndarray:
    """numpy convolve による高速移動平均スムージング"""
    if w <= 1 or len(arr) < w:
        return arr
    kernel = np.ones(w, dtype=np.float32) / w
    return np.convolve(arr, kernel, mode="same")


def derivative(arr: np.ndarray, dt: float) -> np.ndarray:
    """numpy gradient による高速中央差分微分 (単位: /dt)"""
    if len(arr) < 2:
        return np.zeros_like(arr)
    return np.gradient(arr.astype(np.float64), dt).astype(np.float32)


def susp_at_speed_index(susp: np.ndarray, speed_idx: int, ratio: int = 4) -> float:
    """Speed sample index → 対応する SUSP 値 (SUSP は Speed の ratio 倍サンプル)"""
    si = speed_idx * ratio
    if 0 <= si < len(susp):
        return float(susp[si])
    return float("nan")


def susp_mean_in_range(susp: np.ndarray, s_start: int, s_end: int, ratio: int = 4) -> float:
    """Speed sample 範囲 [s_start, s_end] に対応する SUSP の平均"""
    i0  = max(0, s_start * ratio)
    i1  = min(len(susp), s_end * ratio)
    if i0 >= i1:
        return float("nan")
    return float(np.mean(susp[i0:i1]))


def safe_mean(lst):
    lst = [float(v) for v in lst if v is not None and not math.isnan(float(v) if hasattr(v, '__float__') else v)]
    return round(sum(lst) / len(lst), 2) if lst else None


# ── 動力学計算 ────────────────────────────────────────────────────────
def wheel_forces(rider_tag: str, mass_kg: float, ax_ms2: float):
    """
    ホイール荷重計算 (m/s² ← 正=加速, 負=制動)
    Returns (Nf_N, Nr_N)
    """
    p    = RIDER_PARAMS.get(rider_tag, RIDER_PARAMS["DA77"])
    L    = p["wheelbase_m"]
    h    = p["cog_h_m"]
    fw   = p["f_weight_ratio"]
    rw   = 1.0 - fw
    # 静的荷重
    Nf0  = mass_kg * G * fw
    Nr0  = mass_kg * G * rw
    # 動的荷重移動 (制動時はフロントへ移動, ax<0 → -ax>0)
    dN   = mass_kg * (-ax_ms2) * h / L   # 正値=フロントへ移動
    Nf   = Nf0 + dN
    Nr   = Nr0 - dN
    return max(0.0, Nf), max(0.0, Nr)


# ── APEX 検出 ────────────────────────────────────────────────────────
def detect_apexes(speed_kmh: np.ndarray, lap_start: int, lap_end: int,
                  sr: float) -> list:
    """
    ラップ内の全コーナーAPEX を局所極小点(Local Minima)方式で検出。

    【旧方式の問題】
      「速度が閾値以下の連続区間の最小1点」 → 連続コーナー・シケインが
      1点に折りたたまれ、1ラップあたり3〜8コーナーしか検出できなかった。

    【新方式】
      速度トレースの勾配符号が (負 → 正) に変化する点 = 局所極小 を全て検出。
      ラップ最高速の 92% 以下 かつ 隣接APEXと 0.25秒以上 離れているものを採用。
      → 全コーナーの APEX を個別に検出可能 (フィリップアイランド13、アッセン18 等)

    Returns: list of sample indices (global index)
    """
    seg = speed_kmh[lap_start:lap_end]
    if len(seg) < 10:
        return []

    # 0.4 秒ウィンドウでスムージング (ノイズによる偽極小除去)
    smooth_w  = max(3, int(APEX_SMOOTH_WINDOW_S * sr))
    seg_s     = smooth(seg, smooth_w)

    max_speed = float(np.max(seg_s))
    if max_speed < 10:
        return []
    threshold = max_speed * APEX_SPEED_THRESHOLD_RATIO

    # 勾配を計算し、(負 → 正) の符号変化 = 局所極小 を抽出
    grad      = np.gradient(seg_s.astype(np.float64))
    neg_to_pos = (grad[:-1] < 0) & (grad[1:] >= 0)
    local_min  = np.where(neg_to_pos)[0]

    # 速度閾値フィルタ: 最高速の 92% 以下のみ
    valid = local_min[seg_s[local_min] < threshold]

    # プロミネンスウィンドウ (直線速度との落差チェック)
    pw = int(APEX_PROMINENCE_WINDOW_S * sr)

    # 最小間隔フィルタ + プロミネンスフィルタ
    min_gap    = max(5, int(APEX_MIN_GAP_S * sr))
    apexes     = []
    last_local = -min_gap

    for idx in valid:
        idx = int(idx)
        # ±APEX_PROMINENCE_WINDOW_S 秒の最高速との落差を確認
        w0    = max(0, idx - pw)
        w1    = min(len(seg_s), idx + pw)
        depth = float(np.max(seg_s[w0:w1])) - float(seg_s[idx])
        if depth < APEX_MIN_PROMINENCE_KMH:
            continue   # 直線上の微細な速度変動 → 無視
        if idx - last_local >= min_gap:
            apexes.append(lap_start + idx)
            last_local = idx

    return apexes


# ── APEX 検出 v2 — 横G最大点 (ACC_Y Peak) ───────────────────────────
def detect_apexes_accy(accy_raw: np.ndarray,
                       speed_raw: np.ndarray,
                       lap_start: int, lap_end: int,
                       sr: float,
                       accy_ratio: int = 1) -> list:
    """
    横G絶対値最大点 (Peak Lateral G) によるAPEX検出 [推奨・将来標準]

    【物理的根拠】
      バイクがコーナーを走行するとき、ボディ固定座標の横方向加速度は:
        ACC_Y ≈ g·sin(θ) + (v²/R)·cos(θ)
      θ=バンク角、R=コーナー半径、v=速度
      この式の最大点 = バンク角最大 かつ 旋回半径最小 = 幾何学的Apex

    【速度最小点との比較】
      速度最小点はライダーがアクセルを微開け(5〜15%)した後に発生し、
      幾何学的Apexより平均50〜200ms遅れる。
      ACC_Yピークは遅れなく幾何学的Apexを指す。

    【BIKE_ANGLE(バンク角)より信頼性が高い理由】
      - IMU(三軸加速度計)は固体センサーで経年劣化・キャリブレーションズレが少ない
      - バンク角センサーは取付角度誤差・温度ドリフトの影響を受けやすい

    【フォールバック】
      ACC_Y未使用/品質不足 → detect_apexes()(速度最小点)に自動切替

    Args:
        accy_raw  : ACC_Y チャンネル配列 [m/s²]、空配列可
        speed_raw : SPEED_FRONT チャンネル配列 [km/h]
        lap_start : speed_rawのラップ開始インデックス(グローバル)
        lap_end   : speed_rawのラップ終了インデックス(グローバル)
        sr        : speed_rawのサンプルレート [Hz]
        accy_ratio: len(accy_raw)/len(speed_raw) の整数比率

    Returns:
        speed_raw チャンネルのグローバルインデックスリスト
        ACC_Y品質不足の場合は detect_apexes() の結果を返す
    """
    # ── ACC_Y の品質チェック ──────────────────────────────────────
    if len(accy_raw) == 0:
        return detect_apexes(speed_raw, lap_start, lap_end, sr)

    ay0 = int(lap_start * accy_ratio)
    ay1 = int(min(lap_end * accy_ratio, len(accy_raw)))
    if ay1 - ay0 < 20:
        return detect_apexes(speed_raw, lap_start, lap_end, sr)

    seg_ay = accy_raw[ay0:ay1].astype(np.float64)
    abs_ay = np.abs(seg_ay)

    # 最大横Gが閾値未満 → コーナーデータなしと判断してフォールバック
    if float(np.max(abs_ay)) < APEX_ACCY_MIN_MS2 * 1.5:
        return detect_apexes(speed_raw, lap_start, lap_end, sr)

    # ── スムージング ──────────────────────────────────────────────
    sw = max(3, int(APEX_ACCY_SMOOTH_S * sr * accy_ratio))
    abs_ay_s = smooth(abs_ay, sw)

    # ── 局所極大(負→正の勾配符号変化)を抽出 ─────────────────────
    grad = np.gradient(abs_ay_s)
    pos_to_neg = (grad[:-1] >= 0) & (grad[1:] < 0)
    local_max  = np.where(pos_to_neg)[0]

    # ── フィルタ1: 最小横G閾値 ───────────────────────────────────
    valid = local_max[abs_ay_s[local_max] >= APEX_ACCY_MIN_MS2]

    # ── フィルタ2: プロミネンス (周辺との落差) ───────────────────
    pw      = max(10, int(APEX_PROMINENCE_WINDOW_S * sr * accy_ratio))
    min_gap = max(5, int(APEX_ACCY_MIN_GAP_S * sr * accy_ratio))
    apexes  = []
    last_i  = -min_gap

    for idx in valid:
        idx = int(idx)
        w0 = max(0, idx - pw)
        w1 = min(len(abs_ay_s), idx + pw)
        surrounding = np.concatenate([abs_ay_s[w0:idx], abs_ay_s[idx+1:w1]])
        if len(surrounding) == 0:
            continue
        prominence = float(abs_ay_s[idx]) - float(np.max(surrounding))
        if prominence < -APEX_ACCY_PROMINENCE_MS2:   # 周囲より十分突出していない
            continue
        if idx - last_i < min_gap:
            # 同じ間隔内なら横Gが大きい方を優先
            if apexes and float(abs_ay_s[idx]) > float(abs_ay_s[apexes[-1] - ay0]):
                apexes[-1] = ay0 + idx
            continue
        apexes.append(ay0 + idx)   # accy_raw のグローバルインデックス
        last_i = idx

    # ── ACC_Y インデックス → speed_raw インデックスに変換 ────────
    speed_apexes = []
    for ai_global in apexes:
        si = int(round(ai_global / accy_ratio))
        si = max(lap_start, min(lap_end - 1, si))
        speed_apexes.append(si)

    # ── 品質チェック: 速度最小点と検出数の比較 ───────────────────
    fallback = detect_apexes(speed_raw, lap_start, lap_end, sr)
    if len(fallback) > 0:
        ratio_detected = len(speed_apexes) / len(fallback)
        if ratio_detected < APEX_ACCY_QUALITY_RATIO or len(speed_apexes) == 0:
            # 検出数が速度最小点の70%未満 → フォールバック
            return fallback

    return speed_apexes if speed_apexes else fallback


# ── APEX Area 検出 (新定義 2026-04-30, dTPS_A緩和 2026-04-30チーム承認) ──────────
# BRAKE_FRONT -0.6~0.3Bar / GAS 0~6% / dTPS_A -10~100 / SUSP_F 20~140mm / SUSP_R 5~50mm
def detect_apex_area(brake_f, gas, dtps_a, sus_f, sus_r,
                     gas_ratio=2, sus_ratio=4,
                     min_samples=3, merge_gap_samples=20):
    """
    新チームAPEX定義 (2026-04-30):
    5条件が同時成立する連続区間を検出し、各区間の
    SUSP_FRONT/SUSP_REAR の平均値を返す。

    Args:
        brake_f   : BRAKE_FRONT array (1x rate)
        gas       : GAS array (gas_ratio x rate)
        dtps_a    : dTPS_A array (gas_ratio x rate)
        sus_f     : SUSP_FRONT array (sus_ratio x rate)
        sus_r     : SUSP_REAR array (sus_ratio x rate)
        gas_ratio : GAS/dTPS_A vs brake_f のレート比 (default=2)
        sus_ratio : SUSP vs brake_f のレート比 (default=4)
        min_samples    : 有効APEX最小サンプル数(brake_fレート, default=3)
        merge_gap_samples : この間隔以内の区間は同一APEXとしてマージ

    Returns:
        list of dict: [{start, end, mid, susF_avg, susR_avg}, ...]
        すべてbrake_fのインデックス空間
    """
    n = len(brake_f)
    mask = np.zeros(n, dtype=bool)

    for i in range(n):
        gi = min(i * gas_ratio, len(gas) - 1)
        bf = brake_f[i]
        g  = gas[gi]  if len(gas)   > 0 else np.nan
        dg = dtps_a[gi] if len(dtps_a) > 0 else np.nan
        # SUSP: brake_fインデックス→SUSP中央値（4サンプル平均）
        si_start = i * sus_ratio
        si_end   = min(si_start + sus_ratio, len(sus_f))
        sf = float(np.mean(sus_f[si_start:si_end])) if si_end > si_start else np.nan
        sr = float(np.mean(sus_r[si_start:si_end])) if si_end > si_start and len(sus_r) > 0 else np.nan

        if (np.isfinite(bf) and  -0.6 <= bf <= 0.3   and
            np.isfinite(g)  and   0.0 <= g  <= 6.0   and
            np.isfinite(dg) and -10.0 <= dg <= 100.0  and
            np.isfinite(sf) and  20.0 <= sf <= 140.0  and
            np.isfinite(sr) and   5.0 <= sr <= 50.0):
            mask[i] = True

    # 連続区間を抽出
    segments = []
    in_seg = False
    for i in range(n):
        if mask[i] and not in_seg:
            seg_start = i
            in_seg = True
        elif not mask[i] and in_seg:
            segments.append((seg_start, i - 1))
            in_seg = False
    if in_seg:
        segments.append((seg_start, n - 1))

    # 近傍区間をマージ
    if segments:
        merged = [list(segments[0])]
        for s, e in segments[1:]:
            if s - merged[-1][1] <= merge_gap_samples:
                merged[-1][1] = e
            else:
                merged.append([s, e])
        segments = [tuple(m) for m in merged]

    # min_samples フィルタ + 代表値計算
    result = []
    for s, e in segments:
        if (e - s + 1) < min_samples:
            continue
        si_s = s * sus_ratio
        si_e = min((e + 1) * sus_ratio, len(sus_f))
        susF_vals = sus_f[si_s:si_e]
        susR_vals = sus_r[si_s:si_e] if len(sus_r) > 0 else np.array([])
        result.append({
            "start":    s,
            "end":      e,
            "mid":      (s + e) // 2,
            "susF_avg": float(np.mean(susF_vals)) if len(susF_vals) > 0 else float("nan"),
            "susR_avg": float(np.mean(susR_vals)) if len(susR_vals) > 0 else float("nan"),
        })
    return result


# ── ピットレーンリミッター区間検出 ───────────────────────────────────
def detect_pit_limiter_zones(speed_kmh: np.ndarray, lap_start: int, lap_end: int,
                              sr: float) -> list:
    """
    ピットレーンリミッター発動区間の検出。
    Speed F が [56, 64] km/h の範囲に 3秒以上連続して収まっている区間を返す。
    ローリング平均ではなく瞬時速度の帯域チェックを使用
    (リミッターは電子的速度クランプで瞬時に安定するため)。

    Returns: list of (start_idx, end_idx) tuples (global indices)
    """
    seg = speed_kmh[lap_start:lap_end]
    if len(seg) < 10:
        return []
    lo          = PIT_LIMITER_SPEED_KMH - PIT_LIMITER_MARGIN_KMH   # 56 km/h
    hi          = PIT_LIMITER_SPEED_KMH + PIT_LIMITER_MARGIN_KMH   # 64 km/h
    min_samples = max(1, int(PIT_LIMITER_MIN_S * sr))

    # 軽いスムージング (0.2 s) でノイズ除去してから帯域チェック
    smooth_w = max(1, int(0.2 * sr))
    seg_s    = smooth(seg, smooth_w)

    # 帯域内 mask & ランレングス
    mask    = (seg_s >= lo) & (seg_s <= hi)
    changes = np.diff(mask.astype(np.int8), prepend=0, append=0)
    starts  = np.where(changes == 1)[0]
    ends    = np.where(changes == -1)[0]

    zones = []
    for s, e in zip(starts, ends):
        if (e - s) >= min_samples:
            zones.append((lap_start + int(s), lap_start + int(e)))
    return zones


# ── ブレーキング直前検出 ──────────────────────────────────────────────
def detect_brake_entries(speed_kmh: np.ndarray, speed_ms: np.ndarray,
                         lap_start: int, lap_end: int, dt: float) -> list:
    """
    Speed F の急激な減速開始点を検出し、
    その BRAKE_LOOK_AHEAD_SAMPLES 前の index を返す。
    条件: 速度が BRAKE_MIN_SPEED_KMH 以上の状態で急減速が BRAKE_CONFIRM_SAMPLES 以上継続
    Returns: list of sample indices (global) — ブレーキ直前
    """
    seg_kmh = speed_kmh[lap_start:lap_end]
    seg_ms  = speed_ms[lap_start:lap_end]
    if len(seg_ms) < 10:
        return []

    # numpy による高速スムージング + 微分
    seg_s = smooth(seg_ms.astype(np.float64), BRAKE_SMOOTH_WINDOW)
    dv    = derivative(seg_s, dt)   # m/s²

    # マスク: 高速かつ強い減速
    mask_brake = (dv < BRAKE_DECEL_THRESHOLD_MS2) & (seg_kmh >= BRAKE_MIN_SPEED_KMH)

    # ランレングスで連続ブレーキ区間を検出
    changes = np.diff(mask_brake.astype(np.int8), prepend=0, append=0)
    starts  = np.where(changes == 1)[0]
    ends    = np.where(changes == -1)[0]

    entries    = []
    last_entry = -BRAKE_MIN_GAP_SAMPLES

    for s, e in zip(starts, ends):
        duration = int(e) - int(s)
        if duration < BRAKE_CONFIRM_SAMPLES:
            continue
        brake_start = int(s)
        if brake_start - last_entry < BRAKE_MIN_GAP_SAMPLES:
            continue
        look_back = max(0, brake_start - BRAKE_LOOK_AHEAD_SAMPLES)
        entries.append(lap_start + look_back)
        last_entry = brake_start

    return entries


# ── タイヤ内圧チャンネル検索 ─────────────────────────────────────────
def find_tyre_channel(chs: dict, side: str):
    """
    タイヤ内圧チャンネル名を候補リスト→ファジー検索の順で特定する。
    大文字小文字を区別しない照合で 'TyreFront_Press' 等にも対応。
    side: 'F'=フロント, 'R'=リア
    Returns: channel name str (実際のキー名) or None
    """
    candidates = TYRE_CH_CANDIDATES_F if side.upper() == "F" else TYRE_CH_CANDIDATES_R
    # 大文字小文字を無視した完全一致
    chs_upper = {k.upper(): k for k in chs}  # 大文字 → 元のキー名
    for cand in candidates:
        real_key = chs_upper.get(cand.upper())
        if real_key and chs[real_key].get("ext"):
            return real_key
    # ファジー: チャンネル名に TYRE/TIRE + PRESS + サイド識別子を含むもの
    for name in sorted(chs.keys()):
        nu = name.upper()
        has_tyre  = any(kw in nu for kw in ("TYRE", "TIRE"))
        has_press = "PRESS" in nu
        if side.upper() == "F":
            has_side = any(kw in nu for kw in ("_F", "FL", "FRONT"))
        else:
            has_side = any(kw in nu for kw in ("_R", "RL", "REAR"))
        if has_tyre and has_press and has_side and chs[name].get("ext"):
            return name
    return None


def extract_tyre_pressure(press_arr: np.ndarray) -> dict:
    """
    タイヤ内圧の Run 統計量を算出する。

    単位自動検出: 中央値 < 10 の場合は Bar として kPa へ変換（×100）。
    Run開始値:  最初 TYRE_EDGE_RATIO% (最低10サンプル) の中央値
    Run終了値:  最後 TYRE_EDGE_RATIO% (最低10サンプル) の中央値
    平均値:     Run全体の有効サンプル平均
    変化量:     終了値 − 開始値  (kPa)

    Returns: {"start": float|None, "end": float|None, "avg": float|None, "delta": float|None}
    """
    empty = {"start": None, "end": None, "avg": None, "delta": None}
    if len(press_arr) < 4:
        return empty

    # 単位自動検出: kPa (100〜400) で記録されている場合は Bar に変換 (÷100)
    positive = press_arr[press_arr > 0]
    if len(positive) > 0 and float(np.median(positive)) > 10.0:
        press_arr = press_arr / 100.0   # kPa → Bar

    n     = len(press_arr)
    valid = press_arr[(press_arr > TYRE_VALID_MIN) & (press_arr < TYRE_VALID_MAX)]
    if len(valid) < 4:
        return empty

    n_edge    = max(10, int(n * TYRE_EDGE_RATIO))
    start_seg = press_arr[:n_edge]
    end_seg   = press_arr[-n_edge:]

    def seg_median(seg):
        v = seg[(seg > TYRE_VALID_MIN) & (seg < TYRE_VALID_MAX)]
        return float(np.median(v)) if len(v) > 0 else None

    start_val = seg_median(start_seg)
    end_val   = seg_median(end_seg)
    avg_val   = float(np.mean(valid))
    delta_val = round(end_val - start_val, 1) if (start_val is not None and end_val is not None) else None

    return {
        "start": round(start_val, 1) if start_val is not None else None,
        "end":   round(end_val,   1) if end_val   is not None else None,
        "avg":   round(avg_val,   1),
        "delta": delta_val,
    }


def detect_full_braking_sus(
        brake_f_raw: np.ndarray,
        speed_kmh: np.ndarray,
        sus_f_raw: np.ndarray,
        sus_r_raw: np.ndarray,
        lap_start: int,
        lap_end: int,
        brake_scale: float,
        susp_ratio: int,
        sr: float,
) -> list:
    """
    フルブレーキング区間（BRAKE_F ≥ ピーク × FULL_BRAKE_PEAK_RATIO）での
    F/R サスペンションポジションを取得する。

    フルブレーキングの定義:
      1. BRAKE_F が FULL_BRAKE_THRESHOLD_BAR を超えてから 0 に戻るまでを1イベントとする。
      2. イベント内のBRAKE_Fピーク値を取得する。
      3. ピーク値 × FULL_BRAKE_PEAK_RATIO 以上のサンプルを「フルブレーキングエリア」とする。
      4. そのエリア内のF/Rサスストローク平均を記録する。

    brake_scale: len(brake_f_raw) / len(speed_kmh) — 1.0未満でもOK (BRAKE SR < SPEED SR の場合)

    Returns: list of {"susF_mm": float|None, "susR_mm": float|None}
    """
    if len(brake_f_raw) == 0 or len(speed_kmh) == 0:
        return []
    seg_len = lap_end - lap_start
    if seg_len < 10:
        return []

    # speed index → brake index のベクトル化マッピング (float scale で SR差を吸収)
    spd_indices   = np.arange(lap_start, lap_end)
    brake_indices = np.clip((spd_indices * brake_scale).astype(int), 0, len(brake_f_raw) - 1)
    brake_seg     = brake_f_raw[brake_indices]   # speed sample 空間でのbrake値
    speed_seg     = speed_kmh[lap_start:lap_end]

    # ブレーキイベント検出: 閾値超 AND 速度条件
    in_brake = (brake_seg >= FULL_BRAKE_THRESHOLD_BAR) & (speed_seg >= FULL_BRAKE_MIN_SPEED_KMH)
    changes  = np.diff(in_brake.astype(np.int8), prepend=0, append=0)
    starts   = np.where(changes == 1)[0]
    ends     = np.where(changes == -1)[0]

    min_dur_samples = max(1, int(FULL_BRAKE_MIN_DUR_S * sr))
    results = []

    for s, e in zip(starts, ends):
        if (e - s) < min_dur_samples:
            continue
        event_brake = brake_seg[s:e]
        peak        = float(np.max(event_brake))
        if peak < FULL_BRAKE_THRESHOLD_BAR:
            continue
        # フルブレーキングエリア: ピーク圧力 × FULL_BRAKE_PEAK_RATIO 以上
        full_mask       = event_brake >= peak * FULL_BRAKE_PEAK_RATIO
        if not np.any(full_mask):
            continue
        full_spd_indices = spd_indices[s:e][full_mask]  # global speed index

        sF_vals, sR_vals = [], []
        for si in full_spd_indices:
            sF = susp_at_speed_index(sus_f_raw, int(si), susp_ratio)
            sR = susp_at_speed_index(sus_r_raw, int(si), susp_ratio) if len(sus_r_raw) > 0 else float("nan")
            if not math.isnan(sF):
                sF_vals.append(sF)
            if not math.isnan(sR):
                sR_vals.append(sR)

        results.append({
            "susF_mm": round(float(np.mean(sF_vals)), 1) if sF_vals else None,
            "susR_mm": round(float(np.mean(sR_vals)), 1) if sR_vals else None,
        })

    return results


# ── セッション解析メイン ──────────────────────────────────────────────
def analyze_mes(mes_path: Path, event_meta: dict | None = None) -> dict | None:
    """1 MES フォルダを解析して結果 dict を返す。失敗時は None。
    event_meta: _build_event_meta() の戻り値。JA52 HED 補完に使用。
    """
    if event_meta is None:
        event_meta = {}
    base = mes_path.name.replace(".MES", "")

    # ── メタデータ
    hed = parse_hed(mes_path, base)
    if not hed:
        return None

    # ライダー識別 (HED: Rider Number 例 #77 or #52)
    # HED が空の場合はファイル名・フォルダ名からフォールバック
    rider_num = hed.get("Rider Number", "")
    if "77" in rider_num:
        rider_tag = "DA77"
    elif "52" in rider_num or "JA" in rider_num.upper():
        rider_tag = "JA52"
    else:
        # パスから推定 (JA52を優先してチェック)
        fname_up = mes_path.name.upper()
        par_up   = mes_path.parent.name.upper()
        if "JA52" in fname_up or "JA2" in fname_up or "52" in fname_up or par_up in ("52", "JA52"):
            rider_tag = "JA52"
        elif "#77" in fname_up or "DA77" in fname_up or par_up in ("DA77", "77", "DA-77"):
            rider_tag = "DA77"
        else:
            return None   # ライダー不明はスキップ

    # 質量 ([MASS] セクションから優先取得)
    def _to_float(s, default):
        try:
            return float(str(s).split("(")[0].strip() or default)
        except Exception:
            return float(default)
    bike_kg  = _to_float(hed.get("MASS.Bike",  hed.get("MASS.bike",  163)), 163)
    rider_kg = _to_float(hed.get("MASS.Rider", hed.get("MASS.rider", 75)),  75)
    if bike_kg < 50 or bike_kg > 300:   # 数値異常ならデフォルト
        bike_kg = 163
    if rider_kg < 40 or rider_kg > 150:
        rider_kg = 75
    mass_kg  = bike_kg + rider_kg

    # ── ラウンド名の正規化マッピング
    # フォルダ名から抽出した event_key → 統一ラウンド名
    # Data_Bace_TS24_ORIGINAL と SESSION_SUMMARY に合わせた命名規則
    _ROUND_NORM: dict[str, str] = {
        # レースラウンド
        "R01": "ROUND1", "R1": "ROUND1",
        "R02": "ROUND2", "R2": "ROUND2",
        "R03": "ROUND3", "R3": "ROUND3",
        "R04": "ROUND4", "R4": "ROUND4",
        "R05": "ROUND5", "R5": "ROUND5",
        # テスト (フォルダ形式: T02_Jerez, T03_PORTIMAO, T05_CREMONA)
        "T01": "TEST1",  "T1": "TEST1",
        "T02": "TEST2",  "T2": "TEST2",
        "T03": "TEST3",  "T3": "TEST3",
        "T04": "TEST4",  "T4": "TEST4",
        "T05": "TEST5",  "T5": "TEST5",
        "T06": "TEST6",  "T6": "TEST6",
        # テスト (フォルダ形式: 20251203-TEST1-..., 20260121-TEST2-...)
        "TEST1": "TEST1", "TEST2": "TEST2", "TEST3": "TEST3",
        "TEST4": "TEST4", "TEST5": "TEST5", "TEST6": "TEST6",
        # ワークショップ
        "WORKSHOP": "WORKSHOP",
    }

    # セッション情報
    # event (ラウンド名) はフォルダパスから取得し正規化する。
    # HED の Event フィールドは "T04 PHILLIP ISLAND" などブレが多いため使わない。
    _raw_ekey = _event_key_from_path(mes_path)
    event = _ROUND_NORM.get(_raw_ekey or "", "") if _raw_ekey else ""
    if not event:
        # フォルダから取れない場合は HED の Event フィールドを正規化して使用
        # 例: "T04 PHILLIP ISLAND" → T04 → TEST4
        #     "T3_PORTIMAO"        → T3  → TEST3
        #     "R02_PORTIMAO"       → R02 → ROUND2
        hed_event = hed.get("Event", "").strip()
        _m_evt = re.match(r"^(R\d+|T\d+)(?:[^0-9]|$)", hed_event.upper())
        if _m_evt:
            event = _ROUND_NORM.get(_m_evt.group(1), hed_event)
        else:
            # キーそのままで照合 (例: "TEST1", "WORKSHOP")
            event = _ROUND_NORM.get(hed_event.upper(), hed_event)
    session = hed.get("Session", "")
    # ランナンバーはHEDが常に"01"を返すためファイル名末尾の数字を使用
    # 例: D1-#77-03.MES → 3, D2-JA52-09.MES → 9
    run_match = re.search(r"-(\d+)$", base)
    run_no    = int(run_match.group(1)) if run_match else int(hed.get("Run", "1") or 1)
    circuit = hed.get("Circuit", "").upper()
    date_s  = hed.get("Date", "")

    # ── 異形式日付を修正 (例: 25.1127 → 27/11/2025, 26.0215 → 15/02/2026)
    if re.match(r"^\d{2}\.\d{4}$", date_s):
        yy, mmdd = date_s[:2], date_s[3:]
        date_s = f"{mmdd[2:]}/{mmdd[:2]}/20{yy}"   # DD/MM/YYYY

    date_fmt = date_s
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            date_fmt = datetime.strptime(date_s, fmt).strftime("%Y-%m-%d")
            break
        except Exception:
            continue

    # ── セッション種別: ファイル名プレフィックスを HED Session より優先
    sess_map = {
        "FP": "FP",  "F1": "FP",  "F2": "FP",
        "QP": "QP",  "Q1": "QP",  "Q2": "QP",
        "WU": "WUP", "WU1": "WUP","WU2": "WUP",
        "WUP": "WUP","WUP1":"WUP","WUP2":"WUP",
        "R1": "RACE1","R2": "RACE2",
        "RACE1": "RACE1","RACE2": "RACE2",
        "D1": "TEST_D1","D2": "TEST_D2",
        "L1": "TEST_D1","L2": "TEST_D2",
        "SP": "SP",
        "INLAPR1": "RACE1","INLAPR2": "RACE2",
        "D0": "WARMUP","ACCENSIONE": "WARMUP",
    }
    fn_prefix_m = re.match(r"^([A-Za-z0-9]+)", base)
    fn_prefix   = fn_prefix_m.group(1).upper() if fn_prefix_m else ""
    if fn_prefix in sess_map:
        session_type = sess_map[fn_prefix]
    else:
        session_type = sess_map.get(session.upper(), session.upper() or fn_prefix)

    # ── JA52 デフォルト HED 補完
    # Phillip Island/16-02-2026/L1 はJA52ロガーのデフォルト初期値
    # また circuit="?" や空欄の場合もフォルダ情報で補完する
    DEFAULT_CIRCUIT = {"PHILLIP ISLAND", "PHILLIPISLAND"}
    DEFAULT_DATE    = "16/02/2026"
    is_default_meta = (
        (circuit in DEFAULT_CIRCUIT and date_s == DEFAULT_DATE
         and session.upper() in ("L1", "L2", ""))
        or circuit in ("", "?")          # 不明サーキットも補完対象
    )
    if is_default_meta and event_meta:
        ekey = _event_key_from_path(mes_path)
        if ekey and ekey in event_meta:
            em = event_meta[ekey]
            new_circuit = em.get("circuit", "").strip()
            if new_circuit:
                circuit = new_circuit
            new_date = _infer_date_for_session(fn_prefix, em.get("session_dates", {}))
            if new_date:
                date_s = new_date
                # 再パース
                date_fmt = date_s
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
                    try:
                        date_fmt = datetime.strptime(date_s, fmt).strftime("%Y-%m-%d")
                        break
                    except Exception:
                        continue

    # ── 日付ベース補完: event_key が存在しない場合、同日付の他イベントから推定
    if circuit in ("", "?") and event_meta and date_s:
        for em_try in event_meta.values():
            if date_s in em_try["session_dates"].values():
                new_c = em_try.get("circuit", "")
                if new_c:
                    circuit = new_c
                    break

    # ── サーキット名の正規化
    _CIRC_NORM = {
        "PHILLIPISLAND": "PHILLIP ISLAND",
        "PHILLIPISISLAND": "PHILLIP ISLAND",
        "PHILLIP ISLAND": "PHILLIP ISLAND",
        "PHI": "PHILLIP ISLAND",
        "AUSTRALIA": "PHILLIP ISLAND",
    }
    circuit = _CIRC_NORM.get(circuit.upper().strip(), circuit.upper().strip())

    # ── チャンネル定義
    chs = parse_ddd(mes_path, base)
    needed = ["SPEED_FRONT", "SUSP_FRONT", "SUSP_REAR"]
    for ch in needed:
        if ch not in chs or not chs[ch]["ext"]:
            return None   # 必須チャンネルなし

    # ── データ読み込み
    sf_raw    = read_channel(mes_path, base, chs["SPEED_FRONT"])   # km/h
    sus_f_raw = read_channel(mes_path, base, chs["SUSP_FRONT"])    # mm
    sus_r_raw = read_channel(mes_path, base, chs.get("SUSP_REAR", {}))  # mm

    if len(sf_raw) < 10:
        return None

    # SUSP/SPEED サンプル比率
    susp_ratio = max(1, round(len(sus_f_raw) / len(sf_raw))) if len(sus_f_raw) > 0 else 4

    # オプションチャンネル: BRAKE_FRONT（フルブレーキング解析）
    brake_f_raw = np.array([], dtype=np.float32)
    brake_ch_key = next((k for k in chs if k.upper() == "BRAKE_FRONT" and chs[k].get("ext")), None)
    if brake_ch_key:
        brake_f_raw = read_channel(mes_path, base, chs[brake_ch_key])
    # float スケールで対応（BRAKE SR が SPEED SR より低い場合も正しくマッピング）
    brake_scale = len(brake_f_raw) / len(sf_raw) if len(brake_f_raw) > 0 and len(sf_raw) > 0 else 1.0

    # オプションチャンネル: ACC_Y（横G最大点APEX検出 — Primary method）
    acc_y_raw   = np.array([], dtype=np.float32)
    acc_y_ch_key = next((k for k in chs if k.upper() == "ACC_Y" and chs[k].get("ext")), None)
    if acc_y_ch_key:
        acc_y_raw = read_channel(mes_path, base, chs[acc_y_ch_key])
    accy_ratio = max(1, round(len(acc_y_raw) / len(sf_raw))) if len(acc_y_raw) > 0 else 1

    # オプションチャンネル: タイヤ内圧 Front / Rear
    tyre_f_ch  = find_tyre_channel(chs, "F")
    tyre_r_ch  = find_tyre_channel(chs, "R")
    tyre_f_raw = read_channel(mes_path, base, chs[tyre_f_ch]) if tyre_f_ch else np.array([], dtype=np.float32)
    tyre_r_raw = read_channel(mes_path, base, chs[tyre_r_ch]) if tyre_r_ch else np.array([], dtype=np.float32)

    # タイヤ内圧統計量（Run全体で一度だけ計算）
    tyre_f_stats = extract_tyre_pressure(tyre_f_raw)
    tyre_r_stats = extract_tyre_pressure(tyre_r_raw)

    # ── ラップ情報
    n_laps, lap_times_ms = parse_lap(mes_path, base)

    # サンプルレート推定
    if n_laps > 0 and lap_times_ms:
        total_ms = lap_times_ms[-1]
        sr       = len(sf_raw) / (total_ms / 1000.0) if total_ms > 0 else 100.0
    else:
        sr       = 100.0   # デフォルト 100 Hz (LAP なし)

    sr  = max(10.0, min(sr, 500.0))   # 安全クランプ
    dt  = 1.0 / sr                    # 秒/サンプル

    # Speed を m/s に変換 (numpy)
    sf_ms = sf_raw * KMH_TO_MS

    # ── ラップ境界を sample index へ変換
    if n_laps > 0 and lap_times_ms:
        boundaries = []
        prev_ms    = 0
        for t_ms in lap_times_ms:
            s = int(prev_ms / 1000.0 * sr)
            e = int(t_ms   / 1000.0 * sr)
            boundaries.append((s, min(e, len(sf_raw) - 1)))
            prev_ms = t_ms
    else:
        # LAP なし → 全録音を 1 "ラップ" として扱う
        boundaries = [(0, len(sf_raw) - 1)]

    # ── 解析 -------------------------------------------
    apex_results     = []   # (speed_kmh, susF_mm, susR_mm, Nf_N, Nr_N)
    pit_results      = []   # ピットリミッター: (avg_speed, susF_mm, susR_mm)
    brake_results    = []   # (speed_kmh, susF_mm, susR_mm)
    full_brk_results = []   # フルブレーキング: (susF_mm, susR_mm)

    for (lap_start, lap_end) in boundaries:
        lap_duration_s = (lap_end - lap_start) / sr
        if lap_duration_s < MIN_LAP_DURATION_S:
            continue   # アウトラップ/フォーメーションラップ/ピット出入り等を除外
        if lap_end - lap_start < 20:
            continue

        # 1. APEX 検出 (横G最大点 Primary / 速度最小点 Fallback)
        apexes = detect_apexes_accy(acc_y_raw, sf_raw, lap_start, lap_end, sr, accy_ratio)
        for ai in apexes:
            v_kmh = sf_raw[ai]
            if v_kmh < 0:
                continue
            # 加減速度 (m/s²)
            if 0 < ai < len(sf_ms) - 1:
                ax = (sf_ms[ai + 1] - sf_ms[ai - 1]) / (2.0 * dt)
            else:
                ax = 0.0
            Nf, Nr = wheel_forces(rider_tag, mass_kg, ax)
            sF = susp_at_speed_index(sus_f_raw, ai, susp_ratio)
            sR = susp_at_speed_index(sus_r_raw, ai, susp_ratio) if len(sus_r_raw) > 0 else float("nan")
            apex_results.append({
                "speed_kmh": round(v_kmh, 1),
                "susF_mm":   round(sF, 1) if not math.isnan(sF) else None,
                "susR_mm":   round(sR, 1) if not math.isnan(sR) else None,
                "Nf_N":      round(Nf, 1),
                "Nr_N":      round(Nr, 1),
                "ax_ms2":    round(ax, 2),
            })

        # 2. ピットレーンリミッター区間検出
        pit_zones = detect_pit_limiter_zones(sf_raw, lap_start, lap_end, sr)
        for (z0, z1) in pit_zones:
            avg_v   = float(np.mean(sf_raw[z0:z1]))
            sF_mean = susp_mean_in_range(sus_f_raw, z0, z1, susp_ratio)
            sR_mean = susp_mean_in_range(sus_r_raw, z0, z1, susp_ratio) if len(sus_r_raw) > 0 else float("nan")
            pit_results.append({
                "speed_kmh": round(avg_v, 1),
                "susF_mm":   round(sF_mean, 1) if not math.isnan(sF_mean) else None,
                "susR_mm":   round(sR_mean, 1) if not math.isnan(sR_mean) else None,
            })

        # 3. ブレーキ直前検出
        entries = detect_brake_entries(sf_raw, sf_ms, lap_start, lap_end, dt)
        for bi in entries:
            v_kmh = sf_raw[bi] if bi < len(sf_raw) else 0
            sF    = susp_at_speed_index(sus_f_raw, bi, susp_ratio)
            sR    = susp_at_speed_index(sus_r_raw, bi, susp_ratio) if len(sus_r_raw) > 0 else float("nan")
            brake_results.append({
                "speed_kmh": round(v_kmh, 1),
                "susF_mm":   round(sF, 1) if not math.isnan(sF) else None,
                "susR_mm":   round(sR, 1) if not math.isnan(sR) else None,
            })

        # 4. フルブレーキング区間サスペンション (BRAKE_F ≥ peak × FULL_BRAKE_PEAK_RATIO)
        if len(brake_f_raw) > 0:
            fb = detect_full_braking_sus(
                brake_f_raw, sf_raw, sus_f_raw, sus_r_raw,
                lap_start, lap_end, brake_scale, susp_ratio, sr
            )
            full_brk_results.extend(fb)

    # ── 集計
    def agg(lst, key):
        return safe_mean([r[key] for r in lst])

    return {
        "round":                event,
        "date":                 date_fmt,
        "circuit":              circuit,
        "session_type":         session_type,
        "rider":                rider_tag,
        "run_no":               run_no,
        "sample_rate_hz":       round(sr, 1),
        "n_laps_analyzed":      max(1, n_laps) if n_laps > 0 else 1,
        # APEX
        "apex_count":           len(apex_results),
        "apex_speed_avg_kmh":   agg(apex_results, "speed_kmh"),
        "apex_susF_avg_mm":     agg(apex_results, "susF_mm"),
        "apex_susR_avg_mm":     agg(apex_results, "susR_mm"),
        "apex_wheelF_avg_N":    agg(apex_results, "Nf_N"),
        "apex_wheelR_avg_N":    agg(apex_results, "Nr_N"),
        "apex_ax_avg_ms2":      agg(apex_results, "ax_ms2"),
        # ピットレーンリミッター
        "pit_count":            len(pit_results),
        "pit_speed_avg_kmh":    agg(pit_results, "speed_kmh"),
        "pit_susF_avg_mm":      agg(pit_results, "susF_mm"),
        "pit_susR_avg_mm":      agg(pit_results, "susR_mm"),
        # ブレーキ直前
        "brake_entry_count":    len(brake_results),
        "brake_speed_avg_kmh":  agg(brake_results, "speed_kmh"),
        "brake_susF_avg_mm":    agg(brake_results, "susF_mm"),
        "brake_susR_avg_mm":    agg(brake_results, "susR_mm"),
        # フルブレーキング (BRAKE_F ≥ peak × 85%)
        "full_brk_count":       len(full_brk_results),
        "full_brk_susF_avg_mm": agg(full_brk_results, "susF_mm"),
        "full_brk_susR_avg_mm": agg(full_brk_results, "susR_mm"),
        # タイヤ内圧 Front (kPa)
        "tyre_f_start":         tyre_f_stats["start"],
        "tyre_f_end":           tyre_f_stats["end"],
        "tyre_f_avg":           tyre_f_stats["avg"],
        "tyre_f_delta":         tyre_f_stats["delta"],
        # タイヤ内圧 Rear (kPa)
        "tyre_r_start":         tyre_r_stats["start"],
        "tyre_r_end":           tyre_r_stats["end"],
        "tyre_r_avg":           tyre_r_stats["avg"],
        "tyre_r_delta":         tyre_r_stats["delta"],
    }


# ── イベントメタ補完 (JA52 HED が不正な場合 DA77 HED から補完) ─────────
def _event_key_from_path(mes_path: Path) -> str | None:
    """MES パスの親/祖父フォルダ名からイベントキーを抽出 (R01, T02, TEST1 等)
    例: R01 DA77/ → R01, T02_Jerez/DA77/ → T02, 20251203-TEST1-.../ → TEST1
    """
    for folder in (mes_path.parent, mes_path.parent.parent):
        n = folder.name.upper()
        # R0N... or T0N... → 文字列先頭マッチ、後続は非数字で終端
        m = re.match(r"^(R\d+|T\d+)(?:[^0-9]|$)", n)
        if m:
            return m.group(1)
        m = re.search(r"TEST(\d+)", n)
        if m:
            return f"TEST{m.group(1)}"
    return None


def _build_event_meta(root: Path) -> dict:
    """
    DA77/#77 HED ファイルを走査し、
    イベントキー → {circuit, session_dates:{SESSION_PREFIX: date_str}} を返す。
    JA52 のデフォルト HED 補完に使用。
    """
    meta: dict = {}
    for entry in sorted(root.rglob("*.MES")):
        if not entry.is_dir():
            continue
        base = entry.name.replace(".MES", "")
        # DA77/#77 ファイルのみ（JA52/52 を除外）
        # ※ 直近の親フォルダ名のみ参照（祖父フォルダ名にJA52を含む場合も除外しない）
        par_name = entry.parent.name.upper()
        if "52" in base or "JA52" in par_name or par_name in ("52", "JA52"):
            continue
        if not ("77" in base or "77" in entry.parent.name or "DA77" in entry.parent.name.upper()):
            continue

        hed = parse_hed(entry, base)
        if not hed:
            continue
        circuit = hed.get("Circuit", "").strip()
        date    = hed.get("Date", "").strip()
        if not circuit or not date or "?" in date or not re.search(r"\d", date):
            continue

        ekey = _event_key_from_path(entry)
        if not ekey:
            continue

        m2 = re.match(r"^([A-Za-z0-9]+)", base)
        prefix = m2.group(1).upper() if m2 else ""

        # ACCENSIONE / D0 (エンジン暖機) はサーキット判定に使わない
        WARMUP_PFX = {"ACCENSIONE", "D0", "WARMUP"}
        is_warmup = prefix in WARMUP_PFX

        if ekey not in meta:
            meta[ekey] = {"circuit": "", "session_dates": {}, "_from_warmup": True}

        # 回路名: ウォームアップより正規セッションを優先
        if not meta[ekey]["circuit"] or (
                meta[ekey]["_from_warmup"] and not is_warmup):
            meta[ekey]["circuit"] = circuit.upper().strip()
            meta[ekey]["_from_warmup"] = is_warmup

        if prefix and not is_warmup:
            if prefix not in meta[ekey]["session_dates"]:
                meta[ekey]["session_dates"][prefix] = date

    # 内部フラグを除去
    for v in meta.values():
        v.pop("_from_warmup", None)
    return meta


def _infer_date_for_session(fn_prefix: str, session_dates: dict) -> str:
    """
    ファイル名プレフィックスに対応する日付を session_dates から探す。
    完全一致 → FP/QP日 → Race日 → 最初の日付、の順でフォールバック。
    """
    fp_keys   = {"FP", "F1", "F2", "FP1", "FP2"}
    race_keys = {"R1", "R2", "WU", "WU1", "WU2", "WUP", "SP", "INLAPR1", "INLAPR2"}

    if fn_prefix in session_dates:
        return session_dates[fn_prefix]
    if fn_prefix in fp_keys:
        for k in fp_keys:
            if k in session_dates:
                return session_dates[k]
    if fn_prefix in race_keys:
        for k in race_keys:
            if k in session_dates:
                return session_dates[k]
    return list(session_dates.values())[0] if session_dates else ""


# ── MES フォルダ収集 ─────────────────────────────────────────────────
def find_all_mes(root: Path) -> list:
    """DATA 2D 配下の全 .MES フォルダを再帰的に収集"""
    mes_list = []
    for entry in root.rglob("*.MES"):
        if entry.is_dir():
            mes_list.append(entry)
    # 直下 .MES も収集
    for entry in root.iterdir():
        if entry.is_dir() and entry.name.endswith(".MES"):
            if entry not in mes_list:
                mes_list.append(entry)
    return sorted(mes_list)


# ── Excel 書き込み ───────────────────────────────────────────────────
def write_to_excel(results: list, excel_path: Path, sheet_name: str):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ❌ openpyxl not found. pip install openpyxl --break-system-packages")
        return False

    wb = load_workbook(excel_path)

    # シートが既存ならクリア
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── スタイル定義
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")  # 濃紺
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    GRP_FILL  = {
        "info":  PatternFill("solid", fgColor="D6E4F0"),
        "apex":  PatternFill("solid", fgColor="E2EFDA"),
        "slow":  PatternFill("solid", fgColor="FFF2CC"),
        "brake": PatternFill("solid", fgColor="FCE4D6"),
        "fbrk":  PatternFill("solid", fgColor="E8D5F5"),
        "tyreF": PatternFill("solid", fgColor="D1ECF1"),
        "tyreR": PatternFill("solid", fgColor="BEE3EA"),
    }
    GRP_FONT  = {k: Font(bold=True, size=9, color="000000") for k in GRP_FILL}
    CELL_FONT = Font(size=9)
    CENTER    = Alignment(horizontal="center", vertical="center")
    THIN      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── ヘッダー定義
    HEADERS = [
        # (label, key, group, width)
        ("Round",          "round",               "info",  12),
        ("Date",           "date",                "info",  12),
        ("Circuit",        "circuit",             "info",  18),
        ("Session",        "session_type",        "info",  10),
        ("Rider",          "rider",               "info",  8),
        ("Run",            "run_no",              "info",  6),
        ("SR (Hz)",        "sample_rate_hz",      "info",  8),
        ("Laps",           "n_laps_analyzed",     "info",  6),
        # APEX
        ("APEX Count",     "apex_count",          "apex",  10),
        ("APEX Spd (km/h)","apex_speed_avg_kmh",  "apex",  14),
        ("APEX SusF (mm)", "apex_susF_avg_mm",    "apex",  14),
        ("APEX SusR (mm)", "apex_susR_avg_mm",    "apex",  14),
        ("APEX WhlF (N)",  "apex_wheelF_avg_N",   "apex",  14),
        ("APEX WhlR (N)",  "apex_wheelR_avg_N",   "apex",  14),
        ("APEX ax (m/s²)", "apex_ax_avg_ms2",     "apex",  13),
        # ピットレーンリミッター
        ("Pit Count",      "pit_count",           "slow",  10),
        ("Pit Spd (km/h)", "pit_speed_avg_kmh",   "slow",  14),
        ("Pit SusF (mm)",  "pit_susF_avg_mm",      "slow",  14),
        ("Pit SusR (mm)",  "pit_susR_avg_mm",      "slow",  14),
        # ブレーキ直前
        ("Brk Count",      "brake_entry_count",   "brake", 10),
        ("Brk Spd (km/h)", "brake_speed_avg_kmh", "brake", 14),
        ("Brk SusF (mm)",  "brake_susF_avg_mm",   "brake", 14),
        ("Brk SusR (mm)",  "brake_susR_avg_mm",   "brake", 14),
        # フルブレーキング (BRAKE_F ≥ peak × 85%)
        ("FullBrk Cnt",    "full_brk_count",       "fbrk", 10),
        ("FullBrk SusF",   "full_brk_susF_avg_mm", "fbrk", 13),
        ("FullBrk SusR",   "full_brk_susR_avg_mm", "fbrk", 13),
        # タイヤ内圧 Front (Bar)
        ("TyreF St(Bar)",  "tyre_f_start",          "tyreF", 11),
        ("TyreF En(Bar)",  "tyre_f_end",             "tyreF", 11),
        ("TyreF Avg(Bar)", "tyre_f_avg",             "tyreF", 12),
        ("TyreF Δ(Bar)",   "tyre_f_delta",           "tyreF", 10),
        # タイヤ内圧 Rear (Bar)
        ("TyreR St(Bar)",  "tyre_r_start",           "tyreR", 11),
        ("TyreR En(Bar)",  "tyre_r_end",              "tyreR", 11),
        ("TyreR Avg(Bar)", "tyre_r_avg",              "tyreR", 12),
        ("TyreR Δ(Bar)",   "tyre_r_delta",            "tyreR", 10),
    ]

    # ── Row 1: グループ行
    grp_labels = {
        "info":  ("Session Info",           1,  8),
        "apex":  ("APEX Analysis",          9,  15),
        "slow":  ("Pit Lane Limiter",       16, 19),
        "brake": ("Braking Entry",          20, 23),
        "fbrk":  ("Full Braking Sus",       24, 26),
        "tyreF": ("Tyre Pressure Front",    27, 30),
        "tyreR": ("Tyre Pressure Rear",     31, 34),
    }
    ws.row_dimensions[1].height = 18
    for grp, (label, col_s, col_e) in grp_labels.items():
        cell = ws.cell(row=1, column=col_s, value=label)
        cell.fill      = GRP_FILL[grp]
        cell.font      = GRP_FONT[grp]
        cell.alignment = CENTER
        if col_s < col_e:
            ws.merge_cells(
                start_row=1, start_column=col_s,
                end_row=1,   end_column=col_e
            )

    # ── Row 2: 列ヘッダー
    ws.row_dimensions[2].height = 22
    for ci, (label, key, grp, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = THIN
        ws.column_dimensions[get_column_letter(ci)].width = width

    # ── データ行
    for ri, row in enumerate(results, start=3):
        ws.row_dimensions[ri].height = 16
        for ci, (label, key, grp, _) in enumerate(HEADERS, start=1):
            val  = row.get(key)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = CELL_FONT
            cell.alignment = CENTER
            cell.border    = THIN
            # ゼブラ
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

    # ウィンドウ枠固定
    ws.freeze_panes = "A3"

    wb.save(excel_path)
    return True


# ── Main ─────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  TS24 Puccetti — 2D Channel Dynamics Analysis")
    print("=" * 60)

    if not DATA_2D_ROOT.exists():
        print(f"\n❌ DATA 2D フォルダが見つかりません: {DATA_2D_ROOT}")
        return

    if not EXCEL_PATH.exists():
        print(f"\n❌ Excel ファイルが見つかりません: {EXCEL_PATH}")
        return

    # MES 収集
    mes_list = find_all_mes(DATA_2D_ROOT)
    print(f"\n📂 Found {len(mes_list)} MES folders")

    # イベントメタ事前構築 (DA77 HED → JA52 補完用)
    print("🔍 Building event metadata from DA77 HED files...")
    event_meta = _build_event_meta(DATA_2D_ROOT)
    print(f"   → {len(event_meta)} events: {sorted(event_meta.keys())}")

    # 解析
    results = []
    errors  = []
    for i, mes in enumerate(mes_list, 1):
        print(f"  [{i:3d}/{len(mes_list)}] {mes.parent.name}/{mes.name} ... ", end="", flush=True)
        try:
            r = analyze_mes(mes, event_meta)
            if r:
                results.append(r)
                n_apex = r["apex_count"]
                n_pit  = r["pit_count"]
                n_brk  = r["brake_entry_count"]
                print(f"✅ APEX={n_apex} PitLim={n_pit} Brk={n_brk}")
            else:
                print("⏭ skipped (missing channels or metadata)")
        except Exception as e:
            errors.append((str(mes), str(e)))
            print(f"⚠️  {e}")

    print(f"\n✅ Analyzed: {len(results)} sessions  ❌ Errors: {len(errors)}")

    if not results:
        print("\n⚠️  解析結果なし。チャンネルデータを確認してください。")
        return

    # 重複排除 (round+date+circuit+session+rider+run が同一のもの → 後者を優先)
    seen, deduped = {}, []
    for r in results:
        k = (r["round"], r["date"], r["circuit"], r["session_type"], r["rider"], r["run_no"])
        seen[k] = r
    deduped = list(seen.values())
    deduped.sort(key=lambda r: (r["date"], r["rider"], r["run_no"]))

    print(f"\n📝 Writing {len(deduped)} rows → sheet '{OUTPUT_SHEET}' ...")
    ok = write_to_excel(deduped, EXCEL_PATH, OUTPUT_SHEET)

    if ok:
        print(f"\n✅ DYNAMICS_ANALYSIS sheet written to:")
        print(f"   {EXCEL_PATH}")
    else:
        print("\n❌ Excel 書き込み失敗")

    if errors:
        print("\n⚠️  Errors encountered:")
        for path, msg in errors[:10]:
            print(f"   {path}: {msg}")


if __name__ == "__main__":
    main()
