"""
ts24_workbench.py — TS24 Engineer Workbench v2.0
=================================================
PyQt6製ローカルデスクトップアプリ。
Run Browser / Quick Log / Problem Log / Setup Decision / Trend Analysis の5タブ構成。
CSV不要。DBから直接Runを選択してProblemを記録できる。

読み取り: ts24_unified.db, lap_suspension_data.json
書き込み: ts24_unified.db (problem_log / setup_decision_log のみ)

起動: python ts24_workbench.py
      または TS24_Workbench.command をダブルクリック
"""

from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
<<<<<<< Updated upstream
from PyQt6.QtCore import Qt, QFileSystemWatcher
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDoubleSpinBox, QFormLayout,
    QHBoxLayout, QLabel, QMainWindow, QMessageBox, QPushButton, QSizePolicy, QSpinBox,
=======
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDoubleSpinBox, QFileDialog, QFormLayout, QHBoxLayout, QLabel,
    QMainWindow, QMessageBox, QPushButton, QSizePolicy, QSpinBox,
>>>>>>> Stashed changes
    QSplitter, QTabWidget, QTableWidget, QTableWidgetItem,
    QTextEdit, QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
    QLineEdit, QFrame, QScrollArea,
)

# ── パス設定 ──────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
DB_PATH      = SCRIPT_DIR.parent / "02_DATABASE" / "ts24_unified.db"
<<<<<<< Updated upstream
=======
XL_PATH      = SCRIPT_DIR.parent / "02_DATABASE" / "TS24 DB Master.xlsx"
OVERLAY_JSON = SCRIPT_DIR / "lap_overlay_data.json"
TEMPLATES_JSON = SCRIPT_DIR / "turn_templates.json"
>>>>>>> Stashed changes

# ── 定数 ─────────────────────────────────────────────────────────────
PROBLEM_TAGS = [
    "chattering_brake", "front_dive", "line_loss_exit",
    "nervousness", "no_turn_in", "push_rear_exit",
    "understeer_apex", "other",
]

COMPONENTS = [
    "f_preload", "f_comp", "f_reb", "f_spr_l", "f_spr_r",
    "r_preload", "r_comp", "r_reb", "r_spr",
    "ride_hgt", "swing_arm", "f_offset",
    "tyre_front", "tyre_rear", "other",
]

PHASES       = ["PH1", "PH2", "PH3", "PH4", "PH5", "GENERAL"]
SEVERITIES   = ["HIGH", "MEDIUM", "LOW"]
SOURCES      = ["OBSERVATION", "DATA", "RIDER_REPORT"]
CHANGE_TYPES = ["FORK", "SHOCK", "GEOMETRY", "TYRE", "ENGINE", "OTHER"]
RESULT_EVALS = ["POSITIVE", "NEGATIVE", "NEUTRAL", "UNKNOWN"]


# ════════════════════════════════════════════════════════════════════
# ユーティリティ
# ════════════════════════════════════════════════════════════════════

def format_laptime(sec: float) -> str:
    """秒数を 分:秒,00 形式に変換する（表示専用）。

    例: 97.45  → "1:37,45"
        300.63 → "5:00,63"
        65.0   → "1:05,00"
    """
    m = int(sec // 60)
    s = sec % 60
    return f"{m}:{s:05.2f}".replace(".", ",")


# ════════════════════════════════════════════════════════════════════
# DB アクセス層
# ════════════════════════════════════════════════════════════════════

class WorkbenchDB:
    def __init__(self, db_path: Path, xl_path: Path | None = None):
        self.db_path = str(db_path)
        self.xl_path = str(xl_path) if xl_path else str(Path(db_path).parent / "TS24 DB Master.xlsx")
        try:
            with self._conn() as conn:
                self._migrate_problem_log(conn)
        except Exception:
            pass

    def _migrate_problem_log(self, conn: sqlite3.Connection):
        """problem_log テーブルに Phase 2 拡張列を追加（既存DBへの後方互換対応）。"""
        tables = {r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        )}
        if "problem_log" not in tables:
            return
        existing = {row[1] for row in conn.execute("PRAGMA table_info(problem_log)")}
        new_columns = [
            ("distance_start_m", "REAL"),
            ("distance_end_m",   "REAL"),
            ("time_start_s",     "REAL"),
            ("time_end_s",       "REAL"),
            ("data_source_file", "TEXT"),
            ("analysis_note",    "TEXT"),
        ]
        for col_name, col_type in new_columns:
            if col_name not in existing:
                conn.execute(
                    f"ALTER TABLE problem_log ADD COLUMN {col_name} {col_type}"
                )
        conn.commit()

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def get_circuits(self) -> list[str]:
        with self._conn() as conn:
            cur = conn.execute(
                "SELECT DISTINCT circuit FROM runs WHERE circuit IS NOT NULL ORDER BY circuit"
            )
            return [r[0] for r in cur.fetchall()]

    def get_runs(self, circuit: str | None = None) -> list[dict]:
        with self._conn() as conn:
            if circuit:
                cur = conn.execute(
                    """SELECT run_id, circuit, session, rider, run_no, perf_best_lap
                       FROM runs WHERE circuit = ? ORDER BY session, rider, run_no""",
                    (circuit,),
                )
            else:
                cur = conn.execute(
                    """SELECT run_id, circuit, session, rider, run_no, perf_best_lap
                       FROM runs ORDER BY circuit, session, rider, run_no"""
                )
            return [dict(r) for r in cur.fetchall()]

    def get_run(self, run_id: str) -> dict:
        """単一 run の全フィールドを返す。"""
        with self._conn() as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM runs WHERE run_id = ?", (run_id,)).fetchone()
            return dict(row) if row else {}

    def get_next_runs(self, run_id: str) -> list[dict]:
        """同一 circuit・rider の次の run 候補を返す。"""
        with self._conn() as conn:
            cur = conn.execute(
                "SELECT circuit, rider FROM runs WHERE run_id = ?", (run_id,)
            )
            row = cur.fetchone()
            if not row:
                return []
            cur2 = conn.execute(
                """SELECT run_id, session, run_no FROM runs
                   WHERE circuit = ? AND rider = ? AND run_id != ?
                   ORDER BY session, run_no""",
                (row["circuit"], row["rider"], run_id),
            )
            return [dict(r) for r in cur2.fetchall()]

    def get_problem_logs(self, run_id: str) -> list[dict]:
        """problem_log を取得。スキーマ拡張後も安全に動作するよう SELECT * を使用。"""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    "SELECT * FROM problem_log WHERE run_id = ? ORDER BY created_at DESC",
                    (run_id,),
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception as e:
            print("[WorkbenchDB] get_problem_logs error:", e)
            return []

    def add_problem_log(self, data: dict) -> int:
        with self._conn() as conn:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """INSERT INTO problem_log
                   (run_id, round, circuit, session, rider, run_no, lap_no,
                    corner, phase, problem_tag, description, severity, source,
                    distance_start_m, distance_end_m, time_start_s, time_end_s,
                    data_source_file, analysis_note,
                    export_status, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    data.get("run_id"), data.get("round"), data.get("circuit"),
                    data.get("session"), data.get("rider"), data.get("run_no"),
                    data.get("lap_no"), data.get("corner"), data.get("phase"),
                    data.get("problem_tag"), data.get("description"),
                    data.get("severity", "MEDIUM"), data.get("source", "OBSERVATION"),
                    data.get("distance_start_m"), data.get("distance_end_m"),
                    data.get("time_start_s"), data.get("time_end_s"),
                    data.get("data_source_file", ""), data.get("analysis_note", ""),
                    "PENDING", now, now,
                ),
            )
            conn.commit()
            return conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    def delete_problem_log(self, problem_id: int):
        with self._conn() as conn:
            conn.execute("DELETE FROM problem_log WHERE problem_id = ?", (problem_id,))
            conn.commit()

    def get_setup_decisions(self, run_id: str) -> list[dict]:
        with self._conn() as conn:
            cur = conn.execute(
                """SELECT * FROM setup_decision_log WHERE run_id_from = ?
                   ORDER BY created_at DESC""",
                (run_id,),
            )
            return [dict(r) for r in cur.fetchall()]

    def add_setup_decision(self, data: dict) -> int:
        with self._conn() as conn:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """INSERT INTO setup_decision_log
                   (run_id_from, run_id_to, round, circuit, session, rider,
                    change_type, component, from_value, to_value,
                    rationale, expected_effect, actual_effect, result_eval,
                    export_status, created_at, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    data.get("run_id_from"), data.get("run_id_to"),
                    data.get("round"), data.get("circuit"),
                    data.get("session"), data.get("rider"),
                    data.get("change_type"), data.get("component"),
                    data.get("from_value"), data.get("to_value"),
                    data.get("rationale"), data.get("expected_effect"),
                    data.get("actual_effect"), data.get("result_eval", "UNKNOWN"),
                    "PENDING", now, now,
                ),
            )
            conn.commit()
            return conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    def update_setup_decision(self, decision_id: int, actual_effect: str, result_eval: str):
        with self._conn() as conn:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                """UPDATE setup_decision_log
                   SET actual_effect = ?, result_eval = ?, updated_at = ?
                   WHERE decision_id = ?""",
                (actual_effect, result_eval, now, decision_id),
            )
            conn.commit()

    def get_laps(self, run_id: str) -> list[dict]:
        """laps テーブルから指定 run_id の全ラップを返す。"""
        try:
            with self._conn() as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute(
                    """SELECT lap_no, lap_time_s, is_outlap
                       FROM laps WHERE run_id = ? ORDER BY lap_no""",
                    (run_id,),
                ).fetchall()
                return [dict(r) for r in rows]
        except Exception:
            return []

<<<<<<< Updated upstream
=======
    # ── Trend 分析用メソッド ──────────────────────────────────────────

    def get_all_rounds(self) -> list[str]:
        """全ラウンド一覧を返す。"""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    "SELECT DISTINCT round FROM runs WHERE round IS NOT NULL ORDER BY round"
                )
                return [r[0] for r in cur.fetchall()]
        except Exception:
            return []

    def get_trend_laps(self, rider: str | None = None,
                       round_s: str | None = None,
                       session_s: str | None = None) -> list[dict]:
        """条件フィルタ付きで laps を取得。"""
        conds, params = [], []
        if rider:    conds.append("rider = ?");    params.append(rider)
        if round_s:  conds.append("round = ?");    params.append(round_s)
        if session_s: conds.append("session = ?"); params.append(session_s)
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    f"SELECT * FROM laps {where} ORDER BY round, session, run_no, lap_no",
                    params,
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception:
            return []

    def get_trend_runs(self, rider: str | None = None,
                       round_s: str | None = None,
                       session_s: str | None = None) -> list[dict]:
        """条件フィルタ付きで runs（サマリー列）を取得。"""
        conds, params = [], []
        if rider:    conds.append("rider = ?");    params.append(rider)
        if round_s:  conds.append("round = ?");    params.append(round_s)
        if session_s: conds.append("session = ?"); params.append(session_s)
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    f"""SELECT run_id, round, circuit, session, rider, run_no, date,
                               perf_best_lap, perf_avg_lap, perf_n_laps,
                               perf_rank, perf_gap_s, tyre_front, tyre_rear, comment
                        FROM runs {where} ORDER BY round, session, run_no""",
                    params,
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception:
            return []

    def get_trend_runs_detail(self, run_ids: list[str]) -> list[dict]:
        """指定 run_id リストの runs 全列を取得（セットアップ表示用）。"""
        if not run_ids:
            return []
        placeholders = ",".join("?" * len(run_ids))
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    f"SELECT * FROM runs WHERE run_id IN ({placeholders}) ORDER BY round, session, run_no",
                    run_ids,
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception:
            return []

    def get_trend_problems(self, rider: str | None = None,
                           round_s: str | None = None,
                           session_s: str | None = None) -> list[dict]:
        """条件フィルタ付きで problem_log を取得。"""
        conds, params = [], []
        if rider:    conds.append("rider = ?");    params.append(rider)
        if round_s:  conds.append("round = ?");    params.append(round_s)
        if session_s: conds.append("session = ?"); params.append(session_s)
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    f"SELECT * FROM problem_log {where} ORDER BY created_at",
                    params,
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception:
            return []

    def get_lap_suspension(self, rider: str | None = None,
                           round_s: str | None = None,
                           session_s: str | None = None) -> list[dict]:
        """lap_suspension テーブルをフィルタ付きで取得。"""
        conds, params = [], []
        if rider:     conds.append("rider = ?");   params.append(rider)
        if round_s:   conds.append("round = ?");   params.append(round_s)
        if session_s: conds.append("session = ?"); params.append(session_s)
        where = f"WHERE {' AND '.join(conds)}" if conds else ""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    f"SELECT * FROM lap_suspension {where} "
                    f"ORDER BY round, session, run_no, lap_no",
                    params,
                )
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception:
            return []

    def get_perf_correlation(self) -> tuple[list, list, list, list]:
        """PERFORMANCE_CORRELATION シートから per-run データと FAST/SLOW 比較データを返す。
        Returns: (run_headers, run_rows, cmp_headers, cmp_rows)
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.xl_path, read_only=True, data_only=True)
            ws = wb["PERFORMANCE_CORRELATION"]
            all_rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
            wb.close()
        except Exception as e:
            print(f"[get_perf_correlation] load error: {e}")
            return [], [], [], []

        run_hdrs, run_data, cmp_hdrs, cmp_data = [], [], [], []
        section = None
        for r in all_rows:
            non_null = [v for v in r if v is not None]
            if not non_null:
                continue
            # per-run ヘッダー行を検出 (RUN_ID / Rider / Circuit が並ぶ行)
            if "Rider" in str(r) and "Circuit" in str(r) and "Best Lap" in str(r) and len(non_null) >= 8:
                run_hdrs = [str(v).replace("\n", " ").strip() if v else "" for v in r]
                section = "run"
                continue
            # FAST/SLOW 比較ヘッダー行 (★FAST が含まれる)
            if "★FAST" in str(r) or "FAST" in str(r[5] or ""):
                cmp_hdrs = [str(v).replace("\n", " ").strip() if v else "" for v in r]
                section = "cmp"
                continue
            # データ行
            if section == "run" and run_hdrs:
                # col B (index 1) に run_id 形式の文字列がある行
                if r[1] and str(r[1]).count("_") >= 3:
                    d = dict(zip(run_hdrs, r))
                    # Rider カラムが有効な場合のみ
                    if d.get("Rider") in ("DA77", "JA52"):
                        run_data.append(d)
            elif section == "cmp" and cmp_hdrs:
                if r[2] in ("DA77", "JA52"):  # Rider col
                    d = dict(zip(cmp_hdrs, r))
                    cmp_data.append(d)
        return run_hdrs, run_data, cmp_hdrs, cmp_data

    def get_trend_notes(self) -> dict:
        """TREND_ANALYSIS シートからタグデータとエンジニアノートを辞書で返す。
        Returns: {"tags": [(tag, count, phase, meaning), ...],
                  "rider_tags": {"JA52": [...], "DA77": [...]},
                  "notes": [(session_key, note_text), ...]}
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.xl_path, read_only=True, data_only=True)
            ws = wb["TREND_ANALYSIS"]
            all_rows = list(ws.iter_rows(min_row=1, max_row=120, values_only=True))
            wb.close()
        except Exception as e:
            print(f"[get_trend_notes] load error: {e}")
            return {"tags": [], "rider_tags": {}, "notes": []}

        tags, rider_tags, notes = [], {}, []
        section = None
        current_rider = None

        for r in all_rows:
            if not any(r):
                continue
            v0 = str(r[0] or "").strip()
            v2 = str(r[2] or "").strip() if r[2] else ""

            # セクション検出
            if "TOP PROBLEM TAGS" in v0:
                section = "top_tags"; continue
            if "RIDER COMPARISON" in v0:
                section = "rider_cmp"; continue
            if "KEY ENGINEER NOTES" in v0:
                section = "notes"; continue

            if section == "top_tags":
                if v0 == "TAG" or not v0:
                    continue
                count = r[1]; phase = r[2]; meaning = r[3]
                if v0 and count is not None:
                    try:
                        tags.append((v0, int(count), str(phase or ""), str(meaning or "")))
                    except Exception:
                        pass

            elif section == "rider_cmp":
                if "Rider:" in v0:
                    m = __import__("re").search(r"Rider:\s*(\w+)", v0)
                    if m:
                        current_rider = m.group(1)
                        rider_tags[current_rider] = []
                    continue
                if v0 == "TAG" or not v0:
                    continue
                if current_rider and r[1] is not None:
                    circuits = str(r[2] or "")
                    try:
                        rider_tags[current_rider].append(
                            (v0, int(r[1]), circuits)
                        )
                    except Exception:
                        pass

            elif section == "notes":
                # 形式: col0=session_key (e.g. "20260501-ROUND4-DA77"), col2=note
                if v0 and (v2 or len(v0) > 10):
                    notes.append((v0, v2 if v2 else str(r[1] or "")))

        return {"tags": tags, "rider_tags": rider_tags, "notes": notes}

    def lookup_run_id(self, round_s: str, session_s: str, rider_s: str, run_no: int) -> str:
        """ファイル名構成要素 (ROUND/SESSION/RIDER/RUN_NO) から DB の run_id を検索する。"""
        try:
            with self._conn() as conn:
                cur = conn.execute(
                    """SELECT run_id FROM runs
                       WHERE round = ? AND session = ? AND rider = ? AND run_no = ?
                       LIMIT 1""",
                    (round_s, session_s, rider_s, run_no),
                )
                row = cur.fetchone()
                if row:
                    return row[0]
        except Exception as e:
            print("[WorkbenchDB] lookup_run_id error:", e)
        return ""


# ════════════════════════════════════════════════════════════════════
# 波形ビュー (Speed / Brake / Gas — Reference only)
# ════════════════════════════════════════════════════════════════════

class WaveformView(QWidget):
    def __init__(self, db: "WorkbenchDB | None" = None, parent=None):
        super().__init__(parent)
        self._db_ref = db
        self._overlay_data: list[dict] = []
        self._templates: dict = {}
        self._circuit: str = ""
        self._csv_x_mode: str = "progress"   # "distance" | "time" | "progress"
        self._laps_cache: list[dict] = []
        # 2ライダー比較用
        self._laps_cache_b: list[dict] = []   # 比較ライダーのラップキャッシュ
        self._label_a: str = ""               # プライマリライダー名（例: "DA77"）
        self._label_b: str = ""               # 比較ライダー名（例: "JA52"）
        self._offset_b: float = 0.0           # Bライダーの時間オフセット（秒）
        self._problem_tab: "ProblemLogTab | None" = None
        self._run_id_wave: str = ""
        self._setup_ui()
        self._load_static_data()

    def set_problem_tab(self, tab: "ProblemLogTab") -> None:
        """MainWindow から呼ばれ、Problem Log タブへの参照を設定する。"""
        self._problem_tab = tab

    def set_circuit(self, circuit: str) -> None:
        """サーキット名をセット（コーナーテンプレート適用用）。"""
        self._circuit = circuit

    def _setup_ui(self):
        try:
            import pyqtgraph as pg
            self._pg = pg
            self._has_pg = True
        except ImportError:
            self._has_pg = False

        layout = QVBoxLayout(self)

        # Reference warning (hidden in time mode, visible in progress mode)
        self._lbl_warn = QLabel(
            "⚠️  Reference only — time-normalized data, not track-position aligned."
        )
        self._lbl_warn.setStyleSheet("color: #D83B01; font-style: italic; padding: 4px;")
        self._lbl_warn.setFixedHeight(24)
        layout.addWidget(self._lbl_warn)

        # X-axis mode indicator
        self._lbl_xmode = QLabel("X axis: Lap Progress (0–1)")
        self._lbl_xmode.setStyleSheet(
            "color: #107C10; font-size: 10px; padding: 2px 4px;"
            " background: #F0FFF0; border-radius: 3px;"
        )
        self._lbl_xmode.setFixedHeight(22)
        layout.addWidget(self._lbl_xmode)

        # Lap selectors
        sel_row = QHBoxLayout()
        sel_row.addWidget(QLabel("Lap A:"))
        self._combo_a = QComboBox()
        self._combo_a.setMinimumWidth(260)
        sel_row.addWidget(self._combo_a)
        sel_row.addSpacing(16)
        sel_row.addWidget(QLabel("Lap B:"))
        self._combo_b = QComboBox()
        self._combo_b.setMinimumWidth(260)
        sel_row.addWidget(self._combo_b)
        btn = QPushButton("表示更新")
        btn.clicked.connect(self._draw)
        sel_row.addWidget(btn)
        btn_send_log = QPushButton("📋  Problem Log へ送る")
        btn_send_log.setToolTip(
            "選択範囲（青いハイライト）の座標情報を Problem Log に自動入力します。\n"
            "Lap A の run_id / lap_no / time / distance が入力されます。"
        )
        btn_send_log.setStyleSheet(
            "QPushButton { background: #107C10; color: white; padding: 4px 12px;"
            " border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background: #0D6A0D; }"
        )
        btn_send_log.clicked.connect(self._send_to_problem_log)
        sel_row.addSpacing(24)
        sel_row.addWidget(btn_send_log)

        if self._has_pg:
            pg = self._pg
            pg.setConfigOption("background", "w")
            pg.setConfigOption("foreground", "k")

            # 個別 PlotWidget — show/hide 対応
            self._pw_speed = pg.PlotWidget(title="Speed (km/h)")
            self._pw_brake = pg.PlotWidget(title="Brake (bar)")
            self._pw_gas   = pg.PlotWidget(title="Gas (%)")
            self._pw_suspf = pg.PlotWidget(title="SUSP_FRONT (mm)")
            self._pw_suspr = pg.PlotWidget(title="SUSP_REAR (mm)")

            self._pw_speed.setMinimumHeight(120)
            for _pw in [self._pw_brake, self._pw_gas, self._pw_suspf, self._pw_suspr]:
                _pw.setMinimumHeight(80)

            # PlotItem エイリアス（_draw() 等の後方互換）
            self._p_speed = self._pw_speed.getPlotItem()
            self._p_brake = self._pw_brake.getPlotItem()
            self._p_gas   = self._pw_gas.getPlotItem()
            self._p_suspf = self._pw_suspf.getPlotItem()
            self._p_suspr = self._pw_suspr.getPlotItem()

            # X軸リンク（全パネルを Speed に同期）
            for _pw in [self._pw_brake, self._pw_gas, self._pw_suspf, self._pw_suspr]:
                _pw.setXLink(self._pw_speed)

            # _all_plots: PlotWidget タプル（clear/plot/enableAutoRange 共通操作用）
            self._all_plots = (
                self._pw_speed, self._pw_brake, self._pw_gas,
                self._pw_suspf, self._pw_suspr,
            )
            for _pw in self._all_plots:
                _pw.setLabel("bottom", "Lap Progress")
                _pw.showGrid(x=True, y=True, alpha=0.3)
                _pw.setXRange(0, 1)

            # LinearRegionItem（選択範囲ハイライト）
            self._region = pg.LinearRegionItem(
                values=[0, 100],
                brush=pg.mkBrush(0, 120, 212, 30),
                pen=pg.mkPen("#0078D4", width=1.5),
                movable=True,
            )
            self._region.setZValue(10)
            self._pw_speed.addItem(self._region)

            # チャンネルチェックボックス（sel_row に追加）
            sep = QLabel("  |  チャンネル:")
            sep.setStyleSheet("font-size: 10px; color: #666;")
            sel_row.addWidget(sep)
            self._ch_checks: dict = {}
            for _name, _pw in [
                ("Speed",  self._pw_speed),
                ("Brake",  self._pw_brake),
                ("Gas",    self._pw_gas),
                ("SUSP_F", self._pw_suspf),
                ("SUSP_R", self._pw_suspr),
            ]:
                cb = QCheckBox(_name)
                cb.setChecked(True)
                cb.setStyleSheet("font-size: 10px;")
                cb.toggled.connect(lambda checked, w=_pw: w.setVisible(checked))
                sel_row.addWidget(cb)
                self._ch_checks[_name] = (cb, _pw)

            # ── 比較ライダー表示ラベル ──────────────────────────────────
            self._lbl_b_rider = QLabel("")
            self._lbl_b_rider.setStyleSheet(
                "color: #FF8C00; font-weight: bold; font-size: 10px;"
                " background: #FFF3E0; padding: 2px 6px; border-radius: 3px;"
            )
            self._lbl_b_rider.setVisible(False)
            sel_row.addSpacing(16)
            sel_row.addWidget(self._lbl_b_rider)

            # ── Bオフセットコントロール ────────────────────────────────
            lbl_off = QLabel("  Bオフセット:")
            lbl_off.setStyleSheet("font-size: 10px; color: #666;")
            self._offset_spin = QDoubleSpinBox()
            self._offset_spin.setRange(-600.0, 600.0)
            self._offset_spin.setSingleStep(0.5)
            self._offset_spin.setSuffix(" s")
            self._offset_spin.setValue(0.0)
            self._offset_spin.setDecimals(1)
            self._offset_spin.setFixedWidth(90)
            self._offset_spin.setToolTip("比較ライダー(B)の時間軸をずらして位置合わせ")
            self._offset_spin.valueChanged.connect(self._on_offset_changed)

            btn_reset_off = QPushButton("↺")
            btn_reset_off.setFixedWidth(28)
            btn_reset_off.setFixedHeight(22)
            btn_reset_off.setToolTip("オフセットを 0 にリセット")
            btn_reset_off.clicked.connect(lambda: self._offset_spin.setValue(0.0))

            btn_clear_b = QPushButton("✕ 比較解除")
            btn_clear_b.setFixedHeight(22)
            btn_clear_b.setStyleSheet(
                "QPushButton { background: #797673; color: white; padding: 2px 8px;"
                " border-radius: 3px; font-size: 10px; }"
                "QPushButton:hover { background: #5C5A58; }"
            )
            btn_clear_b.setToolTip("比較CSVをクリア")
            btn_clear_b.clicked.connect(self.clear_compare)

            sel_row.addWidget(lbl_off)
            sel_row.addWidget(self._offset_spin)
            sel_row.addWidget(btn_reset_off)
            sel_row.addSpacing(8)
            sel_row.addWidget(btn_clear_b)

            sel_row.addStretch()
            layout.addLayout(sel_row)

            # スクロールエリア（PlotWidget 縦積み）
            self._wave_scroll = QScrollArea()
            self._wave_scroll.setWidgetResizable(True)
            _wave_container = QWidget()
            _wave_vlay = QVBoxLayout(_wave_container)
            _wave_vlay.setSpacing(0)
            _wave_vlay.setContentsMargins(0, 0, 0, 0)
            for _pw in self._all_plots:
                _wave_vlay.addWidget(_pw)
            self._wave_scroll.setWidget(_wave_container)

            # QSplitter: 左(波形スクロール) + 右(Problem入力パネル)
            self._wave_splitter = QSplitter(Qt.Orientation.Horizontal)
            self._wave_splitter.addWidget(self._wave_scroll)
            if self._db_ref is not None:
                self._right_panel = _ProblemRightPanel(
                    db=self._db_ref,
                    on_close=self._close_right_panel,
                )
                self._wave_splitter.addWidget(self._right_panel)
                self._wave_splitter.setStretchFactor(0, 3)
                self._wave_splitter.setStretchFactor(1, 1)
                self._wave_splitter.setSizes([1, 0])
            layout.addWidget(self._wave_splitter, 1)  # stretch=1: 残りの縦スペースをすべて占有
        else:
            sel_row.addStretch()
            layout.addLayout(sel_row)
            layout.addWidget(QLabel(
                "pyqtgraph が見つかりません。\n"
                "pip install pyqtgraph でインストールしてください。"
            ))

    def _load_static_data(self):
        if OVERLAY_JSON.exists():
            try:
                self._overlay_data = json.loads(OVERLAY_JSON.read_text(encoding="utf-8"))
            except Exception:
                self._overlay_data = []
        if TEMPLATES_JSON.exists():
            try:
                self._templates = json.loads(TEMPLATES_JSON.read_text(encoding="utf-8"))
            except Exception:
                self._templates = {}

    def set_run(self, run_id: str, circuit: str):
        self._run_id_wave = run_id
        self._circuit = circuit
        self._csv_x_mode = "progress"
        self._lbl_warn.setVisible(True)
        self._lbl_xmode.setText("X axis: Lap Progress (0–1)")
        self._lbl_xmode.setStyleSheet(
            "color: #107C10; font-size: 10px; padding: 2px 4px;"
            " background: #F0FFF0; border-radius: 3px;"
        )
        if self._has_pg:
            for p in self._all_plots:
                p.setLabel("bottom", "Lap Progress")
        self._combo_a.clear()
        self._combo_b.clear()
        if not self._overlay_data:
            return
        laps = [r for r in self._overlay_data if r.get("run_id") == run_id]
        labels = [
            f"Lap {r.get('lap_no','?')}  {r.get('lap_time_s','?')}s"
            for r in laps
        ]
        self._combo_a.addItems(labels)
        self._combo_b.addItems(labels)
        if len(labels) > 1:
            self._combo_b.setCurrentIndex(1)
        self._laps_cache = laps

    def set_csv_laps(self, laps: list[dict]):
        """CSV インポートからのラップデータを波形に設定する（§0 参考値）。

        laps 要素の形式:
          {"x": [...], "speed": [...], "brake": [...],
           "gas": [...], "susp_front": [...], "susp_rear": [...],
           "x_mode": "time" | "progress", "lap_no": int, "lap_time_s": float}
        """
        self._laps_cache = laps
        self._circuit = ""
        self._csv_x_mode = laps[0].get("x_mode", "progress") if laps else "progress"
        if self._has_pg:
            if self._csv_x_mode == "distance":
                x_label    = "Distance (m)"
                mode_text  = "X axis: Distance (m)  [CSV]"
                mode_style = (
                    "color: #107C10; font-size: 10px; padding: 2px 4px;"
                    " background: #F0FFF0; border-radius: 3px;"
                )
                self._lbl_warn.setVisible(False)
            elif self._csv_x_mode == "time":
                x_label    = "Time (s)"
                mode_text  = "X axis: Time (s)  [CSV]"
                mode_style = (
                    "color: #0078D4; font-size: 10px; padding: 2px 4px;"
                    " background: #EFF6FF; border-radius: 3px;"
                )
                self._lbl_warn.setVisible(False)
            else:
                x_label    = "Lap Progress (0–1)  [fallback]"
                mode_text  = "X axis: Normalized Progress (0–1)  [CSV — Dist/Time not found]"
                mode_style = (
                    "color: #797673; font-size: 10px; padding: 2px 4px;"
                    " background: #FAF9F8; border-radius: 3px;"
                )
                self._lbl_warn.setVisible(True)
            for p in self._all_plots:
                p.setLabel("bottom", x_label)
            self._lbl_xmode.setText(mode_text)
            self._lbl_xmode.setStyleSheet(mode_style)
        self._combo_a.clear()
        self._combo_b.clear()
        labels = []
        for i, r in enumerate(laps):
            lap_no = r.get("lap_no", i + 1)
            lt = r.get("lap_time_s")
            lt_str = format_laptime(float(lt)) if lt else "?:??,-"
            x_mode_r = r.get("x_mode", "progress")
            if x_mode_r == "distance":
                dist_m = r.get("dist_span_m", 0.0)
                labels.append(f"CSV Lap {lap_no}  {dist_m:.0f}m  ({lt_str})")
            else:
                labels.append(f"CSV Lap {lap_no}  {lt_str}")
        self._combo_a.addItems(labels)
        self._combo_b.addItems(labels)
        if len(labels) > 1:
            self._combo_b.setCurrentIndex(1)
        # プライマリCSVが更新されたとき、比較CSVがなければ _combo_b も同期更新
        if not self._laps_cache_b:
            self._update_combo_b()

    # ── 2ライダー比較 API ──────────────────────────────────────────────

    def set_label_a(self, label: str) -> None:
        """プライマリCSV（A）のライダーラベルを設定。"""
        self._label_a = label

    def set_compare_laps(self, laps_b: list, label_b: str) -> None:
        """比較ライダー(B)のラップデータをセット。_combo_b を更新して再描画。"""
        self._laps_cache_b = laps_b
        self._label_b = label_b
        self._update_combo_b()
        self._lbl_b_rider.setText(f"B: {label_b}  ({len(laps_b)} laps)")
        self._lbl_b_rider.setVisible(True)
        self._draw()

    def clear_compare(self) -> None:
        """比較ライダーのデータをクリアし、通常モードに戻す。"""
        self._laps_cache_b = []
        self._label_b = ""
        self._offset_b = 0.0
        self._offset_spin.setValue(0.0)
        self._lbl_b_rider.setVisible(False)
        self._update_combo_b()
        self._draw()

    def _on_offset_changed(self, v: float) -> None:
        """オフセットスピンボックス変更時。"""
        self._offset_b = float(v)
        self._draw()

    def _update_combo_b(self) -> None:
        """_laps_cache_b が存在する場合はそちらを、なければ _laps_cache を _combo_b に表示。"""
        src = self._laps_cache_b if self._laps_cache_b else self._laps_cache
        self._combo_b.blockSignals(True)
        self._combo_b.clear()
        for i, r in enumerate(src):
            lap_no = r.get("lap_no", i + 1)
            lt = r.get("lap_time_s")
            lt_str = format_laptime(float(lt)) if lt else "?:??,-"
            xm = r.get("x_mode", "")
            if xm == "distance":
                dist_m = float(r.get("dist_span_m", 0))
                self._combo_b.addItem(f"CSV Lap {lap_no}  {dist_m:.0f}m  ({lt_str})")
            else:
                self._combo_b.addItem(f"CSV Lap {lap_no}  {lt_str}")
        self._combo_b.blockSignals(False)

    def _send_to_problem_log(self) -> None:
        """選択範囲の座標情報を ProblemLogTab に送り、自動入力させる。"""
        if self._problem_tab is None:
            QMessageBox.warning(self, "未接続", "Problem Log タブが接続されていません。")
            return
        if not self._laps_cache:
            QMessageBox.warning(self, "データなし", "波形データがありません。先に CSV を送信してください。")
            return

        ia = self._combo_a.currentIndex()
        if ia < 0 or ia >= len(self._laps_cache):
            QMessageBox.warning(self, "Lap未選択", "Lap A を選択してください。")
            return

        lap_a = self._laps_cache[ia]
        x_start, x_end = self._region.getRegion()
        if x_start > x_end:
            x_start, x_end = x_end, x_start

        x_mode = self._csv_x_mode

        data: dict = {
            "run_id":  self._run_id_wave,
            "lap_no":  lap_a.get("lap_no"),
            "x_mode":  x_mode,
        }

        if x_mode == "distance":
            data["distance_start_m"] = round(float(x_start), 1)
            data["distance_end_m"]   = round(float(x_end),   1)
            data["time_start_s"]     = None
            data["time_end_s"]       = None
        elif x_mode == "time":
            data["time_start_s"]     = round(float(x_start), 3)
            data["time_end_s"]       = round(float(x_end),   3)
            data["distance_start_m"] = None
            data["distance_end_m"]   = None
        else:
            data["time_start_s"]     = round(float(x_start), 4)
            data["time_end_s"]       = round(float(x_end),   4)
            data["distance_start_m"] = None
            data["distance_end_m"]   = None

        data["data_source_file"] = lap_a.get("source_file", "")

        # 右パネルに送る（波形を維持したまま入力できる）
        if hasattr(self, "_right_panel"):
            self._right_panel.prefill_from_waveform(data)
            self._open_right_panel()

        # Problem Log タブにも送る（互換性維持）
        if self._problem_tab is not None:
            self._problem_tab.prefill_from_waveform(data)

    def _open_right_panel(self) -> None:
        if hasattr(self, "_wave_splitter"):
            total = self._wave_splitter.width()
            self._wave_splitter.setSizes([int(total * 0.65), int(total * 0.35)])

    def _close_right_panel(self) -> None:
        if hasattr(self, "_wave_splitter"):
            total = self._wave_splitter.width()
            self._wave_splitter.setSizes([total, 0])

    def _draw(self):
        if not self._has_pg or not self._laps_cache:
            return
        pg = self._pg
        import numpy as np

        ia = self._combo_a.currentIndex()
        ib = self._combo_b.currentIndex()
        if ia < 0 or ia >= len(self._laps_cache):
            return
        lap_a = self._laps_cache[ia]
        # B ラップ: 比較CSVがあればそちらから、なければ同一CSVから
        if self._laps_cache_b:
            lap_b = self._laps_cache_b[ib] if (0 <= ib < len(self._laps_cache_b)) else None
            pen_b = pg.mkPen("#FF8C00", width=1.5)   # オレンジ = 比較ライダー
        else:
            lap_b = self._laps_cache[ib] if (0 <= ib < len(self._laps_cache)) else None
            pen_b = pg.mkPen("#E74C3C", width=1.5)   # 赤 = 同一CSV内比較（従来）
        colors = {"a": pg.mkPen("#0078D4", width=2), "b": pen_b}
        x_mode = self._csv_x_mode

        for p in self._all_plots:
            p.clear()
        if hasattr(self, "_region"):
            self._pw_speed.addItem(self._region)

        def _get_x(lap):
            ch = lap.get("channels", {})
            # CSV laps use "x" key; overlay laps use "lap_progress"
            return (ch.get("x") or lap.get("x")
                    or ch.get("lap_progress") or lap.get("lap_progress"))

        def _get_y(lap, channel):
            ch = lap.get("channels", {})
            return ch.get(channel) or lap.get(channel)

        def _normalize(xs_raw):
            arr = np.array(xs_raw, dtype=float)
            lo, hi = arr.min(), arr.max()
            return (arr - lo) / (hi - lo) if hi > lo else np.zeros_like(arr)

        def _plot(lap, label, pen, channel, plot_obj, offset_x: float = 0.0):
            xs_raw = _get_x(lap)
            ys_raw = _get_y(lap, channel)
            if not xs_raw or not ys_raw or len(xs_raw) != len(ys_raw):
                return
            if x_mode in ("time", "distance"):
                xs = np.array(xs_raw, dtype=float) + offset_x
            else:
                xs = _normalize(xs_raw)
            ys = np.array(ys_raw, dtype=float)
            plot_obj.plot(x=xs, y=ys, pen=pen, name=label)

        _CHAN_PANELS = [
            ("speed",      self._p_speed),
            ("brake",      self._p_brake),
            ("gas",        self._p_gas),
            ("susp_front", self._p_suspf),
            ("susp_rear",  self._p_suspr),
        ]
        label_a = f"{self._label_a + ' ' if self._label_a else ''}L{lap_a.get('lap_no', '')}"
        offset_apply = self._offset_b if x_mode == "time" else 0.0

        for ch, p in _CHAN_PANELS:
            _plot(lap_a, f"A:{label_a}", colors["a"], ch, p)
            if lap_b:
                label_b = f"{self._label_b + ' ' if self._label_b else ''}L{lap_b.get('lap_no', '')}"
                _plot(lap_b, f"B:{label_b}", colors["b"], ch, p, offset_x=offset_apply)

        # Y auto-range; X range depends on mode
        for p in self._all_plots:
            p.enableAutoRange(axis="y")
            if x_mode in ("time", "distance"):
                p.enableAutoRange(axis="x")
            else:
                p.setXRange(0.0, 1.0, padding=0.01)

        # LinearRegion をデータ範囲の 20%〜40% に再配置
        try:
            if lap_a and hasattr(self, "_region"):
                xs_raw = _get_x(lap_a)
                if xs_raw:
                    xs = (np.array(xs_raw, dtype=float)
                          if x_mode in ("time", "distance")
                          else _normalize(xs_raw))
                    x_min, x_max = float(xs[0]), float(xs[-1])
                    span = x_max - x_min
                    self._region.setRegion([x_min + span * 0.2, x_min + span * 0.4])
        except Exception:
            pass

        # Turn markers — progress mode only
        if x_mode not in ("time", "distance"):
            tmpl_raw = self._templates.get(self._circuit, [])
            if isinstance(tmpl_raw, dict):
                tmpl_raw = [{"name": k, **v} for k, v in tmpl_raw.items()
                             if isinstance(v, dict)]
            for turn in tmpl_raw:
                prog = turn.get("progress")
                if prog is None:
                    continue
                for p in self._all_plots:
                    line = pg.InfiniteLine(
                        pos=prog, angle=90,
                        pen=pg.mkPen("#107C10", width=1, style=Qt.PenStyle.DashLine),
                        label=turn.get("name", ""),
                        labelOpts={"color": "#107C10", "position": 0.9,
                                   "rotateAxis": (1, 0)},
                    )
                    p.addItem(line)


>>>>>>> Stashed changes
# ════════════════════════════════════════════════════════════════════
# Problem Log タブ
# ════════════════════════════════════════════════════════════════════


def _fmt_range(r: dict) -> str:
    """problem_log 行の距離/時間範囲を表示文字列に変換する。"""
    ds = r.get("distance_start_m")
    de = r.get("distance_end_m")
    if ds is not None and de is not None:
        return f"{ds:.0f}→{de:.0f}m"
    ts = r.get("time_start_s")
    te = r.get("time_end_s")
    if ts is not None and te is not None:
        return f"{ts:.1f}→{te:.1f}s"
    return "—"

class _RunSelectorWidget(QWidget):
    """Circuit → Run 選択 UI。CSV ロード不要で DB から直接 Run を選べる。"""

    def __init__(self, db: WorkbenchDB, on_run_selected, parent=None):
        super().__init__(parent)
        self._db = db
        self._cb = on_run_selected
        self._setup_ui()

    def _setup_ui(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(4, 2, 4, 2)
        lay.setSpacing(6)
        lbl_c = QLabel("🗂 Circuit:")
        lbl_c.setStyleSheet("font-size: 10px; font-weight: bold;")
        lay.addWidget(lbl_c)
        self._combo_circ = QComboBox()
        self._combo_circ.setMinimumWidth(120)
        lay.addWidget(self._combo_circ)
        lbl_r = QLabel("Run:")
        lbl_r.setStyleSheet("font-size: 10px; font-weight: bold;")
        lay.addWidget(lbl_r)
        self._combo_run = QComboBox()
        self._combo_run.setMinimumWidth(250)
        lay.addWidget(self._combo_run)
        lay.addStretch()
        self._combo_circ.currentTextChanged.connect(self._on_circ)
        self._combo_run.currentIndexChanged.connect(self._on_run)
        self.setStyleSheet(
            "background: #F0F4F8; border: 1px solid #C8D3DC;"
            " border-radius: 4px; padding: 2px;"
        )
        self._load_circuits()

    def _load_circuits(self):
        try:
            circs = self._db.get_circuits()
        except Exception:
            circs = []
        self._combo_circ.blockSignals(True)
        self._combo_circ.clear()
        self._combo_circ.addItem("— 全て —")
        self._combo_circ.addItems(circs)
        self._combo_circ.blockSignals(False)
        if circs:
            self._combo_circ.setCurrentIndex(1)
        else:
            self._on_circ("— 全て —")

    def _on_circ(self, circuit: str):
        circ = circuit if circuit != "— 全て —" else None
        try:
            runs = self._db.get_runs(circ)
        except Exception:
            runs = []
        self._combo_run.blockSignals(True)
        self._combo_run.clear()
        for r in runs:
            label = f"{r['rider']}  {r['session']} R{r['run_no']}  ({r['run_id']})"
            self._combo_run.addItem(label, userData=r["run_id"])
        self._combo_run.blockSignals(False)
        if self._combo_run.count():
            self._combo_run.setCurrentIndex(0)
            self._emit()

    def _on_run(self, _idx: int):
        self._emit()

    def _emit(self):
        run_id = self._combo_run.currentData()
        if run_id:
            self._cb(run_id)

    def select_run_id(self, run_id: str):
        """外部から特定 run_id を選択状態にする（波形ロード連携）。"""
        for i in range(self._combo_run.count()):
            if self._combo_run.itemData(i) == run_id:
                self._combo_run.setCurrentIndex(i)
                return


class ProblemLogTab(QWidget):
    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._run_id: str = ""
        self._run_meta: dict = {}
        self._wave_prefill: dict = {}
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # ── DB ベース Run セレクタ ────────────────────────────────────
        self._run_sel = _RunSelectorWidget(self._db, self._on_db_run_selected)
        layout.addWidget(self._run_sel)

        # ── 一覧テーブル ──────────────────────────────────
        self._table = QTableWidget(0, 9)
        self._table.setHorizontalHeaderLabels(
            ["ID", "Corner", "Phase", "Tag", "Description", "Severity", "Lap", "Range", "Source"]
        )
        self._table.horizontalHeader().setStretchLastSection(False)
        self._table.setColumnWidth(0, 40)
        self._table.setColumnWidth(1, 60)
        self._table.setColumnWidth(2, 60)
        self._table.setColumnWidth(3, 140)
        self._table.setColumnWidth(4, 220)
        self._table.setColumnWidth(5, 70)
        self._table.setColumnWidth(6, 40)
        self._table.setColumnWidth(7, 160)
        self._table.setColumnWidth(8, 80)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSortingEnabled(True)
        layout.addWidget(self._table, stretch=2)

        # 削除ボタン
        del_btn = QPushButton("選択行を削除")
        del_btn.clicked.connect(self._delete_selected)
        layout.addWidget(del_btn)

        # ── 区切り ────────────────────────────────────────
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)

        # ── 新規追加フォーム ──────────────────────────────
        form_label = QLabel("▼ 新規問題を追加")
        form_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        layout.addWidget(form_label)

        # ── 波形連携: 自動入力エリア ─────────────────────────────────
        self._wave_info_box = QWidget()
        self._wave_info_box.setStyleSheet(
            "background: #EFF6FF; border: 1px solid #0078D4;"
            " border-radius: 4px; padding: 4px;"
        )
        wave_info_lay = QVBoxLayout(self._wave_info_box)
        wave_info_lay.setContentsMargins(8, 4, 8, 4)
        wave_info_lay.setSpacing(2)

        wave_header = QLabel("📊 波形から自動入力（読み取り専用）")
        wave_header.setStyleSheet("color: #0078D4; font-weight: bold; font-size: 10px;")
        wave_info_lay.addWidget(wave_header)

        self._lbl_auto_run   = QLabel("Run: —")
        self._lbl_auto_lap   = QLabel("Lap: —")
        self._lbl_auto_range = QLabel("Range: —")
        for lbl in (self._lbl_auto_run, self._lbl_auto_lap, self._lbl_auto_range):
            lbl.setStyleSheet("color: #004578; font-size: 10px;")
            wave_info_lay.addWidget(lbl)

        btn_clear_wave = QPushButton("✕ 自動入力をクリア")
        btn_clear_wave.setFixedHeight(20)
        btn_clear_wave.setStyleSheet("font-size: 10px; color: #666;")
        btn_clear_wave.clicked.connect(self._clear_wave_prefill)
        wave_info_lay.addWidget(btn_clear_wave)

        self._wave_info_box.setVisible(False)
        layout.addWidget(self._wave_info_box)

        form = QFormLayout()

        self._lbl_run = QLabel("（未選択）")
        form.addRow("Run:", self._lbl_run)

        self._spin_lap = QSpinBox()
        self._spin_lap.setRange(0, 999)
        self._spin_lap.setSpecialValueText("—")
        form.addRow("Lap No (optional):", self._spin_lap)

        self._combo_corner = QComboBox()
        self._combo_corner.addItem("NONE")
        for i in range(1, 18):
            self._combo_corner.addItem(f"T{i}")
        form.addRow("Corner:", self._combo_corner)

        self._combo_phase = QComboBox()
        self._combo_phase.addItems(PHASES)
        form.addRow("Phase:", self._combo_phase)

        self._combo_tag = QComboBox()
        self._combo_tag.addItems(PROBLEM_TAGS)
        form.addRow("Problem Tag:", self._combo_tag)

        self._txt_desc = QTextEdit()
        self._txt_desc.setFixedHeight(60)
        self._txt_desc.setPlaceholderText("詳細説明（任意）")
        form.addRow("Description:", self._txt_desc)

        self._combo_sev = QComboBox()
        self._combo_sev.addItems(SEVERITIES)
        self._combo_sev.setCurrentText("MEDIUM")
        form.addRow("Severity:", self._combo_sev)

        self._combo_src = QComboBox()
        self._combo_src.addItems(SOURCES)
        form.addRow("Source:", self._combo_src)

        layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_add   = QPushButton("追加")
        btn_clear = QPushButton("クリア")
        btn_add.clicked.connect(self._add_entry)
        btn_clear.clicked.connect(self._clear_form)
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    def _on_db_run_selected(self, run_id: str):
        """Run セレクタから呼ばれる。DB からメタを取得して set_run() へ渡す。"""
        try:
            meta = self._db.get_run(run_id)
        except Exception:
            meta = {}
        self._run_id = run_id
        self._run_meta = meta
        self._lbl_run.setText(run_id or "（未選択）")
        self._refresh_table()

    def set_run(self, run_id: str, meta: dict):
        self._run_id = run_id
        self._run_meta = meta
        self._lbl_run.setText(run_id or "（未選択）")
        self._refresh_table()
        if hasattr(self, "_run_sel"):
            self._run_sel.select_run_id(run_id)

    def prefill_from_waveform(self, data: dict) -> None:
        """WaveformView から呼ばれ、座標情報を自動入力する。"""
        self._wave_prefill = data

        run_id = data.get("run_id", "")
        if run_id and self._run_id and run_id != self._run_id:
            QMessageBox.warning(
                self,
                "Run 不一致",
                f"波形のRun ({run_id}) と\n"
                f"Problem Log のRun ({self._run_id}) が一致しません。\n"
                "左パネルで同じ Run を選択してください。",
            )
            self._wave_prefill = {}
            return

        self._lbl_auto_run.setText(f"Run: {run_id or '—'}")
        lap_no = data.get("lap_no")
        self._lbl_auto_lap.setText(f"Lap: {lap_no if lap_no is not None else '—'}")

        x_mode = data.get("x_mode", "")
        if x_mode == "distance":
            ds = data.get("distance_start_m")
            de = data.get("distance_end_m")
            range_str = (f"{ds:.1f}m → {de:.1f}m  ({de - ds:.1f}m)"
                         if ds is not None and de is not None else "—")
        elif x_mode == "time":
            ts = data.get("time_start_s")
            te = data.get("time_end_s")
            range_str = (f"{ts:.3f}s → {te:.3f}s  ({te - ts:.3f}s)"
                         if ts is not None and te is not None else "—")
        else:
            ts = data.get("time_start_s")
            te = data.get("time_end_s")
            range_str = f"progress {ts:.4f} → {te:.4f}" if ts is not None else "—"

        self._lbl_auto_range.setText(f"Range: {range_str}")
        self._wave_info_box.setVisible(True)

        if lap_no is not None:
            self._spin_lap.setValue(int(lap_no))

        src_items = [self._combo_src.itemText(i) for i in range(self._combo_src.count())]
        if "DATA" in src_items:
            self._combo_src.setCurrentText("DATA")

    def _clear_wave_prefill(self) -> None:
        """波形からの自動入力をリセットする。"""
        self._wave_prefill = {}
        self._lbl_auto_run.setText("Run: —")
        self._lbl_auto_lap.setText("Lap: —")
        self._lbl_auto_range.setText("Range: —")
        self._wave_info_box.setVisible(False)

    def _refresh_table(self):
        try:
            print("[ProblemLog] run_id:", self._run_id)
            rows = self._db.get_problem_logs(self._run_id) if self._run_id else []
            print("[ProblemLog] rows:", len(rows))
            if rows:
                print("[ProblemLog] keys:", list(rows[0].keys()))
        except Exception as e:
            print("[ProblemLog] refresh error:", e)
            import traceback; traceback.print_exc()
            rows = []
        self._table.setRowCount(0)
        for ri, r in enumerate(rows):
            try:
                self._table.insertRow(ri)
                vals = [
                    str(r.get("problem_id", "")),
                    str(r.get("corner", "") or ""),
                    str(r.get("phase", "") or ""),
                    str(r.get("problem_tag", "") or ""),
                    str(r.get("description", "") or ""),
                    str(r.get("severity", "") or ""),
                    str(r.get("lap_no", "") if r.get("lap_no") is not None else "—"),
                    _fmt_range(r),
                    str(r.get("source", "") or ""),
                ]
                for ci, val in enumerate(vals):
                    self._table.setItem(ri, ci, QTableWidgetItem(val))
            except Exception as e:
                print(f"[ProblemLog] row {ri} render error:", e)

    def _add_entry(self):
        if not self._run_id:
            QMessageBox.warning(self, "未選択", "左パネルでRunを選択してください。")
            return
        corner_val = self._combo_corner.currentText()
        if corner_val == "NONE":
            corner_val = None
        lap_val = self._spin_lap.value()
        if lap_val == 0:
            lap_val = None
        data = {
            "run_id":     self._run_id,
            "round":      self._run_meta.get("round"),
            "circuit":    self._run_meta.get("circuit"),
            "session":    self._run_meta.get("session"),
            "rider":      self._run_meta.get("rider"),
            "run_no":     self._run_meta.get("run_no"),
            "lap_no":     lap_val,
            "corner":     corner_val,
            "phase":      self._combo_phase.currentText(),
            "problem_tag": self._combo_tag.currentText(),
            "description": self._txt_desc.toPlainText().strip(),
            "severity":   self._combo_sev.currentText(),
            "source":     self._combo_src.currentText(),
            "distance_start_m": self._wave_prefill.get("distance_start_m"),
            "distance_end_m":   self._wave_prefill.get("distance_end_m"),
            "time_start_s":     self._wave_prefill.get("time_start_s"),
            "time_end_s":       self._wave_prefill.get("time_end_s"),
            "data_source_file": self._wave_prefill.get("data_source_file", ""),
            "analysis_note":    "",
        }
        self._db.add_problem_log(data)
        self._clear_form()
        self._clear_wave_prefill()
        self._refresh_table()

    def _clear_form(self):
        self._spin_lap.setValue(0)
        self._combo_corner.setCurrentIndex(0)
        self._combo_phase.setCurrentIndex(0)
        self._combo_tag.setCurrentIndex(0)
        self._txt_desc.clear()
        self._combo_sev.setCurrentText("MEDIUM")
        self._combo_src.setCurrentIndex(0)

    def refresh(self):
        """DB変更時に外部から呼び出される。テーブルを再読み込みする。"""
        self._refresh_table()

    def _delete_selected(self):
        rows = self._table.selectedItems()
        if not rows:
            return
        row = self._table.currentRow()
        pid_item = self._table.item(row, 0)
        if not pid_item:
            return
        try:
            pid = int(pid_item.text())
        except ValueError:
            return
        reply = QMessageBox.question(
            self, "削除確認", f"problem_id={pid} を削除しますか？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._db.delete_problem_log(pid)
            self._refresh_table()


# ════════════════════════════════════════════════════════════════════
# Setup Decision Log タブ
# ════════════════════════════════════════════════════════════════════

class SetupDecisionTab(QWidget):
    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._run_id: str = ""
        self._run_meta: dict = {}
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        # ── DB ベース Run セレクタ ────────────────────────────────────
        self._run_sel = _RunSelectorWidget(self._db, self._on_db_run_selected)
        layout.addWidget(self._run_sel)

        # ── 一覧テーブル ──────────────────────────────────
        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels(
            ["ID", "Component", "From", "To", "Rationale", "Result"]
        )
        self._table.setColumnWidth(0, 40)
        self._table.setColumnWidth(1, 100)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 80)
        self._table.setColumnWidth(4, 280)
        self._table.setColumnWidth(5, 80)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSortingEnabled(True)
        layout.addWidget(self._table, stretch=2)

        # 区切り
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)

        form_label = QLabel("▼ セットアップ変更を記録")
        form_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        layout.addWidget(form_label)

        form = QFormLayout()

        self._lbl_run_from = QLabel("（未選択）")
        form.addRow("Run From:", self._lbl_run_from)

        self._combo_run_to = QComboBox()
        self._combo_run_to.addItem("NEXT（未定）")
        form.addRow("Run To:", self._combo_run_to)

        self._combo_chg_type = QComboBox()
        self._combo_chg_type.addItems(CHANGE_TYPES)
        form.addRow("Change Type:", self._combo_chg_type)

        self._combo_comp = QComboBox()
        self._combo_comp.addItems(COMPONENTS)
        form.addRow("Component:", self._combo_comp)

        row_fv = QHBoxLayout()
        self._edit_from = QLineEdit()
        self._edit_from.setPlaceholderText("例: 18")
        self._edit_to = QLineEdit()
        self._edit_to.setPlaceholderText("例: 16")
        row_fv.addWidget(self._edit_from)
        row_fv.addWidget(QLabel("→"))
        row_fv.addWidget(self._edit_to)
        form.addRow("From → To:", row_fv)

        self._txt_rationale = QTextEdit()
        self._txt_rationale.setFixedHeight(55)
        self._txt_rationale.setPlaceholderText("変更の根拠（データ/ライダー報告）")
        form.addRow("Rationale:", self._txt_rationale)

        self._txt_expected = QTextEdit()
        self._txt_expected.setFixedHeight(55)
        self._txt_expected.setPlaceholderText("期待する効果")
        form.addRow("Expected Effect:", self._txt_expected)

        self._combo_result = QComboBox()
        self._combo_result.addItems(RESULT_EVALS)
        self._combo_result.setCurrentText("UNKNOWN")
        form.addRow("Result Eval:", self._combo_result)

        layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_add   = QPushButton("追加")
        btn_clear = QPushButton("クリア")
        btn_add.clicked.connect(self._add_entry)
        btn_clear.clicked.connect(self._clear_form)
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    def _on_db_run_selected(self, run_id: str):
        try:
            meta = self._db.get_run(run_id)
        except Exception:
            meta = {}
        self.set_run(run_id, meta)

    def set_run(self, run_id: str, meta: dict):
        self._run_id = run_id
        self._run_meta = meta
        self._lbl_run_from.setText(run_id or "（未選択）")
        self._combo_run_to.clear()
        self._combo_run_to.addItem("NEXT（未定）")
        next_runs = self._db.get_next_runs(run_id) if run_id else []
        for nr in next_runs:
            self._combo_run_to.addItem(
                f"{nr['run_id']}  ({nr['session']} R{nr['run_no']})",
                userData=nr["run_id"],
            )
        self._refresh_table()
        if hasattr(self, "_run_sel"):
            self._run_sel.select_run_id(run_id)

    def _refresh_table(self):
        rows = self._db.get_setup_decisions(self._run_id) if self._run_id else []
        self._table.setRowCount(0)
        for r in rows:
            ri = self._table.rowCount()
            self._table.insertRow(ri)
            for ci, val in enumerate([
                r.get("decision_id", ""),
                r.get("component", ""),
                r.get("from_value", ""),
                r.get("to_value", ""),
                r.get("rationale", ""),
                r.get("result_eval", ""),
            ]):
                self._table.setItem(ri, ci, QTableWidgetItem(str(val) if val is not None else "—"))

    def _add_entry(self):
        if not self._run_id:
            QMessageBox.warning(self, "未選択", "左パネルでRunを選択してください。")
            return
        run_to_idx = self._combo_run_to.currentIndex()
        run_to_data = self._combo_run_to.itemData(run_to_idx)
        run_to_val = run_to_data if run_to_data else None

        data = {
            "run_id_from":    self._run_id,
            "run_id_to":      run_to_val,
            "round":          self._run_meta.get("round"),
            "circuit":        self._run_meta.get("circuit"),
            "session":        self._run_meta.get("session"),
            "rider":          self._run_meta.get("rider"),
            "change_type":    self._combo_chg_type.currentText(),
            "component":      self._combo_comp.currentText(),
            "from_value":     self._edit_from.text().strip(),
            "to_value":       self._edit_to.text().strip(),
            "rationale":      self._txt_rationale.toPlainText().strip(),
            "expected_effect": self._txt_expected.toPlainText().strip(),
            "actual_effect":  None,
            "result_eval":    self._combo_result.currentText(),
        }
        self._db.add_setup_decision(data)
        self._clear_form()
        self._refresh_table()

    def refresh(self):
        """DB変更時に外部から呼び出される。テーブルを再読み込みする。"""
        self._refresh_table()

    def _clear_form(self):
        self._combo_chg_type.setCurrentIndex(0)
        self._combo_comp.setCurrentIndex(0)
        self._edit_from.clear()
        self._edit_to.clear()
        self._txt_rationale.clear()
        self._txt_expected.clear()
        self._combo_result.setCurrentText("UNKNOWN")



# ════════════════════════════════════════════════════════════════════
# Run Browser タブ
# ════════════════════════════════════════════════════════════════════

class RunBrowserTab(QWidget):
    """🗺️ Run Browser — DB全Run一覧。Circuit / Rider / Session フィルタ + 行クリックで選択。"""

    run_selected = None  # will be set to a callable

    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._on_run_selected = None  # callback(run_id, meta)
        self._setup_ui()
        self._refresh()

    def set_on_run_selected(self, cb):
        self._on_run_selected = cb

    def refresh(self):
        """DB変更時に外部から呼び出される。Run一覧を再読み込みする。"""
        self._refresh()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(6)

        # ── フィルタ行 ────────────────────────────────────────────────
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("Circuit:"))
        self._combo_circ = QComboBox()
        self._combo_circ.setFixedWidth(130)
        self._combo_circ.addItem("ALL")
        self._combo_circ.currentTextChanged.connect(self._on_filter)
        filter_row.addWidget(self._combo_circ)

        filter_row.addWidget(QLabel("Rider:"))
        self._combo_rider = QComboBox()
        self._combo_rider.setFixedWidth(100)
        self._combo_rider.addItems(["ALL", "DA77", "JA52"])
        self._combo_rider.currentTextChanged.connect(self._on_filter)
        filter_row.addWidget(self._combo_rider)

        filter_row.addWidget(QLabel("Session:"))
        self._combo_session = QComboBox()
        self._combo_session.setFixedWidth(120)
        self._combo_session.addItem("ALL")
        self._combo_session.currentTextChanged.connect(self._on_filter)
        filter_row.addWidget(self._combo_session)

        btn_refresh = QPushButton("🔄 更新")
        btn_refresh.setFixedWidth(70)
        btn_refresh.clicked.connect(self._refresh)
        filter_row.addWidget(btn_refresh)
        filter_row.addStretch()

        self._lbl_count = QLabel("")
        self._lbl_count.setStyleSheet("color: #888; font-size: 10px;")
        filter_row.addWidget(self._lbl_count)
        lay.addLayout(filter_row)

        # ── テーブル ──────────────────────────────────────────────────
        self._table = QTableWidget()
        self._table.setColumnCount(6)
        self._table.setHorizontalHeaderLabels(
            ["Run ID", "Circuit", "Session", "Rider", "Run No", "Best Lap"]
        )
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.cellClicked.connect(self._on_row_clicked)
        lay.addWidget(self._table)

    def _refresh(self):
        try:
            circuits = self._db.get_circuits()
        except Exception:
            circuits = []
        self._combo_circ.blockSignals(True)
        saved_circ = self._combo_circ.currentText()
        self._combo_circ.clear()
        self._combo_circ.addItem("ALL")
        self._combo_circ.addItems(circuits)
        idx = self._combo_circ.findText(saved_circ)
        if idx >= 0:
            self._combo_circ.setCurrentIndex(idx)
        self._combo_circ.blockSignals(False)

        try:
            all_runs = self._db.get_runs()
        except Exception:
            all_runs = []

        sessions = sorted({r.get("session", "") or "" for r in all_runs if r.get("session")})
        self._combo_session.blockSignals(True)
        saved_sess = self._combo_session.currentText()
        self._combo_session.clear()
        self._combo_session.addItem("ALL")
        self._combo_session.addItems(sessions)
        idx = self._combo_session.findText(saved_sess)
        if idx >= 0:
            self._combo_session.setCurrentIndex(idx)
        self._combo_session.blockSignals(False)

        self._populate_table(all_runs)

    def _on_filter(self):
        circ = self._combo_circ.currentText()
        try:
            runs = self._db.get_runs(circuit=circ if circ != "ALL" else None)
        except Exception:
            runs = []
        self._populate_table(runs)

    def _populate_table(self, runs: list[dict]):
        rider_f   = self._combo_rider.currentText()
        session_f = self._combo_session.currentText()

        filtered = [
            r for r in runs
            if (rider_f   == "ALL" or r.get("rider")   == rider_f)
            and (session_f == "ALL" or r.get("session") == session_f)
        ]

<<<<<<< Updated upstream
        self._table.setRowCount(len(filtered))
        for row, r in enumerate(filtered):
            best = r.get("perf_best_lap")
            best_str = format_laptime(float(best)) if best else "—"
            vals = [
                r.get("run_id", ""),
                r.get("circuit", ""),
                r.get("session", ""),
                r.get("rider", ""),
                str(r.get("run_no", "")),
                best_str,
=======
        # Preview table
        prev_lbl = QLabel("▼ データプレビュー（先頭 50 行）")
        prev_lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        layout.addWidget(prev_lbl)
        self._preview = QTableWidget(0, 0)
        self._preview.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self._preview, stretch=1)

        # Send button
        btn_row = QHBoxLayout()
        self._btn_send = QPushButton("📊  波形に送る")
        self._btn_send.setEnabled(False)
        self._btn_send.clicked.connect(self._send)
        self._btn_send.setStyleSheet(
            "QPushButton { background: #0078D4; color: white; padding: 6px 16px;"
            " border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background: #106EBE; }"
            "QPushButton:disabled { background: #ccc; color: #888; }"
        )
        btn_row.addWidget(self._btn_send)
        self._lbl_sent = QLabel("")
        self._lbl_sent.setStyleSheet("color: #107C10; font-size: 10px; padding: 0 8px;")
        btn_row.addWidget(self._lbl_sent)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    # ── CSV 読み込み ─────────────────────────────────────────────────

    def _browse(self):
        default = Path.home() / "Desktop" / "Data TS24 Claude" / "06_CSV"
        if not default.exists():
            default = SCRIPT_DIR.parent
        path, _ = QFileDialog.getOpenFileName(
            self, "CSV ファイルを選択", str(default),
            "CSV Files (*.csv);;All Files (*.*)",
        )
        if not path:
            return
        self._lbl_file.setText(Path(path).name)
        self._load_csv(Path(path))

    def load_file(self, path: str) -> None:
        """外部から CSV パスを渡して即読み込み・波形送信を実行する。"""
        p = Path(path)
        self._lbl_file.setText(p.name)
        if not self._run_id:
            # ファイル名を解析して DB から run_id を検索
            _FNAME_RE = re.compile(
                r"(\d{8})-"
                r"(ROUND\d+|TEST\d+|UNK)-"
                r"([A-Z0-9]+)-?"
                r"(?:RUN(\d+)-)?"
                r"(DA77|JA52|DA\d+|JA\d+)",
                re.IGNORECASE,
            )
            m = _FNAME_RE.search(p.stem)
            if m and self._db:
                _, round_s, session_s, run_no_s, rider_s = m.groups()
                run_no = int(run_no_s) if run_no_s else 1
                db_run_id = self._db.lookup_run_id(
                    round_s.upper(), session_s.upper(), rider_s.upper(), run_no
                )
                self._run_id = db_run_id or p.stem
                print(f"[CsvImportTab] run_id lookup: stem={p.stem!r} → run_id={self._run_id!r}")
            else:
                self._run_id = p.stem
        self._load_csv(p)
        if self._df is not None:
            self._send()

    def _load_csv(self, path: Path):
        df = None
        for enc in ("utf-8-sig", "shift_jis"):
            for sep in (";", ","):
                try:
                    candidate = pd.read_csv(
                        path, encoding=enc, sep=sep,
                        decimal="," if sep == ";" else ".",
                        skiprows=[1], header=0,
                    )
                    # Accept if we get more than 1 column
                    if len(candidate.columns) > 1:
                        df = candidate
                        break
                except Exception:
                    pass
            if df is not None:
                break
        if df is None:
            QMessageBox.critical(self, "CSV 読み込みエラー", f"読み込めませんでした: {path.name}")
            return

        df.columns = [str(c).strip() for c in df.columns]
        self._df = df
        n_rows, n_cols = df.shape
        self._lbl_info.setText(f"{n_rows} 行 × {n_cols} 列 を読み込みました。")

        # Dist column validity check
        dist_col = next(
            (c for c in df.columns if c.lower().strip() in ("dist", "distance")), None
        )
        if dist_col:
            try:
                vals = pd.to_numeric(df[dist_col], errors="coerce").fillna(0).values
                self._lbl_dist.setVisible(float(vals.max()) < 10.0)
            except Exception:
                self._lbl_dist.setVisible(False)
        else:
            self._lbl_dist.setVisible(False)

        # Rebuild channel mapping
        while self._map_layout.rowCount():
            self._map_layout.removeRow(0)
        self._col_combos.clear()

        for col in df.columns:
            combo = QComboBox()
            combo.addItems(self._TARGETS)
            combo.setCurrentText(self._auto_detect(col))
            self._col_combos[col] = combo
            self._map_layout.addRow(f"{col}:", combo)

        # Preview
        preview = df.head(50)
        self._preview.setColumnCount(len(preview.columns))
        self._preview.setHorizontalHeaderLabels(list(preview.columns))
        self._preview.setRowCount(len(preview))
        for ri, row_data in enumerate(preview.itertuples(index=False)):
            for ci, val in enumerate(row_data):
                self._preview.setItem(ri, ci, QTableWidgetItem(str(val)))
        self._preview.resizeColumnsToContents()
        self._btn_send.setEnabled(True)

    def _auto_detect(self, col_name: str) -> str:
        lower = col_name.lower().strip()
        for target, aliases in self.CHANNEL_MAP.items():
            for alias in aliases:
                if alias in lower or lower in alias:
                    return target
        return "(ignore)"

    # ── 波形に送る ────────────────────────────────────────────────────

    def _send(self):
        if self._df is None:
            return

        channel_to_col: dict[str, str] = {}
        for col, combo in self._col_combos.items():
            ch = combo.currentText()
            if ch != "(ignore)":
                channel_to_col[ch] = col

        data_chs = {k for k in channel_to_col if k != "time"}
        if not data_chs:
            QMessageBox.warning(
                self, "マッピング不足",
                "speed / brake などのデータチャンネルを1つ以上割り当ててください。",
            )
            return

        has_time = "time" in channel_to_col
        has_dist = "distance" in channel_to_col

        df = self._df.copy()
        n = len(df)
        if n == 0:
            QMessageBox.warning(self, "データなし", "CSV にデータ行がありません。")
            return

        # x_mode 決定（優先順位: distance > time > progress）
        x_mode = "progress"
        if has_time:
            x_mode = "time"
        if has_dist:
            try:
                d_check = pd.to_numeric(
                    df[channel_to_col["distance"]], errors="coerce"
                ).fillna(0).values
                if float(d_check.max()) > 10.0:
                    x_mode = "distance"
            except Exception:
                pass

        # ── 時間配列の取得 ────────────────────────────────────────────
        if has_time:
            try:
                t_raw = pd.to_numeric(
                    df[channel_to_col["time"]], errors="coerce"
                ).fillna(0).values
            except Exception:
                t_raw = None
        else:
            t_raw = None

        # ── 距離配列の取得 ────────────────────────────────────────────
        d_raw = None
        if x_mode == "distance":
            try:
                d_raw = pd.to_numeric(
                    df[channel_to_col["distance"]], errors="coerce"
                ).fillna(0).values
            except Exception:
                d_raw = None
                x_mode = "time" if t_raw is not None else "progress"

        if t_raw is None and x_mode == "time":
            x_mode = "progress"

        # ── Lap分割（優先順位制御）────────────────────────────────────
        import numpy as np

        circuit_len_m = self._spin_circuit_len.value() if hasattr(self, "_spin_circuit_len") else 0
        split_mode = "unknown"

        # ─ Step A: CSV時間ギャップでセグメント境界を検出（全優先度で共通）
        segments: list[tuple[int, int]] = []
        csv_gap_durations: list[float] = []
        if t_raw is not None:
            seg_start = 0
            for i in range(1, len(t_raw)):
                if (t_raw[i] - t_raw[i - 1]) > 5.0:
                    segments.append((seg_start, i - 1))
                    csv_gap_durations.append(float(t_raw[i] - t_raw[i - 1]))
                    seg_start = i
            segments.append((seg_start, len(t_raw) - 1))
        else:
            segments = [(0, n - 1)]

        lap_indices: list[list[int]] = []

        # ─ 優先1: CSV内のLap列で分割 ─────────────────────────────────
        has_lap_col = "lap_no" in channel_to_col
        if has_lap_col:
            try:
                lap_col = pd.to_numeric(
                    df[channel_to_col["lap_no"]], errors="coerce"
                ).fillna(0).values.astype(int)
                cur_lap = lap_col[0]
                cur: list[int] = [0]
                for i in range(1, len(lap_col)):
                    if lap_col[i] != cur_lap:
                        lap_indices.append(cur)
                        cur = [i]
                        cur_lap = lap_col[i]
                    else:
                        cur.append(i)
                lap_indices.append(cur)
                lap_indices = [seg for seg in lap_indices if len(seg) >= 2]
                split_mode = "lap_col"
            except Exception:
                has_lap_col = False
                lap_indices = []

        # ─ 優先2: DBのlapsテーブルを使った精確分割 ─────────────────
        if not lap_indices and self._run_id and t_raw is not None:
            try:
                db_laps = self._db.get_laps(self._run_id)
                timed_laps = [
                    (r["lap_no"], float(r["lap_time_s"]))
                    for r in db_laps
                    if not r.get("is_outlap") and r.get("lap_time_s")
                ]
                if timed_laps:
                    GAP_TOLERANCE = 2.0
                    gap_lap_nos: set[int] = set()
                    for gap_dur in csv_gap_durations:
                        for lap_no_db, lt in timed_laps:
                            if abs(lt - gap_dur) < GAP_TOLERANCE:
                                gap_lap_nos.add(lap_no_db)
                                break
                    valid_laps = [
                        (lap_no_db, lt)
                        for lap_no_db, lt in timed_laps
                        if lap_no_db not in gap_lap_nos
                    ]
                    if valid_laps:
                        result: list[list[int]] = []
                        lap_cursor = 0
                        for s_start, s_end in segments:
                            t_cursor = float(t_raw[s_start])
                            s_dur = float(t_raw[s_end]) - t_cursor
                            accumulated = 0.0
                            while lap_cursor < len(valid_laps):
                                _, lt = valid_laps[lap_cursor]
                                if accumulated + lt > s_dur + 1.0:
                                    break
                                accumulated += lt
                                lap_cursor += 1
                                t_lap_end = t_cursor + lt
                                start_i = int(np.searchsorted(t_raw, t_cursor, side="left"))
                                end_i = int(np.searchsorted(t_raw, t_lap_end, side="right")) - 1
                                end_i = min(end_i, s_end)
                                if end_i > start_i:
                                    result.append(list(range(start_i, end_i + 1)))
                                t_cursor = float(t_raw[end_i + 1]) if end_i + 1 <= s_end else t_lap_end
                        if result:
                            lap_indices = result
                            split_mode = "db_driven"
            except Exception:
                lap_indices = []

        # ─ 優先3: 固定距離での近似分割 ───────────────────────────────
        if not lap_indices and d_raw is not None and circuit_len_m > 0:
            all_segs: list[list[int]] = []
            start_dist_val = float(d_raw[0])
            cur_d: list[int] = [0]
            for i in range(1, len(d_raw)):
                if (d_raw[i] - start_dist_val) >= circuit_len_m:
                    all_segs.append(cur_d)
                    cur_d = [i]
                    start_dist_val = float(d_raw[i])
                else:
                    cur_d.append(i)
            all_segs.append(cur_d)
            min_span = circuit_len_m * 0.5
            lap_indices = [
                seg for seg in all_segs
                if len(seg) >= 2 and (d_raw[seg[-1]] - d_raw[seg[0]]) >= min_span
>>>>>>> Stashed changes
            ]
            for col, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setData(Qt.ItemDataRole.UserRole, r.get("run_id", ""))
                self._table.setItem(row, col, item)

        self._table.resizeColumnsToContents()
        self._lbl_count.setText(f"{len(filtered)} runs")

    def _on_row_clicked(self, row: int, col: int):
        item = self._table.item(row, 0)
        if not item:
            return
        run_id = item.data(Qt.ItemDataRole.UserRole)
        if not run_id:
            return
        try:
            meta = self._db.get_run(run_id)
        except Exception:
            meta = {"run_id": run_id}
        if self._on_run_selected:
            self._on_run_selected(run_id, meta)


# ════════════════════════════════════════════════════════════════════
# Quick Log タブ
# ════════════════════════════════════════════════════════════════════

class QuickLogTab(QWidget):
    """⚡ Quick Log — CSV不要、30秒でProblemを記録する最小UIタブ。"""

    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._setup_ui()

    def set_run(self, run_id: str, meta: dict) -> None:
        """RunBrowserからの選択を反映する。"""
        self._run_selector.select_run_id(run_id)

    def refresh(self):
        """DB変更時に外部から呼び出される。Run選択コンボを再読み込みする。"""
        if hasattr(self, "_run_selector"):
            self._run_selector._load_circuits()

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 12, 16, 12)
        lay.setSpacing(8)

        title = QLabel("⚡ Quick Problem Log")
        title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        title.setStyleSheet("color: #0078D4;")
        lay.addWidget(title)

        # ── Run セレクタ ────────────────────────────────────────────────────────────
        self._run_selector = _RunSelectorWidget(
            db=self._db,
            on_run_selected=self._on_run_selected,
        )
        lay.addWidget(self._run_selector)

        self._lbl_run_info = QLabel("Run未選択")
        self._lbl_run_info.setStyleSheet("color: #888; font-size: 10px;")
        lay.addWidget(self._lbl_run_info)

        # ── フォーム ────────────────────────────────────────────────────────────
        form = QFormLayout()
        form.setSpacing(6)

        self._spin_lap = QSpinBox()
        self._spin_lap.setRange(0, 99)
        self._spin_lap.setSpecialValueText("—")
        form.addRow("Lap No:", self._spin_lap)

        self._combo_corner = QComboBox()
        self._combo_corner.addItem("NONE")
        for i in range(1, 20):
            self._combo_corner.addItem(f"T{i}")
        form.addRow("Corner:", self._combo_corner)

        self._combo_phase = QComboBox()
        self._combo_phase.addItems(PHASES)
        form.addRow("Phase:", self._combo_phase)

        self._combo_tag = QComboBox()
        self._combo_tag.addItems(PROBLEM_TAGS)
        form.addRow("Problem Tag:", self._combo_tag)

        self._txt_desc = QTextEdit()
        self._txt_desc.setFixedHeight(72)
        self._txt_desc.setPlaceholderText("詳細説明（任意）")
        form.addRow("Description:", self._txt_desc)

        self._combo_sev = QComboBox()
        self._combo_sev.addItems(SEVERITIES)
        form.addRow("Severity:", self._combo_sev)

        lay.addLayout(form)

        # ── ボタン行 ────────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_save = QPushButton("💾  Save Problem")
        btn_save.setFixedHeight(36)
        btn_save.setStyleSheet(
            "QPushButton { background: #107C10; color: white; border-radius: 6px;"
            " font-size: 13px; font-weight: bold; padding: 0 20px; }"
            "QPushButton:hover { background: #0E6B0E; }"
        )
        btn_save.clicked.connect(self._save)
        btn_clear = QPushButton("クリア")
        btn_clear.setFixedHeight(36)
        btn_clear.clicked.connect(self._clear_form)
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        self._lbl_result = QLabel("")
        self._lbl_result.setStyleSheet("color: #107C10; font-size: 10px;")
        lay.addWidget(self._lbl_result)

        lay.addStretch()

        self._current_run_id: str = ""
        self._current_meta: dict = {}

    def _on_run_selected(self, run_id: str) -> None:
        self._current_run_id = run_id
        try:
            meta = self._db.get_run(run_id)
        except Exception:
            meta = {}
        self._current_meta = meta
        circuit  = meta.get("circuit", "")
        rider    = meta.get("rider", "")
        run_no   = meta.get("run_no", "")
        session  = meta.get("session", "")
        self._lbl_run_info.setText(
            f"✅ {circuit}  |  {rider}  |  {session}  Run #{run_no}"
        )
        self._lbl_result.setText("")

    def _save(self) -> None:
        if not self._current_run_id:
            QMessageBox.warning(self, "警告", "Runを先に選択してください。")
            return
        corner_val = self._combo_corner.currentText()
        if corner_val == "NONE":
            corner_val = None
        data = {
            "run_id":      self._current_run_id,
            "round":       self._current_meta.get("round"),
            "circuit":     self._current_meta.get("circuit"),
            "session":     self._current_meta.get("session"),
            "rider":       self._current_meta.get("rider"),
            "run_no":      self._current_meta.get("run_no"),
            "lap_no":      self._spin_lap.value() or None,
            "corner":      corner_val,
            "phase":       self._combo_phase.currentText(),
            "problem_tag": self._combo_tag.currentText(),
            "description": self._txt_desc.toPlainText().strip(),
            "severity":    self._combo_sev.currentText(),
            "source":      "OBSERVATION",
        }
        try:
            self._db.add_problem_log(data)
        except Exception as e:
            QMessageBox.critical(self, "DB Error", str(e))
            return
        tag = data["problem_tag"]
        self._lbl_result.setText(f"✅ 保存完了: {tag}")
        self._clear_form()

    def _clear_form(self) -> None:
        self._spin_lap.setValue(0)
        self._combo_corner.setCurrentIndex(0)
        self._combo_phase.setCurrentIndex(0)
        self._combo_tag.setCurrentIndex(0)
        self._txt_desc.clear()
        self._combo_sev.setCurrentIndex(0)


# ════════════════════════════════════════════════════════════════════
# トレンド分析タブ
# ════════════════════════════════════════════════════════════════════

class TrendAnalysisTab(QWidget):
    """ライダー別・セッション別のパフォーマンストレンドを表示するタブ。"""

    _RIDER_COLORS = {"DA77": "#0078D4", "JA52": "#E86C00"}

    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db = db
        self._has_pg = False
        self._setup_ui()
        self.refresh()

    # ── UI 構築 ─────────────────────────────────────────────────────
    def _setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)

        # ── 左フィルターパネル ──────────────────────────────────
        filter_panel = QWidget()
        filter_panel.setFixedWidth(190)
        fl = QVBoxLayout(filter_panel)
        fl.setContentsMargins(0, 0, 8, 0)
        fl.setSpacing(4)

        fl.addWidget(QLabel("<b>フィルター</b>"))

        fl.addWidget(QLabel("ライダー"))
        self._cb_rider = QComboBox()
        self._cb_rider.addItems(["両方", "DA77", "JA52"])
        fl.addWidget(self._cb_rider)

        fl.addWidget(QLabel("ラウンド"))
        self._cb_round = QComboBox()
        self._cb_round.addItem("全て")
        fl.addWidget(self._cb_round)

        fl.addWidget(QLabel("セッション"))
        self._cb_session = QComboBox()
        self._cb_session.addItems(["全て", "FP", "QP", "WUP1", "WUP2", "RACE1", "RACE2"])
        fl.addWidget(self._cb_session)

        btn_refresh = QPushButton("🔄 データ更新")
        btn_refresh.clicked.connect(self.refresh)
        fl.addWidget(btn_refresh)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        fl.addWidget(sep)

        fl.addWidget(QLabel("<b>サマリー</b>"))
        self._lbl_summary = QLabel("—")
        self._lbl_summary.setWordWrap(True)
        self._lbl_summary.setStyleSheet("font-size: 11px; color: #444;")
        fl.addWidget(self._lbl_summary)
        fl.addStretch()

        root.addWidget(filter_panel)

        # ── 右エリア：サブタブ ────────────────────────────────────
        self._inner_tabs = QTabWidget()

        # 📈 Lap Times
        self._w_laptime = QWidget()
        lt_layout = QVBoxLayout(self._w_laptime)
        lt_layout.setContentsMargins(0, 0, 0, 0)
        try:
            import pyqtgraph as pg
            pg.setConfigOption("background", "w")
            pg.setConfigOption("foreground", "k")
            self._pw_best = pg.PlotWidget(title="Best Lap per Run")
            self._pw_best.showGrid(x=True, y=True, alpha=0.3)
            self._pw_best.setLabel("left", "Lap Time (s)")
            self._pw_best.addLegend(offset=(-10, 10))
            self._pw_all = pg.PlotWidget(title="All Laps — Scatter (outlap除く)")
            self._pw_all.showGrid(x=True, y=True, alpha=0.3)
            self._pw_all.setLabel("left", "Lap Time (s)")
            lt_layout.addWidget(self._pw_best, 3)
            lt_layout.addWidget(self._pw_all, 2)
            self._has_pg = True
        except ImportError:
            lt_layout.addWidget(QLabel("pyqtgraph が未インストールです。"))
        self._inner_tabs.addTab(self._w_laptime, "📈 Lap Times")

        # 📋 Performance Table
        self._tbl_perf = QTableWidget()
        self._tbl_perf.setAlternatingRowColors(True)
        self._tbl_perf.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl_perf.setSortingEnabled(True)
        self._inner_tabs.addTab(self._tbl_perf, "📋 Performance")

        # 🔧 Setup History
        self._tbl_setup = QTableWidget()
        self._tbl_setup.setAlternatingRowColors(True)
        self._tbl_setup.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._inner_tabs.addTab(self._tbl_setup, "🔧 Setup History")

        # ⚠️ Problem Analysis
        self._w_problem = QWidget()
        prob_layout = QVBoxLayout(self._w_problem)
        prob_layout.setContentsMargins(0, 0, 0, 0)
        prob_splitter = QSplitter(Qt.Orientation.Vertical)
        if self._has_pg:
            import pyqtgraph as pg
            self._pw_prob = pg.PlotWidget(title="Problem Tag 頻度")
            self._pw_prob.showGrid(x=False, y=True, alpha=0.3)
            self._pw_prob.setLabel("left", "件数")
            prob_splitter.addWidget(self._pw_prob)
        self._tbl_prob = QTableWidget()
        self._tbl_prob.setAlternatingRowColors(True)
        prob_splitter.addWidget(self._tbl_prob)
        prob_layout.addWidget(prob_splitter)
        self._inner_tabs.addTab(self._w_problem, "⚠️ Problems")

        # 📊 Lap Log (全ラップ詳細)
        self._tbl_laplog = QTableWidget()
        self._tbl_laplog.setAlternatingRowColors(True)
        self._tbl_laplog.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl_laplog.setSortingEnabled(True)
        self._inner_tabs.addTab(self._tbl_laplog, "📊 Lap Log")

        # 🦾 Suspension (速度帯別サスペンション分析)
        self._w_sus = QWidget()
        sus_v = QVBoxLayout(self._w_sus)
        sus_v.setContentsMargins(4, 4, 4, 4)
        sus_v.setSpacing(4)

        # ── フィルターバー ──
        sus_bar = QWidget()
        sus_bar_h = QHBoxLayout(sus_bar)
        sus_bar_h.setContentsMargins(0, 0, 0, 4)
        sus_bar_h.addWidget(QLabel("⚡ 速度帯:"))
        self._cb_speed_zone = QComboBox()
        self._cb_speed_zone.addItems(["全て", "低速 <80km/h", "中速 80-120km/h", "高速 >120km/h"])
        self._cb_speed_zone.currentIndexChanged.connect(self._on_sus_filter_changed)
        sus_bar_h.addWidget(self._cb_speed_zone)
        sus_bar_h.addSpacing(16)
        sus_bar_h.addWidget(QLabel("📊 Y軸:"))
        self._cb_sus_metric = QComboBox()
        self._cb_sus_metric.addItems(
            ["ApexSusF vs LapTime", "ApexSusR vs LapTime",
             "BrkSusF vs LapTime",  "BrkSusR vs LapTime",
             "ApexSusF vs ApexSusR (F/R姿勢)"])
        self._cb_sus_metric.currentIndexChanged.connect(self._on_sus_filter_changed)
        sus_bar_h.addWidget(self._cb_sus_metric)
        sus_bar_h.addStretch()
        self._lbl_sus_zone_info = QLabel("")
        self._lbl_sus_zone_info.setStyleSheet("color:#555; font-size:10px;")
        sus_bar_h.addWidget(self._lbl_sus_zone_info)
        sus_v.addWidget(sus_bar)

        # ── プロットエリア: メイン散布図 + 速度帯別バーチャート ──
        if self._has_pg:
            import pyqtgraph as pg
            sus_plot_split = QSplitter(Qt.Orientation.Horizontal)
            self._pw_sus_main = pg.PlotWidget()
            self._pw_sus_main.showGrid(x=True, y=True, alpha=0.3)
            self._pw_sus_main.addLegend(offset=(-10, 10))
            sus_plot_split.addWidget(self._pw_sus_main)
            self._pw_sus_zone = pg.PlotWidget(title="速度帯別 平均SusF / SusR (DA77|JA52)")
            self._pw_sus_zone.showGrid(x=False, y=True, alpha=0.3)
            self._pw_sus_zone.setLabel("left", "Sus Position (mm)")
            sus_plot_split.addWidget(self._pw_sus_zone)
            sus_plot_split.setSizes([700, 400])
            sus_v.addWidget(sus_plot_split, 2)

        # ── テーブル ──
        self._tbl_sus = QTableWidget()
        self._tbl_sus.setAlternatingRowColors(False)
        self._tbl_sus.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl_sus.setSortingEnabled(True)
        self._tbl_sus.verticalHeader().setDefaultSectionSize(22)
        sus_v.addWidget(self._tbl_sus, 3)
        self._inner_tabs.addTab(self._w_sus, "🦾 Suspension")

        # 🔍 Analysis (FAST/SLOW比較 + 回路別問題トレンド)
        self._w_analysis = QWidget()
        ana_v = QVBoxLayout(self._w_analysis)
        ana_v.setContentsMargins(4, 4, 4, 4)
        ana_split = QSplitter(Qt.Orientation.Vertical)

        # ── 上部: FAST vs SLOW ──
        ana_top = QWidget()
        ana_top_v = QVBoxLayout(ana_top)
        ana_top_v.setContentsMargins(0, 0, 0, 0)
        ana_lbl = QLabel(
            "<b>🏁 FAST vs SLOW サスペンション比較</b>"
            "  <small style='color:#666;'>(上位30% vs 下位30% ラップ | "
            "負Δ = FAST時により沈む = 良い方向)</small>"
        )
        ana_lbl.setStyleSheet("padding:2px;")
        ana_top_v.addWidget(ana_lbl)
        if self._has_pg:
            self._pw_fast_slow = pg.PlotWidget()
            self._pw_fast_slow.showGrid(x=False, y=True, alpha=0.3)
            self._pw_fast_slow.setLabel("left", "FAST-SLOW Δ SusF (mm)")
            ana_top_v.addWidget(self._pw_fast_slow, 2)
        self._tbl_fast_slow = QTableWidget()
        self._tbl_fast_slow.setAlternatingRowColors(True)
        self._tbl_fast_slow.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        ana_top_v.addWidget(self._tbl_fast_slow, 2)
        ana_split.addWidget(ana_top)

        # ── 下部: 回路別問題タグ ──
        ana_bot = QWidget()
        ana_bot_v = QVBoxLayout(ana_bot)
        ana_bot_v.setContentsMargins(0, 0, 0, 0)
        ana_lbl2 = QLabel(
            "<b>🗺️ 回路別 問題タグ トレンド</b>"
            "  <small style='color:#666;'>(セル = DA77件数 | JA52件数)</small>"
        )
        ana_lbl2.setStyleSheet("padding:2px;")
        ana_bot_v.addWidget(ana_lbl2)
        self._tbl_circuit_prob = QTableWidget()
        self._tbl_circuit_prob.setAlternatingRowColors(True)
        self._tbl_circuit_prob.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        ana_bot_v.addWidget(self._tbl_circuit_prob)
        ana_split.addWidget(ana_bot)

        ana_split.setSizes([380, 280])
        ana_v.addWidget(ana_split)
        self._inner_tabs.addTab(self._w_analysis, "🔍 Analysis")

        # 📊 Perf Corr (パフォーマンス相関)
        self._w_perf_corr = QWidget()
        pc_layout = QVBoxLayout(self._w_perf_corr)
        pc_layout.setContentsMargins(4, 4, 4, 4)
        pc_lbl = QLabel("<b>パフォーマンス相関 — セッション別サスペンション vs ラップタイム</b>")
        pc_layout.addWidget(pc_lbl)
        self._tbl_perf_corr = QTableWidget()
        self._tbl_perf_corr.setAlternatingRowColors(True)
        self._tbl_perf_corr.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._tbl_perf_corr.setSortingEnabled(True)
        pc_layout.addWidget(self._tbl_perf_corr, 3)
        pc_sep = QFrame(); pc_sep.setFrameShape(QFrame.Shape.HLine)
        pc_layout.addWidget(pc_sep)
        pc_lbl2 = QLabel("<b>FAST vs SLOW サスペンション比較 (回路別)</b>")
        pc_layout.addWidget(pc_lbl2)
        self._tbl_perf_cmp = QTableWidget()
        self._tbl_perf_cmp.setAlternatingRowColors(True)
        self._tbl_perf_cmp.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        pc_layout.addWidget(self._tbl_perf_cmp, 2)
        self._inner_tabs.addTab(self._w_perf_corr, "📊 Perf Corr")

        # 📝 Session Notes (エンジニアノート + 問題タグ)
        self._w_notes = QWidget()
        notes_layout = QVBoxLayout(self._w_notes)
        notes_layout.setContentsMargins(4, 4, 4, 4)
        notes_splitter = QSplitter(Qt.Orientation.Horizontal)
        # 左: 問題タグ表
        self._tbl_tag_summary = QTableWidget()
        self._tbl_tag_summary.setAlternatingRowColors(True)
        self._tbl_tag_summary.setMaximumWidth(450)
        notes_splitter.addWidget(self._tbl_tag_summary)
        # 右: エンジニアノートテキスト
        from PyQt6.QtWidgets import QTextBrowser
        self._txt_notes = QTextBrowser()
        self._txt_notes.setOpenExternalLinks(False)
        notes_splitter.addWidget(self._txt_notes)
        notes_splitter.setSizes([420, 800])
        notes_layout.addWidget(notes_splitter)
        self._inner_tabs.addTab(self._w_notes, "📝 Session Notes")

        root.addWidget(self._inner_tabs, 1)

        # フィルター変更でコネクト
        self._cb_rider.currentIndexChanged.connect(self._on_filter_changed)
        self._cb_round.currentIndexChanged.connect(self._on_filter_changed)
        self._cb_session.currentIndexChanged.connect(self._on_filter_changed)

    # ── データ更新 ───────────────────────────────────────────────────
    def refresh(self):
        """DBとExcelからデータを再読み込みしてラウンドComboBoxと静的ビューを更新。"""
        rounds = self._db.get_all_rounds()
        prev = self._cb_round.currentText()
        self._cb_round.blockSignals(True)
        self._cb_round.clear()
        self._cb_round.addItem("全て")
        for r in rounds:
            self._cb_round.addItem(r)
        idx = self._cb_round.findText(prev)
        if idx >= 0:
            self._cb_round.setCurrentIndex(idx)
        self._cb_round.blockSignals(False)

        # Excelデータをキャッシュ（フィルター変更時に再ロードしない）
        try:
            self._perf_corr_cache = self._db.get_perf_correlation()
        except Exception:
            self._perf_corr_cache = ([], [], [], [])
        try:
            self._trend_notes_cache = self._db.get_trend_notes()
        except Exception:
            self._trend_notes_cache = {"tags": [], "rider_tags": {}, "notes": []}

        self._update_views()

    def _on_filter_changed(self, _=None):
        self._update_views()

    def _on_sus_filter_changed(self, _=None):
        """Suspensionタブ内の速度帯/メトリック変更時のみ再描画。"""
        sus_data = getattr(self, "_sus_cache", [])
        self._build_suspension_view(sus_data)

    def _get_filters(self) -> tuple:
        rider_sel = self._cb_rider.currentText()
        rider = None if rider_sel == "両方" else rider_sel
        round_sel = self._cb_round.currentText()
        round_s = None if round_sel == "全て" else round_sel
        session_sel = self._cb_session.currentText()
        session_s = None if session_sel == "全て" else session_sel
        return rider, round_s, session_s

    def _update_views(self):
        rider, round_s, session_s = self._get_filters()
        laps     = self._db.get_trend_laps(rider, round_s, session_s)
        runs     = self._db.get_trend_runs(rider, round_s, session_s)
        problems = self._db.get_trend_problems(rider, round_s, session_s)
        sus_data = self._db.get_lap_suspension(rider, round_s, session_s)

        valid_laps = [l for l in laps if not l.get("is_outlap") and l.get("lap_time_s")]
        riders = sorted(set(l.get("rider", "") for l in laps if l.get("rider")))
        self._lbl_summary.setText(
            f"Runs: {len(runs)}\n"
            f"Laps: {len(valid_laps)}\n"
            f"Sus Laps: {len(sus_data)}\n"
            f"Problems: {len(problems)}\n"
            f"Riders: {', '.join(riders)}"
        )

        self._build_laptime_plot(laps, runs)
        self._build_perf_table(laps, runs)
        self._build_setup_table(runs)
        self._build_problem_view(problems)
        self._build_laplog_table(laps)

        # sus_data をキャッシュ（速度帯フィルター変更時に再利用）
        self._sus_cache = sus_data
        self._build_suspension_view(sus_data)

        # Analysis tab (FAST/SLOW + Circuit Problem Trend)
        notes = getattr(self, "_trend_notes_cache", {"tags": [], "rider_tags": {}, "notes": []})
        self._build_analysis_view(sus_data, notes)

        # Perf Corr / Session Notes はキャッシュから
        pc = getattr(self, "_perf_corr_cache", ([], [], [], []))
        self._build_perf_corr_tables(pc, rider, round_s)
        self._build_session_notes(notes, rider)

    # ── ユーティリティ ───────────────────────────────────────────────
    @staticmethod
    def _fmt(s) -> str:
        """秒数を M'SS.000 形式（motorsport標準）に変換。"""
        if s is None:
            return "—"
        try:
            s = float(s)
        except (ValueError, TypeError):
            return str(s)
        m = int(s) // 60
        sec = s - m * 60
        return f"{m}'{sec:06.3f}"

    @staticmethod
    def _fmt_delta(delta_s) -> str:
        """差分秒数を +0.000s 形式に変換。"""
        if delta_s is None:
            return "—"
        try:
            d = float(delta_s)
        except (ValueError, TypeError):
            return "—"
        if abs(d) < 0.001:
            return "—"
        return f"+{d:.3f}s" if d > 0 else f"{d:.3f}s"

    # ── 📈 Lap Time プロット ─────────────────────────────────────────
    def _build_laptime_plot(self, laps: list, runs: list):
        if not self._has_pg:
            return
        import pyqtgraph as pg

        run_ids = [r["run_id"] for r in runs]
        x_map   = {rid: i for i, rid in enumerate(run_ids)}
        x_labels = [
            f"{r.get('session','?')}R{r.get('run_no','?')}\n{r.get('rider','?')}"
            for r in runs
        ]

        self._pw_best.clear()
        self._pw_all.clear()

        for pw in (self._pw_best, self._pw_all):
            pw.getAxis("bottom").setTicks([list(enumerate(x_labels))] if x_labels else [[]])

        for rider, color in self._RIDER_COLORS.items():
            rider_runs = [r for r in runs if r.get("rider") == rider]
            best_x, best_y, all_x, all_y = [], [], [], []

            for r in rider_runs:
                rid = r["run_id"]
                xi = x_map.get(rid)
                if xi is None:
                    continue
                r_laps = [l for l in laps if l["run_id"] == rid
                          and not l.get("is_outlap") and l.get("lap_time_s")]
                if not r_laps:
                    continue
                times = [float(l["lap_time_s"]) for l in r_laps]
                best_x.append(xi)
                best_y.append(min(times))
                jitter = 0.12 if rider == "JA52" else -0.12
                for t in times:
                    all_x.append(xi + jitter)
                    all_y.append(t)

            pen   = pg.mkPen(color, width=2)
            brush = pg.mkBrush(color)

            if best_x:
                self._pw_best.plot(best_x, best_y, pen=pen, name=rider)
                self._pw_best.plot(best_x, best_y, pen=None, symbol="o",
                                   symbolSize=11, symbolBrush=brush,
                                   symbolPen=pg.mkPen("w", width=1.5))

            if all_x:
                r_c = int(color[1:3], 16)
                g_c = int(color[3:5], 16)
                b_c = int(color[5:7], 16)
                self._pw_all.plot(all_x, all_y, pen=None, symbol="o",
                                  symbolSize=7,
                                  symbolBrush=pg.mkBrush(r_c, g_c, b_c, 140),
                                  symbolPen=pg.mkPen(r_c, g_c, b_c, 200),
                                  name=rider)

    # ── 📋 Performance テーブル ──────────────────────────────────────
    def _build_perf_table(self, laps: list, runs: list):
        cols    = ["Round", "Session", "Run", "Rider", "Laps",
                   "Best Lap", "Avg Lap", "Worst Lap",
                   "Gap to Best", "Improvement"]
        headers = cols
        self._tbl_perf.setSortingEnabled(False)
        self._tbl_perf.setColumnCount(len(cols))
        self._tbl_perf.setHorizontalHeaderLabels(headers)

        from collections import defaultdict
        rows_data = []
        for r in runs:
            rid = r["run_id"]
            r_laps = [l for l in laps if l["run_id"] == rid
                      and not l.get("is_outlap") and l.get("lap_time_s")]
            if not r_laps:
                continue
            times = [float(l["lap_time_s"]) for l in r_laps]
            improvement = times[0] - times[-1] if len(times) >= 2 else 0.0
            rows_data.append({
                "round": r.get("round", ""), "session": r.get("session", ""),
                "run_no": r.get("run_no", ""), "rider": r.get("rider", ""),
                "n_laps": len(times), "best": min(times),
                "avg": sum(times) / len(times), "worst": max(times),
                "improvement": improvement,
            })

        # セッション別ベスト（ギャップ計算用）
        session_bests: dict = defaultdict(lambda: float("inf"))
        for rd in rows_data:
            key = (rd["round"], rd["session"])
            session_bests[key] = min(session_bests[key], rd["best"])

        self._tbl_perf.setRowCount(len(rows_data))
        for i, rd in enumerate(rows_data):
            key = (rd["round"], rd["session"])
            gap = rd["best"] - session_bests[key]
            values = [
                rd["round"], rd["session"], f"R{rd['run_no']}", rd["rider"],
                str(rd["n_laps"]),
                self._fmt(rd["best"]), self._fmt(rd["avg"]), self._fmt(rd["worst"]),
                f"+{gap:.3f}s" if gap > 0.001 else "—",
                f"{rd['improvement']:+.3f}s" if abs(rd["improvement"]) > 0.001 else "—",
            ]
            for j, v in enumerate(values):
                item = QTableWidgetItem(str(v))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if gap < 0.001 and j >= 5:  # ベストラン行をハイライト
                    item.setBackground(QColor("#E8F5E9"))
                self._tbl_perf.setItem(i, j, item)

        self._tbl_perf.resizeColumnsToContents()
        self._tbl_perf.horizontalHeader().setStretchLastSection(True)
        self._tbl_perf.setSortingEnabled(True)

    # ── 🔧 Setup History テーブル ────────────────────────────────────
    def _build_setup_table(self, runs: list):
        run_ids  = [r["run_id"] for r in runs]
        detailed = self._db.get_trend_runs_detail(run_ids)

        key_cols = ["session", "run_no", "rider",
                    "tyre_front", "tyre_rear",
                    "f_comp", "f_reb", "f_offset",
                    "r_comp", "r_reb", "ride_hgt",
                    "perf_best_lap", "perf_n_laps", "comment"]
        headers  = ["Session", "Run", "Rider",
                    "Tyre F", "Tyre R",
                    "F Comp", "F Reb", "F Offset",
                    "R Comp", "R Reb", "Ride Hgt",
                    "Best Lap", "Laps", "Comment"]

        self._tbl_setup.setColumnCount(len(headers))
        self._tbl_setup.setHorizontalHeaderLabels(headers)
        self._tbl_setup.setRowCount(len(detailed))

        # 前行との差異を検出してハイライトするため値を収集
        prev_vals: dict = {}
        for i, r in enumerate(detailed):
            for j, col in enumerate(key_cols):
                v = r.get(col)
                if col == "perf_best_lap" and v is not None:
                    display = self._fmt(v)
                elif col == "run_no":
                    display = f"R{v}" if v is not None else "—"
                else:
                    display = str(v) if v is not None else "—"

                item = QTableWidgetItem(display)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                # セットアップ値が前行から変化した場合に黄色ハイライト
                setup_cols = {"f_comp", "f_reb", "f_offset", "r_comp", "r_reb", "ride_hgt",
                              "tyre_front", "tyre_rear"}
                if col in setup_cols and v is not None:
                    if prev_vals.get(col) is not None and str(prev_vals[col]) != str(v):
                        item.setBackground(QColor("#FFF9C4"))
                    prev_vals[col] = v

                self._tbl_setup.setItem(i, j, item)

        self._tbl_setup.resizeColumnsToContents()
        self._tbl_setup.horizontalHeader().setStretchLastSection(True)

    # ── ⚠️ Problem Analysis ─────────────────────────────────────────
    def _build_problem_view(self, problems: list):
        from collections import Counter
        tag_counts = Counter(p.get("problem_tag", "?") for p in problems)
        tags_sorted = sorted(tag_counts.items(), key=lambda x: -x[1])

        # バーチャート
        if self._has_pg:
            import pyqtgraph as pg
            self._pw_prob.clear()
            if tags_sorted:
                x = list(range(len(tags_sorted)))
                y = [cnt for _, cnt in tags_sorted]
                labels = [tag for tag, _ in tags_sorted]
                bargraph = pg.BarGraphItem(x=x, height=y, width=0.6,
                                           brush="#E74C3C", pen=pg.mkPen("w", width=0.5))
                self._pw_prob.addItem(bargraph)
                self._pw_prob.getAxis("bottom").setTicks([list(enumerate(labels))])
                self._pw_prob.setTitle("Problem Tag 頻度")
            else:
                # データなし時のメッセージ
                self._pw_prob.setTitle("")
                msg = pg.TextItem(
                    text="Problem Log にデータがありません\n"
                         "波形タブで問題を選択し「Problem Log へ送る」ボタンで登録できます",
                    color="#AAAAAA", anchor=(0.5, 0.5))
                msg.setPos(0.5, 0.5)
                self._pw_prob.addItem(msg)
                self._pw_prob.setXRange(0, 1)
                self._pw_prob.setYRange(0, 1)

        # テーブル
        cols = ["Tag", "件数", "Max Severity", "Riders", "Sessions", "Corners"]
        self._tbl_prob.setColumnCount(len(cols))
        self._tbl_prob.setHorizontalHeaderLabels(cols)
        self._tbl_prob.setRowCount(len(tags_sorted))

        sev_map: dict = {}
        for p in problems:
            tag = p.get("problem_tag", "?")
            sev_map[tag] = max(sev_map.get(tag, 0), int(p.get("severity") or 0))

        for i, (tag, cnt) in enumerate(tags_sorted):
            tag_probs = [p for p in problems if p.get("problem_tag") == tag]
            riders   = ", ".join(sorted({p.get("rider", "") for p in tag_probs if p.get("rider")}))
            sessions = ", ".join(sorted({p.get("session", "") for p in tag_probs if p.get("session")}))
            corners  = ", ".join(sorted({str(p.get("corner", "")) for p in tag_probs if p.get("corner")}))
            sev = sev_map.get(tag, 0)
            values = [tag, str(cnt), str(sev) if sev else "—", riders, sessions, corners]
            for j, v in enumerate(values):
                item = QTableWidgetItem(v)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if cnt >= 3:
                    item.setBackground(QColor("#FFEBEE"))
                elif cnt >= 2:
                    item.setBackground(QColor("#FFF3E0"))
                self._tbl_prob.setItem(i, j, item)

        self._tbl_prob.resizeColumnsToContents()
        self._tbl_prob.horizontalHeader().setStretchLastSection(True)

    # ── 📊 Lap Log テーブル ─────────────────────────────────────────
    def _build_laplog_table(self, laps: list):
        cols    = ["Round", "Circuit", "Session", "Run", "Lap", "Rider",
                   "Lap Time", "Outlap", "Tyre F", "Tyre R",
                   "Weather", "Track °C", "Air °C"]
        self._tbl_laplog.setSortingEnabled(False)
        self._tbl_laplog.setColumnCount(len(cols))
        self._tbl_laplog.setHorizontalHeaderLabels(cols)
        self._tbl_laplog.setRowCount(len(laps))

        for i, l in enumerate(laps):
            lt = l.get("lap_time_s")
            is_out = bool(l.get("is_outlap"))
            values = [
                l.get("round", ""), l.get("circuit", ""),
                l.get("session", ""), f"R{l.get('run_no','?')}",
                str(l.get("lap_no", "")), l.get("rider", ""),
                self._fmt(lt),
                "✓" if is_out else "",
                l.get("tyre_front", "") or "—",
                l.get("tyre_rear", "") or "—",
                l.get("weather", "") or "—",
                str(l.get("track_temp", "")) or "—",
                str(l.get("air_temp", "")) or "—",
            ]
            for j, v in enumerate(values):
                item = QTableWidgetItem(str(v))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if is_out:
                    item.setForeground(QColor("#999999"))
                # ライダー色分け
                rider = l.get("rider", "")
                if rider in self._RIDER_COLORS and not is_out:
                    c = self._RIDER_COLORS[rider]
                    item.setBackground(QColor(c).lighter(195))
                self._tbl_laplog.setItem(i, j, item)

        self._tbl_laplog.resizeColumnsToContents()
        self._tbl_laplog.horizontalHeader().setStretchLastSection(True)
        self._tbl_laplog.setSortingEnabled(True)

    # ── 🦾 Suspension ビュー (速度帯フィルター対応) ──────────────────
    @staticmethod
    def _valid_sus(sus_data: list) -> list:
        """outlap除去 + 極端な異常値除去の共通フィルター。"""
        return [d for d in sus_data
                if d.get("lap_time_s")
                and float(d["lap_time_s"]) > 85
                and float(d["lap_time_s"]) < 360
                and d.get("lap_no") and int(d["lap_no"]) > 0
                and d.get("apex_spd_avg")
                and float(d["apex_spd_avg"]) > 10]

    def _build_suspension_view(self, sus_data: list):
        """速度帯フィルター付きサスペンション分析ビュー。"""
        zone = getattr(self._cb_speed_zone, "currentText", lambda: "全て")()
        metric = getattr(self._cb_sus_metric, "currentText",
                         lambda: "ApexSusF vs LapTime")()

        valid = self._valid_sus(sus_data)

        # ── 速度帯フィルター ──
        if zone == "低速 <80km/h":
            filtered = [d for d in valid if float(d["apex_spd_avg"]) < 80]
        elif zone == "中速 80-120km/h":
            filtered = [d for d in valid if 80 <= float(d["apex_spd_avg"]) <= 120]
        elif zone == "高速 >120km/h":
            filtered = [d for d in valid if float(d["apex_spd_avg"]) > 120]
        else:
            filtered = valid

        # ── メトリック選択 ──
        if metric == "ApexSusR vs LapTime":
            y_key, y_lbl = "apex_susR_avg", "ApexSusR (mm)"
            x_key, x_lbl = "lap_time_s",    "Lap Time (s)"
        elif metric == "BrkSusF vs LapTime":
            y_key, y_lbl = "brk_susF_avg",  "BrkSusF (mm)"
            x_key, x_lbl = "lap_time_s",    "Lap Time (s)"
        elif metric == "BrkSusR vs LapTime":
            y_key, y_lbl = "brk_susR_avg",  "BrkSusR (mm)"
            x_key, x_lbl = "lap_time_s",    "Lap Time (s)"
        elif metric == "ApexSusF vs ApexSusR (F/R姿勢)":
            y_key, y_lbl = "apex_susF_avg", "ApexSusF (mm)"   # Y軸=フロント
            x_key, x_lbl = "apex_susR_avg", "ApexSusR (mm)"   # X軸=リア
        else:  # デフォルト
            y_key, y_lbl = "apex_susF_avg", "ApexSusF (mm)"
            x_key, x_lbl = "lap_time_s",    "Lap Time (s)"

        # ── ゾーン情報ラベル ──
        parts = []
        for rider in ("DA77", "JA52"):
            rpts = [d for d in filtered if d.get("rider") == rider
                    and d.get("apex_susF_avg") and d.get("apex_susR_avg")]
            if rpts:
                af = sum(float(d["apex_susF_avg"]) for d in rpts) / len(rpts)
                ar = sum(float(d["apex_susR_avg"]) for d in rpts) / len(rpts)
                parts.append(f"{rider}: n={len(rpts)} | ApexSusF={af:.1f}mm | ApexSusR={ar:.1f}mm")
        if hasattr(self, "_lbl_sus_zone_info"):
            self._lbl_sus_zone_info.setText("   ".join(parts))

        # ── メイン散布図 ──
        if self._has_pg and hasattr(self, "_pw_sus_main"):
            import pyqtgraph as pg
            self._pw_sus_main.clear()
            title = f"{y_lbl} vs {x_lbl}"
            if zone != "全て":
                title += f"  [{zone}]"
            self._pw_sus_main.setTitle(title)
            self._pw_sus_main.setLabel("left", y_lbl)
            self._pw_sus_main.setLabel("bottom", x_lbl)

            for rider, color in self._RIDER_COLORS.items():
                pts = [(float(d[x_key]), float(d[y_key]))
                       for d in filtered
                       if d.get("rider") == rider
                       and d.get(x_key) and d.get(y_key)]
                if not pts:
                    continue
                xs, ys = [p[0] for p in pts], [p[1] for p in pts]
                rc, gc, bc = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
                self._pw_sus_main.plot(
                    xs, ys, pen=None, symbol="o", symbolSize=8,
                    symbolBrush=pg.mkBrush(rc, gc, bc, 170),
                    symbolPen=pg.mkPen(rc, gc, bc, 220), name=rider)
                # 平均線
                if xs and y_key != x_key:
                    mean_y = sum(ys) / len(ys)
                    self._pw_sus_main.addItem(pg.InfiniteLine(
                        pos=mean_y, angle=0,
                        pen=pg.mkPen(rc, gc, bc, 120, width=1.5,
                                     style=Qt.PenStyle.DashLine),
                        label=f"{rider} avg:{mean_y:.1f}mm",
                        labelOpts={"color": color, "position": 0.05}))

        # ── 速度帯別バーチャート (右側) ──
        if self._has_pg and hasattr(self, "_pw_sus_zone"):
            import pyqtgraph as pg
            self._pw_sus_zone.clear()
            zone_defs = [
                ("低速\n<80", lambda d: float(d["apex_spd_avg"]) < 80),
                ("中速\n80-120", lambda d: 80 <= float(d["apex_spd_avg"]) <= 120),
                ("高速\n>120", lambda d: float(d["apex_spd_avg"]) > 120),
            ]
            x_ticks, bar_items = [], []
            x_base = 0.0
            for z_lbl, z_fn in zone_defs:
                z_pts = [d for d in valid if d.get("apex_spd_avg") and z_fn(d)]
                for ri, (rider, color) in enumerate(self._RIDER_COLORS.items()):
                    rz = [d for d in z_pts
                          if d.get("rider") == rider
                          and d.get("apex_susF_avg") and d.get("apex_susR_avg")]
                    if rz:
                        avg_f = sum(float(d["apex_susF_avg"]) for d in rz) / len(rz)
                        avg_r = sum(float(d["apex_susR_avg"]) for d in rz) / len(rz)
                        xpos = x_base + ri * 0.45
                        rc, gc, bc = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
                        # SusF (solid), SusR (transparent)
                        self._pw_sus_zone.addItem(pg.BarGraphItem(
                            x=[xpos], height=[avg_f], width=0.4,
                            brush=pg.mkBrush(rc, gc, bc, 220),
                            pen=pg.mkPen("k", width=0.5)))
                        self._pw_sus_zone.addItem(pg.BarGraphItem(
                            x=[xpos], height=[avg_r], width=0.4,
                            brush=pg.mkBrush(rc, gc, bc, 80),
                            pen=pg.mkPen(rc, gc, bc, 180, width=1.5,
                                         style=Qt.PenStyle.DashLine)))
                x_ticks.append((x_base + 0.22, z_lbl))
                x_base += 1.3
            if x_ticks:
                self._pw_sus_zone.getAxis("bottom").setTicks([x_ticks])

        # ── テーブル ──
        db_keys = ["round", "circuit", "session", "run_no", "lap_no", "rider",
                   "lap_time_s", "apex_spd_avg",
                   "apex_susF_avg", "apex_susR_avg",
                   "brk_susF_avg", "brk_susR_avg",
                   "fullbrk_susF", "fullbrk_susR",
                   "lap_susF_mean", "lap_susF_min", "lap_susF_max", "lap_susR_mean"]
        col_hdrs = ["Round", "Circuit", "Sess", "Run", "Lap", "Rider",
                    "Lap Time", "ApexSpd\n(km/h)",
                    "ApexSusF\n(mm)", "ApexSusR\n(mm)",
                    "BrkSusF\n(mm)", "BrkSusR\n(mm)",
                    "FullBrk\nSusF", "FullBrk\nSusR",
                    "SusF\nMean", "SusF\nMin", "SusF\nMax", "SusR\nMean"]

        self._tbl_sus.setSortingEnabled(False)
        self._tbl_sus.setColumnCount(len(col_hdrs))
        self._tbl_sus.setHorizontalHeaderLabels(col_hdrs)
        self._tbl_sus.setRowCount(len(filtered))

        # ライダー別に交互背景色（視認性を高める）
        RIDER_BG = {
            "DA77": ("#1A3A5C", "#FFFFFF"),   # 濃紺テキスト → 明るい青系背景
            "JA52": ("#5C2A00", "#FFFFFF"),   # 濃茶テキスト → 明るいオレンジ系背景
        }
        RIDER_FILL = {
            "DA77": QColor("#D0E8FF"),   # 青系
            "JA52": QColor("#FFE0B2"),   # オレンジ系
        }
        for i, d in enumerate(filtered):
            rider = d.get("rider", "")
            bg = RIDER_FILL.get(rider, QColor("#EEEEEE"))
            fg = QColor(RIDER_BG.get(rider, ("#000000", "#000000"))[0])
            for j, key in enumerate(db_keys):
                v = d.get(key)
                if key == "lap_time_s":
                    display = self._fmt(v)
                elif key == "run_no":
                    display = f"R{v}" if v is not None else "—"
                elif isinstance(v, float):
                    display = f"{v:.2f}"
                elif v is None:
                    display = "—"
                else:
                    display = str(v)
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setBackground(bg)
                item.setForeground(fg)
                self._tbl_sus.setItem(i, j, item)

        self._tbl_sus.resizeColumnsToContents()
        self._tbl_sus.horizontalHeader().setStretchLastSection(True)
        self._tbl_sus.setSortingEnabled(True)

    # ── 🔍 Analysis (FAST/SLOW + 回路別問題トレンド) ─────────────────
    def _build_analysis_view(self, sus_data: list, notes_data: dict):
        """FAST vs SLOW サスペンション比較 + 回路別問題タグ クロス集計。"""
        from collections import defaultdict

        valid = self._valid_sus(sus_data)

        # ════════════════════════════════════════════════
        # 1) FAST vs SLOW 比較 (ライダー×回路 ごと)
        # ════════════════════════════════════════════════
        groups: dict = defaultdict(list)
        for d in valid:
            groups[(d.get("rider"), d.get("circuit"))].append(d)

        def avg_key(lst, key):
            vals = [float(d[key]) for d in lst if d.get(key)]
            return sum(vals) / len(vals) if vals else None

        comp_rows = []
        for (rider, circuit), laps in sorted(groups.items()):
            if len(laps) < 4:
                continue
            times = sorted(float(d["lap_time_s"]) for d in laps)
            n3 = max(1, len(times) // 3)
            fast_t = set(times[:n3])
            slow_t = set(times[-n3:])
            fast_laps = [d for d in laps if float(d["lap_time_s"]) in fast_t]
            slow_laps = [d for d in laps if float(d["lap_time_s"]) in slow_t]

            def delta(fval, sval):
                if fval is None or sval is None:
                    return None
                return round(fval - sval, 2)

            faf = avg_key(fast_laps, "apex_susF_avg")
            saf = avg_key(slow_laps, "apex_susF_avg")
            far = avg_key(fast_laps, "apex_susR_avg")
            sar = avg_key(slow_laps, "apex_susR_avg")
            fbf = avg_key(fast_laps, "brk_susF_avg")
            sbf = avg_key(slow_laps, "brk_susF_avg")
            fbr = avg_key(fast_laps, "brk_susR_avg")
            sbr = avg_key(slow_laps, "brk_susR_avg")

            comp_rows.append({
                "rider": rider or "?", "circuit": circuit or "?",
                "n": len(laps), "n3": n3,
                "fast_best": min(times[:n3]), "slow_best": max(times[-n3:]),
                "fApexF": faf, "sApexF": saf, "dApexF": delta(faf, saf),
                "fApexR": far, "sApexR": sar, "dApexR": delta(far, sar),
                "fBrkF": fbf,  "sBrkF": sbf,  "dBrkF":  delta(fbf, sbf),
                "fBrkR": fbr,  "sBrkR": sbr,  "dBrkR":  delta(fbr, sbr),
            })

        # バーチャート: Δ ApexSusF per rider+circuit
        if self._has_pg and hasattr(self, "_pw_fast_slow"):
            import pyqtgraph as pg
            self._pw_fast_slow.clear()
            x_ticks = []
            for xi, row in enumerate(comp_rows):
                color = self._RIDER_COLORS.get(row["rider"], "#888")
                rc, gc, bc = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
                d_f = row.get("dApexF") or 0
                d_r = row.get("dApexR") or 0
                # SusF bar (solid), SusR bar (behind, lighter)
                self._pw_fast_slow.addItem(pg.BarGraphItem(
                    x=[xi - 0.18], height=[d_f], width=0.32,
                    brush=pg.mkBrush(rc, gc, bc, 220),
                    pen=pg.mkPen("k", width=0.5)))
                self._pw_fast_slow.addItem(pg.BarGraphItem(
                    x=[xi + 0.18], height=[d_r], width=0.32,
                    brush=pg.mkBrush(rc, gc, bc, 90),
                    pen=pg.mkPen(rc, gc, bc, 180, width=1)))
                x_ticks.append((xi, f"{row['rider']}\n{(row['circuit'] or '')[:7]}"))
            self._pw_fast_slow.setTitle(
                "FAST vs SLOW Δ: ■=ApexSusF  □=ApexSusR  (負=FAST時に沈む=良方向)")
            self._pw_fast_slow.addItem(pg.InfiniteLine(
                pos=0, angle=0, pen=pg.mkPen("k", width=1)))
            if x_ticks:
                self._pw_fast_slow.getAxis("bottom").setTicks([x_ticks])
                self._pw_fast_slow.getAxis("bottom").setStyle(tickFont=None)

        # FAST/SLOW テーブル
        fs_hdrs = ["Rider", "Circuit", "N", "n/3",
                   "FAST Best", "SLOW Best",
                   "FAST\nApexF", "SLOW\nApexF", "Δ ApexF",
                   "FAST\nApexR", "SLOW\nApexR", "Δ ApexR",
                   "FAST\nBrkF",  "SLOW\nBrkF",  "Δ BrkF",
                   "FAST\nBrkR",  "SLOW\nBrkR",  "Δ BrkR"]
        self._tbl_fast_slow.setSortingEnabled(False)
        self._tbl_fast_slow.setColumnCount(len(fs_hdrs))
        self._tbl_fast_slow.setHorizontalHeaderLabels(fs_hdrs)
        self._tbl_fast_slow.setRowCount(len(comp_rows))
        self._tbl_fast_slow.verticalHeader().setDefaultSectionSize(22)

        delta_cols = {8, 11, 14, 17}  # Δ列インデックス

        for i, row in enumerate(comp_rows):
            color = self._RIDER_COLORS.get(row["rider"])
            base_bg = QColor(color).lighter(215) if color else QColor("#F5F5F5")

            def fmtv(v):
                if v is None: return "—"
                if isinstance(v, float): return f"{v:.2f}"
                return str(v)

            vals = [
                row["rider"], row["circuit"],
                str(row["n"]), str(row["n3"]),
                self._fmt(row["fast_best"]), self._fmt(row["slow_best"]),
                fmtv(row["fApexF"]), fmtv(row["sApexF"]),
                f"{row['dApexF']:+.2f}" if row.get("dApexF") is not None else "—",
                fmtv(row["fApexR"]), fmtv(row["sApexR"]),
                f"{row['dApexR']:+.2f}" if row.get("dApexR") is not None else "—",
                fmtv(row["fBrkF"]),  fmtv(row["sBrkF"]),
                f"{row['dBrkF']:+.2f}" if row.get("dBrkF") is not None else "—",
                fmtv(row["fBrkR"]),  fmtv(row["sBrkR"]),
                f"{row['dBrkR']:+.2f}" if row.get("dBrkR") is not None else "—",
            ]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if j in delta_cols and v not in ("—", ""):
                    try:
                        dv = float(str(v).replace("+", ""))
                        # 負=FAST時により沈む=良方向 → 緑
                        item.setBackground(QColor("#C8E6C9") if dv < -0.5
                                           else QColor("#FFCDD2") if dv > 0.5
                                           else QColor("#FFF9C4"))
                    except ValueError:
                        item.setBackground(base_bg)
                else:
                    item.setBackground(base_bg)
                self._tbl_fast_slow.setItem(i, j, item)

        self._tbl_fast_slow.resizeColumnsToContents()
        self._tbl_fast_slow.horizontalHeader().setStretchLastSection(True)
        self._tbl_fast_slow.setSortingEnabled(True)

        # ════════════════════════════════════════════════
        # 2) 回路別 問題タグ クロス集計
        # ════════════════════════════════════════════════
        rider_tags = notes_data.get("rider_tags", {})
        circuit_tag: dict = defaultdict(lambda: defaultdict(dict))

        for rider, rtags in rider_tags.items():
            for tag, count, circuits_str in rtags:
                for circ in [c.strip() for c in circuits_str.split(",") if c.strip()]:
                    circuit_tag[circ][tag][rider] = count

        if not circuit_tag:
            self._tbl_circuit_prob.setRowCount(0)
            return

        # タグを全体件数で降順ソート
        all_tags_raw = set()
        for td in circuit_tag.values():
            all_tags_raw.update(td.keys())
        tag_totals = {
            t: sum(circuit_tag[c].get(t, {}).get("DA77", 0)
                   + circuit_tag[c].get(t, {}).get("JA52", 0)
                   for c in circuit_tag)
            for t in all_tags_raw
        }
        tags_ord = sorted(all_tags_raw, key=lambda t: -tag_totals.get(t, 0))
        circuits_ord = sorted(circuit_tag.keys())

        # ヘッダー: Circuit + タグ列 (DA77 | JA52)
        hdr = ["Circuit"] + tags_ord
        self._tbl_circuit_prob.setColumnCount(len(hdr))
        self._tbl_circuit_prob.setHorizontalHeaderLabels(hdr)
        self._tbl_circuit_prob.setRowCount(len(circuits_ord))
        self._tbl_circuit_prob.verticalHeader().setDefaultSectionSize(24)

        for i, circ in enumerate(circuits_ord):
            for j, col in enumerate(hdr):
                if col == "Circuit":
                    item = QTableWidgetItem(circ)
                    item.setBackground(QColor("#E3F2FD"))
                else:
                    td = circuit_tag[circ].get(col, {})
                    da = td.get("DA77", 0)
                    ja = td.get("JA52", 0)
                    total = da + ja
                    if total == 0:
                        display = "—"
                        bg = QColor("#FAFAFA")
                    else:
                        display = f"D:{da}  J:{ja}"
                        bg = (QColor("#FFCDD2") if total >= 2
                              else QColor("#FFF9C4"))
                    item = QTableWidgetItem(display)
                    item.setBackground(bg)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._tbl_circuit_prob.setItem(i, j, item)

        self._tbl_circuit_prob.resizeColumnsToContents()
        self._tbl_circuit_prob.horizontalHeader().setStretchLastSection(True)

    # ── 📊 Perf Corr テーブル ────────────────────────────────────────
    def _build_perf_corr_tables(self, pc_cache: tuple, rider_f: str | None, round_f: str | None):
        """PERFORMANCE_CORRELATION の per-run と FAST/SLOW 比較をテーブル表示。"""
        run_hdrs, run_data, cmp_hdrs, cmp_data = pc_cache

        # ── per-run テーブル ──
        display_run_cols = [
            "RUN_ID", "Rider", "Circuit", "Date", "Session", "Run",
            "Best Lap", "Avg Lap", "N Laps",
            "APEX SusF (mm)", "APEX SusR (mm)",
            "APEX WhlF (N)", "APEX WhlR (N)",
            "APEX Spd (km/h)", "APEX ax (m/s²)",
            "Brk SusF (mm)", "Brk SusR (mm)", "Brk Spd (km/h)",
            "Rank", "Gap (s)", "Tier",
        ]
        # run_hdrs のキーマッピング (header → friendly name)
        hdr_map = {}
        if run_hdrs:
            for h in run_hdrs:
                clean = h.replace("\n", " ").strip()
                hdr_map[clean] = clean

        # フィルタリング
        filtered = run_data
        if rider_f:
            filtered = [d for d in filtered if d.get("Rider") == rider_f]
        if round_f:
            filtered = [d for d in filtered
                        if round_f.upper() in str(d.get(run_hdrs[1] if run_hdrs else "RUN_ID", "")).upper()]

        # テーブル構築
        raw_cols = run_hdrs[1:] if run_hdrs else []  # skip col0 (stale)
        friendly = [c.replace("\n", " ").strip() for c in raw_cols]
        self._tbl_perf_corr.setSortingEnabled(False)
        self._tbl_perf_corr.setColumnCount(len(friendly))
        self._tbl_perf_corr.setHorizontalHeaderLabels(friendly)
        self._tbl_perf_corr.setRowCount(len(filtered))

        tier_colors = {"FAST": "#E8F5E9", "MED": "#FFF9C4", "SLOW": "#FFEBEE"}
        for i, d in enumerate(filtered):
            tier = str(d.get("Tier") or "")
            bg = tier_colors.get(tier.upper(), None)
            for j, col in enumerate(raw_cols):
                v = d.get(col)
                # 数値型の場合 2桁小数
                if col in ("Best Lap", "Avg Lap"):
                    display = str(v) if v else "—"
                elif isinstance(v, float):
                    display = f"{v:.2f}"
                else:
                    display = str(v) if v is not None else "—"
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if bg:
                    item.setBackground(QColor(bg))
                self._tbl_perf_corr.setItem(i, j, item)

        self._tbl_perf_corr.resizeColumnsToContents()
        self._tbl_perf_corr.horizontalHeader().setStretchLastSection(True)
        self._tbl_perf_corr.setSortingEnabled(True)

        # ── FAST/SLOW 比較テーブル ──
        cmp_filtered = cmp_data
        if rider_f:
            cmp_filtered = [d for d in cmp_data
                            if d.get(cmp_hdrs[2] if len(cmp_hdrs) > 2 else "Rider") == rider_f]

        cmp_raw = cmp_hdrs[2:] if len(cmp_hdrs) > 2 else cmp_hdrs
        cmp_friendly = [c.replace("\n", " ").strip() for c in cmp_raw]
        self._tbl_perf_cmp.setSortingEnabled(False)
        self._tbl_perf_cmp.setColumnCount(len(cmp_friendly))
        self._tbl_perf_cmp.setHorizontalHeaderLabels(cmp_friendly)
        self._tbl_perf_cmp.setRowCount(len(cmp_filtered))

        for i, d in enumerate(cmp_filtered):
            rider = d.get(cmp_hdrs[2] if len(cmp_hdrs) > 2 else "", "")
            bc = self._RIDER_COLORS.get(rider)
            for j, col in enumerate(cmp_raw):
                v = d.get(col)
                if isinstance(v, float):
                    display = f"{v:.2f}"
                else:
                    display = str(v) if v is not None else "—"
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if bc:
                    item.setBackground(QColor(bc).lighter(215))
                self._tbl_perf_cmp.setItem(i, j, item)

        self._tbl_perf_cmp.resizeColumnsToContents()
        self._tbl_perf_cmp.horizontalHeader().setStretchLastSection(True)

    # ── 📝 Session Notes ─────────────────────────────────────────────
    def _build_session_notes(self, notes_data: dict, rider_f: str | None):
        """問題タグサマリー + エンジニアノートを表示。"""
        # ── 左: 問題タグテーブル ──
        tags = notes_data.get("tags", [])
        rider_tags = notes_data.get("rider_tags", {})

        # 全タグ + ライダー別件数を統合表示
        combined: dict = {}
        for tag, cnt, phase, meaning in tags:
            combined[tag] = {"total": cnt, "phase": phase, "meaning": meaning,
                             "DA77": 0, "JA52": 0}
        for rider, rtags in rider_tags.items():
            for tag, cnt, circuits in rtags:
                if tag not in combined:
                    combined[tag] = {"total": 0, "phase": "", "meaning": "",
                                     "DA77": 0, "JA52": 0}
                combined[tag][rider] = cnt
                combined[tag]["circuits_" + rider] = circuits

        # ライダーフィルター適用
        if rider_f:
            sorted_tags = sorted(
                [(t, d) for t, d in combined.items() if d.get(rider_f, 0) > 0],
                key=lambda x: -x[1].get(rider_f, 0)
            )
        else:
            sorted_tags = sorted(combined.items(), key=lambda x: -x[1].get("total", 0))

        tag_cols = ["Tag", "Phase", "Total", "DA77", "JA52", "Meaning"]
        self._tbl_tag_summary.setSortingEnabled(False)
        self._tbl_tag_summary.setColumnCount(len(tag_cols))
        self._tbl_tag_summary.setHorizontalHeaderLabels(tag_cols)
        self._tbl_tag_summary.setRowCount(len(sorted_tags))

        for i, (tag, d) in enumerate(sorted_tags):
            total = d.get("total", 0)
            bg = "#FFCDD2" if total >= 8 else "#FFE0B2" if total >= 5 else "#FFF9C4" if total >= 3 else None
            values = [tag, d.get("phase", ""), str(total),
                      str(d.get("DA77", 0)), str(d.get("JA52", 0)), d.get("meaning", "")]
            for j, v in enumerate(values):
                item = QTableWidgetItem(str(v))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if bg:
                    item.setBackground(QColor(bg))
                self._tbl_tag_summary.setItem(i, j, item)

        self._tbl_tag_summary.resizeColumnsToContents()
        self._tbl_tag_summary.horizontalHeader().setStretchLastSection(True)
        self._tbl_tag_summary.setSortingEnabled(True)

        # ── 右: エンジニアノートHTML ──
        notes = notes_data.get("notes", [])
        if rider_f:
            notes = [(k, v) for k, v in notes if rider_f in k.upper()]

        html_parts = ["<html><body style='font-family:Arial;font-size:11px;'>"]
        for session_key, note_text in sorted(notes, reverse=True):
            if not note_text:
                continue
            # セッションキー色分け
            if "ROUND4" in session_key: hdr_col = "#1565C0"
            elif "ROUND3" in session_key: hdr_col = "#2E7D32"
            elif "ROUND2" in session_key: hdr_col = "#6A1B9A"
            elif "ROUND1" in session_key: hdr_col = "#BF360C"
            else: hdr_col = "#424242"

            # ライダー色
            if "DA77" in session_key: rider_col = "#0078D4"
            elif "JA52" in session_key: rider_col = "#E86C00"
            else: rider_col = "#666"

            html_parts.append(
                f"<div style='margin-bottom:12px; border-left:4px solid {hdr_col}; "
                f"padding-left:8px;'>"
                f"<div style='font-weight:bold; color:{rider_col}; margin-bottom:4px;'>"
                f"📋 {session_key}</div>"
                f"<div style='color:#333; white-space:pre-wrap;'>"
                f"{note_text.replace('<', '&lt;').replace('>', '&gt;')}</div>"
                f"</div>"
            )
        html_parts.append("</body></html>")
        self._txt_notes.setHtml("".join(html_parts))


# ════════════════════════════════════════════════════════════════════
# メインウィンドウ
# ════════════════════════════════════════════════════════════════════

class PostureAnalysisTab(QWidget):
    """🎯 姿勢分析タブ
    Pitch = ApexSusF - ApexSusR  (負値=ノーズDOWN=良好なターンイン)
    Heave = (ApexSusF + ApexSusR) / 2  (全体沈み込み量)
    データソース: lap_suspension_data.json (参考値 §0)
    """

    _LAP_SUS = SCRIPT_DIR / "lap_suspension_data.json"
    _COLORS   = {"DA77": "#0078D4", "JA52": "#FF8C00"}

    def __init__(self, db: WorkbenchDB, parent=None):
        super().__init__(parent)
        self._db  = db
        self._df  = None        # pandas DataFrame
        self._circuit_filter = ""
        self._setup_ui()
        self._load_data()

    # ── データ読み込み ──────────────────────────────────────────────

    def refresh(self):
        """DB/JSON変更時に外部から呼び出される。データを再読み込みする。"""
        self._load_data()

    def _load_data(self):
        if not self._LAP_SUS.exists():
            if hasattr(self, "_lbl_status"):
                self._lbl_status.setText(
                    "⚠️  lap_suspension_data.json が見つかりません。"
                    " python lap_suspension_stats.py を実行してください。"
                )
            return
        try:
            raw = json.loads(self._LAP_SUS.read_text(encoding="utf-8"))
            self._df = pd.DataFrame(raw)
            self._df.columns = [c.lower() for c in self._df.columns]
            sf = "apex_susf_avg"
            sr = "apex_susr_avg"
            if sf in self._df.columns and sr in self._df.columns:
                self._df["pitch"] = self._df[sf] - self._df[sr]
                self._df["heave"] = (self._df[sf] + self._df[sr]) / 2.0
            if hasattr(self, "_lbl_status"):
                n = len(self._df)
                riders = self._df["rider"].unique().tolist() if "rider" in self._df.columns else []
                self._lbl_status.setText(
                    f"✅  {n} ラップ読み込み済 | riders: {', '.join(riders)}"
                )
            # サーキット選択コンボ更新
            if "circuit" in self._df.columns:
                circs = sorted(self._df["circuit"].dropna().unique().tolist())
                self._combo_circ.blockSignals(True)
                self._combo_circ.clear()
                self._combo_circ.addItem("全サーキット")
                self._combo_circ.addItems(circs)
                self._combo_circ.blockSignals(False)
            self._update_all()
        except Exception as e:
            if hasattr(self, "_lbl_status"):
                self._lbl_status.setText(f"❌ 読み込みエラー: {e}")

    def _filtered_df(self):
        if self._df is None:
            return None
        df = self._df
        circ = self._combo_circ.currentText()
        if circ and circ != "全サーキット" and "circuit" in df.columns:
            df = df[df["circuit"] == circ]
        return df

    # ── UI 構築 ────────────────────────────────────────────────────

    def _setup_ui(self):
        try:
            import pyqtgraph as pg
            self._pg   = pg
            self._haspg = True
        except ImportError:
            self._haspg = False

        root = QVBoxLayout(self)
        root.setSpacing(4)

        # §0 原則注記
        warn = QLabel("⚠️  §0 参考値 — lap_suspension_data.json (推定値)")
        warn.setStyleSheet("color: #D83B01; font-style: italic; font-size: 10px; padding: 2px;")
        root.addWidget(warn)

        # ツールバー行
        tb = QHBoxLayout()
        self._lbl_status = QLabel("データ読込中…")
        self._lbl_status.setStyleSheet("font-size: 10px; color: #666;")
        tb.addWidget(self._lbl_status, stretch=1)
        tb.addWidget(QLabel("Circuit:"))
        self._combo_circ = QComboBox()
        self._combo_circ.setMinimumWidth(130)
        self._combo_circ.currentTextChanged.connect(self._update_all)
        tb.addWidget(self._combo_circ)
        btn_reload = QPushButton("↺ 再読込")
        btn_reload.setFixedHeight(24)
        btn_reload.clicked.connect(self._load_data)
        tb.addWidget(btn_reload)
        root.addLayout(tb)

        if not self._haspg:
            root.addWidget(QLabel("pyqtgraph が必要です: pip install pyqtgraph"))
            return

        pg = self._pg
        pg.setConfigOption("background", "w")
        pg.setConfigOption("foreground", "k")

        # 2×2 グリッド: QSplitter 縦 × (横スプリッタ 上/下)
        vsplit = QSplitter(Qt.Orientation.Vertical)

        # 上段スプリッタ
        top = QSplitter(Qt.Orientation.Horizontal)
        self._pw_scatter = pg.PlotWidget(title="Pitch vs Lap Time")
        self._pw_phase   = pg.PlotWidget(title="Phase Space (SusF vs SusR)")
        top.addWidget(self._pw_scatter)
        top.addWidget(self._pw_phase)
        top.setStretchFactor(0, 1)
        top.setStretchFactor(1, 1)

        # 下段スプリッタ
        bot = QSplitter(Qt.Orientation.Horizontal)
        self._pw_radar = pg.PlotWidget(title="Rider Fingerprint")
        self._pw_trend = pg.PlotWidget(title="Pitch / Heave Lap推移")
        bot.addWidget(self._pw_radar)
        bot.addWidget(self._pw_trend)
        bot.setStretchFactor(0, 1)
        bot.setStretchFactor(1, 1)

        vsplit.addWidget(top)
        vsplit.addWidget(bot)
        vsplit.setStretchFactor(0, 1)
        vsplit.setStretchFactor(1, 1)
        root.addWidget(vsplit, stretch=1)

        for _pw in (self._pw_scatter, self._pw_phase, self._pw_radar, self._pw_trend):
            _pw.showGrid(x=True, y=True, alpha=0.3)
            _pw.addLegend()

        # Radar は極座標描画のため軸を非表示・正方形
        self._pw_radar.setAspectLocked(True)
        self._pw_radar.hideAxis("bottom")
        self._pw_radar.hideAxis("left")

    # ── 描画 ───────────────────────────────────────────────────────

    def _update_all(self):
        if not self._haspg or self._df is None:
            return
        df = self._filtered_df()
        if df is None or df.empty:
            return
        self._draw_pitch_scatter(df)
        self._draw_phase_space(df)
        self._draw_radar(df)
        self._draw_trend(df)

    def _draw_pitch_scatter(self, df):
        """Panel 1: Pitch vs Lap Time散布図。"""
        pg  = self._pg
        pw  = self._pw_scatter
        pw.clear()
        pw.setLabel("left", "Pitch (mm) = SusF - SusR")
        pw.setLabel("bottom", "Lap Time (s)")
        if "pitch" not in df.columns or "lap_time_s" not in df.columns:
            return
        pw.addLegend()
        for rider, col in self._COLORS.items():
            sub = df[df.get("rider", pd.Series(dtype=str)) == rider] if "rider" in df.columns else df
            if rider not in df.get("rider", pd.Series(dtype=str)).values:
                continue
            sub = df[df["rider"] == rider].dropna(subset=["pitch", "lap_time_s"])
            if sub.empty:
                continue
            pw.plot(
                x=sub["lap_time_s"].values.tolist(),
                y=sub["pitch"].values.tolist(),
                pen=None,
                symbol="o", symbolSize=6,
                symbolBrush=pg.mkBrush(col),
                symbolPen=pg.mkPen(col, width=0.5),
                name=rider,
            )
        # ゼロライン
        pw.addItem(pg.InfiniteLine(pos=0, angle=0, pen=pg.mkPen("#888", width=1,
                                   style=Qt.PenStyle.DotLine)))

    def _draw_phase_space(self, df):
        """Panel 2: SusF vs SusR Phase Space。速いラップ=青、遅い=赤。"""
        pg = self._pg
        pw = self._pw_phase
        pw.clear()
        pw.setLabel("left",   "Apex SusR (mm)")
        pw.setLabel("bottom", "Apex SusF (mm)")
        sf_col = "apex_susf_avg"
        sr_col = "apex_susr_avg"
        lt_col = "lap_time_s"
        if sf_col not in df.columns or sr_col not in df.columns:
            return
        sub = df.dropna(subset=[sf_col, sr_col])
        if sub.empty:
            return
        pw.addLegend()
        for rider, col in self._COLORS.items():
            if "rider" not in df.columns:
                break
            rs = sub[sub["rider"] == rider]
            if rs.empty:
                continue
            symbol = "o" if rider == "DA77" else "t"
            pw.plot(
                x=rs[sf_col].values.tolist(),
                y=rs[sr_col].values.tolist(),
                pen=None,
                symbol=symbol, symbolSize=7,
                symbolBrush=pg.mkBrush(col + "A0"),
                symbolPen=pg.mkPen(col, width=0.5),
                name=rider,
            )
        # 対角線（SusF=SusR）
        lim = max(sub[sf_col].max(), sub[sr_col].max()) * 1.05
        pw.plot([0, lim], [0, lim], pen=pg.mkPen("#CCC", width=1,
                style=Qt.PenStyle.DotLine))

    def _draw_radar(self, df):
        """Panel 3: ライダー指紋レーダーチャート（5軸）。"""
        import math
        pg = self._pg
        pw = self._pw_radar
        pw.clear()
        pw.addLegend()

        METRICS = [
            ("pitch",          "Pitch\n(SusF-SusR)",  True),   # (列名, ラベル, 小さい=良い)
            ("heave",          "Heave\n(avg sink)",    True),
            ("brk_susf_avg",   "BRK\nSusF",           True),
            ("apex_spd_avg",   "Apex\nSpeed",          False),  # 大きい=良い
            ("lap_time_s",     "Lap\nTime",            True),
        ]
        n = len(METRICS)
        angles = [2 * math.pi * i / n - math.pi / 2 for i in range(n)]

        # 各指標の全ライダー平均
        rider_vals: dict[str, list[float]] = {}
        raw_stats: dict[str, dict] = {}
        for rider in self._COLORS:
            if "rider" not in df.columns:
                break
            rs = df[df["rider"] == rider]
            vals = []
            stats = {}
            for col, _lbl, lower_better in METRICS:
                if col in rs.columns:
                    v = rs[col].dropna().mean()
                    stats[col] = float(v) if not pd.isna(v) else 0.0
                else:
                    stats[col] = 0.0
                vals.append(stats[col])
            rider_vals[rider] = vals
            raw_stats[rider] = stats

        if not rider_vals:
            return

        # 各軸を 0-1 正規化（全ライダー横断）
        norm_vals: dict[str, list[float]] = {r: [] for r in rider_vals}
        for i, (col, _lbl, lower_better) in enumerate(METRICS):
            all_v = [rider_vals[r][i] for r in rider_vals if rider_vals[r]]
            mn, mx = min(all_v), max(all_v)
            span = mx - mn if mx != mn else 1.0
            for rider in rider_vals:
                raw = rider_vals[rider][i]
                norm = (raw - mn) / span  # 0=低, 1=高
                # 「小さい=良い」指標は反転して 1=良い になるよう
                if lower_better:
                    norm = 1.0 - norm
                norm_vals[rider].append(norm)

        # グリッド円
        for r in [0.25, 0.5, 0.75, 1.0]:
            xs = [math.cos(a) * r for a in angles] + [math.cos(angles[0]) * r]
            ys = [math.sin(a) * r for a in angles] + [math.sin(angles[0]) * r]
            pw.plot(xs, ys, pen=pg.mkPen("#DDD", width=0.7))

        # 軸線 + ラベル
        for a, (_, lbl, _b) in zip(angles, METRICS):
            pw.plot([0, math.cos(a)], [0, math.sin(a)],
                    pen=pg.mkPen("#AAA", width=0.7))
            ti = pg.TextItem(lbl, anchor=(0.5, 0.5), color="#555")
            ti.setPos(math.cos(a) * 1.22, math.sin(a) * 1.22)
            pw.addItem(ti)

        # ライダーポリゴン
        for rider, col in self._COLORS.items():
            if rider not in norm_vals:
                continue
            nv = norm_vals[rider]
            xs = [math.cos(angles[i]) * nv[i] for i in range(n)] + \
                 [math.cos(angles[0]) * nv[0]]
            ys = [math.sin(angles[i]) * nv[i] for i in range(n)] + \
                 [math.sin(angles[0]) * nv[0]]
            pw.plot(xs, ys, pen=pg.mkPen(col, width=2.5), name=rider,
                    fillLevel=0, brush=pg.mkBrush(col + "28"))

        pw.setXRange(-1.4, 1.4, padding=0)
        pw.setYRange(-1.4, 1.4, padding=0)

    def _draw_trend(self, df):
        """Panel 4: Lap 推移 (Pitch / Heave)。最新ランを自動選択。"""
        pg = self._pg
        pw = self._pw_trend
        pw.clear()
        pw.setLabel("left",   "mm")
        pw.setLabel("bottom", "Lap No")
        pw.addLegend()
        if "pitch" not in df.columns or "lap_no" not in df.columns:
            return
        for rider, col in self._COLORS.items():
            if "rider" not in df.columns:
                break
            rs = df[df["rider"] == rider].sort_values("lap_no")
            if rs.empty:
                continue
            rs = rs.dropna(subset=["lap_no", "pitch", "heave"])
            if rs.empty:
                continue
            laps = rs["lap_no"].values.tolist()
            pw.plot(laps, rs["pitch"].values.tolist(),
                    pen=pg.mkPen(col, width=2),
                    symbol="o", symbolSize=5, symbolBrush=pg.mkBrush(col),
                    name=f"{rider} Pitch")
            pw.plot(laps, rs["heave"].values.tolist(),
                    pen=pg.mkPen(col, width=1.5, style=Qt.PenStyle.DashLine),
                    symbol="t", symbolSize=5, symbolBrush=pg.mkBrush(col + "80"),
                    name=f"{rider} Heave")
        # Pitch=0 ライン
        pw.addItem(pg.InfiniteLine(pos=0, angle=0,
                   pen=pg.mkPen("#888", width=1, style=Qt.PenStyle.DotLine)))


class MainWindow(QMainWindow):
    def __init__(self, db: WorkbenchDB):
        super().__init__()
        self._db = db
        self.setWindowTitle("TS24 Engineer Workbench v2.0")
        self.resize(1400, 800)
        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── 上部ツールバー ────────────────────────────────────────────────────────────
        toolbar = QWidget()
        toolbar.setFixedHeight(40)
        toolbar.setStyleSheet("background: #1E1E1E; border-bottom: 1px solid #333;")
        tb_lay = QHBoxLayout(toolbar)
        tb_lay.setContentsMargins(8, 4, 8, 4)

        lbl_title = QLabel("TS24 Engineer Workbench v2.0")
        lbl_title.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        lbl_title.setStyleSheet("color: #FFFFFF;")
        tb_lay.addWidget(lbl_title)
<<<<<<< Updated upstream
=======

        lbl_circ = QLabel("  Circuit:")
        lbl_circ.setStyleSheet("color: #CCC;")
        tb_lay.addWidget(lbl_circ)
        self._combo_circuit = QComboBox()
        self._combo_circuit.setFixedWidth(120)
        self._combo_circuit.setToolTip("テンプレート（コーナーマーカー）に使用")
        self._combo_circuit.currentTextChanged.connect(self._on_circuit_changed)
        tb_lay.addWidget(self._combo_circuit)

        btn_open_csv = QPushButton("📂  CSVを開く")
        btn_open_csv.setFixedHeight(28)
        btn_open_csv.setStyleSheet(
            "QPushButton { background: #0078D4; color: white; border-radius: 4px; padding: 0 12px; }"
            "QPushButton:hover { background: #106EBE; }"
        )
        btn_open_csv.clicked.connect(self._open_csv)
        tb_lay.addWidget(btn_open_csv)

        btn_compare_csv = QPushButton("📂 比較CSV")
        btn_compare_csv.setFixedHeight(28)
        btn_compare_csv.setToolTip("2人目ライダーのCSVを追加して波形を重ねて表示")
        btn_compare_csv.setStyleSheet(
            "QPushButton { background: #5C2D91; color: white; padding: 4px 10px;"
            " border-radius: 4px; font-weight: bold; }"
            "QPushButton:hover { background: #4A2175; }"
        )
        btn_compare_csv.clicked.connect(self._open_csv_compare)
        tb_lay.addWidget(btn_compare_csv)

>>>>>>> Stashed changes
        tb_lay.addStretch()

        self._lbl_status = QLabel("")
        self._lbl_status.setStyleSheet("color: #888; font-size: 10px;")
        tb_lay.addWidget(self._lbl_status)

        root.addWidget(toolbar)

        # ── タブエリア ────────────────────────────────────────────────────────────
        self._tabs = QTabWidget()

        self._tab_browser = RunBrowserTab(db=self._db)
        self._tab_quick   = QuickLogTab(db=self._db)
        self._tab_problem = ProblemLogTab(db=self._db)
        self._tab_setup   = SetupDecisionTab(db=self._db)
<<<<<<< Updated upstream
        self._tab_posture = PostureAnalysisTab(db=self._db)

        # Run Browser → Quick Log / Problem Log / Setup Decision に連携
        self._tab_browser.set_on_run_selected(self._on_run_selected)

        self._tabs.addTab(self._tab_browser, "🗺️ Run Browser")
        self._tabs.addTab(self._tab_quick,   "⚡ Quick Log")
        self._tabs.addTab(self._tab_problem, "📋 Problem Log")
        self._tabs.addTab(self._tab_setup,   "🔧 Setup Decision")
        self._tabs.addTab(self._tab_posture, "📈 Trend Analysis")
=======
        self._tab_csv     = CsvImportTab(wave_view=self._tab_wave, db=self._db)
        self._tab_trend   = TrendAnalysisTab(db=self._db)
        self._tab_wave.set_problem_tab(self._tab_problem)
        if hasattr(self._tab_wave, "_right_panel"):
            self._tab_wave._right_panel.set_problem_tab(self._tab_problem)
        self._tab_csv._on_loaded = self._on_csv_loaded
        self._tabs.addTab(self._tab_wave,    "📊 波形 (Reference)")
        self._tabs.addTab(self._tab_problem, "⚠️  Problem Log")
        self._tabs.addTab(self._tab_setup,   "🔧 Setup Decision")
        self._tabs.addTab(self._tab_csv,     "📂 2D CSV")
        self._tabs.addTab(self._tab_trend,   "📈 Trend Analysis")
>>>>>>> Stashed changes

        root.addWidget(self._tabs)

        # ── DB ファイル監視 ──────────────────────────────────────────────────────
        self._fs_watcher = QFileSystemWatcher([str(DB_PATH)])
        self._fs_watcher.fileChanged.connect(self._on_db_changed)

    def _on_db_changed(self, _path: str) -> None:
        """DB ファイルが更新されたとき全タブを自動リフレッシュする。"""
        self._lbl_status.setText("🔄 DB更新検出 — リフレッシュ中…")
        for tab in (self._tab_browser, self._tab_quick,
                    self._tab_problem, self._tab_setup, self._tab_posture):
            try:
                tab.refresh()
            except Exception:
                pass
        # watchdog が rename-replace でファイルを再作成する場合、パスが消える
        if str(DB_PATH) not in self._fs_watcher.files():
            self._fs_watcher.addPath(str(DB_PATH))
        self._lbl_status.setText("✅ リフレッシュ完了")

<<<<<<< Updated upstream
    def _on_run_selected(self, run_id: str, meta: dict) -> None:
        """RunBrowserでRunが選択されたとき全タブに伝播する。"""
        self._tab_quick.set_run(run_id, meta)
=======
    # ── ファイル名自動解析 ─────────────────────────────────────────
    _FNAME_RE = re.compile(
        r"(\d{8})-"                          # date: YYYYMMDD
        r"(ROUND\d+|TEST\d+|UNK)-"           # round
        r"([A-Z0-9]+)-?"                     # session (QP/RACE1/FP/WUP...)
        r"(?:RUN(\d+)-)?"                    # optional run number
        r"(DA77|JA52|DA\d+|JA\d+)",          # rider
        re.IGNORECASE,
    )

    def _parse_filename(self, stem: str) -> dict:
        """ファイル名から date/round/session/run_no/rider を抽出。"""
        m = self._FNAME_RE.search(stem)
        if not m:
            return {}
        date_s, round_s, session_s, run_no_s, rider_s = m.groups()
        result = {
            "date":    date_s,
            "round":   round_s.upper(),
            "session": session_s.upper(),
            "rider":   rider_s.upper(),
        }
        if run_no_s:
            result["run_no"] = int(run_no_s)
        return result

    # DB未登録ラウンドのフォールバック (ROUND → CIRCUIT)
    _ROUND_FALLBACK: dict[str, str] = {
        "ROUND1":  "PHILLIP ISLAND",
        "ROUND4":  "BALATON",
        "ROUND5":  "",
        "ROUND10": "",
    }

    def _detect_circuit(self, round_s: str) -> str:
        """ROUND → CIRCUIT をDBから引く。DBにない場合はフォールバックマップを使用。"""
        try:
            conn = __import__("sqlite3").connect(str(DB_PATH))
            cur = conn.cursor()
            cur.execute(
                "SELECT DISTINCT circuit FROM runs WHERE round = ? AND circuit NOT LIKE 'TEST%'",
                (round_s,),
            )
            rows = [r[0] for r in cur.fetchall() if r[0]]
            conn.close()
            # ROUND2は PORTIMAO + CREMONA など混在 → レース系のみ優先
            race_circuits = [c for c in rows if c not in ("CREMONA", "WORKSHOP", "AUSTRALIA")]
            if len(race_circuits) == 1:
                return race_circuits[0]
            if len(rows) == 1:
                return rows[0]
        except Exception:
            pass
        # DB未登録の場合はフォールバック
        return self._ROUND_FALLBACK.get(round_s.upper(), "")

    def _open_csv(self):
        """ファイルダイアログでCSVを選択し、2D CSVタブで読み込んで波形に送る。"""
        default = str(Path.home() / "Desktop" / "Data TS24 Claude" / "06_CSV")
        path, _ = QFileDialog.getOpenFileName(
            self, "CSVファイルを選択", default,
            "CSV files (*.csv);;All files (*)"
        )
        if not path:
            return
        stem = Path(path).stem
        self._lbl_status.setText(f"読込中: {Path(path).name}")

        # ファイル名自動解析
        parsed = self._parse_filename(stem)
        self._tab_wave.set_label_a(parsed.get("rider", ""))

        try:
            self._tab_csv.load_file(path)
            self._tabs.setCurrentWidget(self._tab_wave)

            # サーキット自動検出・ComboBox 更新
            circuit_detected = ""
            if parsed.get("round"):
                circuit_detected = self._detect_circuit(parsed["round"])
                if circuit_detected:
                    idx = self._combo_circuit.findText(
                        circuit_detected, Qt.MatchFlag.MatchFixedString
                    )
                    if idx < 0:
                        # DB未登録サーキット（例: BALATON）を一時的にComboBoxへ追加
                        self._combo_circuit.addItem(circuit_detected)
                        idx = self._combo_circuit.findText(
                            circuit_detected, Qt.MatchFlag.MatchFixedString
                        )
                    if idx >= 0:
                        self._combo_circuit.setCurrentIndex(idx)
                        self._on_circuit_changed(circuit_detected)

            # ステータスバー表示
            parts = []
            if parsed.get("rider"):   parts.append(f"Rider: {parsed['rider']}")
            if parsed.get("session"): parts.append(f"Session: {parsed['session']}")
            if parsed.get("round"):   parts.append(f"Round: {parsed['round']}")
            if circuit_detected:      parts.append(f"Circuit: {circuit_detected} ✅")
            else:                     parts.append("Circuit: 手動選択 ▲")
            self._lbl_status.setText("  |  ".join(parts))
        except Exception as e:
            self._lbl_status.setText(f"エラー: {e}")

    def _open_csv_compare(self) -> None:
        """比較ライダーのCSVを読み込み、WaveformView に渡す。"""
        import pandas as pd
        import numpy as np

        default = str(Path.home() / "Desktop" / "Data TS24 Claude" / "06_CSV")
        path, _ = QFileDialog.getOpenFileName(
            self, "比較CSVファイルを選択", default,
            "CSV files (*.csv);;All files (*)"
        )
        if not path:
            return

        stem = Path(path).stem
        parsed = self._parse_filename(stem)
        label_b = parsed.get("rider", stem[:8])

        try:
            # ── CSV 読み込み（CsvImportTab と同じ文字コード試行）──────────
            df = None
            for enc in ("utf-8-sig", "shift-jis", "utf-8"):
                for sep in (";", ","):
                    try:
                        df = pd.read_csv(
                            path, sep=sep, engine="python",
                            encoding=enc, header=0, skiprows=[1], dtype=str,
                        )
                        if len(df.columns) > 1:
                            break
                    except Exception:
                        continue
                if df is not None and len(df.columns) > 1:
                    break
            if df is None or df.empty:
                QMessageBox.critical(self, "読込失敗", "CSV を読み込めませんでした。")
                return

            # カンマ小数点 → ドット変換 + 数値化
            df = df.apply(
                lambda col: pd.to_numeric(
                    col.str.replace(",", ".", regex=False), errors="coerce"
                )
            )

            # ── チャンネル自動検出（CsvImportTab.CHANNEL_MAP と同一）────
            CHANNEL_MAP = {
                "time":       ["time", "time2d"],
                "distance":   ["dist", "distance"],
                "speed":      ["speed_front", "speed"],
                "brake":      ["brake_front"],
                "gas":        ["gas", "gas_smooth", "tps"],
                "susp_front": ["susp_front"],
                "susp_rear":  ["susp_rear"],
                "lean_angle": ["bike_angle", "lean_angle"],
            }
            cols_lower = {c.lower(): c for c in df.columns}
            col_map: dict = {}
            for ch, aliases in CHANNEL_MAP.items():
                for alias in aliases:
                    if alias in cols_lower:
                        col_map[ch] = cols_lower[alias]
                        break

            if "time" not in col_map:
                QMessageBox.warning(self, "チャンネル不足", "Time 列が検出できませんでした。")
                return

            t_raw = df[col_map["time"]].values

            # ── x_mode 判定 ────────────────────────────────────────────
            d_raw = None
            x_mode = "time"
            if "distance" in col_map:
                d_raw = df[col_map["distance"]].values
                if float(np.nanmax(d_raw)) > 10.0:
                    x_mode = "distance"

            # ── 時間ギャップでセグメント分割 ──────────────────────────
            segments: list = []
            seg_start = 0
            for i in range(1, len(t_raw)):
                if (t_raw[i] - t_raw[i - 1]) > 5.0:
                    segments.append((seg_start, i - 1))
                    seg_start = i
            segments.append((seg_start, len(t_raw) - 1))

            # ── ラップデータ生成（1セグメント = 1ラップとして扱う）──────
            laps_b: list = []
            for lap_no, (s_start, s_end) in enumerate(segments, 1):
                if s_end - s_start < 10:
                    continue
                # 常に lap_time_s を time チャンネルから計算
                lt = round(float(t_raw[s_end]) - float(t_raw[s_start]), 3)

                if x_mode == "distance" and d_raw is not None:
                    x_arr = d_raw[s_start:s_end + 1]
                    x_vals = (x_arr - float(x_arr[0])).tolist()   # Lap内0始まりに正規化
                    dist_span_m = round(float(x_arr[-1]) - float(x_arr[0]), 1)
                else:
                    x_arr = t_raw[s_start:s_end + 1]
                    x_vals = (x_arr - float(x_arr[0])).tolist()   # Lap内0始まりに正規化
                    dist_span_m = 0.0

                ch_data: dict = {"x": x_vals}
                for ch in ["speed", "brake", "gas", "susp_front", "susp_rear", "lean_angle"]:
                    if ch in col_map:
                        ch_data[ch] = df[col_map[ch]].iloc[s_start:s_end + 1].tolist()

                laps_b.append({
                    "lap_no":      lap_no,
                    "lap_time_s":  lt,
                    "dist_span_m": dist_span_m,
                    "x_mode":      x_mode,
                    "channels":    ch_data,
                    "source_file": path,
                })

            if not laps_b:
                QMessageBox.warning(self, "読込失敗", "ラップデータが検出できませんでした。")
                return

            self._tab_wave.set_compare_laps(laps_b, label_b)
            self._lbl_status.setText(
                f"比較CSV: {Path(path).name}  |  Rider B: {label_b}  |  {len(laps_b)} セグメント"
            )

        except Exception as e:
            QMessageBox.critical(self, "CSV読込エラー", str(e))

    def _on_csv_loaded(self, meta: dict) -> None:
        """CSV読み込み完了後に run_meta を全タブに伝播させる。"""
        run_id = meta.get("run_id", "")
>>>>>>> Stashed changes
        self._tab_problem.set_run(run_id, meta)
        self._tab_setup.set_run(run_id, meta)
        self._lbl_status.setText(
            f"Run: {run_id}  |  {meta.get('circuit','')}  {meta.get('rider','')}  "
            f"Run#{meta.get('run_no','')}"
        )
        self._tabs.setCurrentWidget(self._tab_problem)




# ════════════════════════════════════════════════════════════════════
# エントリーポイント
# ════════════════════════════════════════════════════════════════════

def main():
    if not DB_PATH.exists():
        app = QApplication(sys.argv)
        QMessageBox.critical(
            None, "DB not found",
            f"ts24_unified.db が見つかりません:\n{DB_PATH}\n\n"
            "02_DATABASE フォルダを確認してください。",
        )
        sys.exit(1)

    db = WorkbenchDB(DB_PATH, XL_PATH)

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setFont(QFont("Arial", 10))

    window = MainWindow(db)
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
