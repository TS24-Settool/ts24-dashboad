"""
performance_correlation.py — TS24 Puccetti
LAP_TIMES × DYNAMICS_ANALYSIS パフォーマンス相関分析
==================================================
PUCCETTI_DB_MASTER.xlsx の LAP_TIMES シートと DYNAMICS_ANALYSIS シートを
ライダー×サーキット×日付×Runで結合し、
「ラップが速いときの車体姿勢（SusF/R at APEX・Brake）」を算出して
PERFORMANCE_CORRELATION シートに書き出す。

実行方法:
  python3 performance_correlation.py
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import math
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = SCRIPT_DIR.parent / "02_DATABASE" / "TS24 DB Master.xlsx"
MIN_LAP_S  = 80.0   # 80秒未満 = セクタータイム/部分計測 → 除外

def norm_circuit(c):
    c = str(c or '').upper().strip()
    if c in ('PHILLIPISLAND','PHILLIP ISLAND','PHI','AUSTRALIA','WORKSHOP','PHILLIP_ISLAND'):
        return 'PHILLIP ISLAND'
    return c

def norm_session(s):
    s = str(s or '').upper().strip()
    m = {'WUP':'WUP','WUP1':'WUP','WUP2':'WUP',
         'FP':'FP','FP1':'FP','FP2':'FP','F1':'FP','F2':'FP','L1':'FP','L2':'FP',
         'QP':'QP','QP1':'QP','QP2':'QP','Q1':'QP','Q2':'QP',
         'SP':'SP',
         'RACE1':'RACE1','R1':'RACE1','RACE2':'RACE2','R2':'RACE2',
         'TEST_D1':'TEST_D1','DAY1':'TEST_D1','D1':'TEST_D1',
         'TEST_D2':'TEST_D2','DAY2':'TEST_D2','D2':'TEST_D2'}
    return m.get(s,s)

def fmt_lap(s):
    if s is None: return ''
    m=int(s)//60; sec=s-m*60; return f'{m}:{sec:06.3f}'

def sf(v, d=2):
    if v is None or (isinstance(v,float) and math.isnan(v)): return None
    try: return round(float(v),d)
    except: return None

wb = openpyxl.load_workbook(EXCEL_PATH)

# ── LAP_TIMES 読み込み ─────────────────────────────────────────────────
ws_lt = wb['LAP_TIMES']
lt_map = defaultdict(list)  # key=(rider,circ_n,date,run) → [lap_s,...]
for r in range(3, ws_lt.max_row+1):
    v = [ws_lt.cell(r,c).value for c in range(1,16)]
    rnd,circ,date,sess,rider,run,lap,lt,lt_s,outlap = v[:10]
    if not rider or outlap=='YES': continue
    if not isinstance(lt_s,(int,float)) or lt_s < MIN_LAP_S or lt_s > 400: continue
    key = (str(rider), norm_circuit(circ), str(date or ''), int(run or 0))
    lt_map[key].append({'lap_s':float(lt_s),'sess':norm_session(sess),'round':rnd})

lt_sessions = {}
for key, laps in lt_map.items():
    best = min(l['lap_s'] for l in laps)
    avg  = sum(l['lap_s'] for l in laps)/len(laps)
    sess_set = list(set(l['sess'] for l in laps))
    lt_sessions[key] = {'best_s':best,'avg_s':avg,'n_laps':len(laps),'sessions':sess_set}

# ── DYNAMICS 読み込み ─────────────────────────────────────────────────
ws_dy = wb['DYNAMICS_ANALYSIS']
dy_h  = [ws_dy.cell(2,c).value for c in range(1,ws_dy.max_column+1)]
dc    = {h:i+1 for i,h in enumerate(dy_h) if h}
dy_sessions = {}
for r in range(3, ws_dy.max_row+1):
    g = lambda h: ws_dy.cell(r,dc[h]).value if h in dc else None
    rider=g('Rider'); circ=g('Circuit'); run=g('Run'); date=g('Date'); sess=g('Session')
    if not rider: continue
    key = (str(rider), norm_circuit(circ), str(date or ''), int(run or 0))
    dy_sessions[key] = {
        'round':g('Round'),'session_raw':sess,
        'apex_susF':sf(g('APEX SusF (mm)')), 'apex_susR':sf(g('APEX SusR (mm)')),
        'apex_whlF':sf(g('APEX WhlF (N)'),1), 'apex_whlR':sf(g('APEX WhlR (N)'),1),
        'apex_spd':sf(g('APEX Spd (km/h)'),1), 'apex_ax':sf(g('APEX ax (m/s²)')),
        'brk_susF':sf(g('Brk SusF (mm)')), 'brk_susR':sf(g('Brk SusR (mm)')),
        'brk_spd':sf(g('Brk Spd (km/h)'),1), 'apex_count':g('APEX Count'),
    }

# ── 結合 ─────────────────────────────────────────────────────────────
matched = []
for key, lt in lt_sessions.items():
    if key not in dy_sessions: continue
    dy = dy_sessions[key]
    rider,circ,date,run = key
    matched.append({'rider':rider,'circuit':circ,'date':date,'run':run,
                    'round':dy['round'],'session_raw':dy['session_raw'],
                    'lt_sessions':'/'.join(lt['sessions']),
                    **lt, **{k:v for k,v in dy.items() if k not in ('round','session_raw')}})

print(f'結合成功: {len(matched)} セッション')

# ── グループ内ランク ─────────────────────────────────────────────────
groups = defaultdict(list)
for m in matched:
    groups[(m['rider'],m['circuit'])].append(m)

for (r,c), rows in groups.items():
    rows.sort(key=lambda x: x['best_s'])
    n = len(rows)
    for i,row in enumerate(rows):
        row['rank']=i+1; row['rank_total']=n
        row['gap_to_best']=round(row['best_s']-rows[0]['best_s'],3)
        pct = i/max(n-1,1)
        if n<3: row['tier']='FAST' if i==0 else 'SLOW'
        elif pct<=0.33: row['tier']='FAST'
        elif pct>=0.67: row['tier']='SLOW'
        else: row['tier']='MED'

# ── サマリー ─────────────────────────────────────────────────────────
def wmean(rows,key):  # 値ありのもの平均
    vals=[r[key] for r in rows if r.get(key) is not None]
    return round(sum(vals)/len(vals),2) if vals else None

summary=[]
for (rider,circ),rows in sorted(groups.items()):
    fast=[r for r in rows if r['tier']=='FAST']
    slow=[r for r in rows if r['tier']=='SLOW']
    s={'rider':rider,'circuit':circ,'n_sess':len(rows),
       'best_lap':rows[0]['best_s'],'best_lap_f':fmt_lap(rows[0]['best_s'])}
    for m in('apex_susF','apex_susR','apex_whlF','apex_whlR','apex_spd','brk_susF','brk_susR'):
        s[f'fast_{m}']=wmean(fast,m); s[f'slow_{m}']=wmean(slow,m); s[f'all_{m}']=wmean(rows,m)
        fa=s[f'fast_{m}']; sl=s[f'slow_{m}']
        s[f'd_{m}']=round(fa-sl,2) if (fa is not None and sl is not None) else None
    summary.append(s)

# ── Excel 書き込み ────────────────────────────────────────────────────
SHEET='PERFORMANCE_CORRELATION'
if SHEET in wb.sheetnames: del wb[SHEET]
ws=wb.create_sheet(SHEET)

def fill(h): return PatternFill('solid',fgColor=h)
def fnt(bold=False,sz=9,color='000000'): return Font(bold=bold,size=sz,color=color)
def brd(): return Border(left=Side(style='thin'),right=Side(style='thin'),
                         top=Side(style='thin'),bottom=Side(style='thin'))
CTR=Alignment(horizontal='center',vertical='center',wrap_text=True)
LFT=Alignment(horizontal='left',vertical='center',wrap_text=True)

F={'title':fill('1F3864'),'hdrA':fill('2E75B6'),'hdrB':fill('375623'),
   'FAST':fill('C6EFCE'),'MED':fill('FFEB9C'),'SLOW':fill('FFC7CE'),
   'zA':fill('EBF3FB'),'zB':fill('F2F9EE'),'zS':fill('FFF5F5'),
   'gA':fill('D6E4F0'),'gB':fill('E2EFDA'),'gC':fill('FCE4D6'),'gD':fill('FFF2CC'),
   'sum_F':fill('A9D18E'),'sum_S':fill('FF9999'),'sum_D':fill('FFD966'),
   'sum_hdr':fill('1E4620'),'white':fill('FFFFFF')}

# Row1: タイトル
ws.merge_cells('A1:T1')
c=ws.cell(1,1,'TS24 Puccetti — Performance Correlation  |  Best Lap Time × Suspension Posture at APEX & Braking Entry')
c.fill=F['title']; c.font=fnt(True,12,'FFFFFF'); c.alignment=CTR
ws.row_dimensions[1].height=26

# Row2: グループヘッダー
ws.row_dimensions[2].height=18
for cols, label, fk in [('A2:E2','Session Info','gA'),('F2:H2','Lap Time','gB'),
    ('I2:N2','APEX Posture','gB'),('O2:Q2','Braking Posture','gC'),('R2:T2','Performance','gD')]:
    ws.merge_cells(cols); c=ws.cell(2,ord(cols[0])-64,label)
    c.fill=F[fk]; c.font=fnt(True,9); c.alignment=CTR

# Row3: 列ヘッダー
COLS=[
    (1,'Rider',7,'gA'),(2,'Circuit',15,'gA'),(3,'Date',11,'gA'),
    (4,'Session',9,'gA'),(5,'Run',5,'gA'),
    (6,'Best Lap',10,'gB'),(7,'Avg Lap',10,'gB'),(8,'N Laps',7,'gB'),
    (9,'APEX\nSusF(mm)',10,'gB'),(10,'APEX\nSusR(mm)',10,'gB'),
    (11,'APEX\nWhlF(N)',10,'gB'),(12,'APEX\nWhlR(N)',10,'gB'),
    (13,'APEX\nSpd(km/h)',11,'gB'),(14,'APEX\nax(m/s²)',10,'gB'),
    (15,'Brk\nSusF(mm)',10,'gC'),(16,'Brk\nSusR(mm)',10,'gC'),(17,'Brk\nSpd(km/h)',11,'gC'),
    (18,'Rank',8,'gD'),(19,'Gap(s)',8,'gD'),(20,'Tier',8,'gD'),
]
ws.row_dimensions[3].height=34
for col,label,width,_ in COLS:
    c=ws.cell(3,col,label); c.fill=F['hdrA']; c.font=fnt(True,9,'FFFFFF')
    c.alignment=CTR; c.border=brd()
    ws.column_dimensions[get_column_letter(col)].width=width

# データ行
all_sorted=sorted(matched,key=lambda x:(x['rider'],x['circuit'],x['best_s']))
for ri,m in enumerate(all_sorted,start=4):
    ws.row_dimensions[ri].height=16
    tier=m['tier']
    base=F['zA'] if ri%2==0 else F['white']
    vals={1:m['rider'],2:m['circuit'],3:m['date'],4:m['lt_sessions'],5:m['run'],
          6:fmt_lap(m['best_s']),7:fmt_lap(m['avg_s']),8:m['n_laps'],
          9:m['apex_susF'],10:m['apex_susR'],11:m['apex_whlF'],12:m['apex_whlR'],
          13:m['apex_spd'],14:m['apex_ax'],
          15:m['brk_susF'],16:m['brk_susR'],17:m['brk_spd'],
          18:f"{m['rank']}/{m['rank_total']}",19:m['gap_to_best'],20:tier}
    for col,val in vals.items():
        c=ws.cell(ri,col,val); c.alignment=CTR; c.border=brd()
        c.font=fnt(sz=9)
        c.fill=F[tier] if col==20 else base
    # Tierセルは色付き
    ws.cell(ri,20).font=fnt(True,9); ws.cell(ri,20).fill=F[tier]

detail_end=3+len(all_sorted)

# ── サマリーセクション ──────────────────────────────────────────────
SR=detail_end+3
ws.row_dimensions[SR].height=26
ws.merge_cells(f'A{SR}:T{SR}')
c=ws.cell(SR,1,'TARGET SETUP REFERENCE  ——  FAST vs SLOW セッション平均比較  |  Δ = FAST − SLOW')
c.fill=F['sum_hdr']; c.font=fnt(True,11,'FFFFFF'); c.alignment=LFT

SH=SR+1; ws.row_dimensions[SH].height=38
SUM_COLS=[
    (1,'Rider',7),(2,'Circuit',15),(3,'N\nSess',6),(4,'Best\nLap',10),
    (5,'★FAST\nSusF(mm)',11),(6,'★FAST\nSusR(mm)',11),(7,'★FAST\nBrkF(mm)',11),(8,'★FAST\nBrkR(mm)',11),
    (9,'★FAST\nWhlF(N)',11),(10,'★FAST\nWhlR(N)',11),(11,'★FAST\nSpd(km/h)',11),
    (12,'SLOW\nSusF(mm)',11),(13,'SLOW\nSusR(mm)',11),(14,'SLOW\nBrkF(mm)',11),(15,'SLOW\nBrkR(mm)',11),
    (16,'Δ SusF\n(mm)',10),(17,'Δ SusR\n(mm)',10),(18,'Δ BrkF\n(mm)',10),(19,'Δ BrkR\n(mm)',10),
    (20,'Setup Direction',16),
]
for col,label,w in SUM_COLS:
    c=ws.cell(SH,col,label); c.fill=F['sum_hdr']; c.font=fnt(True,9,'FFFFFF')
    c.alignment=CTR; c.border=brd()
    ws.column_dimensions[get_column_letter(col)].width=max(ws.column_dimensions[get_column_letter(col)].width,w)

def delta_cell(ws,r,col,val):
    c=ws.cell(r,col,val); c.font=fnt(True,9); c.alignment=CTR; c.border=brd()
    if val is None: c.fill=F['white']
    elif val>0.5:   c.fill=fill('C6EFCE')   # 緑:Fast時により圧縮
    elif val<-0.5:  c.fill=fill('FFC7CE')   # 赤:Fast時により伸長
    else:           c.fill=F['sum_D']

def dir_text(df,dr,dbf,dbr):
    parts=[]
    for label,d in [('Apex-SusF',df),('Apex-SusR',dr),('BrkF',dbf),('BrkR',dbr)]:
        if d is None: continue
        if abs(d)<0.3: parts.append(f'{label}:≈same')
        elif d>0: parts.append(f'{label}:↑{d:+.1f}mm when fast')
        else:     parts.append(f'{label}:{d:+.1f}mm when fast')
    return '  |  '.join(parts) if parts else '—'

for si,s in enumerate(summary,start=SH+1):
    ws.row_dimensions[si].height=24
    bg=F['zA'] if si%2==0 else F['white']
    def sc(col,val,bf=None):
        c=ws.cell(si,col,val); c.font=fnt(sz=9); c.alignment=CTR; c.border=brd()
        c.fill=bf if bf else bg
    sc(1,s['rider']); sc(2,s['circuit']); sc(3,s['n_sess']); sc(4,s['best_lap_f'])
    sc(5,s['fast_apex_susF'],F['sum_F']); sc(6,s['fast_apex_susR'],F['sum_F'])
    sc(7,s['fast_brk_susF'],F['sum_F']);  sc(8,s['fast_brk_susR'],F['sum_F'])
    sc(9,s['fast_apex_whlF'],F['sum_F']); sc(10,s['fast_apex_whlR'],F['sum_F'])
    sc(11,s['fast_apex_spd'],F['sum_F'])
    sc(12,s['slow_apex_susF'],F['sum_S']); sc(13,s['slow_apex_susR'],F['sum_S'])
    sc(14,s['slow_brk_susF'],F['sum_S']);  sc(15,s['slow_brk_susR'],F['sum_S'])
    df=s['d_apex_susF']; dr=s['d_apex_susR']
    dbf=s['d_brk_susF']; dbr=s['d_brk_susR']
    delta_cell(ws,si,16,df); delta_cell(ws,si,17,dr)
    delta_cell(ws,si,18,dbf); delta_cell(ws,si,19,dbr)
    c=ws.cell(si,20,dir_text(df,dr,dbf,dbr))
    c.font=fnt(sz=8,bold=(df is not None and abs(df)>0.5))
    c.alignment=LFT; c.border=brd(); c.fill=F['sum_D']

ws.freeze_panes='A4'
wb.save(EXCEL_PATH)

# ── コンソール出力 ─────────────────────────────────────────────────
print(f'\n{"="*70}')
print('  PERFORMANCE CORRELATION — Summary')
print(f'{"="*70}')
for s in summary:
    print(f"\n  [{s['rider']} @ {s['circuit']}]  n={s['n_sess']}  Best: {s['best_lap_f']}")
    print(f"    ★ FAST avg:  Apex SusF={s['fast_apex_susF']}mm  SusR={s['fast_apex_susR']}mm  "
          f"BrkF={s['fast_brk_susF']}mm  BrkR={s['fast_brk_susR']}mm  "
          f"WhlF={s['fast_apex_whlF']}N")
    print(f"    ✗ SLOW avg:  Apex SusF={s['slow_apex_susF']}mm  SusR={s['slow_apex_susR']}mm  "
          f"BrkF={s['slow_brk_susF']}mm  BrkR={s['slow_brk_susR']}mm  "
          f"WhlF={s['slow_apex_whlF']}N")
    print(f"    Δ (F-S):     SusF={s['d_apex_susF']}mm  SusR={s['d_apex_susR']}mm  "
          f"BrkF={s['d_brk_susF']}mm  BrkR={s['d_brk_susR']}mm")
print(f'\n✅ PERFORMANCE_CORRELATION シート書き込み完了')
