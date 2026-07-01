#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_master_db.py — TS24 DB 再構築 (DB_REBUILD_SPEC_v1.0)
権威: Run構造/日付=Report, セットアップ=Original(絶対値), ラップタイム+3エリア=2D
ID: run_id={date}_{round}_{circuit}_{session}_{rider}_R{run}, lap_id={run_id}_L{n}

使い方:
  python3 build_master_db.py --event 20260501-ROUND4-JA52   # 1イベント検証(DB書込なし, 表示のみ)
  python3 build_master_db.py --all                          # 全イベント -> ts24_master.db
  python3 build_master_db.py --all --out /path/scratch.db   # スクラッチDBへ出力(本番DB非変更)

2D outing 探索は3レイアウトを横断 (discover_outings):
  (A) 通常ネスト <base>.MES/  (B) リネームコピー <...>.MES - Copia/  (C) ラッパー無し loose
内部 .DDD/.LAP の stem を base に採用するためフォルダ命名に非依存。新レイアウト(B/C)は
HED 自己申告サーキットがイベントと矛盾する場合のみ除外 (gated_outings, 誤ファイル対策)。
"""
import sys, re, importlib.util
from pathlib import Path
from collections import defaultdict
import numpy as np

SCRIPT_DIR = Path(__file__).parent
ROOT       = SCRIPT_DIR.parent
DATA_2D    = ROOT / "DATA 2D"
ORIGINAL   = ROOT / "04_REFERENCE" / "Data_Base_TS24_ORIGINAL.xlsx"
REPORTS    = ROOT / "01_REPORTS"
OUT_DB     = ROOT / "02_DATABASE" / "ts24_master.db"

# ── parse_2d_channels 低レベル関数 ──
spec = importlib.util.spec_from_file_location("p2d", SCRIPT_DIR / "parse_2d_channels.py")
p2d  = importlib.util.module_from_spec(spec); spec.loader.exec_module(p2d)
parse_hed, parse_ddd, parse_lap, read_channel = p2d.parse_hed, p2d.parse_ddd, p2d.parse_lap, p2d.read_channel

# ── 3エリア定義 (DB_REBUILD_SPEC §5, ロガー校正単位) ──
# DELTA_GAS: Tatsuki確定の本質=「アクセルを開ける方向(正)の入力のみ抽出」=dTPS_A>0。
#   進入のアクセルオフ(負)を除外するのが目的。元指定の 20-60 は2Dツール側の単位での値で、
#   .MES生チャンネル dTPS_A(%/s, -634..547) とは単位が異なるため、正方向(>0)で実装。
#   2Dが手元に戻ったら DELTA_GAS_MIN を再調整可能。
DELTA_GAS_MIN = 1e-6   # >0 = 開け方向
AREAS = {
    "MID_CORNER":   {"BRAKE_FRONT":(-0.3,3.0), "THROTTLE":(-0.5,5.0), "DELTA_GAS":(DELTA_GAS_MIN,1e12),
                     "SUSP_FRONT":(50.0,100.0), "SUSP_REAR":(8.0,40.0)},
    "FULL_BRAKING": {"BRAKE_FRONT":(9.0,20.0), "SUSP_FRONT":(90.0,130.0), "SUSP_REAR":(-0.5,2.0)},
    "CORNER_EXIT":  {"BRAKE_FRONT":(-0.5,0.0), "THROTTLE":(50.0,100.0),
                     "SUSP_FRONT":(0.0,70.0), "SUSP_REAR":(2.0,30.0)},
}
MIN_LAP_S = 30.0

# ── 3フェーズ×F/R×方向 サス速度マトリクス (§43/§44 設計・Tatsuki GO 2026-07-01) ──
# 既存 brk_f_dive_spd_*(凍結・peak=max) と ce_r_spd_*(abs) は不変。以下22列を「追加のみ」。
#   フェーズ: brk=FULL_BRAKING / apex=MID_CORNER / ce=CORNER_EXIT
#   方向: dive=圧縮(v>0) / reb=伸び(-v,v<0)。avg=mean(n>=5) / peak=p95(n>=10)。n未満はNULL(0で代用しない)。
#   相対ダンピング速度指数(グリッドM微分)。校正済み絶対mm/sではない・車速km/hと混同禁止。
PEAK_NMIN = 10   # peak(p95) は方向サンプル n>=10（avg の NMIN_Z=5 より厳格：小nでp95がmaxへ退化するため）
PHASE_SPD_NEW_COLS = [
    "brk_f_reb_spd_avg",  "brk_f_reb_spd_peak",
    "brk_r_dive_spd_avg", "brk_r_dive_spd_peak",
    "brk_r_reb_spd_avg",  "brk_r_reb_spd_peak",
    "apex_f_dive_spd_avg","apex_f_dive_spd_peak",
    "apex_f_reb_spd_avg", "apex_f_reb_spd_peak",
    "apex_r_dive_spd_avg","apex_r_dive_spd_peak",
    "apex_r_reb_spd_avg", "apex_r_reb_spd_peak",
    "ce_f_dive_spd_avg",  "ce_f_dive_spd_peak",
    "ce_f_reb_spd_avg",   "ce_f_reb_spd_peak",
    "ce_r_dive_spd_avg",  "ce_r_dive_spd_peak",
    "ce_r_reb_spd_avg",   "ce_r_reb_spd_peak",
]

# ── 正規化 ──
def circuit_canon(c):
    u = re.sub(r"[^A-Z0-9]", "", str(c or "").upper())
    t = {"PHILIPISLAND":"PHILLIPISLAND","PHILLIPISLAND":"PHILLIPISLAND","PHILLIPISISLAND":"PHILLIPISLAND",
         "BALATON":"BALATON","BALATONPARK":"BALATON","MOTORLANDARAGON":"ARAGON","ARAGON":"ARAGON",
         "WORKSHOP":"PHILLIPISLAND","AUSTRALIA":"PHILLIPISLAND","MAGNYCOURS":"MAGNYCOURS"}
    return t.get(u, u)

def session_canon_2d(base, rnd):
    pre = re.match(r"^([A-Za-z]+\d*)", base); pre = pre.group(1).upper() if pre else ""
    md = re.match(r"^(D|L)(\d)", pre); day = f"DAY{md.group(2)}" if md else None
    if re.match(r"^F", pre):  return "FP"
    if re.match(r"^Q", pre):  return "QP"
    if re.match(r"^WUP?2", pre) or pre=="WU2": return "WUP2"
    if re.match(r"^WUP?1", pre) or pre=="WU1": return "WUP1"
    if re.match(r"^WUP?$", pre) or pre=="WU":  return "WUP1"
    if re.match(r"^R1", pre) or pre=="RACE1":  return "RACE1"
    if re.match(r"^R2", pre) or pre=="RACE2":  return "RACE2"
    if re.match(r"^SP", pre): return "SP"
    if rnd.startswith("TEST") and day: return f"{rnd}_{day}"
    if rnd.startswith("TEST"): return f"{rnd}_DAY1"
    return pre

def session_canon_orig(s):
    u = re.sub(r"\s+"," ", str(s or "").strip().upper())
    m = re.match(r"^TEST\s*(\d)\s*DAY\s*(\d)$", u)
    return f"TEST{m.group(1)}_DAY{m.group(2)}" if m else re.sub(r"\s+","",u)

# ── Original (絶対セットアップ) ──
ORIG_FIELDS = [  # (col_idx, field)
 (4,"weather"),(5,"track_temp"),(6,"air_temp"),(7,"fork_type"),(8,"f_set_c"),(9,"f_set_r"),
 (10,"f_tos_spring"),(11,"f_tos_length"),(12,"f_spr_l"),(13,"f_spr_r"),(14,"f_preload"),
 (15,"f_oil_level"),(16,"f_comp"),(17,"f_reb"),(18,"f_offset"),(19,"f_offset2"),
 (20,"f_hgt_top"),(21,"f_hgt_bot"),(22,"shock_type"),(23,"r_set_c"),(24,"r_set_r"),(25,"r_spr"),
 (26,"r_preload"),(27,"r_comp"),(28,"r_reb"),(29,"r_tos_spring"),(30,"r_tos_length"),
 (31,"shock_len"),(32,"link"),(33,"ride_hgt"),(34,"swing_arm"),(35,"tyre_front"),(36,"tyre_rear")]

def load_original():
    import openpyxl
    wb = openpyxl.load_workbook(ORIGINAL, read_only=True, data_only=True); ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    hdr = [str(h).strip().upper() if h else "" for h in rows[0]]
    iR,iC,iS,iN = hdr.index("RIDER"),hdr.index("CIRCUIT"),hdr.index("SESSION"),hdr.index("RUN")
    runs = defaultdict(list)   # (rider,circ,sess) -> [ {run, setup..} ] (重複=複数シーズン)
    for r in rows[1:]:
        if not r or r[iR] is None: continue
        rider=str(r[iR]).strip(); circ=circuit_canon(r[iC]); sess=session_canon_orig(r[iS])
        try: run=int(float(r[iN]))
        except Exception: run=1
        setup={f:r[i] for i,f in ORIG_FIELDS}
        runs[(rider,circ,sess)].append({"run":run, **setup})
    return runs

# ── イベント探索 ──
EVENT_RE = re.compile(r"^(\d{8})-(ROUND\d+|TEST\d+)-(DA77|JA52|JA25)$", re.I)
def discover_events():
    evs={}
    for d in sorted(DATA_2D.iterdir()):
        if not d.is_dir(): continue
        m = EVENT_RE.match(d.name)
        if not m: continue
        date,rnd,rider = m.group(1), m.group(2).upper(), m.group(3).upper().replace("JA25","JA52")
        evs[d.name]={"date":date,"round":rnd,"rider":rider,"dir":d,
                     "report":_find_report(rider,rnd,date)}
    return evs

def _find_report(rider,rnd,date):
    p = REPORTS / rider
    if not p.exists(): return None
    for f in p.glob(f"{date}-{rnd}-{rider}.xlsx"):
        return f
    for f in p.glob(f"*-{rnd}-{rider}.xlsx"):
        return f
    return None

def circuit_from_report(report):
    if not report or not report.exists(): return ""
    import openpyxl
    wb = openpyxl.load_workbook(report, read_only=True, data_only=True)
    if "DAY1" in wb.sheetnames:
        ws=wb["DAY1"]
        for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            for j,c in enumerate(row):
                if c and str(c).strip().upper()=="CIRCUIT" and j+2 < len(row):
                    return str(row[j+2] or "").strip()
    return ""

def circuit_from_2d(event_dir):
    """イベントフォルダ内の *.line ファイル名からサーキットを推定"""
    cand=list(event_dir.rglob("*.line"))
    for f in cand:
        nm=f.stem.strip()
        if nm and nm.lower() not in ("ring",):
            return nm
    return ""

def event_circuit(ev):
    c = circuit_from_report(ev["report"]) or circuit_from_2d(ev["dir"])
    return circuit_canon(c)

# ── 2D: 1 .MES のラップ抽出 + 3エリアメトリクス ──
KMH_TO_MS = 1/3.6
def _key(chs, *names):
    for n in names:
        for k in chs:
            if k.upper()==n and chs[k].get("ext"): return k
    return None

def _lap_timebase(mes_path, base):
    """LAP マーカーの時間基準 (= vals[1]) を返す。1秒あたりの marker 単位数。
    .LAP は little-endian uint32: vals[0]=n_laps, vals[1]=time-base(=1秒あたりの単位数),
    vals[2]=0, vals[3:]=累積ラップ終了マーカー。
      - 旧来の lean export (2026系・2025 ROUND1 等): vals[1]=1000 → marker は ms。
      - 238ch Dorna/WSBK export (2025 ROUND10/TEST1 等): vals[1]=400 → marker は 400Hz の
        サンプル数 (= 1/400 秒単位)。1000 で割ると 2.5倍 短いラップになる split-lap バグの原因。
    全 .MES を走査した結果 vals[1] は 400 か 1000 の二値のみ。異常値はデフォルト 1000 に倒す。
    Time2D/CourAccu/GPS は 238ch ファイルでは DDD スケールが信頼できず壊れているため、
    SPEED_FRONT 積分距離を独立検証に用い vals[1] 規約が full-lap (≈トラック長) を生むことを確認済み:
      Aragon ROUND10 → /400 で flying ≈114s・周回距離≈4990m (実5077m), 旧/1000 では誤って45s。
    """
    import struct as _st
    p = mes_path / f"{base}.LAP"
    if not p.exists(): return 1000.0
    raw = p.read_bytes(); ni = len(raw)//4
    if ni < 3: return 1000.0
    tb = _st.unpack(f"<{ni}I", raw[:ni*4])[1]
    # 妥当な時間基準のみ採用 (10〜2000 単位/秒)。範囲外は ms 想定でフォールバック。
    return float(tb) if 10 <= tb <= 2000 else 1000.0

def _parse_hed_time(v):
    """HED の "1:44.604" / "104.604" を秒(float)に。不正は None。"""
    if not v: return None
    m = re.match(r"(\d+)[:'](\d{1,2})[.,](\d+)", str(v).strip())
    if m: return int(m.group(1))*60 + int(m.group(2)) + float("0."+m.group(3))
    try: return float(str(v).replace(",", "."))
    except Exception: return None

def extract_outing(mes_path, base=None):
    """returns {laps:[{lap_no,lap_time_s,metrics:{area:{...}}}], nlaps}
    base = 内部 .DDD/.LAP/.HED ファイルの stem。省略時のみディレクトリ名から推定
    (通常ネスト <base>.MES だけで成立)。loose / renamed-copy レイアウトでは
    フォルダ名と内部 base が一致しないため、呼び出し側が base を明示する。"""
    if base is None:
        base = mes_path.name[:-4]
    chs = parse_ddd(mes_path, base)
    if "SPEED_FRONT" not in chs or "SUSP_FRONT" not in chs: return None
    sf  = read_channel(mes_path, base, chs["SPEED_FRONT"])
    suf = read_channel(mes_path, base, chs["SUSP_FRONT"])
    sur = read_channel(mes_path, base, chs.get("SUSP_REAR",{})) if "SUSP_REAR" in chs else np.array([])
    if len(sf)<10: return None
    bk_k = _key(chs,"BRAKE_FRONT"); gk = _key(chs,"GAS_SMOOTH","GAS","TPS","TPS_A")
    dk = _key(chs,"DTPS_A","DTPS","TPS_DELTA","DELTA_TPS")
    bk  = read_channel(mes_path,base,chs[bk_k]) if bk_k else np.array([])
    gas = read_channel(mes_path,base,chs[gk]) if gk else np.array([])
    dt  = read_channel(mes_path,base,chs[dk]) if dk else np.array([])
    n_laps, lap_ms = parse_lap(mes_path, base)
    if not (n_laps>0 and lap_ms): return None
    # LAP マーカーの時間基準 (1秒あたりの単位数) をヘッダ vals[1] から取得。
    # 旧実装は /1000.0 固定 (ms 前提) だったため、400Hz サンプル単位で記録された
    # 238ch export (2025 ROUND10/TEST1) で lap_time が 2.5倍 短くなる split-lap バグが発生。
    # フォーマット差をデータから自動検出 (vals[1]=400 vs 1000) して正しい秒数を得る。
    timebase = _lap_timebase(mes_path, base)   # 400 or 1000 (単位/秒)
    total_units = lap_ms[-1]                    # 末尾=録画全体の累積マーカー (単位)
    sr = max(10.0, min(len(sf)/(total_units/timebase) if total_units>0 else 100.0, 500.0))
    # per-channel global resampler: map a [0,1] lap window to each channel
    def lap_window(arr, t0, t1):
        if len(arr)==0: return np.array([])
        a=int(t0*len(arr)); b=int(t1*len(arr))
        b=min(max(b,a+1), len(arr))
        return arr[a:b].astype(np.float64)
    laps=[]; prev=0
    for i,t in enumerate(lap_ms):
        # lap_time[s] = (累積マーカー差) / 時間基準。t0/t1 は録画内の正規化比率なので単位不問。
        lap_t=(t-prev)/timebase; t0=prev/total_units; t1=t/total_units; prev=t
        if lap_t < MIN_LAP_S: continue
        # resample all channels of this lap to common length M
        segs={"SPEED_FRONT":lap_window(sf,t0,t1),"SUSP_FRONT":lap_window(suf,t0,t1),
              "SUSP_REAR":lap_window(sur,t0,t1),"BRAKE_FRONT":lap_window(bk,t0,t1),
              "THROTTLE":lap_window(gas,t0,t1),"DELTA_GAS":lap_window(dt,t0,t1)}
        M = max((len(v) for v in segs.values() if len(v)>0), default=0)
        if M < 20: continue
        grid=np.linspace(0,1,M)
        R={}
        for nm,v in segs.items():
            R[nm] = np.interp(grid, np.linspace(0,1,len(v)), v) if len(v)>0 else None
        metrics={}
        for area,conds in AREAS.items():
            mask=np.ones(M,dtype=bool); ok=True
            for ch,(lo,hi) in conds.items():
                if R.get(ch) is None: ok=False; break
                mask &= (R[ch]>=lo)&(R[ch]<=hi)
            cnt=int(mask.sum()) if ok else 0
            def avg(nm):
                return round(float(R[nm][mask].mean()),2) if ok and cnt>0 and R.get(nm) is not None else None
            metrics[area]={"n":cnt,"susf":avg("SUSP_FRONT"),"susr":avg("SUSP_REAR"),
                           "speed":avg("SPEED_FRONT"),"brake":avg("BRAKE_FRONT"),"thr":avg("THROTTLE")}
        # ── サス速度(ダンピング指標) + Rear-light(ブレーキバランス) (Tatsukiアイデア 2026-06-19) ──
        # 位置[mm]の微分→速度[mm/s]。圧縮(+)=Diving, 伸び(-)=Rebound のピーク(max)で評価し
        # 圧/伸ダンピング設定の判断材料にする(位置balance=バネ/ジオメトリ, 速度=ダンピング)。
        def _spd(arr):
            if arr is None or len(arr) < 3 or lap_t <= 0: return (None, None)
            v = np.gradient(arr) / (lap_t / len(arr))   # mm/s
            comp = v[v > 0]; reb = -v[v < 0]
            return (round(float(comp.max()), 1) if comp.size else None,
                    round(float(reb.max()), 1) if reb.size else None)
        f_dive, f_reb = _spd(R.get("SUSP_FRONT"))
        r_dive, r_reb = _spd(R.get("SUSP_REAR"))
        # Rear "0mm"(伸び切り=軽い)滞在率: ブレーキ区間(BRAKE_FRONT>=5bar)で SUSP_REAR<=1mm の割合[%]
        # 大=リアが浮きフロントタイヤのみで停止 → ブレーキバランス指標。
        rear_light_brk = None
        if R.get("SUSP_REAR") is not None and R.get("BRAKE_FRONT") is not None:
            bmask = R["BRAKE_FRONT"] >= 5.0
            if int(bmask.sum()) >= 5:
                rear_light_brk = round(float(np.mean(R["SUSP_REAR"][bmask] <= 1.0)) * 100, 1)
        # ── ゾーン限定サス速度 + PH1-2 リア0mm累積秒 (Tatsuki 2026-06-20, setup判断グレード) ──
        # 速度はグリッドR(M点)上の位置微分[mm/s]。既存 f_dive_spd と同一手法のためデータセット内で
        # 相互比較可(校正済み絶対mm/sではない=相対ダンピング速度指数)。n<5 は NULL(0は実測ゼロ速度と
        # 誤読されるため厳禁)。Front圧縮(+)=Diving のみ。CE Rearは絶対値(忙しさ)。
        NMIN_Z = 5
        brk_f_dive_avg = brk_f_dive_peak = None
        ce_r_spd_avg = ce_r_spd_peak = None
        ph12_rear0_s = None
        dtg = (lap_t / M) if (lap_t and M) else None   # グリッド1点あたり秒
        def _vel(arr):
            if arr is None or len(arr) < 3 or not dtg: return None
            return np.gradient(arr) / dtg              # mm/s, 長さM
        def _zone_mask(area):
            m = np.ones(M, dtype=bool)
            for ch,(lo,hi) in AREAS[area].items():
                if R.get(ch) is None: return None
                m &= (R[ch]>=lo)&(R[ch]<=hi)
            return m
        vf = _vel(R.get("SUSP_FRONT")); vr = _vel(R.get("SUSP_REAR"))
        # ① Hard Brake(FULL_BRAKING)内 フロント圧縮方向(diving,+)速度 avg/peak
        fb_mask = _zone_mask("FULL_BRAKING")
        if vf is not None and fb_mask is not None and int(fb_mask.sum()) >= NMIN_Z:
            comp = vf[fb_mask]; comp = comp[comp > 0]   # 圧縮方向のみ
            if comp.size >= NMIN_Z:
                brk_f_dive_avg = round(float(comp.mean()), 1)
                brk_f_dive_peak = round(float(comp.max()), 1)
        # ② Corner Exit(CORNER_EXIT)内 リア速度絶対値 avg/peak
        ce_mask = _zone_mask("CORNER_EXIT")
        if vr is not None and ce_mask is not None and int(ce_mask.sum()) >= NMIN_Z:
            av = np.abs(vr[ce_mask]); av = av[np.isfinite(av)]
            if av.size >= NMIN_Z:
                ce_r_spd_avg = round(float(av.mean()), 1)
                ce_r_spd_peak = round(float(av.max()), 1)
        # ③ PH1-2(ブレーキ進入相 BRAKE_FRONT>=0.3bar)で SUSP_REAR<=0mm の累積秒
        if R.get("BRAKE_FRONT") is not None and R.get("SUSP_REAR") is not None and dtg:
            pm = R["BRAKE_FRONT"] >= 0.3
            if int(pm.sum()) >= 1:
                cnt0 = int(np.sum((R["SUSP_REAR"] <= 0.0) & pm))
                ph12_rear0_s = round(cnt0 * dtg, 3)
        # ── 3フェーズ×F/R×方向 サス速度マトリクス (§44 / Tatsuki GO 2026-07-01) ──
        # 既存 vf/vr/fb_mask/ce_mask を再利用（既存列の値は不変）。mc_mask=MID_CORNER を追加。
        # dive=v>0(圧縮) / reb=-v(v<0,伸び)。avg=mean(n>=5) / peak=p95(n>=10)。既存 brk_f_dive は凍結のため除外。
        mc_mask = _zone_mask("MID_CORNER")
        def _dir_stat(v, mask, positive):
            if v is None or mask is None or int(mask.sum()) < NMIN_Z:
                return (None, None)
            vz = v[mask]; vz = vz[np.isfinite(vz)]
            s = vz[vz > 0] if positive else -vz[vz < 0]
            a = round(float(s.mean()), 1) if s.size >= NMIN_Z else None
            p = round(float(np.percentile(s, 95)), 1) if s.size >= PEAK_NMIN else None
            return (a, p)
        _psm = []
        for _pk, _mask in (("brk", fb_mask), ("apex", mc_mask), ("ce", ce_mask)):
            for _sk, _v in (("f", vf), ("r", vr)):
                for _pos in (True, False):           # True=dive, False=reb
                    if _pk == "brk" and _sk == "f" and _pos:
                        continue                     # brk_f_dive_spd_* は既存列を凍結
                    _a, _p = _dir_stat(_v, _mask, _pos)
                    _psm.append(_a); _psm.append(_p)
        phase_spd_matrix = tuple(_psm)               # PHASE_SPD_NEW_COLS 順の22値
        laps.append({"lap_no":i+1,"lap_time_s":round(lap_t,3),
                     "susf_mean":round(float(R["SUSP_FRONT"].mean()),2) if R.get("SUSP_FRONT") is not None else None,
                     "susf_max":round(float(R["SUSP_FRONT"].max()),2) if R.get("SUSP_FRONT") is not None else None,
                     "susr_mean":round(float(R["SUSP_REAR"].mean()),2) if R.get("SUSP_REAR") is not None else None,
                     "f_dive_spd":f_dive,"f_reb_spd":f_reb,"r_dive_spd":r_dive,"r_reb_spd":r_reb,
                     "rear_light_brk":rear_light_brk,
                     "brk_f_dive_spd_avg":brk_f_dive_avg,"brk_f_dive_spd_peak":brk_f_dive_peak,
                     "ce_r_spd_avg":ce_r_spd_avg,"ce_r_spd_peak":ce_r_spd_peak,"ph12_rear0_s":ph12_rear0_s,
                     "phase_spd_matrix":phase_spd_matrix,
                     "metrics":metrics})
    # HED Laptimes.Fastest を権威に、物理的に不可能な(=記録最速より速い) stray マーカー由来の
    # 偽ラップを除外。例: D2-#77-09(Jerez) は .LAP に 84.5s の部分ラップ(1周の81%)が混入し
    # best_lap を汚すが、2D 自身も HED Fastest(=104.6s) からは除外している。クリーンな outing は
    # 実測最速==HED Fastest のため fastest*0.97 フィルタは発火せず退行ゼロ。
    fastest_s = _parse_hed_time(parse_hed(mes_path, base).get("Fastest lap"))
    if fastest_s and fastest_s > 0:
        laps = [l for l in laps if l["lap_time_s"] >= fastest_s * 0.97]
    return {"laps":laps,"nlaps":len(laps)} if laps else None

# ── イベント単位の組み立て (検証表示) ──
NOISE = re.compile(r"ACCENSIONE|RD\d+-S\d+|-KAW_|^D0-", re.I)

# ── 2D outing 探索 (3レイアウト横断) ─────────────────────────────────────
# 各 outing は (mes_path: ディレクトリ, base: 内部ファイル stem) で表す。
# parse_ddd/parse_lap/read_channel/parse_hed は全て {mes_path}/{base}.<ext> を読むため、
# base さえ正しければディレクトリ命名規則に依存しない。
#   (A) nested : <event>/.../<base>.MES/<base>.DDD                 ← 現行と同一集合
#   (B) copia  : <event>/.../<folder>.MES - Copia/<base>.DDD       ← folder と base が別名
#   (C) loose  : <event>/<base>.DDD (.MES ラッパー無しの直置き)
# A は現行 rglob('*.MES') と完全一致させ既存挙動を一切変えない。B/C は A で取りこぼした
# 分だけ追加し、内部 base が A と重複するものは除外 (= 52_Copiy 等の複製を二重計上しない)。
def _outing_base_in(d):
    """ディレクトリ d 直下の .DDD ファイル stem を返す (無ければ None)。"""
    ddds = sorted(d.glob("*.DDD"))
    return ddds[0].name[:-4] if ddds else None

def discover_outings(event_dir):
    """(mes_path, base, tier) のリストを返す。tier ∈ {nested, copia, loose}。"""
    out=[]; tier1=set(); seen=set()
    # Tier A — 通常ネスト .MES ディレクトリ (現行 rglob 挙動を厳密維持)
    for d in sorted(event_dir.rglob("*.MES")):
        if not d.is_dir() or "_COPY" in str(d) or NOISE.search(d.name): continue
        base=d.name[:-4]; key=(str(d),base)
        if key in seen: continue
        seen.add(key); tier1.add(base); out.append((d,base,"nested"))
    # Tier B — リネームコピーされた .MES フォルダ (".MES - Copia" 等、末尾に空白+接尾辞)
    for d in sorted(event_dir.rglob("*.MES *")):
        if not d.is_dir() or "_COPY" in str(d) or NOISE.search(d.name): continue
        base=_outing_base_in(d)
        if not base or base in tier1: continue        # 複製 (nested と同一 outing) → 除外
        key=(str(d),base)
        if key in seen: continue
        seen.add(key); out.append((d,base,"copia"))
    # Tier C — .MES ラッパーを持たない loose なチャンネルファイル群
    for ddd in sorted(event_dir.rglob("*.DDD")):
        d=ddd.parent
        if d.name.endswith(".MES") or ".MES " in d.name: continue   # A/B で処理済み
        base=ddd.name[:-4]
        if base in tier1: continue
        if not (d/f"{base}.LAP").exists(): continue                 # LAP マーカー必須
        if "_COPY" in str(ddd) or NOISE.search(ddd.name): continue
        key=(str(d),base)
        if key in seen: continue
        seen.add(key); out.append((d,base,"loose"))
    return out

def _hed_meta(mes_path, base):
    """HED 自己申告の (circuit_canon, track_length, fastest_lap)。欠落は空文字。"""
    h=parse_hed(mes_path, base)
    circ=circuit_canon(h.get("Circuit") or h.get("GENERAL.Circuit") or "")
    tlen=h.get("Track Length") or h.get("GENERAL.Track Length") or ""
    fast=h.get("Laptimes.Fastest lap") or ""
    return circ, tlen, fast

def gated_outings(ev, ev_circ):
    """discover_outings + 物理整合ゲート → [(mes_path, base)]。
    新レイアウト (copia/loose) に限り、ファイル自身の HED サーキットがイベントの
    サーキットと矛盾する outing を除外する (誤ファイル混入対策, 2026-06-18)。
    nested は HED の Circuit 欄が高頻度で陳腐化 (例: ROUND10 Aragon の HED='Portimao')
    しているため一切ゲートしない = 既存イベントを退行させない。"""
    kept=[]
    for mes_path, base, tier in discover_outings(ev["dir"]):
        if tier in ("copia","loose"):
            hc, tlen, fast = _hed_meta(mes_path, base)
            if ev_circ and hc and hc != ev_circ:
                print(f"  [SKIP 2D 誤ファイル] {ev['dir'].name}/{base} (tier={tier}): "
                      f"HED circuit={hc} (len={tlen}, fastest={fast}) != event={ev_circ} → 取込除外")
                continue
        kept.append((mes_path, base))
    return kept

def list_session_outings(event_dir):
    """session -> [(mes_path, run_seq_from_filename)]"""
    out=defaultdict(list)
    for p in event_dir.rglob("*.MES"):
        if "_COPY" in str(p) or NOISE.search(p.name): continue
        base=p.name[:-4]
        rnd = ""  # session_canon_2d only needs TEST flag; derive from event later
        out[p].append(base)
    return out

def build_event(name, ev, orig):
    print(f"\n{'='*70}\nEVENT {name}  date={ev['date']} round={ev['round']} rider={ev['rider']}")
    circ = event_circuit(ev)
    ev["circuit"]=circ
    print(f"  circuit = {circ}   report={ev['report'].name if ev['report'] else None}")
    rnd = ev["round"]
    # gather outings per session (3レイアウト横断 + 誤ファイルゲート)
    sess_outings=defaultdict(list)
    for mes_path, base in gated_outings(ev, circ):
        sess_outings[session_canon_2d(base, rnd)].append((mes_path, base))
    for sess in sorted(sess_outings):
        n_orig = len(orig.get((ev["rider"],circ,sess),[]))
        outs=[]
        for p,base in sorted(sess_outings[sess], key=lambda pb: pb[1]):
            r=extract_outing(p, base)
            if r: outs.append((p, base, r["nlaps"], r))
        # ラップ数の多い上位 n_orig 本を採用 (n_orig=0なら全部参考表示)
        outs_sorted=sorted(outs, key=lambda x:-x[2])
        keep = outs_sorted[:n_orig] if n_orig>0 else outs_sorted
        keep = sorted(keep, key=lambda x:x[1])   # 時系列(base名)順
        print(f"  -- session {sess}: Original_runs={n_orig}, 2D_outings={len(outs)} -> 採用{len(keep)}本")
        for run_i,(p,base,nl,r) in enumerate(keep, start=1):
            best=min((l['lap_time_s'] for l in r['laps']), default=None)
            mc=[l['metrics']['MID_CORNER'] for l in r['laps']]
            fb=[l['metrics']['FULL_BRAKING'] for l in r['laps']]
            ce=[l['metrics']['CORNER_EXIT'] for l in r['laps']]
            def m(rows,k):
                vals=[x[k] for x in rows if x[k] is not None]; return round(sum(vals)/len(vals),1) if vals else None
            print(f"     RUN{run_i} <- {base}.MES  laps={nl} best={best}s")
            print(f"        MID  susF={m(mc,'susf')} susR={m(mc,'susr')} brk={m(mc,'brake')} thr={m(mc,'thr')} (avg n/lap={m(mc,'n')})")
            print(f"        FULLBRK susF={m(fb,'susf')} susR={m(fb,'susr')} brk={m(fb,'brake')} (avg n/lap={m(fb,'n')})")
            print(f"        EXIT susF={m(ce,'susf')} susR={m(ce,'susr')} thr={m(ce,'thr')} (avg n/lap={m(ce,'n')})")

# ── Report: per-run コメント + Weekend Summary + START/END setup ──
def _scanon_report(raw, run):
    """Reportのsession表記 -> canonical (session_canon_2d/orig と一致させる)"""
    # TEST: "TEST2 DAY1" -> "TEST2_DAY1" (build側 session_canon_2d/orig と同形式に。
    #   これが無いと "TEST2DAY1"≠"TEST2_DAY1" でTESTイベントのコメントが全て紐付かない)
    mt=re.match(r"^TEST\s*(\d)\s*DAY\s*(\d)", str(raw or "").upper())
    if mt: return f"TEST{mt.group(1)}_DAY{mt.group(2)}"
    u=re.sub(r"\s+","",str(raw or "").upper())
    if u.startswith("F"): return "FP"
    if u.startswith("Q"): return "QP"
    if u in ("WU1","WUP1") or u=="WU": return "WUP1"
    if u in ("WU2","WUP2"): return "WUP2"
    if u in ("R1","RACE1"): return "RACE1"
    if u in ("R2","RACE2"): return "RACE2"
    if u.startswith("SP"): return "SP"
    if u.startswith("D1") or u.startswith("L1"): return None
    return u or None

def parse_report(report):
    """returns dict: comments{(sess,run):text}, weekend_summary, start_setup, end_setup"""
    res={"comments":{}, "weekend":"", "start_setup":"", "end_setup":""}
    if not report or not report.exists(): return res
    import openpyxl
    wb=openpyxl.load_workbook(report, read_only=True, data_only=True)
    for sh in ("DAY1","DAY2"):
        if sh not in wb.sheetnames: continue
        ws=wb[sh]
        rows={i:r for i,r in enumerate(ws.iter_rows(min_row=1,max_row=60,values_only=True),start=1)}
        r7=rows.get(7); r48=rows.get(48)
        if not r7: continue
        maxc=len(r7)
        for c in range(4, maxc-1):  # session label cols (0-indexed)
            sraw=r7[c];
            if sraw is None: continue
            runv=r7[c+1] if c+1<maxc else None
            try: run=int(float(runv))
            except Exception: continue
            sess=_scanon_report(sraw, run)
            if not sess: continue
            cmt = r48[c] if (r48 and c<len(r48)) else None
            if cmt and str(cmt).strip():
                res["comments"][(sess,run)] = str(cmt).strip()
    if "REPORT" in wb.sheetnames:
        ws=wb["REPORT"]
        rows={i:r for i,r in enumerate(ws.iter_rows(min_row=1,max_row=40,values_only=True),start=1)}
        # weekend summary = 列7(0-indexed) r8以降の非空テキスト連結
        wk=[]
        for i in range(8,33):
            r=rows.get(i)
            if r and len(r)>7 and r[7] and len(str(r[7]).strip())>5:
                wk.append(str(r[7]).strip())
        res["weekend"]=" / ".join(dict.fromkeys(wk))
        # start/end setup: 列3(START), 列5(END) の行11-31を key=label でまとめる
        def snap(col):
            out=[]
            for i in range(11,32):
                r=rows.get(i)
                if not r: continue
                label=r[1] if len(r)>1 else None
                val=r[col] if len(r)>col else None
                if label and val not in (None,""):
                    out.append(f"{str(label).strip()}={str(val).strip()}")
            return "; ".join(out)
        res["start_setup"]=snap(3); res["end_setup"]=snap(5)
    return res

# ── 順位: 既存 ts24_unified.db の race_results を再利用 ──
def load_positions():
    import sqlite3
    src = ROOT/"02_DATABASE"/"ts24_unified.db"
    pos={}
    if not src.exists(): return pos
    try:
        c=sqlite3.connect(src)
        for rnd,circ,sess,rnum,p in c.execute(
          "SELECT round,circuit,session_type,rider_num,position FROM race_results WHERE rider_num IN (52,77)"):
            rider="JA52" if str(rnum)=="52" else "DA77"
            key=(rider, circuit_canon(circ), session_canon_orig(sess), _round_num(rnd))
            pos[key]=p
        c.close()
    except Exception as e:
        print("[WARN] race_results読込失敗:",e)
    return pos

def _round_num(r):
    m=re.search(r"(\d+)", str(r or "")); return m.group(1) if m else ""

# ── タグ: problem_library + キーワード簡易付与 ──
TAG_KEYWORDS = {
 "nervousness":["nervous"], "push_rear_exit":["r push","rear push","push f","r pushing f","push the f"],
 "line_loss_exit":["loose line","lose line","go outside","go always long","run wide","line at exit"],
 "no_turn_in":["no turn","not turn","difficult to make a turn","turn in"],
 "no_grip":["no grip","lack of grip","less grip","without grip","not feel grip","do not feel grip"],
 "chattering_brake":["chatter"], "front_dive":["dive","f dive"],
 "no_confidence_brake":["can not stop","cannot stop","difficult to stop","hard to stop","stop the bike"],
 "understeer_apex":["understeer","no turn"], "general_nervous":["really nervous","bike is nervous"],
}
def tag_comment(text):
    if not text: return []
    t=text.lower(); out=[]
    for tag,kws in TAG_KEYWORDS.items():
        if any(k in t for k in kws): out.append(tag)
    return out

# ── DB スキーマ ──
SETUP_COLS = [f for _,f in ORIG_FIELDS]
SCHEMA = f"""
DROP TABLE IF EXISTS runs; DROP TABLE IF EXISTS laps;
DROP TABLE IF EXISTS lap_metrics; DROP TABLE IF EXISTS events;
DROP TABLE IF EXISTS performance; DROP TABLE IF EXISTS tags; DROP TABLE IF EXISTS run_tags;
DROP TABLE IF EXISTS lap_suspension;
CREATE TABLE events(
  event_id TEXT PRIMARY KEY, date TEXT, round TEXT, rider TEXT, circuit TEXT, report_file TEXT,
  weekend_summary TEXT, start_setup TEXT, end_setup TEXT);
CREATE TABLE runs(
  run_id TEXT PRIMARY KEY, rider TEXT, circuit TEXT, round TEXT, session TEXT, run_no INTEGER,
  date TEXT, event_id TEXT, source TEXT, has_2d INTEGER, n_laps INTEGER, best_lap_s REAL, perf_best_lap REAL, comment TEXT,
  {', '.join(c+' TEXT' for c in SETUP_COLS)}, updated_at TEXT, created_at TEXT);
CREATE TABLE laps(
  lap_id TEXT PRIMARY KEY, run_id TEXT, lap_no INTEGER, lap_time_s REAL,
  susf_mean REAL, susf_max REAL, susr_mean REAL, mes_file TEXT,
  f_dive_spd REAL, f_reb_spd REAL, r_dive_spd REAL, r_reb_spd REAL, rear_light_brk REAL,
  is_outlap INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
CREATE TABLE lap_metrics(
  lap_id TEXT, area TEXT, n INTEGER, susf REAL, susr REAL, speed REAL, brake REAL, thr REAL,
  PRIMARY KEY(lap_id, area));
CREATE TABLE performance(
  run_id TEXT PRIMARY KEY, rider TEXT, circuit TEXT, round TEXT, session TEXT, run_no INTEGER,
  best_lap_s REAL, run_avg_lap_s REAL, session_position INTEGER, n_laps INTEGER);
CREATE TABLE tags(tag TEXT PRIMARY KEY, category TEXT, complaint_en TEXT);
CREATE TABLE run_tags(run_id TEXT, tag TEXT, source TEXT, PRIMARY KEY(run_id, tag));
CREATE TABLE lap_suspension(
  lap_id TEXT PRIMARY KEY, run_id TEXT, round TEXT, circuit TEXT, session TEXT,
  rider TEXT, run_no INTEGER, lap_no INTEGER, date TEXT, lap_time_s REAL, lap_time_fmt TEXT,
  apex_count INTEGER, apex_spd_avg REAL, apex_susF_avg REAL, apex_susR_avg REAL,
  wf_f_apex_n REAL, wf_r_apex_n REAL,
  brk_count INTEGER, brk_spd_avg REAL, brk_susF_avg REAL, brk_susR_avg REAL,
  wf_f_brk_n REAL, wf_r_brk_n REAL,
  fullbrk_count INTEGER, fullbrk_susF REAL, fullbrk_susR REAL,
  ce_count INTEGER, ce_spd_avg REAL, ce_susF_avg REAL, ce_susR_avg REAL, wf_f_ce_n REAL, wf_r_ce_n REAL,
  f_dive_spd REAL, f_reb_spd REAL, r_dive_spd REAL, r_reb_spd REAL, rear_light_brk REAL,
  brk_f_dive_spd_avg REAL, brk_f_dive_spd_peak REAL, ce_r_spd_avg REAL, ce_r_spd_peak REAL, ph12_rear0_s REAL,
  {', '.join(c+' REAL' for c in PHASE_SPD_NEW_COLS)},
  lap_susF_mean REAL, lap_susF_min REAL, lap_susF_max REAL, lap_susR_mean REAL,
  updated_at TEXT DEFAULT (datetime('now')));
CREATE INDEX idx_runs_evt ON runs(event_id);
CREATE INDEX idx_laps_run ON laps(run_id);
CREATE INDEX idx_lm_lap ON lap_metrics(lap_id);
CREATE INDEX idx_lapsus_run ON lap_suspension(run_id);
CREATE INDEX idx_lapsus_circuit ON lap_suspension(circuit);
CREATE INDEX idx_lapsus_rider ON lap_suspension(rider);
"""

def _sessions_in_2d(ev, rnd):
    """session -> [(mes_path, base)]。3レイアウト横断 + 新レイアウト誤ファイルゲート。
    ev["circuit"] を先に設定しておくこと (ゲートのイベント基準サーキットに使用)。"""
    out=defaultdict(list)
    for mes_path, base in gated_outings(ev, ev.get("circuit","")):
        out[session_canon_2d(base, rnd)].append((mes_path, base))
    return out

# ── lap_suspension 再生成 (新 run_id・3エリア lap_metrics から射影) ──
def _build_lap_suspension(conn, extra_by_lapid=None, matrix_by_lapid=None):
    """laps + lap_metrics + runs(バネレート) から lap_suspension を再構築。
    extra_by_lapid: lap_id -> (brk_f_dive_avg, brk_f_dive_peak, ce_r_avg, ce_r_peak, ph12_rear0_s)。
    matrix_by_lapid: lap_id -> PHASE_SPD_NEW_COLS 順の22値タプル (§44 追加のみ・既存列は不変)。
    ゾーン限定サス速度は laps に持たず lap_suspension のみへ射影 (Tatsuki 2026-06-20)。
    apex ← MID_CORNER, brk/fullbrk ← FULL_BRAKING (旧 brake進入エリアは廃止のため同値)。
    WheelForce Proxy: WF_F=susF×(F_SPR_L+F_SPR_R)/2, WF_R=susR×R_SPR×0.5 (LR=2.0→MR=0.5)。
    lap_susF_min は新 laps に列が無いため NULL。"""
    def _f(v):
        try: return float(str(v))
        except (TypeError, ValueError): return None
    def _fmt(s):
        if s is None: return None
        m=int(s//60); return f"{m}:{s-60*m:06.3f}"
    # lap_metrics 索引
    mt={}
    for lid,area,n,susf,susr,speed,brake,thr in conn.execute(
            "SELECT lap_id,area,n,susf,susr,speed,brake,thr FROM lap_metrics"):
        mt[(lid,area)]=(n,susf,susr,speed)
    rows=[]
    for (lid,run_id,lap_no,lt,smean,smax,srmean,rnd,circ,sess,rider,run_no,date,fl,fr,rr,
         fdive,freb,rdive,rreb,rlb) in conn.execute("""
            SELECT l.lap_id,l.run_id,l.lap_no,l.lap_time_s,l.susf_mean,l.susf_max,l.susr_mean,
                   r.round,r.circuit,r.session,r.rider,r.run_no,r.date,r.f_spr_l,r.f_spr_r,r.r_spr,
                   l.f_dive_spd,l.f_reb_spd,l.r_dive_spd,l.r_reb_spd,l.rear_light_brk
            FROM laps l JOIN runs r ON l.run_id=r.run_id"""):
        apex=mt.get((lid,"MID_CORNER")); fb=mt.get((lid,"FULL_BRAKING")); ce=mt.get((lid,"CORNER_EXIT"))
        fl_,fr_,rr_=_f(fl),_f(fr),_f(rr)
        fspr=((fl_ or 0)+(fr_ or 0))/2 if (fl_ or fr_) else None
        def wf_f(s): return round(s*fspr,1) if (s is not None and fspr) else None
        def wf_r(s): return round(s*rr_*0.5,1) if (s is not None and rr_) else None
        a_n,a_sf,a_sr,a_spd = apex if apex else (None,None,None,None)
        b_n,b_sf,b_sr,b_spd = fb   if fb   else (None,None,None,None)
        c_n,c_sf,c_sr,c_spd = ce   if ce   else (None,None,None,None)
        ex=(extra_by_lapid or {}).get(lid,(None,None,None,None,None))
        mx=(matrix_by_lapid or {}).get(lid,(None,)*len(PHASE_SPD_NEW_COLS))
        rows.append((lid,run_id,rnd,circ,sess,rider,run_no,lap_no,date,lt,_fmt(lt),
                     a_n,a_spd,a_sf,a_sr, wf_f(a_sf),wf_r(a_sr),
                     b_n,b_spd,b_sf,b_sr, wf_f(b_sf),wf_r(b_sr),
                     b_n,b_sf,b_sr,
                     c_n,c_spd,c_sf,c_sr, wf_f(c_sf),wf_r(c_sr),
                     fdive,freb,rdive,rreb,rlb,
                     ex[0],ex[1],ex[2],ex[3],ex[4],
                     smean,None,smax,srmean, *mx))
    # 22新列は末尾に付与（named INSERT なので物理列順に依存しない）。placeholder 数は列数から算出。
    _cols = ("lap_id,run_id,round,circuit,session,rider,run_no,lap_no,date,lap_time_s,lap_time_fmt,"
             "apex_count,apex_spd_avg,apex_susF_avg,apex_susR_avg,wf_f_apex_n,wf_r_apex_n,"
             "brk_count,brk_spd_avg,brk_susF_avg,brk_susR_avg,wf_f_brk_n,wf_r_brk_n,"
             "fullbrk_count,fullbrk_susF,fullbrk_susR,"
             "ce_count,ce_spd_avg,ce_susF_avg,ce_susR_avg,wf_f_ce_n,wf_r_ce_n,"
             "f_dive_spd,f_reb_spd,r_dive_spd,r_reb_spd,rear_light_brk,"
             "brk_f_dive_spd_avg,brk_f_dive_spd_peak,ce_r_spd_avg,ce_r_spd_peak,ph12_rear0_s,"
             "lap_susF_mean,lap_susF_min,lap_susF_max,lap_susR_mean,"
             + ",".join(PHASE_SPD_NEW_COLS))
    _nc = len(_cols.split(","))
    conn.executemany(
        f"INSERT INTO lap_suspension ({_cols}) VALUES ({','.join(['?']*_nc)})", rows)
    return len(rows)

# サーキット長(m)。is_outlap 物理下限ガード用。
TRACK_M = {"PHILLIPISLAND":4448, "PORTIMAO":4592, "ASSEN":4555, "BALATON":4115,
           "MOST":4212, "ARAGON":5077, "JEREZ":4423, "CREMONA":3768, "MISANO":4226,
           "MAGNYCOURS":4411, "ESTORIL":4182, "DONINGTON":4023}
MAX_AVG_KMH = 200.0   # これを超える平均速度を要する短ラップは物理的に不可能=stray

def _recompute_is_outlap(conn):
    """is_outlap を頑健に再計算 (監査2026-06-18: 相対min×1.15のみでは下限/上限の両汚染を許す)。
    順序: ①stray下限除去(物理) ②GRID/FORMATION除去 ③stray除外後のrun_minで相対×1.15
          ④単一ラップrunの上限絶対ガード(circuit基準×1.25超=非代表ラップ)。
    これにより MOST R2(56.2 stray反転) / PORTIMAO GRID03・SP-#77-03(単一遅ラップbest) を恒久解消。"""
    # サーキット基準(代表的速ラップ): 各circuitの「物理可能ラップ」の下位10%tile相当(=速い側)
    circ_laps = defaultdict(list)
    for circ, lt in conn.execute("""SELECT r.circuit, l.lap_time_s FROM laps l JOIN runs r USING(run_id)
                                     WHERE l.lap_time_s IS NOT NULL"""):
        tl = TRACK_M.get(circ)
        if tl and lt < tl / (MAX_AVG_KMH/3.6):   # 物理不可能ラップは基準算出から除外
            continue
        circ_laps[circ].append(lt)
    circ_ref = {}
    for circ, lst in circ_laps.items():
        lst.sort(); circ_ref[circ] = lst[max(0, len(lst)//10)] if lst else None  # P10 ≈ 代表速ラップ

    runs = conn.execute("SELECT run_id, circuit FROM runs").fetchall()
    for run_id, circ in runs:
        tl = TRACK_M.get(circ); floor = tl / (MAX_AVG_KMH/3.6) if tl else None
        laps = conn.execute("SELECT lap_id, lap_time_s, mes_file FROM laps WHERE run_id=?", (run_id,)).fetchall()
        if not laps: continue
        flags = {}        # lap_id -> is_outlap
        clean = []        # (lap_id, t)
        for lid, t, mes in laps:
            if t is None: flags[lid] = 1; continue
            is_grid = bool(mes) and ("GRID" in str(mes).upper() or "FORMATION" in str(mes).upper())
            is_stray = floor is not None and t < floor          # ①物理下限(>200km/h)
            if is_stray or is_grid:                              # ②GRID/FORMATION
                flags[lid] = 1
            else:
                clean.append((lid, t))
        if clean:
            mn = min(t for _, t in clean)
            ref = circ_ref.get(circ)
            for lid, t in clean:
                out = 1 if t > mn * 1.15 else 0                  # ③相対(stray除外後)
                if len(clean) == 1 and ref and t > ref * 1.25:    # ④単一ラップの上限絶対ガード
                    out = 1
                flags[lid] = out
        for lid, fl in flags.items():
            conn.execute("UPDATE laps SET is_outlap=? WHERE lap_id=?", (fl, lid))

def build_all(out_db=None):
    import sqlite3
    db_path = Path(out_db) if out_db else OUT_DB
    orig = load_original()
    evs  = discover_events()
    # 各イベントの circuit と 2Dセッション
    for name,ev in evs.items():
        ev["circuit"]=event_circuit(ev)
        ev["sessions"]=_sessions_in_2d(ev, ev["round"])
    # (rider,circuit,session) -> [event names date順] (2Dにそのsessionがある場合)
    rcs_events=defaultdict(list)
    for name,ev in sorted(evs.items(), key=lambda kv: kv[1]["date"]):
        for s in ev["sessions"]:
            rcs_events[(ev["rider"],ev["circuit"],s)].append(name)
    # Original プール (consume用) : (rider,circ,sess) -> [run dict] (sheet順)
    pool={k:list(v) for k,v in orig.items()}

    conn=sqlite3.connect(db_path)
    for stmt in SCHEMA.strip().split(";"):
        if stmt.strip(): conn.execute(stmt)
    runs_rows=[]; laps_rows=[]; lm_rows=[]; ev_rows=[]
    extra_by_lapid={}   # lap_id -> ゾーン限定サス速度5値 (lap_suspension へ射影, laps は不変)
    matrix_by_lapid={}  # lap_id -> PHASE_SPD_NEW_COLS 順の22値 (§44 追加のみ, laps は不変)
    stat=defaultdict(int)

    for name,ev in sorted(evs.items(), key=lambda kv: kv[1]["date"]):
        rider,circ,rnd,date=ev["rider"],ev["circuit"],ev["round"],ev["date"]
        rep=parse_report(ev["report"])
        ev_rows.append((name,date,rnd,rider,circ, ev["report"].name if ev["report"] else None,
                        rep["weekend"], rep["start_setup"], rep["end_setup"]))
        for sess,paths in ev["sessions"].items():
            kkey=(rider,circ,sess)
            M=len(orig.get(kkey,[])); E=len(rcs_events.get(kkey,[]))
            per_event = (max(1, round(M/E)) if E>0 else 0) if M>0 else 0
            # 2D outings 抽出 (paths = [(mes_path, base)])
            outs=[]
            for p,base in sorted(paths, key=lambda pb: pb[1]):
                r=extract_outing(p, base)
                if r: outs.append((p,base,r))
            outs_sorted=sorted(outs,key=lambda x:-x[2]["nlaps"])
            if per_event>0:
                keep=sorted(outs_sorted[:per_event], key=lambda x:x[1])
                n_runs=per_event
            else:
                keep=sorted(outs_sorted, key=lambda x:x[1])  # 2D_ONLY: 全keep
                n_runs=len(keep)
            for i in range(n_runs):
                setup = pool.get(kkey,[None])
                srow = pool[kkey].pop(0) if pool.get(kkey) else None
                run_no=i+1
                run_id=f"{date}_{rnd}_{circ}_{sess}_{rider}_R{run_no}"
                outing = keep[i] if i < len(keep) else None   # (mes_path, base, result)
                has2d=1 if outing else 0
                source = "ORIGINAL+2D" if (srow and outing) else ("ORIGINAL" if srow else "2D_ONLY")
                setup_vals=[ (str(srow[c]) if srow and srow.get(c) is not None else None) for c in SETUP_COLS ]
                nlaps=outing[2]["nlaps"] if outing else 0
                best=min((l["lap_time_s"] for l in outing[2]["laps"]), default=None) if outing else None
                cmt=rep["comments"].get((sess,run_no))
                runs_rows.append((run_id,rider,circ,rnd,sess,run_no,date,name,source,has2d,nlaps,best,best,cmt,*setup_vals))
                stat[source]+=1
                if outing:
                    mes_file=f"{outing[1]}.MES"   # base 由来 (nested では従来の {dir名} と一致)
                    for l in outing[2]["laps"]:
                        lap_id=f"{run_id}_L{l['lap_no']}"
                        laps_rows.append((lap_id,run_id,l["lap_no"],l["lap_time_s"],
                                          l["susf_mean"],l["susf_max"],l["susr_mean"],mes_file,
                                          l["f_dive_spd"],l["f_reb_spd"],l["r_dive_spd"],l["r_reb_spd"],l["rear_light_brk"]))
                        extra_by_lapid[lap_id]=(l["brk_f_dive_spd_avg"],l["brk_f_dive_spd_peak"],
                                                l["ce_r_spd_avg"],l["ce_r_spd_peak"],l["ph12_rear0_s"])
                        matrix_by_lapid[lap_id]=l["phase_spd_matrix"]
                        for area,mt in l["metrics"].items():
                            lm_rows.append((lap_id,area,mt["n"],mt["susf"],mt["susr"],mt["speed"],mt["brake"],mt["thr"]))
                        stat["laps"]+=1

    # 残ったOriginal (どのイベントにも消費されず) = setup-only
    for kkey,remain in pool.items():
        rider,circ,sess=kkey
        for srow in remain:
            run_no=srow["run"]
            run_id=f"NA_{circ}_{sess}_{rider}_R{run_no}"
            setup_vals=[ (str(srow[c]) if srow.get(c) is not None else None) for c in SETUP_COLS ]
            runs_rows.append((run_id,rider,circ,"",sess,run_no,None,None,"ORIGINAL_NO2D",0,0,None,None,None,*setup_vals))
            stat["ORIGINAL_NO2D"]+=1

    # run_id 重複検出 + dedup
    from collections import Counter as _C2
    idc=_C2(r[0] for r in runs_rows)
    dups={k:c for k,c in idc.items() if c>1}
    if dups:
        print(f"\n[WARN] run_id重複 {len(dups)}種:")
        for k,c in list(dups.items())[:20]: print(f"   x{c}  {k}")
    seen=set(); deduped=[]
    for r in runs_rows:
        if r[0] in seen: continue
        seen.add(r[0]); deduped.append(r)
    runs_rows=deduped

    # 列を明示指定(末尾の updated_at/created_at/is_outlap は後段パスで設定するため省略)
    RUN_COLS = ["run_id","rider","circuit","round","session","run_no","date","event_id",
                "source","has_2d","n_laps","best_lap_s","perf_best_lap","comment"] + SETUP_COLS
    LAP_COLS = ["lap_id","run_id","lap_no","lap_time_s","susf_mean","susf_max","susr_mean","mes_file",
                "f_dive_spd","f_reb_spd","r_dive_spd","r_reb_spd","rear_light_brk"]
    conn.executemany(f"INSERT INTO events VALUES ({','.join('?'*9)})", ev_rows)
    conn.executemany(f"INSERT INTO runs ({','.join(RUN_COLS)}) VALUES ({','.join('?'*len(RUN_COLS))})", runs_rows)
    conn.executemany(f"INSERT INTO laps ({','.join(LAP_COLS)}) VALUES ({','.join('?'*len(LAP_COLS))})", laps_rows)
    conn.executemany(f"INSERT INTO lap_metrics VALUES ({','.join('?'*8)})", lm_rows)
    conn.commit()

    # ── is_outlap を頑健に再計算 (stray下限/GRID/相対/単一ラップ上限ガード) ──
    _recompute_is_outlap(conn); conn.commit()

    # ── performance (best/avg lap + session_position) ── best/avg は is_outlap=0 のみで集計 ──
    positions=load_positions()
    perf=[]; tagrows=[]
    for r in conn.execute("SELECT run_id,rider,circuit,round,session,run_no,n_laps,comment FROM runs").fetchall():
        run_id,rider,circ,rnd,sess,run_no,nlaps,comment=r
        times=[t[0] for t in conn.execute(
            "SELECT lap_time_s FROM laps WHERE run_id=? AND is_outlap=0 AND lap_time_s IS NOT NULL ORDER BY lap_no",(run_id,)).fetchall()]
        best=avg=None; nfly=0
        if times:
            mn=min(times)
            fly=[t for t in times if t <= mn*1.07]   # best+7%以内=レースペース
            if not fly: fly=times
            best=round(min(fly),3); avg=round(sum(fly)/len(fly),3); nfly=len(fly)
            conn.execute("UPDATE runs SET best_lap_s=?, perf_best_lap=? WHERE run_id=?",(best,best,run_id))
        else:
            conn.execute("UPDATE runs SET best_lap_s=NULL, perf_best_lap=NULL WHERE run_id=?",(run_id,))
        spos=positions.get((rider,circ,sess,_round_num(rnd)))
        perf.append((run_id,rider,circ,rnd,sess,run_no,best,avg,spos,nfly))
        for tg in tag_comment(comment):
            tagrows.append((run_id,tg,"comment_auto"))
    conn.executemany(f"INSERT OR REPLACE INTO performance VALUES ({','.join('?'*10)})", perf)
    # tags master (problem_library 再利用)
    try:
        import sqlite3 as _sq
        src=_sq.connect(ROOT/"02_DATABASE"/"ts24_unified.db")
        seen=set()
        for row in src.execute("SELECT tags, phase_name, complaint_en FROM problem_library"):
            for tg in str(row[0] or "").split(","):
                tg=tg.strip()
                if tg and tg not in seen:
                    seen.add(tg); conn.execute("INSERT OR REPLACE INTO tags VALUES (?,?,?)",(tg,row[1],row[2]))
        src.close()
    except Exception as e: print("[WARN] tags master:",e)
    conn.executemany("INSERT OR REPLACE INTO run_tags VALUES (?,?,?)", tagrows)
    conn.commit()

    # ── タイムスタンプ恒久化 (created_at/updated_at) ──
    conn.execute("UPDATE runs SET created_at=datetime('now'), updated_at=datetime('now')")
    conn.execute("UPDATE laps SET created_at=datetime('now'), updated_at=datetime('now')")

    # ── lap_suspension 再生成 (新 run_id・3エリア lap_metrics から射影) ──
    #   apex      ← MID_CORNER, brk/fullbrk ← FULL_BRAKING(旧 brake進入エリアは廃止のため同一)
    #   WheelForce Proxy: WF_F = susF×(F_SPR_L+F_SPR_R)/2,  WF_R = susR×R_SPR×0.5 (LR=2.0→MR=0.5)
    #   lap_susF_min は新 laps に列が無いため NULL。
    n_lapsus = _build_lap_suspension(conn, extra_by_lapid, matrix_by_lapid)
    conn.commit()

    print(f"\n===== {db_path.name} 構築完了 =====")
    print(f"events={len(ev_rows)}  runs={len(runs_rows)}  laps={len(laps_rows)}  lap_metrics={len(lm_rows)}  lap_suspension={n_lapsus}")
    print(f"performance={len(perf)}  run_tags={len(tagrows)}  comment付きruns={sum(1 for r in runs_rows if r[13])}")
    print("runs内訳:", dict(stat))
    print(f"順位付きperf:", conn.execute("SELECT count(*) FROM performance WHERE session_position IS NOT NULL").fetchone()[0])
    # 検証: 平均
    for area in AREAS:
        row=conn.execute("SELECT round(avg(susf),1),round(avg(susr),1),round(avg(brake),2),round(avg(thr),1),count(*) FROM lap_metrics WHERE area=? AND n>0",(area,)).fetchone()
        print(f"  {area:13} susF={row[0]} susR={row[1]} brake={row[2]} thr={row[3]} (laps={row[4]})")
    # ── 受入ゲート(SPEC §8拡張・監査2026-06-18): 2D best vs PDF best のΔ>1.5s を検出 ──
    # is_outlap/best汚染(MOST stray, PORTIMAOグリッド等)の回帰検出。0件が合格。
    # race_results は unified.db 側(cutoverで保持)にあるため別接続で参照。
    try:
        import sqlite3 as _sq3
        src=_sq3.connect(ROOT/"02_DATABASE"/"ts24_unified.db")
        pdfbest={}
        for rnd,sess,rnum,bl in src.execute(
                "SELECT round,session_type,rider_num,best_lap_s FROM race_results WHERE best_lap_s IS NOT NULL"):
            rider={"52":"JA52","77":"DA77"}.get(str(rnum))
            if rider: pdfbest[(rnd,sess,rider)]=bl
        src.close()
        # PDFはセッション最速(全run横断)。2D側も (round,session,rider) 単位の最速に集約して比較する
        # (per-run比較だと最速runでないだけのrunを誤検出するため)。
        sess2d={}
        for best,rnd,sess,rider in conn.execute(
                "SELECT best_lap_s,round,session,rider FROM performance WHERE best_lap_s IS NOT NULL"):
            k=(rnd,sess,rider)
            if k not in sess2d or best<sess2d[k]: sess2d[k]=best
        bad=[]
        for k,best in sess2d.items():
            pb=pdfbest.get(k)
            if pb is not None and abs(best-pb)>1.5: bad.append((k,best,pb,round(abs(best-pb),3)))
        bad.sort(key=lambda x:-x[3])
        print(f"\n[受入ゲート] 2D(session最速) vs PDF best Δ>1.5s: {len(bad)}件 {'✅合格' if not bad else '⚠要確認'}")
        for r in bad[:10]: print(f"   ⚠ {r[0]} 2D={r[1]} PDF={r[2]} Δ={r[3]}")
    except Exception as e:
        print(f"\n[受入ゲート] race_results参照不可(cutover後に検証):{e}")
    conn.close()

def main():
    args=sys.argv[1:]
    if "--all" in args:
        out = args[args.index("--out")+1] if "--out" in args else None
        build_all(out); return
    orig=load_original()
    evs=discover_events()
    if "--event" in args:
        name=args[args.index("--event")+1]
        if name not in evs: print(f"event not found: {name}\n候補:{list(evs)}"); return
        build_event(name, evs[name], orig)
    else:
        print("events:", list(evs))

if __name__=="__main__":
    main()
